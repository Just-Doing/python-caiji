from urllib.request import urlopen
from bs4 import BeautifulSoup
import http.client
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json

http.client._MAXHEADERS = 1000

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head.strip() in info:
			workSheet.cell(rowIndex, cellIndex).value = str(info[head.strip()]).strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1



fileName="D:\\p2.txt"
excelFileName="D:\\products1.xlsx"
wb = Workbook()
workSheet = wb.active
with open(fileName,'r') as file_to_read:
	index = 1
	while True:
		lines = file_to_read.readline()
		if not lines:
			break
			pass
		productInfo={}
		IUPAC_Name = ""
		InChI = ""
		InChIKey = ""
		Formula_Name = ""
		CASinfoStr = ""
		OtherCASinfoStr = ""
		NSCNumberInfoStr = ""
		ICSCNumberInfoStr = ""
		EC_NumberInfoStr = ""
		RTECSNumberInfoStr = ""
		UNNumberInfoStr = ""
		UNIIInfoStr = ""
		MeSHEntryTermsStr = ""
		DepositorSuppliedSynonymStr = ""
		PhysicalDescriptionInfoStr = ""
		ColorFormInfoStr = ""
		OdorInfoStr = ""
		MeltingPointInfoStr = ""
		SolubilityInfoStr = ""
		DensityInfoStr = ""
		VaporPressureInfoStr = ""
		OctanolWaterPartitionCoefficientInfoStr = ""
		StabilityShelfLifeInfoStr = ""
		DecompositionInfoStr = ""
		CorrosivityInfoStr = ""
		pHInfoStr = ""
		SurfaceTensionInfoStr = ""
		OtherExperimentalPropertiesInfoStr = ""
		WikipediaInfoStr = ""
		CanonicalSMILES = ""
		getCidUrl = "https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/" + lines.strip() + "/cids/JSON"
		cidRes = requests.get(getCidUrl)
		cidData = cidRes.json().get("IdentifierList")
		if cidData:
			cids = cidData["CID"]
			cid=cids[0]
			productDataInfoUrl = "https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound/"+str(cid)+"/JSON/"
			productJsonInfoData = requests.get(productDataInfoUrl)
			productJsonInfo = productJsonInfoData.json().get("Record")
			if productJsonInfo:
				sections = productJsonInfo["Section"]
				# Names and Identifiers
				nameAndIdenti = list(filter(lambda o: o["TOCHeading"]=="Names and Identifiers", sections))
				Computed_Descriptors = list(filter(lambda o: o["TOCHeading"]=="Computed Descriptors", nameAndIdenti[0]["Section"]))
				IUPAC_NameInfoList = list(filter(lambda o: o["TOCHeading"]=="IUPAC Name", Computed_Descriptors[0]["Section"]))
				if len(IUPAC_NameInfoList):
					IUPAC_NameInfo = IUPAC_NameInfoList[0]["Information"]
					for o in IUPAC_NameInfo: IUPAC_Name += o["Value"]["StringWithMarkup"][0]["String"]+";"
				
				InChIInfoList = list(filter(lambda o: o["TOCHeading"]=="InChI", Computed_Descriptors[0]["Section"]))
				if len(InChIInfoList):
					InChIInfo = InChIInfoList[0]["Information"]
					for o in InChIInfo: InChI += o["Value"]["StringWithMarkup"][0]["String"]+";"

				InChIKeyInfoList = list(filter(lambda o: o["TOCHeading"]=="InChI Key", Computed_Descriptors[0]["Section"]))
				if len(InChIKeyInfoList):
					InChIKeyInfo = InChIKeyInfoList[0]["Information"]
					for o in InChIKeyInfo: InChIKey += o["Value"]["StringWithMarkup"][0]["String"]+";"

				CanonicalSMILESInfoList = list(filter(lambda o: o["TOCHeading"]=="Canonical SMILES", Computed_Descriptors[0]["Section"]))
				if len(CanonicalSMILESInfoList):
					CanonicalSMILESInfo = CanonicalSMILESInfoList[0]["Information"]
					for o in CanonicalSMILESInfo: CanonicalSMILES += o["Value"]["StringWithMarkup"][0]["String"]+";"
					
				Molecular_FormulaList = list(filter(lambda o: o["TOCHeading"]=="Molecular Formula", nameAndIdenti[0]["Section"]))
				if len(Molecular_FormulaList):
					Molecular_Formula = Molecular_FormulaList[0]["Information"]
					for o in Molecular_Formula: Formula_Name += o["Value"]["StringWithMarkup"][0]["String"]+";"

				OtherIdentifiers = list(filter(lambda o: o["TOCHeading"]=="Other Identifiers", nameAndIdenti[0]["Section"]))
				
				if len(OtherIdentifiers):
					CAS_NameInfoList = list(filter(lambda o: o["TOCHeading"]=="CAS", OtherIdentifiers[0]["Section"]))
					if len(CAS_NameInfoList):
						CAS_NameInfo = CAS_NameInfoList[0]["Information"]
						CASinfoList = list(filter(lambda o: o.get("Name") and o["Name"]=="CAS", CAS_NameInfo))
						OtherCASinfoList = list(filter(lambda o: o.get("Name") and o["Name"]=="Other CAS", CAS_NameInfo))
						if len(CASinfoList):
							CASinfo = CASinfoList[0]["Value"]["StringWithMarkup"]
							for CAS in CASinfo: CASinfoStr += CAS["String"]+";"
						if len(OtherCASinfoList):
							OtherCASinfo=OtherCASinfoList[0]["Value"]["StringWithMarkup"]
							for OtherCAS in OtherCASinfo: OtherCASinfoStr += OtherCAS["String"]+";"
					EC_NumberInfo = list(filter(lambda o: o["TOCHeading"]=="EC Number", OtherIdentifiers[0]["Section"]))
					if len(EC_NumberInfo):
						for EC_number in EC_NumberInfo[0]["Information"][0]["Value"]["StringWithMarkup"]: EC_NumberInfoStr+=EC_number["String"]

					ICSCNumberInfo = list(filter(lambda o: o["TOCHeading"]=="ICSC Number", OtherIdentifiers[0]["Section"]))
					if len(ICSCNumberInfo):
						for ICSCNumber in ICSCNumberInfo[0]["Information"][0]["Value"]["StringWithMarkup"]: ICSCNumberInfoStr+=ICSCNumber["String"]
						
					NSCNumberInfo = list(filter(lambda o: o["TOCHeading"]=="NSC Number", OtherIdentifiers[0]["Section"]))
					if len(NSCNumberInfo):
						for NSCNumber in NSCNumberInfo[0]["Information"][0]["Value"]["StringWithMarkup"]: NSCNumberInfoStr+=NSCNumber["String"]
						
					RTECSNumberInfo = list(filter(lambda o: o["TOCHeading"]=="RTECS Number", OtherIdentifiers[0]["Section"]))
					if len(RTECSNumberInfo):
						for RTECSNumber in RTECSNumberInfo[0]["Information"][0]["Value"]["StringWithMarkup"]: RTECSNumberInfoStr+=RTECSNumber["String"]
						
					UNNumberInfo = list(filter(lambda o: o["TOCHeading"]=="UN Number", OtherIdentifiers[0]["Section"]))
					if len(UNNumberInfo):
						for UNNumber in UNNumberInfo[0]["Information"][0]["Value"]["StringWithMarkup"]: UNNumberInfoStr+=UNNumber["String"]
						
					UNIIInfo = list(filter(lambda o: o["TOCHeading"]=="UNII", OtherIdentifiers[0]["Section"]))
					if len(UNIIInfo):
						for UNII in UNIIInfo[0]["Information"][0]["Value"]["StringWithMarkup"]: UNIIInfoStr+=UNII["String"]
					
					WikipediaInfo = list(filter(lambda o: o["TOCHeading"]=="Wikipedia", OtherIdentifiers[0]["Section"]))
					if len(WikipediaInfo):
						for Wikipedia in WikipediaInfo[0]["Information"][0]["Value"]["StringWithMarkup"]: WikipediaInfoStr+=Wikipedia["String"]
				
				Synonyms = list(filter(lambda o: o["TOCHeading"]=="Synonyms", nameAndIdenti[0]["Section"]))
				if len(Synonyms):
					MeSHEntryTerms=list(filter(lambda o: o["TOCHeading"]=="MeSH Entry Terms", Synonyms[0]["Section"]))
					if len(MeSHEntryTerms):
						for MeSHEntryTerm in MeSHEntryTerms[0]["Information"][0]["Value"]["StringWithMarkup"]: MeSHEntryTermsStr+= MeSHEntryTerm["String"]+";"
						
					DepositorSuppliedSynonyms=list(filter(lambda o: o["TOCHeading"]=="Depositor-Supplied Synonyms", Synonyms[0]["Section"]))
					if len(DepositorSuppliedSynonyms):
						for DepositorSuppliedSynonym in DepositorSuppliedSynonyms[0]["Information"][0]["Value"]["StringWithMarkup"]: DepositorSuppliedSynonymStr+= DepositorSuppliedSynonym["String"]+";"
				#	Chemical and Physical Properties	
				ChemicalandPhysicalProperties = list(filter(lambda o: o["TOCHeading"]=="Chemical and Physical Properties", sections))
				ComputedProperties = list(filter(lambda o: o["TOCHeading"]=="Computed Properties", ChemicalandPhysicalProperties[0]["Section"]))
				for key in ComputedProperties[0]["Section"]: 
					pName = key["TOCHeading"]
					pValue = ""
					unit = key["Information"][0]["Value"].get("Unit")
					Numbers = key["Information"][0]["Value"].get("Number")
					Strings = key["Information"][0]["Value"].get("StringWithMarkup")
					if Numbers:
						for number in Numbers: pValue += str(number) + ";"
					if Strings:
						for strValue in Strings: pValue += strValue["String"] + ";"
					if len(pValue): pValue = pValue[0:len(pValue)-1]
					pValue= pValue + (unit if(unit) else "")
					productInfo[pName]=pValue
				
				ExperimentalProperties = list(filter(lambda o: o["TOCHeading"]=="Experimental Properties", ChemicalandPhysicalProperties[0]["Section"]))
				if len(ExperimentalProperties):
					PhysicalDescriptionList = list(filter(lambda o: o["TOCHeading"]=="Physical Description", ExperimentalProperties[0]["Section"]))
					if len(PhysicalDescriptionList):
						PhysicalDescriptionInfo = PhysicalDescriptionList[0]["Information"]
						for PhysicalDescription in PhysicalDescriptionInfo: PhysicalDescriptionInfoStr+=PhysicalDescription["Value"]["StringWithMarkup"][0]["String"]+";"
						
					ColorFormList = list(filter(lambda o: o["TOCHeading"]=="Color/Form", ExperimentalProperties[0]["Section"]))
					if len(ColorFormList):
						ColorFormInfo = ColorFormList[0]["Information"]
						for ColorForm in ColorFormInfo: ColorFormInfoStr+=ColorForm["Value"]["StringWithMarkup"][0]["String"]+";"
				
					OdorList = list(filter(lambda o: o["TOCHeading"]=="Odor", ExperimentalProperties[0]["Section"]))
					if len(OdorList):
						OdorInfo = OdorList[0]["Information"]
						for Odor in OdorInfo: OdorInfoStr += Odor["Value"]["StringWithMarkup"][0]["String"]+";"
						
					MeltingPointList = list(filter(lambda o: o["TOCHeading"]=="Melting Point", ExperimentalProperties[0]["Section"]))
					if len(MeltingPointList):
						MeltingPointInfo = MeltingPointList[0]["Information"]
						for MeltingPoint in MeltingPointInfo:
							StringWithMarkupValue = MeltingPoint["Value"].get("StringWithMarkup")
							if StringWithMarkupValue: MeltingPointInfoStr += StringWithMarkupValue[0]["String"]+";"
							
						
					SolubilityList = list(filter(lambda o: o["TOCHeading"]=="Solubility", ExperimentalProperties[0]["Section"]))
					if len(SolubilityList):
						SolubilityInfo = SolubilityList[0]["Information"]
						for Solubility in SolubilityInfo:
							StringWithMarkupValue = Solubility["Value"].get("StringWithMarkup")
							if StringWithMarkupValue: SolubilityInfoStr += StringWithMarkupValue[0]["String"]+";"
						
					DensityList = list(filter(lambda o: o["TOCHeading"]=="Density", ExperimentalProperties[0]["Section"]))
					if len(DensityList):
						DensityInfo = DensityList[0]["Information"]
						for Density in DensityInfo: DensityInfoStr += Density["Value"]["StringWithMarkup"][0]["String"]+";"
						
					VaporPressureList = list(filter(lambda o: o["TOCHeading"]=="Vapor Pressure", ExperimentalProperties[0]["Section"]))
					if len(VaporPressureList):
						VaporPressureInfo = VaporPressureList[0]["Information"]
						for VaporPressure in VaporPressureInfo: VaporPressureInfoStr += VaporPressure["Value"]["StringWithMarkup"][0]["String"]+";"
						
					OctanolWaterPartitionCoefficientList = list(filter(lambda o: o["TOCHeading"]=="Octanol/Water Partition Coefficient", ExperimentalProperties[0]["Section"]))
					if len(OctanolWaterPartitionCoefficientList):
						OctanolWaterPartitionCoefficientInfo = OctanolWaterPartitionCoefficientList[0]["Information"]
						for OctanolWaterPartitionCoefficient in OctanolWaterPartitionCoefficientInfo:
							StringWithMarkupValue = OctanolWaterPartitionCoefficient["Value"].get("StringWithMarkup")
							if StringWithMarkupValue: OctanolWaterPartitionCoefficientInfoStr += StringWithMarkupValue[0]["String"]+";"
						
					StabilityShelfLifeList = list(filter(lambda o: o["TOCHeading"]=="Stability/Shelf Life", ExperimentalProperties[0]["Section"]))
					if len(StabilityShelfLifeList):
						StabilityShelfLifeInfo = StabilityShelfLifeList[0]["Information"]
						for StabilityShelfLife in StabilityShelfLifeInfo: StabilityShelfLifeInfoStr += StabilityShelfLife["Value"]["StringWithMarkup"][0]["String"]+";"
						
					DecompositionList = list(filter(lambda o: o["TOCHeading"]=="Decomposition", ExperimentalProperties[0]["Section"]))
					if len(DecompositionList):
						DecompositionInfo = DecompositionList[0]["Information"]
						for Decomposition in DecompositionInfo: DecompositionInfoStr += Decomposition["Value"]["StringWithMarkup"][0]["String"]+";"
						
					CorrosivityList = list(filter(lambda o: o["TOCHeading"]=="Corrosivity", ExperimentalProperties[0]["Section"]))
					if len(CorrosivityList):
						CorrosivityInfo = CorrosivityList[0]["Information"]
						for Corrosivity in CorrosivityInfo: CorrosivityInfoStr += Corrosivity["Value"]["StringWithMarkup"][0]["String"]+";"
						
					pHList = list(filter(lambda o: o["TOCHeading"]=="pH", ExperimentalProperties[0]["Section"]))
					if len(pHList):
						pHInfo = pHList[0]["Information"]
						for pH in pHInfo: pHInfoStr += pH["Value"]["StringWithMarkup"][0]["String"]+";"
						
					SurfaceTensionList = list(filter(lambda o: o["TOCHeading"]=="Surface Tension", ExperimentalProperties[0]["Section"]))
					if len(SurfaceTensionList):
						SurfaceTensionInfo = SurfaceTensionList[0]["Information"]
						for SurfaceTension in SurfaceTensionInfo: SurfaceTensionInfoStr += SurfaceTension["Value"]["StringWithMarkup"][0]["String"]+";"
						
					OtherExperimentalPropertiesList = list(filter(lambda o: o["TOCHeading"]=="Other Experimental Properties", ExperimentalProperties[0]["Section"]))
					if len(OtherExperimentalPropertiesList):
						OtherExperimentalPropertiesInfo = OtherExperimentalPropertiesList[0]["Information"]
						for OtherExperimentalProperties in OtherExperimentalPropertiesInfo:
							StringWithMarkupValue = OtherExperimentalProperties["Value"].get("StringWithMarkup")
							if StringWithMarkupValue: OtherExperimentalPropertiesInfoStr += StringWithMarkupValue[0]["String"]+";"
			productInfo["matched"]=1
		else: productInfo["matched"]=0
		headers=[
			'Keyword','matched','IUPAC Name','InChI','InChI Key','Canonical+F1 SMILES','Molecular Formula','CAS','EC Number','ICSC Number','NSC Number',
			'RTECS Number','UN Number','UNII','Wikipedia','MeSH Entry Terms','Depositor-Supplied Synonyms','Computed Properties','Property Name','Molecular Weight',
			'Hydrogen Bond Donor Count','Hydrogen Bond Acceptor Count','Rotatable Bond Count','Exact Mass','Monoisotopic Mass','Topological Polar Surface Area',
			'Heavy Atom Count','Formal Charge','Complexity','Isotope Atom Count','Defined Atom Stereocenter Count','Undefined Atom Stereocenter Count',
			'Defined Bond Stereocenter Count','Undefined Bond Stereocenter Count','Covalently-Bonded Unit Count','Compound Is Canonicalized','Physical Description',
			'Color/Form','Odor','Melting Point','Solubility','Density','Vapor Pressure','Octanol/Water Partition Coefficient','Stability/Shelf Life','Decomposition',
			'Corrosivity','pH','Surface Tension','Other Experimental Properties'
		];
		productInfo['Keyword']=lines
		productInfo['IUPAC Name']=IUPAC_Name
		productInfo['InChI']=InChI
		productInfo['InChI Key']=InChIKey
		productInfo['Canonical+F1 SMILES']=CanonicalSMILES
		productInfo['Molecular Formula']=Formula_Name
		productInfo['CAS']=CASinfoStr
		productInfo['EC Number']=EC_NumberInfoStr
		productInfo['ICSC Number']=ICSCNumberInfoStr
		productInfo['NSC Number']=NSCNumberInfoStr
		productInfo['RTECS Number']=RTECSNumberInfoStr
		productInfo['UN Number']=UNNumberInfoStr
		productInfo['UNII']=UNIIInfoStr
		productInfo['Wikipedia']=WikipediaInfoStr
		productInfo['MeSH Entry Terms']=MeSHEntryTermsStr
		productInfo['Depositor-Supplied Synonyms']=DepositorSuppliedSynonymStr
		productInfo['Computed Properties']=""
		productInfo['Property Name']="Property Value"
		productInfo['Physical Description']=PhysicalDescriptionInfoStr
		productInfo['Color/Form']=ColorFormInfoStr
		productInfo['Odor']=OdorInfoStr
		productInfo['Melting Point']=MeltingPointInfoStr
		productInfo['Solubility']=SolubilityInfoStr
		productInfo['Density']=DensityInfoStr
		productInfo['Vapor Pressure']=VaporPressureInfoStr
		productInfo['Octanol/Water Partition Coefficient']=OctanolWaterPartitionCoefficientInfoStr
		productInfo['Stability/Shelf Life']=StabilityShelfLifeInfoStr
		productInfo['Decomposition']=DecompositionInfoStr
		productInfo['Corrosivity']=CorrosivityInfoStr
		productInfo['pH']=pHInfoStr
		productInfo['Surface Tension']=SurfaceTensionInfoStr
		productInfo['Other Experimental Properties']=OtherExperimentalPropertiesInfoStr
		writeExcel(workSheet, headers, index, productInfo)
		print(str(index)+"====================="+IUPAC_Name)
		index=index+1
wb.save(excelFileName)
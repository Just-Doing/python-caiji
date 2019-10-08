from urllib.request import urlopen
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy

http.client._MAXHEADERS = 1000


def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text()

def urllib_download(IMAGE_URL, imageName):
	from urllib.request import urlretrieve
	urlretrieve(IMAGE_URL, imageName)  

def getHtmlFromUrl(url):
	try:
		html = urlopen(url).read()
		return html
	except:
		print("重试"+url)
		getHtmlFromUrl(url)

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1

def getProductObj(url, pInfo, pType):
	pHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(pHtml, "html.parser",from_encoding="utf-8")
	nameSope = sope.find("h1",attrs={"itemprop":"name"})
	lactideSope = sope.find("h2",attrs={"itemprop":"description"})
	SynonymSope = sope.find("p",attrs={"class":"synonym"})
	productInfoDiv = sope.find("div",attrs={"class":"productInfo"})
	LinearFormulaSope = productInfoDiv.find("ul",attrs={"class":"clearfix"})
	Formulas=[]
	if LinearFormulaSope != None:
		Formulas = LinearFormulaSope.find_all("li")
	LinearFormula=""
	CASNumber=""
	MolecularWeight=""
	for Formula in Formulas:
		if(getNodeText(Formula).strip().find("Empirical Formula (Hill Notation)")>-1):
			LinearFormula = getNodeText(Formula)
		if(getNodeText(Formula).strip().find("CAS Number")>-1):
			CASNumber = getNodeText(Formula)
			print(CASNumber)
		if(getNodeText(Formula).strip().find("Molecular Weight")>-1):
			MolecularWeight = getNodeText(Formula)
			print(MolecularWeight)
	LinearFormulaSope = productInfoDiv.find("ul",attrs={"class":"clearfix"})
	PropertiesTr = sope.find_all("tr")
	DescriptionDiv = sope.find("div",attrs={"class":"descriptionContent"})
	descriptionTitle = []
	if(DescriptionDiv != None):
		descriptionTitle = DescriptionDiv.find_all("h4")
	
	SafetyInfoTitle = sope.find_all("div", attrs={"class":"safetyLeft"})
		
	ArticlesDiv = sope.find("div",attrs={"id":"productDetailProtocols"})
		
	pInfo["t1"]=pType['t1'] if 't1' in pType else ''
	pInfo["t2"]=pType['t2'] if 't2' in pType else ''
	pInfo["t3"]=pType['t3'] if 't3' in pType else ''
	pInfo["CASNumber"]=CASNumber
	pInfo["MolecularWeight"]=MolecularWeight
	pInfo["LinearFormula"]=LinearFormula
	pInfo["name"]= getNodeText(nameSope)
	pInfo["Synonym"]=getNodeText(SynonymSope)
	pInfo["Articles"]=getNodeText(ArticlesDiv)
	
	imageAreaSope = sope.find("div",attrs={"class":"image prodImage"})
	imageSope = imageAreaSope.find("img",attrs={"itemprop":"image"})
	if imageSope != None:
		imgUrl = 'https://www.sigmaaldrich.com'+imageSope["src"]
		# urllib_download(imgUrl,pInfo["name"].replace("/","").replace("\n","").replace("\r","").replace("\\","")+'-1.png')
	for propertTr in PropertiesTr:
		pTitle = propertTr.find("td", attrs={"class":'lft'})
		pValue = propertTr.find("td", attrs={"class":'rgt'})
		if(getNodeText(pTitle).strip() == "Related Categories"):
			relateNode = pValue.find_all("a")
			pInfo["RelatedCategories"]=""
			for relate in relateNode:
				pInfo["RelatedCategories"]= pInfo["RelatedCategories"]+getNodeText(relate)+","
		if(getNodeText(pTitle).strip() == "assay"):
			pInfo["assay"]=getNodeText(pValue)
		if(getNodeText(pTitle).strip() == "form"):
			pInfo["form"]=getNodeText(pValue)
		if(getNodeText(pTitle).strip() == "concentration"):
			pInfo["concentration"]=getNodeText(pValue)
		if(getNodeText(pTitle).strip() == "refractive index"):
			pInfo["refractiveindex"]=getNodeText(pValue)
		if(getNodeText(pTitle).strip() == "bp"):
			pInfo["bp"]=getNodeText(pValue)
		
		if(getNodeText(pTitle).strip() == "storage temp."):
			pInfo["storagetemp"]=getNodeText(pValue)
	for desTitle in descriptionTitle:
		if(getNodeText(desTitle).strip() == "Application"):
			pInfo["Application"]=getNodeText(desTitle.nextSibling.nextSibling)
		if(getNodeText(desTitle).strip() == "Packaging"):
			pInfo["Packaging"]=getNodeText(desTitle.nextSibling.nextSibling)
	for SafetyTitle in SafetyInfoTitle:
		if(getNodeText(SafetyTitle).strip() == "RIDADR"):
			pInfo["RIDADR"]=getNodeText(SafetyTitle.nextSibling.nextSibling)
		if(getNodeText(SafetyTitle).strip() == "WGK Germany"):
			pInfo["WGKGermany"]=getNodeText(SafetyTitle.nextSibling.nextSibling)
		if(getNodeText(SafetyTitle).strip() == "Flash Point(F)"):
			pInfo["FlashPointF"]=getNodeText(SafetyTitle.nextSibling.nextSibling)
		if(getNodeText(SafetyTitle).strip() == "Flash Point(C)"):
			pInfo["FlashPointC"]=getNodeText(SafetyTitle.nextSibling.nextSibling)
			
	return pInfo

def getProductInfo(sope, pType, products):
	productTable = sope.find_all(name="table",attrs={"class":"opcTable"})
	for tb in productTable:
		pInfo = {}
		tbody = tb.find("tbody")
		pLinkTrs = tbody.find_all(name="a",attrs={"class":"OPCPDLink"})
		for pNode in pLinkTrs:
			pLink=pNode.get('href')
			if(pLink!= None and pLink.find("/catalog/product/")==0):
				pUrl= "https://www.sigmaaldrich.com"+pLink
				pInfo["plink"] = pUrl
				pInfo = getProductObj(pUrl, pInfo, pType)
				products.append(pInfo.copy())
				print(len(products))

def getProductSope( url, pType, dep, products):
	tp = pType.copy()
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	linkArea=sope.find(name="ul",attrs={"class":"opcsectionlist"})
	if linkArea != None:
		links = linkArea.find_all(name="a")
		dep = dep + 1
		for link in links:
			tp["t"+str(dep)] = link.get_text()
			getProductSope("https://www.sigmaaldrich.com/"+link["href"], tp, dep, products)
	else:
		getProductInfo(sope, tp, products)
	
	
excelFileName="sigma4-4.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
url1 = "https://www.sigmaaldrich.com/materials-science/material-science-products.html?TablePage=111766251"
url2 = "https://www.sigmaaldrich.com/materials-science/material-science-products.html?TablePage=111775084"
url3 = "https://www.sigmaaldrich.com/materials-science/material-science-products.html?TablePage=16390004"
url4 = "https://www.sigmaaldrich.com/materials-science/material-science-products.html?TablePage=16371222"
# getProductSope("https://www.sigmaaldrich.com/materials-science/material-science-products.html?TablePage=103936134",{},0,[])
getProductSope(url4, {}, 0, products)

headers=["t1",'t2','t3','name','Synonym','CASNumber','LinearFormula','MolecularWeight','RelatedCategories','assay','form','concentration','refractiveindex','bp','storagetemp','Application','Packaging',
'RIDADR','WGKGermany','FlashPointF','FlashPointC','Articles','plink']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1

print("flish")	

wb.save(excelFileName)
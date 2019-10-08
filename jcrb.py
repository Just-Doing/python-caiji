from urllib.request import urlopen
import urllib
import requests
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy

http.client._MAXHEADERS = 1000


def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()


def getHtmlFromUrl(url, type="get", para={}):
	from urllib import request, parse
	try:
		header_selfdefine={
			 'Content-Type':'application/x-www-form-urlencoded'
		}
		if type == "post":
			r =requests.post(url, data=para, headers=header_selfdefine)
			return r.content
		else:
			request_obj=urllib.request.Request(url=url)
			response_obj=urllib.request.urlopen(request_obj)
			html_code=response_obj.read().decode('utf-8')
			return html_code
	except:
		print("重试"+url)
		getHtmlFromUrl(url, type, para)

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1


def getProductInfo(url, products):
	productHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
	pInfo = {}
	infoItems = sope.find_all("tr")
	for infoItem in infoItems:
		infoTitle = infoItem.find("th")
		if getNodeText(infoTitle) == "Cell line name":
			pInfo["Celllinename"] = getNodeText(infoItem.find("td"))
		if getNodeText(infoTitle) == "Accession":
			pInfo["Accession"] = getNodeText(infoItem.find("td"))
		if getNodeText(infoTitle) == "Comments":
			pInfo["Comments"] = getNodeText(infoItem.find("td"))
		if getNodeText(infoTitle) == "Species of origin":
			pInfo["Species"] = getNodeText(infoItem.find("td"))
		if getNodeText(infoTitle) == "Hierarchy":
			pInfo["Hierarchy"] = getNodeText(infoItem.find("td"))
		if getNodeText(infoTitle) == "Sex of cell":
			pInfo["Sex"] = getNodeText(infoItem.find("td"))
		if getNodeText(infoTitle) == "Age at sampling":
			pInfo["Age"] = getNodeText(infoItem.find("td"))
		if getNodeText(infoTitle) == "Category":
			pInfo["Category"] = getNodeText(infoItem.find("td"))
		if getNodeText(infoTitle) == "Cell line collections":
			cellLineCollects = re.split(r'[;\r\n]\s*',getNodeText(infoItem.find("td")))
			if len(cellLineCollects) >0:
				pInfo["clc1"] = cellLineCollects[0]
			if len(cellLineCollects) >1:
				pInfo["clc2"] = cellLineCollects[1]
			if len(cellLineCollects) >2:
				pInfo["clc3"] = cellLineCollects[2]
			if len(cellLineCollects) >3:
				pInfo["clc4"] = cellLineCollects[3]
			if len(cellLineCollects) >4:
				pInfo["clc5"] = cellLineCollects[4]
			if len(cellLineCollects) >5:
				pInfo["clc6"] = cellLineCollects[5]
			if len(cellLineCollects) >6:
				pInfo["clc7"] = cellLineCollects[6]
			if len(cellLineCollects) >7:
				pInfo["clc8"] = cellLineCollects[7]
			if len(cellLineCollects) >8:
				pInfo["clc9"] = cellLineCollects[8]
			if len(cellLineCollects) >9:
				pInfo["clc10"] = cellLineCollects[9]
			if len(cellLineCollects) >10:
				pInfo["clc11"] = cellLineCollects[10]
	products.append(pInfo.copy())
	print(len(products))
				
	

def getProductList(url, products):
	productListHtml = getHtmlFromUrl(url, 'post', "input=JCRB")
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	
	listArea = sope.find_all("tr")
	for linkArea in listArea:
		link = linkArea.find("a")
		getProductInfo("https://web.expasy.org"+link["href"], products)

excelFileName="D:\\jcrb.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
url = "https://web.expasy.org/cgi-bin/cellosaurus/search"
# purl="https://web.expasy.org/cellosaurus/CVCL_2267"
# getProductInfo(purl, products)
getProductList(url, products)
headers=['Celllinename','Accession','Comments','Species','Hierarchy','Sex','Age','Category','clc1','clc2','clc3','clc4','clc5','clc6','clc7','clc8','clc9','clc10','clc11']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
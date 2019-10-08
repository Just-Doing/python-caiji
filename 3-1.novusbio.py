from urllib.request import urlopen
from bs4 import BeautifulSoup
import http.client
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import _thread

http.client._MAXHEADERS = 1000

def getHtmlFromUrl(url):
	try:
		html = urlopen(url).read()
		return html
	except:
		print("重试"+url)
		getHtmlFromUrl(url)

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head.strip() in info:
			workSheet.cell(rowIndex, cellIndex).value = str(info[head.strip()]).strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1



def theardFun(startPage, endPage,excelFname):
	excelFileName="D:\\"+excelFname+".xlsx"
	wb = Workbook()
	workSheet = wb.active
	pageIndex=startPage
	headers=[
		'catalog','product name',"product link","publications","application","Species","Species link"
	]
	index=1
	while pageIndex < endPage:
		productUrl = "https://www.novusbio.com/product-type/primary-antibodies?clonality=Monoclonal&page="+str(pageIndex)
		productHtml = getHtmlFromUrl(productUrl)
		if productHtml!=None and len(productHtml)>0:
			htmlSoup = BeautifulSoup(productHtml, "html.parser", from_encoding="utf-8")
			pNodes = htmlSoup.findAll(name="div", attrs={"class":"new-search-result search-result-wrapper"})
			for pNode in pNodes:
				pInfo = {}
				pLinkNode = pNode.find(name="a",attrs={"class":"ecommerce_link"})
				pCatNode = pNode.find(name="div",attrs={"class":"catalog_number_wrapper not3column"})
				pName = pLinkNode.text
				pUrl="https://www.novusbio.com"+pLinkNode["href"]
				cat = pCatNode.text
				
				pInfoHtml = getHtmlFromUrl(pUrl)
				if pInfoHtml!=None and len(pInfoHtml)>0:
					pInfoSoup = BeautifulSoup(pInfoHtml,"html.parser",from_encoding="utf-8")
					specNodes = pInfoSoup.findAll(name="tr",attrs={"class":"lined revfil firstten"})
					specIndex=0
					pInfo["catalog"]=cat
					pInfo["product name"]=pName
					pInfo["product link"]=pUrl
					if len(specNodes) > 0:
						specCount = 0
						if len(specNodes)>5:
							specCount=5 
						else:
							specCount=len(specNodes)
						while specIndex< specCount:
							specNode=specNodes[specIndex]
							publicationsNode=specNode.find(name="a",attrs={"rel":"nofollow"})
							specTdNodes = specNode.findAll(name="td")
							publications = specTdNodes[0].text if len(specTdNodes)>0 else ""
							url = publicationsNode["href"] if publicationsNode != None else ""
							
							
							application=specTdNodes[1].text if len(specTdNodes)>1 else ""
							Species=specTdNodes[2].text if len(specTdNodes)>2 else ""
							
							pInfo["catalog"]=cat if specIndex==0 else ""
							pInfo["product name"]=pName if specIndex==0 else ""
							pInfo["product link"]=pUrl if specIndex==0 else ""
							pInfo["publications"]=publications
							pInfo["application"]=application
							pInfo["Species"]=Species
							pInfo["Species link"]=url
							writeExcel(workSheet,headers,index,pInfo)
							index = index + 1
							specIndex=specIndex+1
							print(str(pageIndex)+"_"+str(index)+pUrl)
					else :
						writeExcel(workSheet,headers,index,pInfo)
						index = index + 1
		pageIndex=pageIndex+1
	wb.save(excelFileName)
	
try:
	_thread.start_new_thread( theardFun, (1,268, "product_novu1" ) )
	_thread.start_new_thread( theardFun, (268,536, "product_novu2" ) )
	_thread.start_new_thread( theardFun, (536,804, "product_novu3" ) )
	_thread.start_new_thread( theardFun, (804,1072, "product_novu4" ) )
	_thread.start_new_thread( theardFun, (1072,1340, "product_novu5" ) )
	_thread.start_new_thread( theardFun, (1340,1608, "product_novu6" ) )
	_thread.start_new_thread( theardFun, (1608,1876, "product_novu7" ) )
	_thread.start_new_thread( theardFun, (1876,2149, "product_novu8" ) )
except:
	print ("Error: 无法启动线程")
while 1:
   pass
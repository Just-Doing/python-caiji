from urllib.request import urlopen
import urllib
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy

http.client._MAXHEADERS = 1000


def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

headers = {"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"}
def down_pic(url, path):
    try:
        req = urllib.request.Request(url, headers=headers)
        data = urllib.request.urlopen(req).read()
        with open(path, 'wb') as f:
            f.write(data)
            f.close()
    except Exception as e:
        print(str(e))

def getHtmlFromUrl(url):
	try:
		header_selfdefine={
			 'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:59.0) Gecko/20100101 Firefox/59.0',
			 'Accept': '*/*',
		}

		request_obj=urllib.request.Request(url=url,headers=header_selfdefine)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read().decode('utf-8')
		return html_code
	except:
		print("重试"+url)
		getHtmlFromUrl(url)

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1

def getProductObj(url, pInfo, pType):
	pHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(pHtml, "html.parser",from_encoding="utf-8")
	nameSope = sope.find("h1",attrs={"itemprop":"name"})
	
		
	pInfo["t1"]=pType['t1'] if 't1' in pType else ''
	pInfo["t2"]=pType['t2'] if 't2' in pType else ''
	pInfo["t3"]=pType['t3'] if 't3' in pType else ''
	pInfo["name"]= getNodeText(nameSope)
	
	imageAreaSope = sope.find("div",attrs={"class":"image prodImage"})
	imageSope = imageAreaSope.find("img",attrs={"itemprop":"image"})
	if imageSope != None:
		imgUrl = 'https://www.sigmaaldrich.com'+imageSope["src"]
		downloadImg('https://www.sigmaaldrich.com'+imgUrl,pInfo["name"].replace("/","").replace("\\","").replace("\n","")+'-1.png')
	
	return pInfo

def getProductInfo(url, pType, products,txtFile):
	productHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
	pInfo = {}
	pNameArea = sope.find("h1", attrs={"id":"page-title"})
	productInfo = getNodeText(pNameArea).split("\n")
	pName = productInfo[0]
	cas = productInfo[1].split("Product Number")[0].replace("（CAS RN：","")
	pInfo["name"]=pName
	pInfo["cas"]=cas
	Synonym=""
	SynonymArea=sope.find("table",attrs={"class":"syg-tbl"})
	SynonymTrs = SynonymArea.find_all("tr")
	for SynonymTr in SynonymTrs:
		itemTitle = SynonymTr.find("th")
		itemValue = SynonymTr.find("td")
		if(getNodeText(itemTitle) == "Synonym"):
			Synonym = Synonym+getNodeText(itemValue)+","
	pInfo["Synonym"]=Synonym
	pInfo["t1"]=pType['t1'] if 't1' in pType else ''
	pInfo["t2"]=pType['t2'] if 't2' in pType else ''
	pInfo["t3"]=pType['t3'] if 't3' in pType else ''
	imageArea = sope.find("td", attrs={"class":"td-img"})
	img = imageArea.find("img")
	if(img != None):
		txtFile.write('https://www.sigmaaldrich.com'+img["src"]+"========"+(cas if cas!='' else pName)+"\n")
		
	pInfoItems = sope.find_all("th", attrs={"class":"base-th"})
	
	for pInfoItem in pInfoItems:
		if(getNodeText(pInfoItem) == "Purity/Analysis Method"):
			pInfo["Purity"]=getNodeText(pInfoItem.nextSibling.nextSibling)
		if(getNodeText(pInfoItem) == "Storage Temperature"):
			pInfo["StorageTemperature"]=getNodeText(pInfoItem.nextSibling.nextSibling)
		if(getNodeText(pInfoItem) == "M.F. / M.W."):
			pInfo["MFMW"]=getNodeText(pInfoItem.nextSibling.nextSibling)
		if(getNodeText(pInfoItem) == "Related CAS RN"):
			pInfo["RelatedCASRN"]=getNodeText(pInfoItem.nextSibling.nextSibling)
		if(getNodeText(pInfoItem) == "Total Nitrogen"):
			pInfo["TotalNitrogen"]=getNodeText(pInfoItem.nextSibling.nextSibling)
		if(getNodeText(pInfoItem) == "Ash Content"):
			pInfo["AshContent"]=getNodeText(pInfoItem.nextSibling.nextSibling)
		if(getNodeText(pInfoItem) == "Drying loss"):
			pInfo["Dryingloss"]=getNodeText(pInfoItem.nextSibling.nextSibling)

	products.append(pInfo.copy())
	print(len(products))
				
	

def getProductList(url,pType,products,txtFile):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	listArea = sope.find("dl", attrs={"class":"chem-name"})
	list = listArea.find_all("dt")
	for linkArea in list:
		link = linkArea.find("a")
		getProductInfo("https://www.tcichemicals.com"+link["href"],pType,products,txtFile)

def getpType(url, pType, dep, products,txtFile):
	tp = pType.copy()
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	typeLinkAreas=sope.find_all(name="h3",attrs={"class":"sub-titleA"})
	if len(typeLinkAreas) > 0:
		firstLink = typeLinkAreas[0].find("a")
		#如果 只有一个type区域 并且没有连接 则视为列表页
		if firstLink == None and len(typeLinkAreas) == 1:
			getProductList(url, tp, products,txtFile)
		else:
			dep = dep + 1
			for linkArea in typeLinkAreas:
				link = linkArea.find("a")
				if link != None:
					tp["t"+str(dep)] = getNodeText(link)
					getpType("https://www.tcichemicals.com/"+link["href"], tp, dep, products,txtFile)
	else:
		getProductList(url, tp, products,txtFile)


txtFile = open('D://Tcich1.txt','w')
excelFileName="D:\\Tcich1.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
url = "https://www.tcichemicals.com/eshop/en/us/category_index/13104"
# getpType(url, {}, 0, products,txtFile)
getProductList(url, {},  products,txtFile)
headers=["t1",'t2','t3','name','Synonym','cas','Purity','StorageTemperature','MFMW','RelatedCASRN','TotalNitrogen','AshContent','Dryingloss']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
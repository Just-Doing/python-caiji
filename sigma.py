from urllib.request import urlopen
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy

http.client._MAXHEADERS = 1000


def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text()

def urllib_download(IMAGE_URL, imageName):
	try:
		from urllib.request import urlretrieve
		urlretrieve(IMAGE_URL, imageName)   
	except:
		print("重试图片下载"+IMAGE_URL)
		urllib_download(IMAGE_URL, imageName) 

def getHtmlFromUrl(url):
	try:
		proxy="127.0.0.1:1080"
		# Build ProxyHandler object by given proxy
		proxy_support=urllib.request.ProxyHandler({'http':proxy})
		# Build opener with ProxyHandler object
		opener = urllib.request.build_opener(proxy_support)
		# Install opener to request
		urllib.request.install_opener(opener)
		html = urlopen(url).read()
		return html
	except:
		print("重试"+url)
		getHtmlFromUrl(url)

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1

def getProductObj(url, pInfo, pType):
	pHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(pHtml, "html.parser",from_encoding="utf-8")
	nameSope = sope.find("h1",attrs={"itemprop":"name"})
	lactideSope = sope.find("h2",attrs={"itemprop":"description"})
	SynonymSope = sope.find("p",attrs={"class":"synonym"})
	productInfoDiv = sope.find("div",attrs={"class":"productInfo"})
	LinearFormulaSope = productInfoDiv.find("ul",attrs={"class":"clearfix"})
	Formulas=[]
	if LinearFormulaSope != None:
		Formulas = LinearFormulaSope.find_all("li")
	LinearFormula=""
	for Formula in Formulas:
		if(getNodeText(Formula).strip().find("Linear Formula")>-1):
			LinearFormula = getNodeText(Formula)
			print(LinearFormula)
	PropertiesTr = sope.find_all("tr")
	
	DescriptionDiv = sope.find("div",attrs={"class":"descriptionContent"})
	descriptionTitle = []
	if(DescriptionDiv != None):
		descriptionTitle = DescriptionDiv.find_all("h4")
	
	SafetyInfoTitle = sope.find_all("div", attrs={"class":"safetyLeft"})
		
	ArticlesDiv = sope.find("div",attrs={"id":"productDetailProtocols"})
	# ArticlesDivTitle = []
	# if(ArticlesDiv != None):
		# ArticlesTitle = ArticlesDiv.find_all("div", attrs={"class":"article-container"})
		
	pInfo["t1"]=pType['t1'] if 't1' in pType else ''
	pInfo["t2"]=pType['t2'] if 't2' in pType else ''
	pInfo["t3"]=pType['t3'] if 't3' in pType else ''
	pInfo["name"]= getNodeText(nameSope)
	pInfo["lactide"]=getNodeText(lactideSope)
	pInfo["Synonym"]=getNodeText(SynonymSope)
	pInfo["LinearFormula"]=LinearFormula
	pInfo["Articles"]=getNodeText(ArticlesDiv)
	
	imageAreaSope = sope.find("div",attrs={"class":"image prodImage"})
	imageSope = imageAreaSope.find("img",attrs={"itemprop":"image"})
	if imageSope != None:
		imgUrl = 'https://www.sigmaaldrich.com'+imageSope["src"]
		urllib_download(imgUrl,pInfo["name"].replace("/","").replace("\\","").replace("\n","")+'-1.png')
	for propertTr in PropertiesTr:
		pTitle = propertTr.find("td", attrs={"class":'lft'})
		pValue = propertTr.find("td", attrs={"class":'rgt'})
		if(getNodeText(pTitle).strip() == "Related Categories"):
			relateNode = pValue.find_all("a")
			pInfo["RelatedCategories"]=""
			for relate in relateNode:
				pInfo["RelatedCategories"]= pInfo["RelatedCategories"]+getNodeText(relate)+","
		if(getNodeText(pTitle).strip() == "form"):
			pInfo["form"]=getNodeText(pValue)
		if(getNodeText(pTitle).strip() == "feed ratio"):
			pInfo["feedratio"]=getNodeText(pValue)
		if(getNodeText(pTitle).strip() == "viscosity"):
			pInfo["viscosity"]=getNodeText(pValue)
		if(getNodeText(pTitle).strip() == "storage temp."):
			pInfo["storagetemp"]=getNodeText(pValue)
	for desTitle in descriptionTitle:
		if(getNodeText(desTitle).strip() == "Application"):
			pInfo["Application"]=getNodeText(desTitle.nextSibling.nextSibling)
		if(getNodeText(desTitle).strip() == "Packaging"):
			pInfo["Packaging"]=getNodeText(desTitle.nextSibling.nextSibling)
	for SafetyTitle in SafetyInfoTitle:
		if(getNodeText(SafetyTitle).strip() == "RIDADR"):
			pInfo["RIDADR"]=getNodeText(SafetyTitle.nextSibling.nextSibling)
		if(getNodeText(SafetyTitle).strip() == "WGK Germany"):
			pInfo["WGKGermany"]=getNodeText(SafetyTitle.nextSibling.nextSibling)
		if(getNodeText(SafetyTitle).strip() == "Flash Point(F)"):
			pInfo["FlashPointF"]=getNodeText(SafetyTitle.nextSibling.nextSibling)
		if(getNodeText(SafetyTitle).strip() == "Flash Point(C)"):
			pInfo["FlashPointC"]=getNodeText(SafetyTitle.nextSibling.nextSibling)
			
	return pInfo

def getProductInfo(sope, pType, products):
	productTable = sope.find_all(name="table",attrs={"class":"opcTable"})
	for tb in productTable:
		pInfo = {}
		tbody = tb.find("tbody")
		pLinkTrs = tbody.find_all(name="a",attrs={"class":"OPCPDLink"})
		for pNode in pLinkTrs:
			pLink=pNode.get('href')
			if(pLink!= None and pLink.find("/catalog/product/")==0):
				pUrl= "https://www.sigmaaldrich.com"+pLink
				pInfo["plink"] = pUrl
				pInfo = getProductObj(pUrl, pInfo, pType)
				products.append(pInfo)
				print(len(products))
	

def getProductSope(url, pType, dep, products):
	tp = pType.copy()
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	linkArea=sope.find(name="ul",attrs={"class":"opcsectionlist"})
	if linkArea != None:
		links = linkArea.find_all(name="a")
		dep = dep + 1
		for link in links:
			tp["t"+str(dep)] = link.get_text()
			getProductSope("https://www.sigmaaldrich.com/"+link["href"], tp, dep, products)
	else:
		getProductInfo(sope, tp, products)
	
excelFileName="sigma1.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
url = "https://www.sigmaaldrich.com/materials-science/material-science-products.html?TablePage=20202255"
getProductSope(url, {}, 0, products)

# productListHtml = getHtmlFromUrl("https://www.sigmaaldrich.com/materials-science/material-science-products.html?TablePage=112226924")
# sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
# getProductInfo(workSheet,sope, {'t1':'t1','t2':'t2','t3':'t3'})
headers=["t1",'t2','t3','name','lactide','Synonym','LinearFormula','RelatedCategories','form','feedratio','viscosity','storagetemp','Application','Packaging',
	'RIDADR','WGKGermany','FlashPointF','FlashPointC','Articles','plink']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
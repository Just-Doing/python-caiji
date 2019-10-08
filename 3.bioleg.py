from urllib.request import urlopen
from bs4 import BeautifulSoup
import http.client
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import _thread
from urllib import parse

http.client._MAXHEADERS = 1000

def getHtmlFromUrl(url):
	try:
		html = urlopen(url).read()
		return html
	except Exception:
		print(Exception)
		print("重试"+url)
		getHtmlFromUrl(url)

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head.strip() in info:
			workSheet.cell(rowIndex, cellIndex).value = str(info[head.strip()]).strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1



def theardFun(startPage, endPage,excelFname):
	excelFileName="c:\\"+excelFname+".xlsx"
	wb = Workbook()
	workSheet = wb.active
	pageIndex=startPage
	headers=[
		'product name','clone','Previously',"product link","Application References","pubmed link"
	]
	index=1
	# urls=[
		# "https://www.biolegend.com/en-us/products/pe-dazzle%E2%84%A2-594-annexin-v",
		# "https://www.biolegend.com/en-us/products/biotin-anti-beta-amyloid--x-42-antibody-11183",
		# "https://www.biolegend.com/en-us/products/fitc-anti-mouse-i-ak-abetak-antibody-2"
	# ]
	# for u in urls:
		# productHtml = getHtmlFromUrl(u)
		# print(u)
	while pageIndex < endPage:
		productUrl = "https://www.biolegend.com/en-us/search-results?GroupID=&PageNum="+str(pageIndex)+"&PageSize=200&Category=PRIM_AB"
		productHtml = getHtmlFromUrl(productUrl)
		if productHtml!=None and len(productHtml)>0:
			htmlSoup = BeautifulSoup(productHtml, "html.parser", from_encoding="utf-8")
			pNodes = htmlSoup.find(name="ul", attrs={"id":"productsHolder"}).findAll(name="li",attrs={"class":"col-xs-12"})
			for pNode in pNodes:
				pInfo = {}
				pLinkNode = pNode.find(name="a",attrs={"itemprop":"name"})
				if pLinkNode!=None:
					pUrl="https://www.biolegend.com"+parse.quote(pLinkNode["href"])
					
					pInfoHtml = getHtmlFromUrl(pUrl)
					pubMed=[]
					if pInfoHtml!=None and len(pInfoHtml)>0:
						pName = pLinkNode.text
						pInfoSoup = BeautifulSoup(pInfoHtml,"html.parser",from_encoding="utf-8")
						pInfo["product name"]=pName
						pInfo["product link"]=pUrl
						
						productDetailNodes = pInfoSoup.findAll(name="dt")
						if len(productDetailNodes) > 0:
							for dNode in productDetailNodes:
								s = dNode.get_text().strip()
								if s.find("Application References")>-1:
									pInfo["Application References"]=dNode.findNext("dd").get_text()
									pubMed = dNode.findNext("dd").findAll(name="a")
								if s.find("Previously")>-1:
									pInfo["Previously"]=dNode.findNext("dd").get_text()
								if s.find("Clone")>-1:
									pInfo["clone"]="'"+dNode.findNext("dd").get_text()
					specIndex=0
					if len(pubMed) > 0:
						specCount = 0
						if len(pubMed)>5:
							specCount=5 
						else:
							specCount=len(pubMed)
						while specIndex< specCount:
							specNode=pubMed[specIndex]
							if specIndex>0:
								pInfo["product name"]=""
								pInfo["product link"]=""
								pInfo["Application References"]=""
								pInfo["clone"]=""
								pInfo["Previously"]=""
							if specNode.text.find("PubMed") < 0:
								pInfo["pubmed link"]=specNode.text.strip()
								writeExcel(workSheet,headers,index,pInfo)
								print(str(pageIndex)+"_"+str(index)+specNode.text.strip())
							else :
								url = specNode["href"] if specNode != None else ""
								pInfo["pubmed link"]=url
								writeExcel(workSheet,headers,index,pInfo)
								index = index + 1
								specIndex=specIndex+1
								print(str(pageIndex)+"_"+str(index)+pUrl)
					else :
						writeExcel(workSheet,headers,index,pInfo)
						index = index + 1
		pageIndex=pageIndex+1
	wb.save(excelFileName)
	
try:
	_thread.start_new_thread( theardFun, (1,45, "product_novu1" ) )
	_thread.start_new_thread( theardFun, (45,97, "product_novu2" ) )
except:
	print ("Error: 无法启动线程")
while 1:
   pass
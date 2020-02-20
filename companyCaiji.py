from urllib.request import urlopen
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy

http.client._MAXHEADERS = 1000


def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

def urllib_download(IMAGE_URL, imageName):
	from urllib.request import urlretrieve
	fileName = imageName+".png"
	print(fileName)
	urlretrieve(IMAGE_URL, fileName)  

def getHtmlFromUrl(url):
	try:
		html = urlopen(url).read()
		return html
	except:
		print("重试"+url)
		getHtmlFromUrl(url)

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1

def getProducts(url, products):
	pHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(pHtml, "html.parser",from_encoding="utf-8")
	nameSope = sope.find("table",attrs={"class":"rightLinks"})
    productTr = nameSope.find_all("tr")
    for pro in productTr:
        pNameTd = pro.find("td", attrs={"class": "company"})
        locationTd = pro.find("td", attrs={"class": "location"})
        descriptionTd = pro.find("td", attrs={"class": "description"})
        pNameLink = pNameTd.find("a")
        pInfo["name"]=getNodeText(pNameLink)
        pInfo["locationTd"]=getNodeText(locationTd)
        pInfo["description"]=getNodeText(descriptionTd)
        pInfo["url"]=pNameLink["href"]
        print(pInfo)
        #products.append(pInfo.copy())

excelFileName="polysciences1.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
url = "https://biopharmguy.com/links/company-by-location-stem-cells.php"
print(url)
getProducts(url, products)
# pinfo = getProductObj("https://www.polysciences.com/default/26270", {}, "")
# print(pinfo)
headers=["pType",'name','description','Synonyms','refrence1','refrence2','refrence3','refrence4','FTIR','Inherent','Acid','Lactide','Viscosity','Glass','Soluble','CAS','Hazards','Handling','Storage','plink']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1

print("flish")	

wb.save(excelFileName)
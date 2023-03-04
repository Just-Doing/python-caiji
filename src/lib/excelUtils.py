
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


def writeExcelRow(workSheet, headers, rowIndex, data):
    cellIndex = 1
    for head in headers:
        try:
            if head in data:
                content = ILLEGAL_CHARACTERS_RE.sub(r'', str(data[head]))
                workSheet.cell(rowIndex, cellIndex).value = content.strip()
            else:
                workSheet.cell(rowIndex, cellIndex).value = ""
            cellIndex = cellIndex+1
        except:
            print("errIndex:"+rowIndex)


def writeExcel(workSheet, headers, products):
    for index, head in enumerate(headers):
        workSheet.cell(1, index+1).value = head.strip()

    for index, p in enumerate(products):
        writeExcelRow(workSheet, headers, index + 2, p)

def generateExcel(fileName, products, headers):
    excelFileName = fileName
    wb = Workbook()
    workSheet = wb.active
    writeExcel(workSheet, headers, products)
    wb.save(excelFileName)
    print("flish")

def generateExcelMultipleSheet(fileName, data):
    excelFileName = fileName
    wb = Workbook()
    for sheetName in data:
        sheet1 = wb.create_sheet(title=sheetName["name"])
        writeExcel(sheet1, sheetName["header"], sheetName["data"])
    wb.save(excelFileName)
    print("flish")

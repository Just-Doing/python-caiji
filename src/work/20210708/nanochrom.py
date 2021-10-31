from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):

	url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
	request_obj=urllib.request.Request(url=url)
	response_obj=urllib.request.urlopen(request_obj)
	html_code=response_obj.read()
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	
def getRenderdHtmlFromUrl(url):

	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	return BeautifulSoup(browser.page_source, "html.parser",from_encoding="utf-8")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, application, products):
	print(str(len(products)) + url)
	sope = getHtmlFromUrl(url)
	navArea = sope.find("div", attrs={"id":"Breadcrumb"})
	pName = sope.find("h1", attrs={"id":"PageTitleDivId"})
	introduction2Area = sope.find("div", attrs={"id":"menu1"})
	introduction2s= introduction2Area.find_all("p")
	introduction2=""
	if len(introduction2Area.contents)>0:
		if introduction2Area.contents[0].name=="strong":
			introduction2 =  getNodeText(introduction2Area.find("strong"))
		else:
			if len(introduction2s) == 0:
				introduction2 = getNodeText(introduction2Area)
			else:
				introduction2 = getNodeText(introduction2s[0])
				if len(introduction2) == 0 and len(introduction2s)>1:
					introduction2 = getNodeText(introduction2s[1])
				if len(introduction2) == 0 and len(introduction2s)>2:
					introduction2 = getNodeText(introduction2s[2])
				if len(introduction2) == 0 and len(introduction2s)>3:
					introduction2 = getNodeText(introduction2s[3])
				if len(introduction2) == 0 and len(introduction2s)>4:
					introduction2 = getNodeText(introduction2s[4])
	imgs = introduction2Area.find_all("img")
	imgNames=""
	if len(imgs)>0:
		for index,img in enumerate(imgs):
			imgName = getNodeText(pName)+"-"+str(index)+".jpg"
			urllib_download("http://www.nanochrom.com"+img["src"], imgName)
			imgNames= imgNames + imgName+";"
	pInfo = {
		"link": url,
		"nav":getNodeText(navArea).replace(" ","").replace("\r","").replace("\n",""),
		"pName":getNodeText(pName),
		"introduction": pName.nextSibling,
		"introduction2":introduction2,
		"imgNames": imgNames,
		"application": application
	}
	products.append(pInfo.copy())

def getProductList(url, products):
	sope = getHtmlFromUrl(url)
	application =getNodeText(sope.find("div", attrs={"class":"userDefined"}))
	pList = sope.find_all("div", attrs={"id":"Products"})
	for p in pList:
		pLinks = p.find_all("a")
		if len(pLinks)==2:
			getProductInfo("http://www.nanochrom.com"+pLinks[1]["href"],application, products)

def getProductType(url, products):
	sope = getHtmlFromUrl(url)
	
	navtypeArea = sope.find("div", attrs={"id":"PrimaryNav"})
	typeArea = navtypeArea.find("ul", attrs={"class":"megaMenuNav"})
	types = typeArea.find_all("li", attrs={"class":"title"})
	for type in types:
		typeLink = type.find("a")
		getProductList("http://www.nanochrom.com"+typeLink["href"], products)

excelFileName="nanochrom.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("http://www.nanochrom.com/np_en/productshow/40.html", products)

getProductType("http://www.nanochrom.com/np_en/index.html", products)
# getProductList("http://www.nanochrom.com/np_en/productlists/12.html", products)

headers=[
	'link','application','nav','pName','introduction','introduction2','imgNames'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
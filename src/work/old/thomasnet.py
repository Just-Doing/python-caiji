from urllib.request import urlopen
import urllib
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy

http.client._MAXHEADERS = 1000


def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		html = urlopen(url).read()
		return html
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getHtmlFromUrl(url)
		else:
			retryCount=0
			return None

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1


def getProductInfo(url, pInfo, products):

	print(str(len(products)) + url)
	productHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
	
	websiteLink = sope.find(name="a", attrs={"title": "Visit Website"})
	addressInfo = sope.find(name="p", attrs={"class": "addrline"})
	pInfo["website"] = websiteLink["href"] if websiteLink != None else ""
	pInfo["address"] = getNodeText(addressInfo)
	products.append(pInfo.copy())
				
	

def getProductList(url, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	
	listArea = sope.find_all(name="header", attrs={"class":"profile-card__header" })
	for linkArea in listArea:
		cHeader = linkArea.find(name="h2", attrs={"class":"profile-card__title"})
		covidInfo = linkArea.find(name="span",attrs={"class":"supplier-badge--mobile-icon supplier-badge supplier-badge--sos"})
	
		link = cHeader.find("a")

		pInfo = {
			"cName": getNodeText(link),
			"covid": "是" if getNodeText(covidInfo) == "PathCOVID-19 Response" else "否"
		}
		getProductInfo("https://www.thomasnet.com"+link["href"], pInfo, products)

excelFileName="D:\\jcrb.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
i = 1
while i < 3:
	url = "https://www.thomasnet.com/nsearch.html?act=C&cov=NA&heading=42031716&pg=" + str(i)
	getProductList(url, products)
	i += 1
headers=['cName','website','covid','address']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
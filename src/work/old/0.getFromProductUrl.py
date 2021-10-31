from urllib.request import urlopen
from bs4 import BeautifulSoup
import http.client
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import _thread
from urllib import parse

http.client._MAXHEADERS = 1000

def getHtmlFromUrl(url):
	try:
		html = urlopen(url).read()
		return html
	except Exception:
		print(Exception)
		print("重试"+url)
		getHtmlFromUrl(url)

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head.strip() in info:
			workSheet.cell(rowIndex, cellIndex).value = str(info[head.strip()]).strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1


def getProductFromUrl(productUrl):
	productHtml = getHtmlFromUrl(productUrl)
	if productHtml!=None and len(productHtml)>0:
		htmlSoup = BeautifulSoup(productHtml, "html.parser", from_encoding="utf-8")
		pNodes = htmlSoup.find(name="div", attrs={"id":"productDetailHero"})
		pInfo = {}
		pNameNode = pNodes.find(name="h1",attrs={"itemprop":"name"})
		molecularNode = pNodes.find(name="h2",attrs={"itemprop":"description"})
		synonymNode = pNodes.find(name="p",attrs={"class":"synonym"})
		imgNode = pNodes.find(name="img",attrs={"itemprop":"image"})
		casNumbNode = pNodes.find(name="ul",attrs={"class":"clearfix"})
		
		return pInfo
		
excelFileName="c:\\"+excelFname+".xlsx"
wb = Workbook()
workSheet = wb.active
headers=[
	'product name','clone','Previously',"product link","Application References","pubmed link"
]
index=1
urls=[
	"https://www.sigmaaldrich.com/catalog/product/aldrich/38534?lang=en&region=US",
	"https://www.sigmaaldrich.com/catalog/product/aldrich/363502?lang=en&region=US",
	"https://www.sigmaaldrich.com/catalog/product/aldrich/440752?lang=en&region=US",
	"https://www.sigmaaldrich.com/catalog/product/aldrich/440744?lang=en&region=US",
	"https://www.sigmaaldrich.com/catalog/product/aldrich/704105?lang=en&region=US"
]
for productUrl in urls:
	pInfo = getProductFromUrl(productUrl)
		
	writeExcel(workSheet,headers,index,pInfo)
	index = index + 1
wb.save(excelFileName)
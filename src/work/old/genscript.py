from urllib.request import urlopen
from selenium import webdriver
import urllib
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy

http.client._MAXHEADERS = 1000


def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url):
	global retryCount
	try:
		html = urlopen(url).read()
		return BeautifulSoup(html, "html.parser",from_encoding="utf-8")
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getHtmlFromUrl(url)
		else:
			retryCount=0
			return None
		
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1


def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	pInfo["link"]=url
	scope = getHtmlFromUrl(url)
	if scope != None:
		sizeArea = scope.find(name="div", attrs={"id":"add-to-cart"})
		sizeTrs = sizeArea.find_all("tr")
		size1Tr = sizeTrs[2]
		pInfo["size1"] = getNodeText(size1Tr.find_all("td")[1])
		pInfo["price1"] = getNodeText(size1Tr.find_all("td")[2])
		if len(sizeTrs) > 3:
			size2Tr = sizeTrs[3]
			pInfo["size2"] = getNodeText(size2Tr.find_all("td")[1])
			pInfo["price2"] = getNodeText(size2Tr.find_all("td")[2])
		if len(sizeTrs) > 4:
			size3Tr = sizeTrs[4]
			pInfo["size3"] = getNodeText(size3Tr.find_all("td")[1])
			pInfo["price3"] = getNodeText(size3Tr.find_all("td")[2])
		overviewTable = scope.find(name="div", attrs={"class":"s0 auto-module"}).find(name="table")
		protiesTable = scope.find(name="div", attrs={"class":"s1 auto-module"}).find(name="table")
		overviewTrs = overviewTable.find_all("tr")
		for overviewTr in overviewTrs:
			tds = overviewTr.find_all("td")
			title = getNodeText(tds[0])
			value = getNodeText(tds[1])
			if title == 'Synonyms':
				pInfo["Synonyms"] = value
			if title == 'Description':
				pInfo["Description"] = value
			if title == 'Host Species':
				pInfo["HostSpecies"] = value
			if title == 'Antigen Species':
				pInfo["AntigenSpecies"] = value
			if title == 'Conjugation':
				pInfo["Conjugation"] = value
			if title == 'Immunogen':
				pInfo["Immunogen"] = value
			if title == 'Fusion Partner':
				pInfo["FusionPartner"] = value
			if title == 'Purification':
				pInfo["Purification"] = value
			if title == 'Accession No':
				pInfo["AccessionNo"] = value
			if title == 'Species':
				pInfo["Species"] = value
			if title == 'Source':
				pInfo["Source"] = value
			if title == 'Biological Activity':
				pInfo["BiologicalActivity"] = value
			if title == 'Sequence':
				pInfo["Sequence"] = value
		
		
		protiesTrs = protiesTable.find_all("tr")
		for protiesTr in protiesTrs:
			tds = protiesTr.find_all("td")
			if len(tds) > 1:
				title = getNodeText(tds[0])
				value = getNodeText(tds[1])
				if title == 'Subclass':
					pInfo["Subclass"] = value
				if title == 'Appearance':
					pInfo["Appearance"] = value
				if title == 'Clone ID':
					pInfo["CloneID"] = value
				if title == 'Concentration':
					pInfo["Concentration"] = value
				if title == 'Reconstitution':
					pInfo["Reconstitution"] = value
				if title == 'Specificity':
					pInfo["Specificity"] = value
				if title == 'Species Reactivity':
					pInfo["SpeciesReactivity"] = value
				if title == 'Predicted Band Size':
					pInfo["PredictedBandSize"] = value
				if title == 'Observed Band Size':
					pInfo["ObservedBandSize"] = value
				if title == 'Storage':
					pInfo["Storage"] = value
				if title == 'Note':
					pInfo["Note"] = value
					
				if title == 'Measured Molecular Weight':
					pInfo["MeasuredMolecularWeight"] = value
				if title == 'Purity':
					pInfo["Purity"] = value
				if title == 'Dimers':
					pInfo["Dimers"] = value
				if title == 'Formulation':
					pInfo["Formulation"] = value
				if title == 'Reconstitution':
					pInfo["Reconstitution"] = value
				if title == 'Quantitation':
					pInfo["Quantitation"] = value
				if title == 'Endotoxin Level':
					pInfo["EndotoxinLevel"] = value
				if title == 'Physical Appearance':
					pInfo["PhysicalAppearance"] = value
				if title == 'Sequence Analysis':
					pInfo["SequenceAnalysis"] = value
				if title == 'UV':
					pInfo["UV"] = value
				if title == 'QC':
					pInfo["QC"] = value
				if title == 'Storage':
					pInfo["Storage"] = value
				if title == 'Usage':
					pInfo["Usage"] = value
		pInfo["Applications"] = getNodeText(scope.find(name="div", attrs={"class":"s2 auto-module"}))
		products.append(pInfo.copy())
	

def getProductList(url, products, type):
	sope = getHtmlFromUrl(url)
	tableArea = sope.find_all(name="div", attrs={"class":"product_table_cart_div" })
	for table in tableArea:
		trs = table.find_all(name="tr")
		cat = ""
		for tr in trs:
			link = tr.find("a")
			if link != None:
				currentCat = getNodeText(tr.find_all(name="td")[1]).split('-')[0]
				if cat != currentCat:
					cat = currentCat
					pInfo = {
						"type": type,
						"cat": cat,
						"name": getNodeText(link)
					}
					getProductInfo("https://www.genscript.com/"+link["href"], pInfo, products)




excelFileName="mouser.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
urls = [
	'https://www.genscript.com/cell_marker_antibody.html?src=leftbar',
	'https://www.genscript.com/loading_control_antibody.html?src=leftbar',
	'https://www.genscript.com/cytokine_and_growth_factor_antibody.html?src=leftbar',
	'https://www.genscript.com/gpcrs_antibody.html?src=leftbar',
	'https://www.genscript.com/H1N1.html?src=leftbar',
	'https://www.genscript.com/phospho_specific_antibody.html?src=leftbar',
	'https://www.genscript.com/disease-antibody-list.html?src=leftbar',
	'https://www.genscript.com/immune-checkpoint-antibodies.html?src=leftbar',
	'https://www.genscript.com/anti-idiotype.html?src=leftbar'
]

urls1 = [
	'https://www.genscript.com/cytokines.html?src=leftbar',
	'https://www.genscript.com/growth_factors.html?src=leftbar',
	'https://www.genscript.com/chemokines.html?src=leftbar',
	'https://www.genscript.com/hormones.html?src=leftbar',
	'https://www.genscript.com/neurotrophins.html?src=leftbar',
	'https://www.genscript.com/enzymes_and_inhibitors.html?src=leftbar',
	'https://www.genscript.com/other_catalog_proteins.html?src=leftbar'
]

for url in urls:
	getProductList(url, products, 'type1')


for url in urls1:
	getProductList(url, products, 'type2')
	
headers=[
	'type','name','cat','link','size1','price1','size2','price2','size3','price3',
	'Synonyms','Description','HostSpecies','AntigenSpecies','Conjugation','Immunogen','FusionPartner',
	'Purification','Subclass',
	'type',
	'AccessionNo','Species','Source','BiologicalActivity','Sequence',
	'type',
	'Appearance','CloneID','Concentration','Reconstitution','Specificity',
	'SpeciesReactivity','PredictedBandSize','ObservedBandSize','Storage','Note','Applications',
	'type',
	'MeasuredMolecularWeight','Purity','Dimers','Formulation','Reconstitution','Quantitation','EndotoxinLevel',
	'PhysicalAppearance','SequenceAnalysis','UV','QC','Storage','Note','Usage'
]
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
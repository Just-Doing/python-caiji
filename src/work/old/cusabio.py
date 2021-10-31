from urllib.request import urlopen
import urllib
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000


def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		reponse = urlopen(url)
		if reponse.getcode() == 200:
			html = reponse.read()
			return html
		else:
			return ''
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getHtmlFromUrl(url)
		else:
			retryCount=0
			return None

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1


def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	productHtml = getHtmlFromUrl(url)
	if productHtml != None:
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
		
		pInfo["link"] = url
		pInfo["nav"] = getNodeText(sope.find(name="ul", attrs={"class":"bread nav-bg"}))
		attrAreas = sope.find_all(name="table", attrs={"class": "pdts-detail-table"})
		for attrArea in attrAreas:
			attrInfos = attrArea.find_all("tr")
			for attr in attrInfos:
				tds = attr.find_all("td")
				if len(tds) > 1:
					attrLabel = getNodeText( tds[0])
					attrVal = getNodeText( tds[1])
					pInfo[attrLabel] = attrVal
		products.append(pInfo.copy())
				

def getProductList(url, type, products):
	print(url)
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	pTableDivs = sope.find_all("div", attrs={"class":"table-responsive"})
	for pTableDiv in pTableDivs:
		ptrs = pTableDiv.find_all("tr")
		for productItem in ptrs:
			link = productItem.find("a")
			pInfo = {
				"name": getNodeText(link)
			}
			if link != None and link["href"] != '':
				getProductInfo(link["href"], pInfo, products)

def getProductTypeList(url, type, products):
	print(url)
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	pTableDiv = sope.find("div", attrs={"class":"table-responsive"})
	proTr = pTableDiv.find_all("tr")
	for productItem in proTr:
		link = productItem.find("a")
		if link!=None and link["href"] != '':
			getProductList(link["href"], getNodeText(link), products)


def getProductPage(url, type, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	page = sope.find(name="ul", attrs={"class":"pagination-condensed" })
	getProductTypeList(url, type, products)
	if page !=None:
		lis = page.find_all("li")
		for li in lis:
			link = li.find("a")
			if link!=None and link["href"] != '':
				getProductTypeList(link["href"], type, products)

excelFileName="cusabio.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
pageindex = 0
# getProductInfo('https://www.cusabio.com/Polyclonal-Antibody/Actin-β-Polyclonal-Antibody-11090938.html', {}, products)
productListHtml = getHtmlFromUrl("https://www.cusabio.com/pathway/Apoptosis.html")
sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
types = sope.find("div", attrs={"class":"drop-menu submenu-wrap path-submenu"})
lis = types.find_all("li")

for li in lis:
	url = li.find("a")
	if pageindex == 2:
		getProductList(url["href"], '', products)
	else:
		getProductPage(url["href"], getNodeText(url), products)
	pageindex+=1

headers=['link','nav','name','Code','Size','Uniprot No.','Target Names','Target Name','Abbreviation','Species','Sample Types',
'Detection Range','Sensitivity','Alternative Names','Assay Time','Sample Volume','Detection Wavelength','Research Area',
'Assay Principle','Measurement','Precision','Raised in','Linearity','Typical Data','Troubleshootingand FAQs','Recovery','Species Reactivity',
'Immunogen','Immunogen Species','Isotype','Purification Method','Concentration','Buffer','Tested Applications',
'Protocols','Troubleshooting and FAQs','Storage','Lead Time','Function','Gene References into Functions',
'Involvement in disease','Subcellular Location','Tissue Specificity','Protein Families','Database Links']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
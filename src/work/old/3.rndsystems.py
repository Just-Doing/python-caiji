from urllib.request import urlopen
from bs4 import BeautifulSoup
import http.client
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json

http.client._MAXHEADERS = 1000

def getHtmlFromUrl(url):
	try:
		html = urlopen(url).read()
		return html
	except:
		print("重试"+url)
		getHtmlFromUrl(url)
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head.strip() in info:
			workSheet.cell(rowIndex, cellIndex).value = str(info[head.strip()]).strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1



fileName="D:\\p2.txt"
excelFileName="D:\\products1.xlsx"
wb = Workbook()
workSheet = wb.active
pageIndex=163
headers=[
	'catalog','product name',"product link","citation","citation url"
]
index=1
while pageIndex < 340:
	productUrl = "https://www.rndsystems.com/cn/search?keywords=primary%20antibodies&clonality=Monoclonal&numResults=50&page="+str(pageIndex)
	productHtml = getHtmlFromUrl(productUrl)
	htmlSoup = BeautifulSoup(productHtml, "html.parser", from_encoding="utf-8")
	pNodes = htmlSoup.findAll(name="div", attrs={"class":"search_result"})
	print(len(pNodes))
	for pNode in pNodes:
		pInfo = {}
		pLinkNode = pNode.find(name="a",attrs={"class":"ecommerce_link"})
		pCatNode = pNode.find(name="a",attrs={"class":"search-result-datasheet-link"})
		pName = pLinkNode.text
		cat = pCatNode.text
		pUrl=pLinkNode["href"]
		pInfoHtml = getHtmlFromUrl(pUrl)
		if pInfoHtml!=None and len(pInfoHtml)>0:
			pInfoSoup = BeautifulSoup(pInfoHtml,"html.parser",from_encoding="utf-8")
			citationContenter = pInfoSoup.find(name="ol",attrs={"id":"citations"})
			pInfo["catalog"]=cat 
			pInfo["product name"]=pName
			pInfo["product link"]=pUrl
			if citationContenter != None:
				citationsNodes = citationContenter.findAll(name="li")
				citIndex=0
				if len(citationsNodes) > 0:
					citCount = 0
					if len(citationsNodes)>5:
						citCount=5 
					else:
						citCount=len(citationsNodes)
					while citIndex< citCount:
						citNode=citationsNodes[citIndex]
						linkNode=citNode.find(name="a",attrs={"class":"safe_link"})
						tile = linkNode.text
						url = linkNode["href"]
						
						
						pInfo["catalog"]=cat if citIndex==0 else ""
						pInfo["product name"]=pName if citIndex==0 else ""
						pInfo["product link"]=pUrl if citIndex==0 else ""
						pInfo["citation"]=tile
						pInfo["citation url"]=url
						writeExcel(workSheet,headers,index,pInfo)
						index = index + 1
						citIndex=citIndex+1
						print(str(pageIndex)+"_"+str(index)+pUrl)
			else: 
				writeExcel(workSheet,headers,index,pInfo)
				index = index + 1
	pageIndex=pageIndex+1
wb.save(excelFileName)
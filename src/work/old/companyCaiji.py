from urllib.request import urlopen
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy

http.client._MAXHEADERS = 1000


def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

def urllib_download(IMAGE_URL, imageName):
	from urllib.request import urlretrieve
	fileName = imageName+".png"
	print(fileName)
	urlretrieve(IMAGE_URL, fileName)  

def getHtmlFromUrl(url):
	try:
		html = urlopen(url).read()
		return html
	except:
		print("重试"+url)
		getHtmlFromUrl(url)

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1

def getProducts(url, products):
	pHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(pHtml, "html.parser",from_encoding="utf-8")
	nameSope = sope.find("table",attrs={"class":"rightLinks"})
	proTrs=nameSope.find_all(name="tr")
	for pro in proTrs:
		pInfo={}
		pNameTd = pro.find("td", attrs={"class": "company"})
		locationTd = pro.find("td", attrs={"class": "location"})
		descriptionTd = pro.find("td", attrs={"class": "description"})
		if pNameTd != None:
			pNameLink = pNameTd.find_all(name="a")
			if len(pNameLink)==2:
				print(pNameLink)
				pInfo["name"]=getNodeText(pNameLink[1])
				pInfo["location"]=getNodeText(locationTd)
				pInfo["description"]=getNodeText(descriptionTd)
				pInfo["url"]=pNameLink[1]["href"]
				products.append(pInfo.copy())
excelFileName="d:/polysciences1.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
url = "https://biopharmguy.com/links/company-by-location-stem-cells.php"
print(url)
getProducts(url, products)
# pinfo = getProductObj("https://www.polysciences.com/default/26270", {}, "")
# print(pinfo)
headers=['url', "name",'location','description']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1

print("flish")	

wb.save(excelFileName)
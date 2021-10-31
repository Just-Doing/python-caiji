from urllib.request import urlopen
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import math
from bs4.element import NavigableString 

http.client._MAXHEADERS = 1000


def getNodeText(node):
	if(node == None):
		return ""
	else:
		if isinstance(node, NavigableString):
			return node
		else:
			return node.get_text()

def urllib_download(IMAGE_URL, imageName):
	try:
		from urllib.request import urlretrieve
		urlretrieve(IMAGE_URL, imageName)   
	except:
		print("retry"+IMAGE_URL)
		urllib_download(IMAGE_URL, imageName)
		
retryCount = 0
loadCount = 0
def getHtmlFromUrl(url):
	global retryCount
	try:
		html = urlopen(url).read()
		return html
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getHtmlFromUrl(url)
		else:
			retryCount=0
			return None

def getJsonFromUrl(url):
	global retryCount
	try:
		html = urlopen(url).read()
		return json.loads(html)
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getJsonFromUrl(url)
		else:
			retryCount=0
			return None

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = ILLEGAL_CHARACTERS_RE.sub(r'', info[head].strip())
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1

def getProductSope( url, products):
	productListHtml = BeautifulSoup( getHtmlFromUrl(url), "html.parser",from_encoding="utf-8")
	productList = productListHtml.find_all(name="div",attrs={"class":"list-link"})
	for product in productList:
		pLink = product.find("a")
		infoUrl = "https://www.atwineries.com"+pLink["href"]
		print(str(len(products)) + "===" + infoUrl)
		productInfo = BeautifulSoup( getHtmlFromUrl(infoUrl), "html.parser",from_encoding="utf-8")
		if productInfo!= None:
			homeUrl = ""
			productInfo.find(name="section", attrs={"class": "contact-info-section"})
			if productInfo != None:
				webSet = productInfo.find(name="a", attrs={"rel":"noopener noreferrer"})
				if webSet!=None:
					homeUrl = webSet["href"]
			pInfo = {
				"pName": getNodeText(pLink),
				"webSet": homeUrl
			}
			products.append(pInfo)
	
	
excelFileName="company.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

i=1
while i < 10:
	url = "https://www.atwineries.com/wineries/united-states/?page="+str(i)
	getProductSope( url, products)
	i += 1


	


headers=["pName",'webSet']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
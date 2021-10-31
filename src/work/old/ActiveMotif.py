from urllib.request import urlopen
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json

http.client._MAXHEADERS = 1000


def urllib_download(IMAGE_URL, pName):
    from urllib.request import urlretrieve
    urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')   

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1


excelFileName="D:\\ActiveMotif.xlsx"
wb = Workbook()
workSheet = wb.active
url = "https://www.activemotif.com.cn/catalog/ajax/990?page=1&step=59"
productListHtml = urlopen(url).read()
decode_json = json.loads(productListHtml.decode(encoding='utf-8'))
headers=[
	'product Type','Product Name','Aliases','Figure','Application','Chemical Properties','Description','Contents','Storage Conditions','Tag 1','Tag 2','Tag 3'
];
index=1
for productInfo in decode_json["data"]:
	print(index)
	info={}
	linkHtml = BeautifulSoup(productInfo["attributes"]["html_name"], "html.parser",from_encoding="utf-8")
	Appl = productInfo["attributes"]["appl."]
	productUrl = "https://www.activemotif.com.cn"+linkHtml.find("a")["href"]
	productHtml = urlopen(productUrl).read()
	productHtmlSoup = BeautifulSoup(productHtml, "html.parser", from_encoding="utf-8")
	baseInfoNode = productHtmlSoup.find(name="div", attrs={"id":"title"})
	# productName
	productName = baseInfoNode.find("h1").get_text()
	aliaseNode = baseInfoNode.find(name="td", attrs={"class":"aliases"})
	# aliase
	aliase = aliaseNode.get_text() if(aliaseNode != None) else ""
	
	applicationNode = productHtmlSoup.find(name="tr", attrs={"valign":"top"})
	# applicationInfo
	applicationInfo=applicationNode.get_text()
	attrInfoNodes = productHtmlSoup.find_all(name="h3", attrs={"class":"attribute-title"})
	for attrNode in attrInfoNodes:
		if(attrNode.get_text().strip() == "Chemical Properties"):
			chemicalProperties = attrNode.next_sibling.next_sibling.get_text()
		if(attrNode.get_text().strip() == "Description"):
			description = attrNode.next_sibling.next_sibling.get_text()
		if(attrNode.get_text().strip() == "Contents"):
			contents = attrNode.next_sibling.next_sibling.get_text()
		if(attrNode.get_text().strip() == "Storage Conditions"):
			storageConditions = attrNode.next_sibling.next_sibling.get_text()
	
	
	mainContent = productHtmlSoup.find(name="div", attrs={"id":"main-content"})
	imgNodes = mainContent.find_all(name="img", attrs={"class":"catalog_image"})
	if(len(imgNodes) > 0):
		count=1
		for imgNode in imgNodes:
			imgSrc = "https://www.activemotif.com.cn"+imgNode["data-src"]
			urllib_download(imgSrc, productName+str(count))
			count+=1
	info["Product Name"]=productName
	info["Aliases"]=Appl
	info["Application"]=applicationInfo
	info["Chemical Properties"]=chemicalProperties
	info["Description"]=description
	info["Contents"]=contents
	info["Storage Conditions"]=storageConditions
	info["Tag 1"]="Epigenetics Small Molecules"
	info["Tag 2"]=applicationInfo
	info["Tag 3"]=Appl
	
	writeExcel(workSheet, headers, index, info)
	index=index+1

wb.save(excelFileName)
from urllib.request import urlopen
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter

http.client._MAXHEADERS = 1000

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head.strip() in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head.strip()].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1


def getProductInfo(productInfoHtml, index, workSheet, typeNum):
	headers=[
		'Product Name','Gene Synonyms','Description','This product includes the following','Parental Cell Line','Adherent Suspension',
		'Gene ID NCBI','Genomic Location','Genome Assembly','Organism','Storage','HGNC Symbol','Expression Level','The TPM expression value indicates this gene is',
		'Guide RNA Sequence','Targeted Transcript','Targeted Region','Mutation','Transcript Plot Name','PCR FWD','PCR BWD','Sequencing Primer','Sequencing Result',
		'Genotype','Biosafety Level','Media Type','Freeze Medium','Revival','Growth Properties','tag1','tag2','tag3','tag4'
	];
	info={}
	productscope = BeautifulSoup(productInfoHtml, "html.parser", from_encoding="utf-8");
	productspec = productscope.find(name="div", attrs={"class":"product-spec"})
	if(productspec != None):
		description = productspec.find(name="div", attrs={"itemprop":"description"})
		descriptionStr = description.get_text() if(description!=None) else ""
	else:
		descriptionStr=""
	# 基本信息
	productNode = productscope.find(name="h1", attrs={"itemprop":"name"})
	productName = productNode.get_text() if productNode!=None else ""
	if(len(productName) > 0):
		synonymsInfoNode = productspec.find(name="p")
		synonymsInfo = (synonymsInfoNode.get_text() if(synonymsInfoNode != None) else "").split("Gene Synonyms")
		descriptionInfo = descriptionStr.split("This product includes the following:")[0]
		includesFollowing = descriptionStr.split("This product includes the following:")
		info['Product Name']=productName
		info['Gene Synonyms']=synonymsInfo[1] if(len(synonymsInfo) > 1) else synonymsInfo[0]
		info['Description']=descriptionInfo
		info['This product includes the following']=includesFollowing[1] if(len(includesFollowing) > 1) else ""
		tag1="Epigenetics Cell Lines" if(typeNum=="1") else "Deubiquitination Cell Lines" if(typeNum=="2") else "DNA Damage Cell Lines" if(typeNum=="3") else "Epigenetics Cell Lines" if(typeNum=="4") else "Epigenetics Cell Lines"
		info['tag1']=tag1
		info['tag2']="Genetically Modified Cell Lines"
		info['tag3']=productName.split("knockout")[0].split("Human")[1]
		print(typeNum=="2")
		info['tag4']="" if(typeNum=="1") else "" if (typeNum=="2") else "" if(typeNum=="3") else "Bromodomain" if(typeNum=="4") else "Histone Acetylation"
		#Technical Information
		technicalNode = productscope.find(name="div", attrs={"class":"product-collateral"})
		infoTypes = technicalNode.find_all(name="dl",attrs={"class":"spec-list"})
		for infoType in infoTypes:
			typeName = infoType.find("dt").get_text().strip()
			typeValue = infoType.find("dd").get_text()
			info[typeName]=typeValue
		writeExcel(workSheet, headers, index, info)


fileName="D:\\list.txt"
excelFileName="D:\\products.xlsx"
wb = Workbook()
workSheet = wb.active
with open(fileName,'r') as file_to_read:
	index = 1
	type=1
	while True:
		print(index)
		lines = file_to_read.readline()
		if not lines:
			break
			pass
		if(lines.find("type=")==0):
			type=lines.split("type=")[1].strip()
		else:
			if(lines.strip().endswith("have:2")):
				productUrl = "https://www.horizondiscovery.com/catalogsearch/result/?q="+lines.split('===')[0]+"&cat=15"
				print(productUrl)
				productInfoHtml = urlopen(productUrl).read()
				getProductInfo(productInfoHtml, index, workSheet, type)
			else:
				productUrl = lines.split("========")
				if(len(productUrl) == 2):
					productInfoHtml = urlopen(productUrl[1]).read()
					getProductInfo(productInfoHtml, index, workSheet, type)
		index=index+1
wb.save(excelFileName)
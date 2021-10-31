from urllib.request import urlopen
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy
import math
from bs4.element import NavigableString 

http.client._MAXHEADERS = 1000


def getNodeText(node):
	if(node == None):
		return ""
	else:
		if isinstance(node, NavigableString):
			return node
		else:
			return node.get_text()

def urllib_download(IMAGE_URL, imageName):
	try:
		from urllib.request import urlretrieve
		urlretrieve(IMAGE_URL, imageName)   
	except:
		print("retry"+IMAGE_URL)
		urllib_download(IMAGE_URL, imageName)
		
retryCount = 0
loadCount = 0
def getHtmlFromUrl(url):
	global retryCount
	try:
		html = urlopen(url).read()
		return html
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getHtmlFromUrl(url)
		else:
			retryCount=0
			return None

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1

def getProductObj(url, products):
	global loadCount
	loadCount += 1
	print(str(loadCount)+"----"+url)
	pInfo = {}
	pHtml = getHtmlFromUrl(url)
	if pHtml != None:
		sope = BeautifulSoup(pHtml, "html.parser",from_encoding="utf-8")
		cName = sope.find(name="div", attrs={"class":"logo"})
		add = sope.find(name="div", attrs={"class":"add"})
		startTimeIcon = sope.find(name="img", attrs={"src":"http://resource.156zs.com/www/mrfc.png"})
		startTime = startTimeIcon.nextSibling.nextSibling if startTimeIcon!=None else None
		title = sope.find(name="div", attrs={"class":"title"})
		start = sope.find(name="div", attrs={"class":"start line_info"})
		end = sope.find(name="div", attrs={"class":"end line_info"})
		cAddr = ""
		cPhone = ""
		cResponser = ""
		startPoint = ""
		endPoint = ""
		startStationAddr = ""
		startStationPhone = ""
		startStationConator = ""
		endStationAddr = ""
		endStationPhone = ""
		endStationConator = ""
		for companyInfo in add.children:
			infoValue = getNodeText(companyInfo).replace(" ","").replace("\n","")
			if "地址" in infoValue:
				cAddr = infoValue
			if "电话" in infoValue:
				cPhone = infoValue
			if "负责人" in infoValue:
				cResponser = infoValue
		linePoints = title.find_all(name= "h2")
		for inx,point in enumerate(linePoints):
			if inx < len(linePoints) - 1:
				startPoint += getNodeText(point)+","
			else:
				endPoint = getNodeText(point)
		
		
		startStationInfo = start.find(name="div",attrs={"class": "address"}).find_all(name= "div");
		for inx,stationInfo in enumerate(startStationInfo):
			stationInfoValue = getNodeText(stationInfo).replace("\n","")
			if inx == 0:
				startStationAddr = stationInfoValue.replace(" ","")
			if "电话" in stationInfoValue:
				startStationPhone = stationInfoValue
			if "联系人" in stationInfoValue:
				startStationConator = stationInfoValue
		
		endStationInfo = end.find(name="div",attrs={"class": "address"}).find_all(name= "div");
		for inx,stationInfo in enumerate(endStationInfo):
			stationInfoValue = getNodeText(stationInfo).replace("\n","")
			if inx == 0:
				endStationAddr = stationInfoValue.replace(" ","")
			if "电话" in stationInfoValue:
				endStationPhone = stationInfoValue
			if "联系人" in stationInfoValue:
				endStationConator = stationInfoValue
		
		products.append({
			"cName":getNodeText(cName.find(name="strong")),                    # 公司名称                
			"startTime": getNodeText(startTime),   # 发车时间
			"cAddr":cAddr,                                                     # 公司地址
			"cPhone":cPhone,                                                   # 公司电话
			"cResponser":cResponser,                                           # 公司负责人
			"startPoint": startPoint,                                          #始发点
			"endPoint": endPoint,                                              #到达点		
			"startStationAddr":startStationAddr,                               # 始发站 地址
			"startStationPhone": startStationPhone,                            # 始发站  电话
			"startStationConator": startStationConator,                        # 始发站 联系人
			"endStationAddr": endStationAddr,                                  # 到达站 地址
			"endStationPhone": endStationPhone,                                # 到达站  电话
			"endStationConator": endStationConator                             # 到达站 联系人
		})
		if loadCount % 1000 == 0:
			saveExcel(products)


def getProductSope( url, products):
	productListHtml = getHtmlFromUrl(url)
	if productListHtml != None:
		sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
		if sope != None:
			links = sope.find_all(name="li", attrs={"class":"-info-"})
			for link in links:
				url = link.find(name="a")
				getProductObj(url["href"], products)
	
def saveExcel(products):
	excelFileName="company.xlsx"
	wb = Workbook()
	workSheet = wb.active
	headers=["cName",'startTime','cAddr','cPhone','cResponser','startPoint',"endPoint","startStationAddr","startStationPhone","startStationConator","endStationAddr","endStationPhone","endStationConator"]
	rindex = 1
	for p in products:
		writeExcel(workSheet, headers, rindex, p)
		rindex = rindex+1
	wb.save(excelFileName)

products = []
provances = json.loads('[{"id":23,"name":"四川","city":[{"id":235,"name":"成都","district":[{"id":2040,"name":"锦江区"},{"id":2041,"name":"青羊区"},{"id":2042,"name":"金牛区"},{"id":2043,"name":"武侯区"},{"id":2044,"name":"成华区"},{"id":2045,"name":"龙泉驿区"},{"id":2046,"name":"青白江区"},{"id":2047,"name":"新都区"},{"id":2048,"name":"温江区"},{"id":2049,"name":"金堂"},{"id":2050,"name":"双流"},{"id":2051,"name":"郫县"},{"id":2052,"name":"大邑"},{"id":2053,"name":"蒲江"},{"id":2054,"name":"新津"},{"id":2055,"name":"都江堰"},{"id":2056,"name":"彭州"},{"id":2057,"name":"邛崃"},{"id":2058,"name":"崇州"}]},{"id":236,"name":"自贡","district":[{"id":2059,"name":"自流井区"},{"id":2060,"name":"贡井区"},{"id":2061,"name":"大安区"},{"id":2062,"name":"沿滩区"},{"id":2063,"name":"荣县"},{"id":2064,"name":"富顺"}]},{"id":237,"name":"攀枝花","district":[{"id":2065,"name":"东区"},{"id":2066,"name":"西区"},{"id":2067,"name":"仁和区"},{"id":2068,"name":"米易"},{"id":2069,"name":"盐边"}]},{"id":238,"name":"泸州","district":[{"id":2070,"name":"江阳区"},{"id":2071,"name":"纳溪区"},{"id":2072,"name":"龙马潭区"},{"id":2073,"name":"泸县"},{"id":2074,"name":"合江"},{"id":2075,"name":"叙永"},{"id":2076,"name":"古蔺"}]},{"id":239,"name":"德阳","district":[{"id":2077,"name":"旌阳区"},{"id":2078,"name":"中江"},{"id":2079,"name":"罗江"},{"id":2080,"name":"广汉"},{"id":2081,"name":"什邡"},{"id":2082,"name":"绵竹"}]},{"id":240,"name":"绵阳","district":[{"id":2083,"name":"涪城区"},{"id":2084,"name":"游仙区"},{"id":2085,"name":"三台"},{"id":2086,"name":"盐亭"},{"id":2087,"name":"安县"},{"id":2088,"name":"梓潼"},{"id":2089,"name":"北川"},{"id":2090,"name":"平武"},{"id":2091,"name":"江油"}]},{"id":241,"name":"广元","district":[{"id":2092,"name":"市中区"},{"id":2093,"name":"元坝区"},{"id":2094,"name":"朝天区"},{"id":2095,"name":"旺苍"},{"id":2096,"name":"青川"},{"id":2097,"name":"剑阁"},{"id":2098,"name":"苍溪"}]},{"id":242,"name":"遂宁","district":[{"id":2099,"name":"船山区"},{"id":2100,"name":"安居区"},{"id":2101,"name":"蓬溪"},{"id":2102,"name":"射洪"},{"id":2103,"name":"大英"}]},{"id":243,"name":"内江","district":[{"id":2104,"name":"市中区"},{"id":2105,"name":"东兴区"},{"id":2106,"name":"威远"},{"id":2107,"name":"资中"},{"id":2108,"name":"隆昌"}]},{"id":244,"name":"乐山","district":[{"id":2109,"name":"市中区"},{"id":2110,"name":"沙湾区"},{"id":2111,"name":"五通桥区"},{"id":2112,"name":"金口河区"},{"id":2113,"name":"犍为"},{"id":2114,"name":"井研"},{"id":2115,"name":"夹江"},{"id":2116,"name":"沐川"},{"id":2117,"name":"峨边"},{"id":2118,"name":"马边"},{"id":2119,"name":"峨眉山"}]},{"id":245,"name":"南充","district":[{"id":2120,"name":"顺庆区"},{"id":2121,"name":"高坪区"},{"id":2122,"name":"嘉陵区"},{"id":2123,"name":"南部"},{"id":2124,"name":"营山"},{"id":2125,"name":"蓬安"},{"id":2126,"name":"仪陇"},{"id":2127,"name":"西充"},{"id":2128,"name":"阆中"}]},{"id":246,"name":"眉山","district":[{"id":2129,"name":"东坡区"},{"id":2130,"name":"仁寿"},{"id":2131,"name":"彭山"},{"id":2132,"name":"洪雅"},{"id":2133,"name":"丹棱"},{"id":2134,"name":"青神"}]},{"id":247,"name":"宜宾","district":[{"id":2135,"name":"翠屏区"},{"id":2136,"name":"宜宾"},{"id":2137,"name":"南溪"},{"id":2138,"name":"江安"},{"id":2139,"name":"长宁"},{"id":2140,"name":"高县"},{"id":2141,"name":"珙县"},{"id":2142,"name":"筠连"},{"id":2143,"name":"兴文"},{"id":2144,"name":"屏山"}]},{"id":248,"name":"广安","district":[{"id":2145,"name":"广安区"},{"id":2146,"name":"岳池"},{"id":2147,"name":"武胜"},{"id":2148,"name":"邻水"},{"id":2149,"name":"华蓥"}]},{"id":249,"name":"达州","district":[{"id":2150,"name":"通川区"},{"id":2151,"name":"达县"},{"id":2152,"name":"宣汉"},{"id":2153,"name":"开江"},{"id":2154,"name":"大竹"},{"id":2155,"name":"渠县"},{"id":2156,"name":"万源"}]},{"id":250,"name":"雅安","district":[{"id":2157,"name":"雨城区"},{"id":2158,"name":"名山"},{"id":2159,"name":"荥经"},{"id":2160,"name":"汉源"},{"id":2161,"name":"石棉"},{"id":2162,"name":"天全"},{"id":2163,"name":"芦山"},{"id":2164,"name":"宝兴"}]},{"id":251,"name":"巴中","district":[{"id":2165,"name":"巴州区"},{"id":2166,"name":"通江"},{"id":2167,"name":"南江"},{"id":2168,"name":"平昌"}]},{"id":252,"name":"资阳","district":[{"id":2169,"name":"雁江区"},{"id":2170,"name":"安岳"},{"id":2171,"name":"乐至"},{"id":2172,"name":"简阳"}]},{"id":253,"name":"阿坝","district":[{"id":2173,"name":"汶川"},{"id":2174,"name":"理县"},{"id":2175,"name":"茂县"},{"id":2176,"name":"松潘"},{"id":2177,"name":"九寨沟"},{"id":2178,"name":"金川"},{"id":2179,"name":"小金"},{"id":2180,"name":"黑水"},{"id":2181,"name":"马尔康"},{"id":2182,"name":"壤塘"},{"id":2183,"name":"阿坝县"},{"id":2184,"name":"若尔盖"},{"id":2185,"name":"红原"}]},{"id":254,"name":"甘孜","district":[{"id":2186,"name":"康定"},{"id":2187,"name":"泸定"},{"id":2188,"name":"丹巴"},{"id":2189,"name":"九龙"},{"id":2190,"name":"雅江"},{"id":2191,"name":"道孚"},{"id":2192,"name":"炉霍"},{"id":2193,"name":"甘孜"},{"id":2194,"name":"新龙"},{"id":2195,"name":"德格"},{"id":2196,"name":"白玉"},{"id":2197,"name":"石渠"},{"id":2198,"name":"色达"},{"id":2199,"name":"理塘"},{"id":2200,"name":"巴塘"},{"id":2201,"name":"乡城"},{"id":2202,"name":"稻城"},{"id":2203,"name":"得荣"}]},{"id":255,"name":"凉山","district":[{"id":2204,"name":"西昌"},{"id":2205,"name":"木里"},{"id":2206,"name":"盐源"},{"id":2207,"name":"德昌"},{"id":2208,"name":"会理"},{"id":2209,"name":"会东"},{"id":2210,"name":"宁南"},{"id":2211,"name":"普格"},{"id":2212,"name":"布拖"},{"id":2213,"name":"金阳"},{"id":2214,"name":"昭觉"},{"id":2215,"name":"喜德"},{"id":2216,"name":"冕宁"},{"id":2217,"name":"越西"},{"id":2218,"name":"甘洛"},{"id":2219,"name":"美姑"},{"id":2220,"name":"雷波"}]}]}]');
addressList = [] ;
for prov in provances:
	for city in prov["city"]:
		addressList.append({"id": str(city["id"]), "name": city["name"], "level": "1"});
		for add in city["district"]:
			addressList.append({"id": str(add["id"]), "name": add["name"], "level": "2"});
			
for startAdd in addressList:
	for endAdd in addressList:
		listUrl = "http://line.156zs.com/"+startAdd["level"]+"_"+startAdd["id"]+"/"+endAdd["level"]+"_"+endAdd["id"]+"/line.html"
		#读取列表
		getProductSope( listUrl, products)
		
saveExcel(products)
print("flish")	


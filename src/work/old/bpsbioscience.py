from urllib.request import urlopen
import urllib
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy

http.client._MAXHEADERS = 1000


def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		html = urlopen(url).read()
		return html
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getHtmlFromUrl(url)
		else:
			retryCount=0
			return None

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1


def getProductInfo(url, pInfo, products):

	print(str(len(products)) + url)
	productHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
	
	pInfo["link"] = url
	attrArea = sope.find(name="div", attrs={"class": "additional-attributes-wrapper"})
	attrInfos = attrArea.find_all(name="div", attrs={"class": "attributes"})
	for attr in attrInfos:
		attrLabel = getNodeText( attr.find("div", attrs={"class":"label"}))
		attrVal = getNodeText(attr.find("div", attrs={"class":"data"}))
		pInfo[attrLabel] = attrVal
	products.append(pInfo.copy())
				
	

def getProductList(url, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	
	listArea = sope.find_all(name="div", attrs={"class":"product-item-details" })
	for productItem in listArea:
		link = productItem.find("a", attrs={"class":"product-item-link"})
		pInfo = {}
		pInfo['name'] = getNodeText(link)
		pInfo['cat'] = getNodeText(productItem.find("div", attrs={"class":"product-sku"}))
		pInfo['size'] = getNodeText(productItem.find("div", attrs={"class":"product-size"}))
		pInfo['price'] = getNodeText(productItem.find("span", attrs={"class":"price"}))
		getProductInfo(link["href"], pInfo, products)

excelFileName="D:\\jcrb.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
i = 1

while i < 3:
	url = "https://bpsbioscience.com/product-types/cell-based-assay-kits/cell-signaling-pathway?p=" + str(i)
	getProductList(url, products)
	i += 1
headers=['link','cat','name','size','price','Description','Materials Required But Not Supplied','Applications','Supplied As','Storage/Stability','Shipping Temperature',
'Background','Synonym(s)','Concentration','Format','Instructions for Use','Notes','Warnings']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000


def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"}

		request_obj=urllib.request.Request(url=url,headers=headers)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getHtmlFromUrl(url)
		else:
			retryCount=0
			return None
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1


def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	productHtml = getHtmlFromUrl(url)
	if productHtml != None:
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
		
		pInfo["link"] = url
		description = sope.find(name="div", attrs={"class": "description"})
		cat = sope.find(name="span", attrs={"class": "sku_wrapper"})
		price = sope.find(name="p", attrs={"class": "price"})
		pInfo["description"] = getNodeText(description)
		pInfo["cat"] = getNodeText(cat)
		pInfo["price"] = getNodeText(price)
		
		keyfeatures = sope.find(name="div", attrs={"class": "checklist_sh_1 ck_sh"})
		pInfo["keyfeatures"] = getNodeText(keyfeatures)
		detailArea = sope.find(name="table", attrs={"class": "easy-table easy-table-default table-striped"})
		detailTrs = detailArea.find_all("tr")
		for detailTr in detailTrs:
			tds = detailTr.find_all("td")
			if len(tds) > 1:
				title = getNodeText(tds[0])
				value = getNodeText(tds[1])
				pInfo[title] = value
		
		howItWork = sope.find(name="div", attrs={"id": "tab-wootab_500"})
		pInfo["howItWork"] = getNodeText(howItWork)
		targetproteins = sope.find(name="div", attrs={"id": "tab-wootab_501"})
		pInfo["targetproteins"] = getNodeText(targetproteins)
		
		products.append(pInfo.copy())
				

def getProductList(url, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	productListArea = sope.find("aside", attrs={"id":"secondary"})
	if productListArea != None:
		linkAreas = productListArea.find_all("li")
		for linkArea in linkAreas:
			link = linkArea.find("a")
			pInfo={
				"name":getNodeText(link)
			}
			getProductInfo(link["href"], pInfo, products)


excelFileName="fullmoonbio.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
# getProductList('https://www.fullmoonbio.com/products/antibody-array/', products)
getProductInfo('https://www.fullmoonbio.com/product/cancer-biomarker-antibody-array/',{}, products)
headers=['link','name','description','cat','price','keyfeatures','Number of Antibodies:','Number of Replicates:','Reactivity:','Internal Controls:','Detection Method:',
'Slide dimensions:','Spot diameter:','Size:','Storage Condition:','howItWork','targetproteins']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
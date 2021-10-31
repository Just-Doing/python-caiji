from urllib.request import urlopen
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import math
from bs4.element import NavigableString 

http.client._MAXHEADERS = 1000


def getNodeText(node):
	if(node == None):
		return ""
	else:
		if isinstance(node, NavigableString):
			return node
		else:
			return node.get_text().strip()

def urllib_download(IMAGE_URL, imageName):
	try:
		from urllib.request import urlretrieve
		urlretrieve(IMAGE_URL, imageName)   
	except:
		print("retry"+IMAGE_URL)
		urllib_download(IMAGE_URL, imageName)
		
retryCount = 0
loadCount = 0
def getHtmlFromUrl(url):
	global retryCount
	try:
		html = urlopen(url).read()
		return html
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getHtmlFromUrl(url)
		else:
			retryCount=0
			return None

def getJsonFromUrl(url):
	global retryCount
	try:
		html = urlopen(url).read()
		return json.loads(html)
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getJsonFromUrl(url)
		else:
			retryCount=0
			return None

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = ILLEGAL_CHARACTERS_RE.sub(r'', info[head].strip())
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1

def getProductSope( url, products):
	productListHtml = BeautifulSoup( getHtmlFromUrl(url), "html.parser",from_encoding="utf-8")
	productList = productListHtml.find_all(name="div",attrs={"class":"list_title"})
	
	for product in productList:
		pLink = product.find("a")
		if pLink != None:
			infoUrl = "http://www.elisakits.cn/"+pLink["href"]
			print(str(len(products)) + "===" + infoUrl)
			productInfoHtml = BeautifulSoup( getHtmlFromUrl(infoUrl), "html.parser",from_encoding="utf-8")
			if productInfoHtml!= None:
				productTable = productInfoHtml.find(name="div", attrs={"class": "pro_intro_warp"})
				if productTable!= None:
					infoTrList = productTable.find_all(name="tr")
					pInfo = { }
					for tr in infoTrList:
						if getNodeText(tr.find("td")) == '目录号':
							pInfo["number"] = getNodeText(tr.find_all("td")[1])
						if getNodeText(tr.find("td")) == '细胞英文(简称）':
							pInfo["细胞英文"] = getNodeText(tr.find_all("td")[1])
						if getNodeText(tr.find("td")) == '细胞名称':
							pInfo["细胞名称"] = getNodeText(tr.find_all("td")[1])
						if getNodeText(tr.find("td")) == '背景资料':
							pInfo["背景资料"] = getNodeText(tr.find_all("td")[1])
						if getNodeText(tr.find("td")) == '细胞来源':
							pInfo["细胞来源"] = getNodeText(tr.find_all("td")[1])
						if getNodeText(tr.find("td")) == '代次':
							pInfo["代次"] = getNodeText(tr.find_all("td")[1])
						if getNodeText(tr.find("td")) == '规格':
							pInfo["规格"] = getNodeText(tr.find_all("td")[1])
						if getNodeText(tr.find("td")) == '细胞数':
							pInfo["细胞数"] = getNodeText(tr.find_all("td")[1])
						if getNodeText(tr.find("td")) == '价格':
							pInfo["价格"] = getNodeText(tr.find_all("td")[1])
						if getNodeText(tr.find("td")) == '生物安全级别':
							pInfo["生物安全级别"] = getNodeText(tr.find_all("td")[1])
						if getNodeText(tr.find("td")) == '组织来源':
							pInfo["组织来源"] = getNodeText(tr.find_all("td")[1])
						if getNodeText(tr.find("td")) == '细胞形态':
							pInfo["细胞形态"] = getNodeText(tr.find_all("td")[1])
						if getNodeText(tr.find("td")) == '细胞活力':
							pInfo["细胞活力"] = getNodeText(tr.find_all("td")[1])
						if getNodeText(tr.find("td")) == '细胞检测':
							pInfo["细胞检测"] = getNodeText(tr.find_all("td")[1])
						if getNodeText(tr.find("td")) == '培养条件':
							pInfo["培养条件"] = getNodeText(tr.find_all("td")[1])
						if getNodeText(tr.find("td")) == '传代方法':
							pInfo["传代方法"] = getNodeText(tr.find_all("td")[1])
						if getNodeText(tr.find("td")) == '冻存条件':
							pInfo["冻存条件"] = getNodeText(tr.find_all("td")[1])
				
					products.append(pInfo)
	
	
excelFileName="company.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

i=1
while i < 38:
	url = "http://www.elisakits.cn/Index/product/ccid/147/p/"+str(i)+".html"
	getProductSope( url, products)
	i += 1


	


headers=["number",'细胞英文','细胞名称','背景资料','细胞来源','代次','规格','细胞数','价格','生物安全级别','组织来源','细胞形态','细胞活力','细胞检测','培养条件','传代方法','冻存条件']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
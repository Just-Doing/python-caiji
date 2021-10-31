from urllib.request import urlopen
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy
import math
from bs4.element import NavigableString 

http.client._MAXHEADERS = 1000


def getNodeText(node):
	if(node == None):
		return ""
	else:
		if isinstance(node, NavigableString):
			return node
		else:
			return node.get_text()

def urllib_download(IMAGE_URL, imageName):
	try:
		from urllib.request import urlretrieve
		urlretrieve(IMAGE_URL, imageName)   
	except:
		print("retry"+IMAGE_URL)
		urllib_download(IMAGE_URL, imageName)
		
retryCount = 0
loadCount = 0
def getHtmlFromUrl(url):
	global retryCount
	try:
		html = urlopen(url).read()
		return html
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getHtmlFromUrl(url)
		else:
			retryCount=0
			return None

def getJsonFromUrl(url):
	global retryCount
	try:
		html = urlopen(url).read()
		return json.loads(html)
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getJsonFromUrl(url)
		else:
			retryCount=0
			return None

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1

def getProductSope( url, products):
	productList = getJsonFromUrl(url)
	for product in productList["response"]["docs"]:
		priceUrl = "https://www.caymanchem.com/solr/cchProductVariant/select?q=catalogNum:("+product["catalogNum"]+")&wt=json&rows=100000&sort=amount%20asc"
		priceInfo = getJsonFromUrl(priceUrl)
		priceValue = ""
		if priceInfo!= None:
			for price in priceInfo["response"]["docs"]:
				priceValue += "$"+str(price["amount"]) + "/" + price["name"]+","
			pInfo = {
				"pName": product["alphaNameSort"],
				"catalogNum": product["catalogNum"],
				"synonyms": "\n".join(product["synonyms"]) if "synonyms" in product else "",
				"purity": product["purity"] if "purity" in product else "",
				"endotoxinTesting": product["endotoxinTesting"] if "endotoxinTesting" in product else "",
				"source": product["source"] if "source" in product else "",
				"aminoAcids": product["aminoAcids"] if "aminoAcids" in product else "",
				"mw": product["mw"] if "mw" in product else "",
				"formulation": product["formulation"] if "formulation" in product else "",
				"uniProtAccessionNumber": product["uniProtAccessionNumber"] if "uniProtAccessionNumber" in product else "",
				"storageTemp": product["storageTemp"] if "storageTemp" in product else "",
				"Stability": product["stabilityDisplay"] if "stabilityDisplay" in product else "",
				"price": priceValue,
			}
			products.append(pInfo)
	
	
excelFileName="company.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

i=0
while i < 7:
	url = "https://www.caymanchem.com/solr/cchProduct/select?q=*:*&qf=catalogNum^2000%20exactname^5000%20exactSynonyms^4000%20edgename^4000%20synonymsPlain^2000%20formalNameDelimited^1500%20vendorItemNumber^4000%20casNumber^10000%20name^1500%20ngram_name^1000%20delimited_name^1500%20tagline^0.01%20keyInformation^0.01%20keywords^200%20inchi^20000%20inchiKey^20000%20smiles^20000%20ngram_synonym^400%20ngram_general^0.01&rows=25&defType=edismax&q.op=AND&enableElevation=true&bq=&facet=true&facet.field=newProduct&facet.field=raptas&facet.limit=100000&facet.mincount=1&wt=json&fq=(websiteNotSearchable:false%20AND%20europeOnly:false%20AND%20%20!raptas:RAP000101)%20AND%20(raptas:RAP000081)&start="+str((i*25))+"&bust=jg90a16lrs&version=2.2&sort=activationDate%20desc"
	getProductSope( url, products)
	i += 1


	


headers=["pName",'catalogNum','synonyms','purity','endotoxinTesting','source',"aminoAcids","mw","formulation","uniProtAccessionNumber","Stability","storageTemp","price"]
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
from urllib.request import urlopen
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy
import math
from bs4.element import NavigableString 

http.client._MAXHEADERS = 1000


def getNodeText(node):
	if(node == None):
		return ""
	else:
		if isinstance(node, NavigableString):
			return node
		else:
			return node.get_text()

def urllib_download(IMAGE_URL, imageName):
	try:
		from urllib.request import urlretrieve
		urlretrieve(IMAGE_URL, imageName)   
	except:
		print("retry"+IMAGE_URL)
		urllib_download(IMAGE_URL, imageName)
		
retryCount = 0
loadCount = 0
def getHtmlFromUrl(url):
	global retryCount
	try:
		html = urlopen(url).read()
		return html
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getHtmlFromUrl(url)
		else:
			retryCount=0
			return None

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1

def getProductObj(url, products):
	global loadCount
	loadCount += 1
	print(str(loadCount)+"----"+url)
	pInfo = {}
	pHtml = getHtmlFromUrl(url)
	if pHtml != None:
		sope = BeautifulSoup(pHtml, "html.parser",from_encoding="utf-8")
		nameSope = sope.find("h1",attrs={"class":"txtblue detailtitle"})
		streetSope = nameSope.nextSibling.nextSibling
		addressSope = streetSope.nextSibling.nextSibling
		if addressSope.nextSibling!= None:
			phoneSope = addressSope.nextSibling.nextSibling
			if getNodeText(phoneSope) == "Phone:" or getNodeText(phoneSope) == "Toll Free Number:":
				pInfo['phoneValue'] = phoneSope.nextSibling
			else:
				pInfo['addressValue'] = getNodeText(addressSope) +","+ getNodeText(phoneSope)
		pInfo['pName'] = getNodeText(nameSope)
		pInfo['street'] = getNodeText(addressSope)
		
		infos = sope.findAll("td", attrs={"class": "compname"})
		for info in infos:
			if(getNodeText(info) == 'Email:'):
				pInfo['email'] =getNodeText(info.nextSibling.nextSibling)
			if(getNodeText(info) == 'Website:'):
				pInfo['url'] =getNodeText(info.nextSibling.nextSibling)
		products.append(pInfo)
	else:
		pInfo['pName'] = 'can not read product page！！'
		pInfo['url'] = url
		products.append(pInfo)


def getProductSope( url, products):
	productListHtml = getHtmlFromUrl(url)
	if productListHtml != None:
		sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
		if sope != None:
			links = sope.find_all(name="a")
			for link in links:
				getProductObj("http://www.medicalproductguide.com"+link["href"], products)
	
	
excelFileName="company.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
##  896    157  488    31
i1=0
while i1 < math.ceil(896/10):
	url = "http://www.medicalproductguide.com/search/ajax_search?op=company&start=0&q=ivd&ctr=0&pid=0&manufacturer=0&distributor=1&importer=0&exporter=0&servicer=0&sells_directly=0&sells_thru_distrib=0&start="+str(i1*10)
	getProductSope( url, products)
	i1 += 1

i2=0
while i2 < math.ceil(157/10):
	url = "http://www.medicalproductguide.com/search/ajax_search?op=company&start=0&q=test%20kit&ctr=0&pid=0&manufacturer=0&distributor=1&importer=0&exporter=0&servicer=0&sells_directly=0&sells_thru_distrib=0&start="+str(i2*10)
	getProductSope( url, products)
	i2 += 1

i3=0
while i3 < math.ceil(488/10):
	url = "http://www.medicalproductguide.com/search/ajax_search?op=company&start=0&q=Immunoassay&ctr=0&pid=0&manufacturer=0&distributor=1&importer=0&exporter=0&servicer=0&sells_directly=0&sells_thru_distrib=0&start="+str(i3*10)
	getProductSope( url, products)
	i3 += 1

i4=0
while i4 < math.ceil(31/10):
	url = "http://www.medicalproductguide.com/search/ajax_search?op=company&start=0&q=virus%20assay&ctr=0&pid=0&manufacturer=0&distributor=1&importer=0&exporter=0&servicer=0&sells_directly=0&sells_thru_distrib=0&start="+str(i4*10)
	getProductSope( url, products)
	i4 += 1
	


headers=["pName",'street','addressValue','phoneValue','email','url']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
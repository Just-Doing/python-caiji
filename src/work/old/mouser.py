from urllib.request import urlopen
import urllib
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy

http.client._MAXHEADERS = 1000


def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url):
	global retryCount
	try:
		header_selfdefine={
			'cookie': 'CARTCOOKIEUUID=7c8a7bd3-ccd0-43a8-8c0d-37df2dc66b30; __neoui=1f701bbc-b60d-45e6-b8bc-4ecc2599240c; _ga=GA1.2.1928951230.1598065637; _gcl_au=1.1.664548637.1598065639; D_ZID=B1EC3CA0-4C4C-327B-AFA1-5BF88CAD7E9C; D_ZUID=C43F7323-A905-38A1-B61E-CC9A6543B65E; D_HID=5E200188-EF24-392D-AF9F-EA060EB22010; D_SID=118.113.201.17:OR3C2puhH1VLClDig8XFC3W6cp5hXbfp7TUIY50NYMk; _cs_c=1; apt_api=7ad42e752a9f9447777a9abf960e0e70; QSI_SI_1ADCODbKz4tsBBH_intercept=true; LPVID=c1NzAzZjk0Y2JhZmE1Mzg4; preferences=ps=www&pl=en-US&pc_www=USDu; _abck=0B6828596C8606CDC68DCAC101DCEE40~0~YAAQhykRApcWY+VzAQAAF4S0JQSi+n2SpOtTpmflmTFj/THaII6vG1GxZRKQr0sZhqiytejQoniN54VaMSSFB9fjUopMVzNGW3FPhj6Hxx+x0sFO/uRzPz7fUVQhmiQT8LNih/uiN2M7nL4suPr6QhSzENCcwfDvFdFU8bAtygfJK49BYRHhjwwQxToS2Zqmap/YYzrF4ci7CSgH8ZGzUlSsI5oCf68iPwobCi8Ng+uMTeiduH5w5o35FLIbjNUrgOV7peZeL5nM6rnLNW4al3qgPo/6Bivg0IkS5sZhYQpmsroery6R+QQkpHlDtAXX39Iq5fdOEw==~-1~-1~-1; __atuvc=5%7C35; ASP.NET_SessionId=yle5v3omjyvuewk0obf1l3li; __RequestVerificationToken=CsndURKT3QO2bXZ_oBjCllIVSc-I8B7k84Q0-ZCdWSe10iMpXjxplUCclvtyvMaAhsWG9If3gso4RIedWpeA5DhESyg1; akacd_Default_PR=3775904897~rv=91~id=12c455417c4af636082997c86b0920ee; bm_sz=A0A5A46744FEB43E83E3839984AA8DD5~YAAQO0sfRT2y0upzAQAA3mkrKwh8LkwKO9qbL7dW3Hs+1umA46JGiW4LviAL8JUyDPgRse3p0gfddRwG73F6aUoaRLyf0R722Y95wHqVzXfxLwNoHCyuQ36Lv/+peo9u34RgAHfXxWcYzoqs1nL22VBDse7ubK6hNbXlowXc4NAJJevyirPdwN8snsjfDkI=; _gid=GA1.2.816257705.1598452104; bm_mi=57A6A56471CEB3ADF448656809BCAEC0~n2RtawbUKGAK6NoC6EvCgM7jaVtkZ9xNdiDpShy1wFR+xc9EZ5UwUuRiYCQFOMvnxIekJ4YWX4e3uKJyr08he44/p75NKC1NLv+/CPZcWyWYk2/+pdf2Ro7qtjhPD4FuzK1fupFFINCinSsUXLa8iRS4x8sG1F0dlxbG6errSM2TPk4BC/lAaIx8c28pCW56yymYqarkAO8ShjrCEZXTM8LcZFjumatwZxT1S7phusTjD3f3yyZpWSoI9eWg3dULvcpbno46MSxv1SFi832Fkg==; _cs_mk=0.34714178515010197_1598452110090; TLS_SUPPORTED=1.2; ak_bmsc=5F137107DB81948E1A401DAACF90D81D451F4B3B285100008371465FB527C141~plwK0AnLJ4D08qBUm8YM4lcwWzsyZGbm/hvfCdR31tL+hsWxJzfZEg9F3Tzqq5LOc+NlPlFNeInmxi7iK1IFFRVBkjKgOLGhENkhikC3hVsDa0CSlocmtusIa/Fna8ms0E2ULD7O5nxQWlCFIDYG2HbB71zx05uDLDTXtw1pZ37fncJ/PtsQxpGyyz23U5f2du6A62dMfWunVbELMJ6qkYtVTYD1/rKw+xa/aSLcmW/RC72f55C/O972TniHTj2jtL5Bru3OoB16b4FX5NheunWArzIX5dWgmndGrPTGb4pg/CGN39T90l7aoQ1IZxagPP6TNcfzF6vN1tWkDvVLzb2w==; QSI_HistorySession=https%3A%2F%2Fwww.mouser.com%2FElectronic-Components%2F~1598452543620; _gat=1; D_IID=C9ABAE03-CC5B-3FE4-826F-BA70E5E40E7E; D_UID=1F84E2BF-2EC0-3ABD-AF3F-61EC7FDA14C8; _cs_id=ed0b8ab7-eb15-a838-e44f-b151bda55a96.1598065641.4.1598453296.1598452111.1.1632229641139.Lax.0; _cs_s=2.3; RT="z=1&dm=mouser.com&si=d29f16a3-9e6b-469f-b945-0c94132d9f55&ss=kebh3tun&sl=2&tt=95wy&bcn=%2F%2F173e255e.akstat.io%2F&ld=puv9&r=089307d06c7ce54cdbef3fb9eaed372f&ul=puva"',
			'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.135 Safari/537.36',
			'Content-Type': 'text/html'
		}
		request_obj = urllib.request.Request(url=url, headers=header_selfdefine,method='GET')
		response_obj = urllib.request.urlopen(request_obj)
		html_code=response_obj.read().decode('utf-8')
		
		return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getHtmlFromUrl(url)
		else:
			retryCount=0
			return None

def getHtmlFromUrl1(url):
	global retryCount
	try:
		header_selfdefine={
			'cookie': 'CARTCOOKIEUUID=7c8a7bd3-ccd0-43a8-8c0d-37df2dc66b30; __neoui=1f701bbc-b60d-45e6-b8bc-4ecc2599240c; _ga=GA1.2.1928951230.1598065637; _gcl_au=1.1.664548637.1598065639; D_ZID=B1EC3CA0-4C4C-327B-AFA1-5BF88CAD7E9C; D_ZUID=C43F7323-A905-38A1-B61E-CC9A6543B65E; D_HID=5E200188-EF24-392D-AF9F-EA060EB22010; D_SID=118.113.201.17:OR3C2puhH1VLClDig8XFC3W6cp5hXbfp7TUIY50NYMk; _cs_c=1; apt_api=7ad42e752a9f9447777a9abf960e0e70; QSI_SI_1ADCODbKz4tsBBH_intercept=true; LPVID=c1NzAzZjk0Y2JhZmE1Mzg4; preferences=ps=www&pl=en-US&pc_www=USDu; ASP.NET_SessionId=yle5v3omjyvuewk0obf1l3li; __RequestVerificationToken=CsndURKT3QO2bXZ_oBjCllIVSc-I8B7k84Q0-ZCdWSe10iMpXjxplUCclvtyvMaAhsWG9If3gso4RIedWpeA5DhESyg1; akacd_Default_PR=3775904897~rv=91~id=12c455417c4af636082997c86b0920ee; bm_sz=A0A5A46744FEB43E83E3839984AA8DD5~YAAQO0sfRT2y0upzAQAA3mkrKwh8LkwKO9qbL7dW3Hs+1umA46JGiW4LviAL8JUyDPgRse3p0gfddRwG73F6aUoaRLyf0R722Y95wHqVzXfxLwNoHCyuQ36Lv/+peo9u34RgAHfXxWcYzoqs1nL22VBDse7ubK6hNbXlowXc4NAJJevyirPdwN8snsjfDkI=; _gid=GA1.2.816257705.1598452104; bm_mi=57A6A56471CEB3ADF448656809BCAEC0~n2RtawbUKGAK6NoC6EvCgM7jaVtkZ9xNdiDpShy1wFR+xc9EZ5UwUuRiYCQFOMvnxIekJ4YWX4e3uKJyr08he44/p75NKC1NLv+/CPZcWyWYk2/+pdf2Ro7qtjhPD4FuzK1fupFFINCinSsUXLa8iRS4x8sG1F0dlxbG6errSM2TPk4BC/lAaIx8c28pCW56yymYqarkAO8ShjrCEZXTM8LcZFjumatwZxT1S7phusTjD3f3yyZpWSoI9eWg3dULvcpbno46MSxv1SFi832Fkg==; TLS_SUPPORTED=1.2; ak_bmsc=5F137107DB81948E1A401DAACF90D81D451F4B3B285100008371465FB527C141~plwK0AnLJ4D08qBUm8YM4lcwWzsyZGbm/hvfCdR31tL+hsWxJzfZEg9F3Tzqq5LOc+NlPlFNeInmxi7iK1IFFRVBkjKgOLGhENkhikC3hVsDa0CSlocmtusIa/Fna8ms0E2ULD7O5nxQWlCFIDYG2HbB71zx05uDLDTXtw1pZ37fncJ/PtsQxpGyyz23U5f2du6A62dMfWunVbELMJ6qkYtVTYD1/rKw+xa/aSLcmW/RC72f55C/O972TniHTj2jtL5Bru3OoB16b4FX5NheunWArzIX5dWgmndGrPTGb4pg/CGN39T90l7aoQ1IZxagPP6TNcfzF6vN1tWkDvVLzb2w==; D_IID=C9ABAE03-CC5B-3FE4-826F-BA70E5E40E7E; D_UID=1F84E2BF-2EC0-3ABD-AF3F-61EC7FDA14C8; LPSID-12757882=Sl887Fm3Qf-OSy8JLm-wTQ; _hjid=8afe5949-db9a-4e2d-9adb-e0a7ed2dd0ae; _hjAbsoluteSessionInProgress=1; __atuvc=6%7C35; __atuvs=5f467636bade8b1f000; _hjAbsoluteSessionInProgress=1; _cs_id=ed0b8ab7-eb15-a838-e44f-b151bda55a96.1598065641.4.1598454003.1598452111.1.1632229641139.Lax.0; _cs_s=4.3; _abck=0B6828596C8606CDC68DCAC101DCEE40~0~YAAQFn8lF9UaEepzAQAA9H5IKwRF13MQ288T0u/Y+HINWeODY70+Fg4jvYnEk+tNQFHxFW6bDmoPiz3fM7gVcGnNhozlDlyaSNzHv3x3/nohoucSgxXjYxJrzS7j1mkLys5jodVU+FZSXL2/I4VMLna0Hgs/1dO8zYxyXKTpuS/P6vvtZOvrKjtCgW93P6xaU0z6+KiVe6fd9v3hjcJRbLjjA6XATofR9uehpZP2xGv5zavjsuW9cPmDHuCH3FLzlsWc/DC2y7/P5b0HfQrTS6gruxxinY2ElxwSmqixorl+D9Kzft1FQKGvKZ21f4zfWiI/PDh+vQ==~-1~-1~-1; QSI_HistorySession=https%3A%2F%2Fwww.mouser.com%2FElectronic-Components%2F~1598452543620%7Chttps%3A%2F%2Fwww.mouser.com%2FSensors%2FOptical-Sensors%2F_%2FN-6g7q8~1598453316520%7Chttps%3A%2F%2Fwww.mouser.com%2FProductDetail%2FAmphenol-Advanced-Sensors%2FZTP-148SRC1%3Fqs%3DsGAEpiMZZMs0JOhy9PM0URuSbPQkQZeU5TJ8b%252FNVJewHvhBRP7KHvA%253D%253D~1598453339926%7Chttps%3A%2F%2Fwww.mouser.com%2FSensors%2FOptical-Sensors%2F_%2FN-6g7q8~1598454036551; RT="z=1&dm=mouser.com&si=d29f16a3-9e6b-469f-b945-0c94132d9f55&ss=kebh3tun&sl=5&tt=ablq&bcn=%2F%2F173e255e.akstat.io%2F&obo=1&ul=1fjow"',
			'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.135 Safari/537.36',
			'Content-Type': 'text/html'
		}
		request_obj = urllib.request.Request(url=url, headers=header_selfdefine,method='GET')
		response_obj = urllib.request.urlopen(request_obj)
		html_code=response_obj.read().decode('utf-8')
		
		return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getHtmlFromUrl(url)
		else:
			retryCount=0
			return None

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1


def getProductInfo(url, pInfo, products):

	print(str(len(products)) + url)
	sope = getHtmlFromUrl(url)
	Manufacture = sope.find(name="a", attrs={"id": "lnkManufacturerName"})
	Model = sope.find(name="span", attrs={"id": "spnMouserPartNumFormattedForProdInfo"})
	pInfo["name"] = getNodeText(Manufacture)  + " " + getNodeText(Model)
	pInfo["Manufacture"] = getNodeText(Manufacture)
	pInfo["Model"] =getNodeText(Model)
	
	ProductCompliance = sope.find(name="div", attrs={"id": "collapseProductCompliance"})
	pInfo["ProductComplianceValue"] = getNodeText(ProductCompliance)
	
	spnDescription = sope.find(name="span", attrs={"id": "spnDescription"})
	pInfo["Description"] = getNodeText(spnDescription)
	
	
	collapseProductSpecs = sope.find(name="div", attrs={"id": "collapseProductSpecs"})
	pInfo["Specifications"] = getNodeText(collapseProductSpecs)
	
	collapseAlsoBought = sope.find(name="div", attrs={"id": "collapseAlsoBought"})
	pInfo["CustomersAlsoBought"] = getNodeText(collapseAlsoBought).replace('\n', ' ').replace('      ', '\n').replace('\n6\n', '\n')
	
	detailfeaturedesc = sope.find(name="div", attrs={"id": "detail-feature-desc"})
	pInfo["MoreInformation"] = getNodeText(detailfeaturedesc)
	
	pdppricingtable = sope.find(name="div", attrs={"id": "pdp-pricing-table"})
	pInfo["Pricing"] = getNodeText(pdppricingtable)
	print(pInfo)
	products.append(pInfo.copy())
				
	

def getProductList(url, products, type1, type2):
	sope = getHtmlFromUrl1(url)
	
	listArea = sope.find_all(name="div", attrs={"class":"mfr-part-num " })
	for linkArea in listArea:
		link = linkArea.find(name="a")
		pInfo = {
			"type1": type1,
			"type2": type2,
			'url': "https://www.mouser.com/"+link["href"]
		}
		print(str(len(products)) + link["href"])
		products.append(pInfo.copy())
		# getProductInfo("https://www.mouser.com/"+link["href"], pInfo, products)



excelFileName="mouser.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

url = "https://www.mouser.com/Electronic-Components"
typeListHtml = getHtmlFromUrl(url)
sensorsLink = typeListHtml.find(name="a", attrs={"id": "lnkTopLvlCategory_254539_14"})
testMeasurementLink = typeListHtml.find(name="a", attrs={"id": "lnkTopLvlCategory_254559_15"})

sensorsChildTypeList = sensorsLink.parent.nextSibling.nextSibling.find_all(name="a");
testMeasurementChildTypeList = testMeasurementLink.parent.nextSibling.nextSibling.find_all(name="a");

for typeLink in sensorsChildTypeList:
	typeUrl = typeLink["href"]
	getProductList("https://www.mouser.com"+typeUrl, products, 'Sensors', getNodeText(typeLink))


for typeLink in testMeasurementChildTypeList:
	typeUrl = typeLink["href"]
	getProductList("https://www.mouser.com"+typeUrl, products, 'Test & Measurement', getNodeText(typeLink))



headers=['type1','type2','name','Manufacture','Model','ProductComplianceValue','Description','Specifications','CustomersAlsoBought','MoreInformation','Pricing']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
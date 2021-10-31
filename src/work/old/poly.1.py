from urllib.request import urlopen
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy

http.client._MAXHEADERS = 1000


def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

def urllib_download(IMAGE_URL, imageName):
	from urllib.request import urlretrieve
	fileName = imageName+".png"
	print(fileName)
	urlretrieve(IMAGE_URL, fileName)  

def getHtmlFromUrl(url):
	try:
		html = urlopen(url).read()
		return html
	except:
		print("重试"+url)
		getHtmlFromUrl(url)

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1

def getProductObj(url, pInfo, pType):
	pHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(pHtml, "html.parser",from_encoding="utf-8")
	nameSope = sope.find("h2",attrs={"class":"product-name"})
	pInfo["name"]=getNodeText(nameSope)
	pInfo["pType"]=pType
	imageArea = sope.find("img",attrs={"id":"image-main"})
	# urllib_download(imageArea["src"], pInfo["name"].replace("/","").replace("\n","").replace("\r","").replace(".","").replace(":","").replace("#","").replace("<","").replace(">",""))
	infoGroup = sope.find_all("div",attrs={"class":"box-collateral box-description"})
	for group in infoGroup:
		groupName = getNodeText(group.find("h2"))
		pinGroup = group.find_all("p")
		if groupName =="Description":
			description = ""
			for p in pinGroup:
				title = p.find("strong")
				if title == None:
					description=description+getNodeText(p)
				else:
					break
			pInfo["description"]=description
			refrence = group.find_all("p", attrs={"style":"font-size:11px;"})
			if len(refrence)>0:
				pInfo["refrence1"]=getNodeText(refrence[0])
			if len(refrence)>1:
				pInfo["refrence2"]=getNodeText(refrence[1])
			if len(refrence)>2:
				pInfo["refrence3"]=getNodeText(refrence[2])
			if len(refrence)>3:
				pInfo["refrence4"]=getNodeText(refrence[3])
			
	
	infoTitles = sope.find_all("strong")
	for infotitle in infoTitles:
		if getNodeText(infotitle)=="FTIR":
			pInfo["FTIR"]=infotitle.nextSibling
		if getNodeText(infotitle)=="Inherent Viscosity (ηinh)":
			pInfo["Inherent"]=infotitle.nextSibling
		if getNodeText(infotitle)=="Acid Number":
			pInfo["Acid"]=infotitle.nextSibling
		if getNodeText(infotitle)=="Lactide/Glycolide":
			pInfo["infotitle"]=infotitle.nextSibling
		if getNodeText(infotitle)=="Viscosity Molecular Weight (Mη)":
			pInfo["Viscosity"]=infotitle.nextSibling
		if getNodeText(infotitle)=="Glass Transition Temp. (Tg)":
			pInfo["Glass"]=infotitle.nextSibling
		if getNodeText(infotitle)=="Soluble in":
			pInfo["Soluble"]=infotitle.nextSibling
		if getNodeText(infotitle)=="Synonyms":
			pInfo["Synonyms"]=infotitle.nextSibling
	CasArea = sope.find("div",attrs={"class":"short-description"})
	if CasArea!=None:
		titles = CasArea.find_all("dt")
		for title in titles:
			if getNodeText(title)=="CAS#:":
				pInfo["CAS"]=getNodeText(title.nextSibling.nextSibling)
			if getNodeText(title)=="Hazards:":
				pInfo["Hazards"]=getNodeText(title.nextSibling.nextSibling)
			if getNodeText(title)=="Handling:":
				pInfo["Handling"]=getNodeText(title.nextSibling.nextSibling)
			if getNodeText(title)=="Storage:":
				pInfo["Storage"]=getNodeText(title.nextSibling.nextSibling)
	return pInfo

def getProductType(url, products):
	pHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(pHtml, "html.parser",from_encoding="utf-8")
	typeArea = sope.find(name="ul",attrs={"class":"menu2"})
	typeLinks = typeArea.find_all("a")
	for typeLink in typeLinks:
		pInfo = {}
		typeUrl= typeLink["href"]
		pType = getNodeText(typeLink)
		pListHtml = getHtmlFromUrl(typeUrl)
		listSope = BeautifulSoup(pListHtml, "html.parser",from_encoding="utf-8")
		pListArea = listSope.find("tbody", attrs={"class":"product-tbl"})
		if pListArea != None:
			pLinkList = pListArea.find_all("a")
			for pLink in pLinkList:
				if pLink != None:
					pUrl= pLink["href"]
					pInfo["plink"] = pUrl
					pInfo = getProductObj(pUrl, pInfo, pType)
					products.append(pInfo.copy())
					print(len(products))
		else:
			pListAreaOl = listSope.find("ol", attrs={"id":"products-list"})
			pLinkList = pListAreaOl.find_all("li")
			for pLinkArea in pLinkList:
				pLink = pLinkArea.find("a", attrs={"class":"product-image"})
				if pLink != None:
					pUrl= pLink["href"]
					pInfo["plink"] = pUrl
					pInfo = getProductObj(pUrl, pInfo, pType)
					products.append(pInfo.copy())
					print(len(products))

	
excelFileName="polysciences1.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
url = "https://www.polysciences.com/default/catalog-products/monomers-polymers/polymers/biodegradable-polymers/poly-dl-lactide-glycolide-polymers"

getProductType(url, products)
# pinfo = getProductObj("https://www.polysciences.com/default/26270", {}, "")
# print(pinfo)
headers=["pType",'name','description','Synonyms','refrence1','refrence2','refrence3','refrence4','FTIR','Inherent','Acid','Lactide','Viscosity','Glass','Soluble','CAS','Hazards','Handling','Storage','plink']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1

print("flish")	

wb.save(excelFileName)
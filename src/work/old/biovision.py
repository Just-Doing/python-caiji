from urllib.request import urlopen
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re

http.client._MAXHEADERS = 1000


def urllib_download(IMAGE_URL, imageName):
    from urllib.request import urlretrieve
    urlretrieve(IMAGE_URL, imageName)   

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1


excelFileName="D:\\biovision.xlsx"
wb = Workbook()
workSheet = wb.active
url = "https://www.biovision.com/products/epigenetics/epigenetic-modulators-a-z.html?limit=all"
productListHtml = urlopen(url).read()
headers=[
	'product Type','Product Name','Figure','target','effect','Size','Alternate Name','Appearance','CAS','Molecular Formula','Molecular Weight','Purity','Solubility','SMILES','InChi','InChi Key','PubChem CID','MDL Number',
	'Handling','StorageConditions','Shipping Conditions','USAGE','Description','Tag 1','Tag 2','Tag 3','tag4'
];
index=1
sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
links=sope.find_all(name="h2",attrs={"class":"product-name"})
for link in links:
	info={}
	productUrl = link.find("a")["href"]
	productHtml = urlopen(productUrl).read()
	productHtmlSoup = BeautifulSoup(productHtml, "html.parser", from_encoding="utf-8")
	# productName
	productName = productHtmlSoup.find(name="h1", attrs={"itemprop":"name"}).get_text()
	if(len(productName) > 0):
		TargetAndEffectStr = productHtmlSoup.find(name="div", attrs={"itemprop":"description"}).get_text()
		TargetAndEffect = TargetAndEffectStr.split(" ")
		if(len(TargetAndEffect) > 2):
			# effect
			effect=TargetAndEffect[len(TargetAndEffect)-1]
			reg=re.compile(r'[-,$()#+&*]')
			tesuStr = re.findall(reg, effect)
			if(len(tesuStr) > 0):
				info["tag4"]="1"
				effect = TargetAndEffectStr
				target = TargetAndEffectStr
			else:
				info["tag4"]="0"
				# target
				target = TargetAndEffectStr.replace(effect, "").replace(TargetAndEffect[0], "")
		else:
			# effect
			effect=TargetAndEffectStr
			info["tag4"]="1"
			# target
			target = TargetAndEffectStr
		print(TargetAndEffectStr)
		attrInfoTable = productHtmlSoup.find(name="table", attrs={"id":"product-attribute-specs-table"})
		for attrNode in attrInfoTable.find_all("tr"):
			if(attrNode.find("th").get_text().strip() == "Size"):
				Size = attrNode.find("td").get_text().strip()
			if(attrNode.find("th").get_text().strip() == "Alternate Name"):
				AlternateName = attrNode.find("td").get_text().strip()
			if(attrNode.find("th").get_text().strip() == "Appearance"):
				Appearance = attrNode.find("td").get_text().strip()
			if(attrNode.find("th").get_text().strip() == "CAS #"):
				CAS = attrNode.find("td").get_text().strip()
			if(attrNode.find("th").get_text().strip() == "Molecular Formula"):
				MolecularFormula = attrNode.find("td").get_text().strip()
			if(attrNode.find("th").get_text().strip() == "Molecular Weight"):
				MolecularWeight = attrNode.find("td").get_text().strip()
			if(attrNode.find("th").get_text().strip() == "Purity"):
				Purity = attrNode.find("td").get_text().strip()
			if(attrNode.find("th").get_text().strip() == "Solubility"):
				Solubility = attrNode.find("td").get_text().strip()
			if(attrNode.find("th").get_text().strip() == "SMILES"):
				SMILES = attrNode.find("td").get_text().strip()
			if(attrNode.find("th").get_text().strip() == "InChi"):
				InChi = attrNode.find("td").get_text().strip()
			if(attrNode.find("th").get_text().strip() == "InChi Key"):
				InChiKey = attrNode.find("td").get_text().strip()
			if(attrNode.find("th").get_text().strip() == "PubChem CID"):
				PubChemCID = attrNode.find("td").get_text().strip()
			if(attrNode.find("th").get_text().strip() == "MDL Number"):
				MDLNumber = attrNode.find("td").get_text().strip()
			if(attrNode.find("th").get_text().strip() == "Handling"):
				Handling = attrNode.find("td").get_text().strip()
			if(attrNode.find("th").get_text().strip() == "Storage Conditions"):
				StorageConditions = attrNode.find("td").get_text().strip()
			if(attrNode.find("th").get_text().strip() == "Shipping Conditions"):
				ShippingConditions = attrNode.find("td").get_text().strip()
			if(attrNode.find("th").get_text().strip() == "USAGE"):
				USAGE = attrNode.find("td").get_text().strip()
	
		imgNode = productHtmlSoup.find(name="a", attrs={"id":"zoom-btn"})
		imgSrc = imgNode["href"]
		imgName = productName.replace("/","").replace("\\","")+'-1.jpg'
		urllib_download(imgSrc, imgName)
		info["Figure"]=imgName
		info["target"]=target
		info["effect"]=effect
		mainContent = productHtmlSoup.find(name="div", attrs={"id":"extra_tabs_description_contents"})
		descriptionNodes = mainContent.find(name="div", attrs={"class":"std"})
		Description=descriptionNodes.get_text()
		info["Product Name"]=productName
		info["Size"]=Size
		info["Alternate Name"]=AlternateName
		info["Appearance"]=Appearance
		info["CAS"]=CAS
		info["Molecular Formula"]=MolecularFormula
		info["Molecular Weight"]=MolecularWeight
		info["Purity"]=Purity
		info["Solubility"]=Solubility
		info["SMILES"]=SMILES
		info["InChi"]=InChi
		info["InChi Key"]=InChiKey
		info["PubChem CID"]=PubChemCID
		info["MDL Number"]=MDLNumber
		info["Handling"]=Handling
		info["StorageConditions"]=StorageConditions
		info["Shipping Conditions"]=ShippingConditions
		info["USAGE"]=USAGE
		info["Description"]=Description
		info["Tag 1"]="Epigenetics Small Molecules"
		info["Tag 2"]=target
		info["Tag 3"]=effect
		writeExcel(workSheet, headers, index, info)
		index=index+1

wb.save(excelFileName)
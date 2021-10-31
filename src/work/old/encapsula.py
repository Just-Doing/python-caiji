from urllib.request import urlopen
import urllib
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy

http.client._MAXHEADERS = 1000


def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		html = urlopen(url).read()
		return html
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getHtmlFromUrl(url)
		else:
			retryCount=0
			return None

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1


def getProductInfo(url, pInfo, products):
	pInfo["link"] = url
	print(str(len(products)) + url)
	productHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
	
	productDec = sope.find(name="div", attrs={"class": "product_desc"})
	productDecProp = productDec.find_all("h3")
	for prop in productDecProp:
		if getNodeText(prop) == "Description":
			pInfo["description"] = getNodeText(prop.next_sibling.next_sibling)
		if getNodeText(prop) == "Appearance":
			pInfo["Appearance"] = getNodeText(prop.next_sibling.next_sibling)
	
	
	productStorgeProp = productDec.find_all("h4")
	for storgeProp in productStorgeProp:
		if getNodeText(storgeProp) == "Storage":
			pInfo["Storage"] = getNodeText(storgeProp.next_sibling.next_sibling)
		if getNodeText(storgeProp) == "Shelf Life":
			pInfo["ShelfLife"] = getNodeText(storgeProp.next_sibling.next_sibling)
			
	tablepressTables = sope.find_all(name="table", attrs={"class":"tablepress"})
	pInfo["LipidComposition"] = ""
	pInfo["BuffersandLiposome"] = ""
	for tab in tablepressTables:
		tbHeaderTitle = tab.find("th")
		if getNodeText(tbHeaderTitle).find("Lipid Composition") > -1 and pInfo["LipidComposition"] == "":
			tbody = tab.find("tbody").find_all("tr")
			for tr in tbody:
				tds = tr.find_all("td")
				if len(tds) == 4:
					title = tr.find("td")
					value = tds[3]
					pInfo["LipidComposition"] = pInfo["LipidComposition"] + getNodeText(title) + ":" + getNodeText(value) + "\n"
				
		if getNodeText(tbHeaderTitle).find("Buffers and Liposome") > -1 and pInfo["BuffersandLiposome"] == "":
			tbody = tab.find("tbody").find_all("tr")
			for tr in tbody:
				tds = tr.find_all("td")
				if tds != None and len(tds) > 1:
					title = tr.find("td")
					value = tds[1]
					pInfo["BuffersandLiposome"] = pInfo["BuffersandLiposome"] + getNodeText(title) + ":" + getNodeText(value) + "\n"
			
			
	products.append(pInfo.copy())


def getProductList(pInfo, scope, products):
	prodLink = scope.find_all("a")
	for link in prodLink:
		getProductInfo(link["href"], pInfo, products)


def getProductType(url, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	
	typeArea = sope.find(name="ul", attrs={"class":"level-1" })
	type1s = typeArea.find_all(name="ul", attrs={"class":"level-2"})
	for type1 in type1s:
		pInfo={}
		pInfo["type1"]= getNodeText(type1.previous_sibling)
		type2s = type1.find_all(name="ul", attrs={"class":"level-3"})
		if len(type2s) == 0:
			getProductList(pInfo, type1, products)
			
		for type2 in type2s:
			pInfo["type2"]= getNodeText(type2.previous_sibling)
			type3s = type2.find_all(name="ul", attrs={"class":"level-4"})
			if len(type3s) == 0:
				getProductList(pInfo, type2, products)
				
			for type3 in type3s:
				pInfo["type3"]= getNodeText(type3.previous_sibling)
				type4s = type3.find_all(name="ul", attrs={"class":"level-5"})
				if len(type4s) == 0:
					getProductList(pInfo, type3, products)
				
				for type4 in type4s:
					pInfo["type4"]= getNodeText(type4.previous_sibling)
					getProductList(pInfo, type4, products)

excelFileName="D:\\encapsula.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

urls = [
	'https://encapsula.com/product-category/liposomal-doxorubicin/',
	'https://encapsula.com/product-category/liposomes-for-dnarna-delivery/',
	'https://encapsula.com/product-category/fluorescent-liposomes/',
	'https://encapsula.com/product-category/lyophilized-atp-liposomes-atpsome/'
]

# urls = [
	# 'https://encapsula.com/product-category/liposomal-doxorubicin'
# ]
# getProductInfo("https://encapsula.com/products/surface-reactive-liposomes-immunosome/fluorescent-immunoliposomes-immunofluor/folate-fluorescent-liposomes/immunofluor-folate/", {}, [])

for url in urls:
	getProductType(url, products)
headers=['link','type1','type2','type3','type4','description','Appearance','Storage','ShelfLife','LipidComposition','BuffersandLiposome']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
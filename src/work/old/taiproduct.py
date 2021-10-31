from urllib.request import urlopen
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import math
from bs4.element import NavigableString 

http.client._MAXHEADERS = 1000


def getNodeText(node):
	if(node == None):
		return ""
	else:
		if isinstance(node, NavigableString):
			return node
		else:
			return node.get_text().strip()

def urllib_download(IMAGE_URL, imageName):
	try:
		from urllib.request import urlretrieve
		urlretrieve(IMAGE_URL, imageName)   
	except:
		print("retry"+IMAGE_URL)
		urllib_download(IMAGE_URL, imageName)
		
retryCount = 0
loadCount = 0
def getHtmlFromUrl(url):
	global retryCount
	try:
		html = urlopen(url).read()
		return html
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getHtmlFromUrl(url)
		else:
			retryCount=0
			return None

def getJsonFromUrl(url):
	global retryCount
	try:
		html = urlopen(url).read()
		return json.loads(html)
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getJsonFromUrl(url)
		else:
			retryCount=0
			return None

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = ILLEGAL_CHARACTERS_RE.sub(r'', info[head].strip())
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1

def getProductSope( url, products):
	productListHtml = BeautifulSoup( getHtmlFromUrl(url), "html.parser",from_encoding="utf-8")
	productInfoTb = productListHtml.find(name="table",attrs={"class":"peptide"})
	infoTrList = productInfoTb.find_all(name="tr")
	pInfo = { }
	for tr in infoTrList:
		if getNodeText(tr.find("td")) == 'Peptide accession number':
			pInfo["number"] = getNodeText(tr.find_all("td")[1])
		if getNodeText(tr.find("td")) == 'Peptide name':
			pInfo["Peptidename"] = getNodeText(tr.find_all("td")[1])
		if getNodeText(tr.find("td")) == 'Organism':
			pInfo["Organism"] = getNodeText(tr.find_all("td")[1])
		if getNodeText(tr.find("td")) == 'Length':
			pInfo["Length"] = getNodeText(tr.find_all("td")[1])
		if getNodeText(tr.find("td")) == 'Monoisotopic mass(without PTMs)':
			pInfo["PTMs"] = getNodeText(tr.find_all("td")[1])
		if getNodeText(tr.find("td")) == 'Sequence':
			pInfo["Sequence"] = getNodeText(tr.find_all("td")[1])
		if getNodeText(tr.find("td")) == 'Peptide family':
			pInfo["Peptidefamily"] = getNodeText(tr.find_all("td")[1])
		if getNodeText(tr.find("td")) == 'Uniprot accession':
			pInfo["Uniprotaccession"] = getNodeText(tr.find_all("td")[1])
		if getNodeText(tr.find("td")) == 'Modification':
			pInfo["Modification"] = getNodeText(tr.find_all("td")[1])

	products.append(pInfo)
	
excelFileName="company.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

i=1
while i < 20027:
	
	url = "http://www.peptides.be/index.php?p=peptide&peptide=PEP"+str(i).zfill(5)
	print(url)
	getProductSope( url, products)
	i += 1


	


headers=["number",'Peptidename','Organism','Length','PTMs','Sequence','Peptidefamily','Uniprotaccession','Modification']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
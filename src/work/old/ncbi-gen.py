from urllib.request import urlopen
from bs4 import BeautifulSoup
import http.client
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re

http.client._MAXHEADERS = 1000

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head.strip() in info:
			workSheet.cell(rowIndex, cellIndex).value = str(info[head.strip()]).strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

def getHtmlFromUrl(url):
	try:
		html = urlopen(url).read()
		return html
	except:
		print("request:"+url)
		getHtmlFromUrl(url)

fileName="list.txt"
excelFileName="products.xlsx"
wb = Workbook()
workSheet = wb.active
with open(fileName,'r') as file_to_read:
	index = 1
	while True:
		lines = file_to_read.readline()
		if not lines:
			break
			pass
		productInfo={}
		genId = ""
		fullName = ""
		pHtml = getHtmlFromUrl(lines.strip())
		if pHtml!=None:
			htmlSoup = BeautifulSoup(pHtml, "html.parser", from_encoding="utf-8")
			
			headers=[
				'genId','fullName'
			];
			genIdHtml = htmlSoup.find(name="span",attrs={"class":"geneid"})
			if genIdHtml != None:
				genId=getNodeText(genIdHtml)
			
			propertyHtmlWaper = htmlSoup.find(name="div",attrs={"id":"summaryDiv"})
			if propertyHtmlWaper != None:
				dtList = propertyHtmlWaper.find_all(name="dt")
				for dt in dtList:
					propertyName = getNodeText(dt).replace("\n","").replace("\r","").replace(" ","")
					if propertyName=="OfficialFullName":
						fullName=getNodeText(dt.nextSibling.nextSibling)
		
		productInfo['genId']=genId.split(",")[0].replace("Gene ID:","")
		productInfo['fullName']=fullName.split("provided by")[0]
		
		writeExcel(workSheet, headers, index, productInfo)
		print(str(index)+"====="+genId)
		index=index+1
wb.save(excelFileName)
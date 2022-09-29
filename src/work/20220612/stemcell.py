from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import requests
from requests.cookies import RequestsCookieJar
import cfscrape
import ssl

ssl._create_default_https_context = ssl._create_unverified_context
http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	print('download:'+IMAGE_URL)
	
	opener = urllib.request.build_opener()
	opener.addheaders = [('User-agent', 'Mozilla/5.0')]
	urllib.request.install_opener(opener)
	urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	headers = {
		"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.82 Safari/537.36",
		"cookie":"_ga=GA1.2.1846208171.1605273760; href=https%3A%2F%2Fwww.sinobiological.com%2Fresearch%2Ftargeted-therapy; accessId=5aff5fb0-84db-11e8-a3b3-d368cce40a8e; _gcl_au=1.1.1660157260.1645016298; Hm_lvt_2d911fad88dfe99ff8bbf610824f3e67=1645016298; sbb=%252be43ohTbVTr09K%252bxQlr1%252bK0onQvF%252bMIXgZM%252bveGXMHU%253dXTaJopSyq01ihC4cD5jOfAa8AEgsWX792EAllASK%252bEKohz0p3JxSEJFavoBnvUmw4fhBpwMcWGZ2Qy%252bRRF2U%252bMSxcQdMfdwOcT%252bR%252bo7qyEU%252br8SBQuGE8GJWgDFeSTZ4QS0HvJFVazETAoyuKMwGHYRoD68%252f7qno5Bg%252bEH9sSXM4upMLtz%252f4IdNkjX6GD0JYHbiUh%252blGTwi25Iz3IKocTDD58DE1yYiY3DxeifN7Qz6OxtXX21lrBpnvgDu9ANN%252f7TTxWWMmOIjxVG772o%252bYGkE9AMxcU5O4cIrT9cubm6dAdgw6n%252fQRZpTVxNv2TGHdHZblPNcfu4dTWVsL3aqaag%253d%253d; _gid=GA1.2.832211649.1645016298; _ce.s=v11.rlc~1645016301520; pageViewNum=13; Hm_lpvt_2d911fad88dfe99ff8bbf610824f3e67=1645017042; Currency=RMB; LocationCode=CN"
	}

	scraper = cfscrape.create_scraper()
	html_code = scraper.get(url,headers=headers).text
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")
	chrome_options.add_argument('user-agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.130 Safari/537.36"')
	chrome_options.add_argument("cookie=__hstc=240517788.35783e1d438e8f99e34188727b050107.1648374652176.1648374652176.1648374652176.1; hubspotutk=35783e1d438e8f99e34188727b050107; __hssrc=1; _gid=GA1.2.122553597.1648374652; _gcl_au=1.1.15847883.1648374654; _fbp=fb.1.1648374655847.2011294960; __hssc=240517788.6.1648374652176; _ga=GA1.1.457442853.1648374652; _ga_SBEXK5LM3N=GS1.1.1648374653.1.1.1648376932.0")
	chrome_options.add_argument("--no-sandbox")

	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	try:
		sizeEle = browser.find_element_by_class_name('radio_swatch')
		if sizeEle != None:
			sizeEle.click()
	except:
		print('no ele')
	return BeautifulSoup(browser.page_source, "html.parser",from_encoding="utf-8")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, products):
	print(str(len(products)) + url)
	sope = getRenderdHtmlFromUrl(url)
	cat = sope.find("div", attrs={"class":"product attribute sku"})
	name = sope.find("h1", attrs={"class":"page-title"})
	desc = sope.find("div", attrs={"class":"description"})
	pName = getNodeText(name)
	catNo = getNodeText(cat).replace("Catalog #","").replace("\n","")
	pInfo={
		"link": url,
		"Product Name": pName,
		"cat": catNo,
		"Overview": getNodeText(desc),
		"pdf": catNo+".pdf"
	}
	specs = sope.find_all("div", attrs={"class":"attribute"})
	for spec in specs:
		title = spec.find("div", attrs={"class":"header"})
		value = spec.find("div", attrs={"class":"value"})
		if title!=None:
			pInfo[getNodeText(title)] = getNodeText(value)
	pdfarea =  sope.find("table", attrs={"class":"scientific-resources-table"})
	if pdfarea!=None:
		pdfLink = pdfarea.find("a")
		if pdfLink!=None:
			pdfHref = pdfLink["href"]
			urllib_download(pdfHref, pInfo["cat"]+".pdf")

	print(pInfo)
	products.append(pInfo.copy())

def getProductList(url, products):
	sope = getHtmlFromUrl(url)
	pListAreas = sope.find_all("a", attrs={"class":"product photo product-item-photo"})
	for pLink in pListAreas:
		getProductInfo(pLink["href"], products)


excelFileName="stemcell.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("https://www.stemcell.com/products/product-types/cytokines/human-recombinant-amphiregulin.html", products)
for pIndex in range(1,19):
	getProductList("https://www.stemcell.com/products/product-types/cytokines.html?p="+str(pIndex)+"&product_list_order=name", products)


headers=[
	'link','Product Name','cat','Overview','Subtype','Alternative Names','Cell Type','Species'
	,'Area of Interest','Molecular Weight','Purity'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
from urllib.request import urlopen
import urllib
from attr import attrs
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import requests
from requests.cookies import RequestsCookieJar
import cfscrape
import ssl

ssl._create_default_https_context = ssl._create_unverified_context
http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	print('download:'+IMAGE_URL)
	
	opener = urllib.request.build_opener()
	opener.addheaders = [('User-agent', 'Mozilla/5.0')]
	urllib.request.install_opener(opener)
	urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	headers = {
		"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.82 Safari/537.36",
		"cookie":"_ga=GA1.2.1846208171.1605273760; href=https%3A%2F%2Fwww.sinobiological.com%2Fresearch%2Ftargeted-therapy; accessId=5aff5fb0-84db-11e8-a3b3-d368cce40a8e; _gcl_au=1.1.1660157260.1645016298; Hm_lvt_2d911fad88dfe99ff8bbf610824f3e67=1645016298; sbb=%252be43ohTbVTr09K%252bxQlr1%252bK0onQvF%252bMIXgZM%252bveGXMHU%253dXTaJopSyq01ihC4cD5jOfAa8AEgsWX792EAllASK%252bEKohz0p3JxSEJFavoBnvUmw4fhBpwMcWGZ2Qy%252bRRF2U%252bMSxcQdMfdwOcT%252bR%252bo7qyEU%252br8SBQuGE8GJWgDFeSTZ4QS0HvJFVazETAoyuKMwGHYRoD68%252f7qno5Bg%252bEH9sSXM4upMLtz%252f4IdNkjX6GD0JYHbiUh%252blGTwi25Iz3IKocTDD58DE1yYiY3DxeifN7Qz6OxtXX21lrBpnvgDu9ANN%252f7TTxWWMmOIjxVG772o%252bYGkE9AMxcU5O4cIrT9cubm6dAdgw6n%252fQRZpTVxNv2TGHdHZblPNcfu4dTWVsL3aqaag%253d%253d; _gid=GA1.2.832211649.1645016298; _ce.s=v11.rlc~1645016301520; pageViewNum=13; Hm_lpvt_2d911fad88dfe99ff8bbf610824f3e67=1645017042; Currency=RMB; LocationCode=CN"
	}

	scraper = cfscrape.create_scraper()
	html_code = scraper.get(url,headers=headers).text
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")
	chrome_options.add_argument('user-agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.130 Safari/537.36"')
	chrome_options.add_argument("cookie=__hstc=240517788.35783e1d438e8f99e34188727b050107.1648374652176.1648374652176.1648374652176.1; hubspotutk=35783e1d438e8f99e34188727b050107; __hssrc=1; _gid=GA1.2.122553597.1648374652; _gcl_au=1.1.15847883.1648374654; _fbp=fb.1.1648374655847.2011294960; __hssc=240517788.6.1648374652176; _ga=GA1.1.457442853.1648374652; _ga_SBEXK5LM3N=GS1.1.1648374653.1.1.1648376932.0")
	chrome_options.add_argument("--no-sandbox")

	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	try:
		sizeEle = browser.find_element_by_class_name('radio_swatch')
		if sizeEle != None:
			sizeEle.click()
	except:
		print('no ele')
	return BeautifulSoup(browser.page_source, "html.parser",from_encoding="utf-8")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url,type, cat, products):
	print(str(len(products)) + url)
	sope = getHtmlFromUrl(url)
	pInfo={
		"link": url,
		"type": type,
		"cat":cat,
		"Product Name": getNodeText(sope.find("h1", attrs={"class":"layui-col-md10 layui-col-sm8 layui-col-xs12"}))
	}
	h2s = sope.find_all("h2", attrs={"class":"title"})
	for h2 in h2s:
		title = getNodeText(h2)
		if title == "分子别名（Synonym）":
			pInfo["分子别名（Synonym）"] = getNodeText(h2.findNextSibling("p"))
		if title == "表达区间及表达系统（Source）":
			pInfo["表达区间及表达系统（Source）"] = getNodeText(h2.findNextSibling("p"))
		if title == "蛋白结构（Molecular Characterization）":
			imgArea = h2.findNextSibling("div")
			if imgArea!=None:
				img = imgArea.find("img")
				if imgArea!=None:
					imgName = cat+".jpg"
					pInfo["蛋白结构（Molecular Characterization）"] = imgName
					urllib_download("https://www.acrobiosystems.cn"+img["src"], imgName)
		
		if title == "内毒素（Endotoxin）":
			pInfo["内毒素（Endotoxin）"] = getNodeText(h2.findNextSibling("p"))
		if title == "纯度（Purity）":
			pInfo["纯度（Purity）"] = getNodeText(h2.findNextSibling("p"))
		if title == "制剂（Formulation）":
			pInfo["制剂（Formulation）"] = getNodeText(h2.findNextSibling("p"))
		if title == "重构方法（Reconstitution）":
			pInfo["重构方法（Reconstitution）"] = getNodeText(h2.findNextSibling("p"))
			pInfo["重构方法（Reconstitution）"] += getNodeText(h2.findNextSibling("strong"))
		if title == "存储（Storage）":
			pInfo["存储（Storage）"] = getNodeText(h2.nextSibling)
			pInfo["存储（Storage）"] += getNodeText(h2.nextSibling.nextSibling)
			pInfo["存储（Storage）"] += getNodeText(h2.nextSibling.nextSibling.nextSibling)
			pInfo["存储（Storage）"] += getNodeText(h2.findNextSibling("ol"))
		
		if title == "背景（Background）":
			pInfo["背景（Background）"] = getNodeText(h2.findNextSibling("p"))
			
	print(pInfo)
	products.append(pInfo.copy())

def getProductList(url, type, products):
	sope = getHtmlFromUrl(url)
	tbody = sope.find("tbody", attrs={"class":"az_tb"})
	trs = tbody.find_all("tr")
	for tr in trs:
		pLink = tr.find("td").find("a")
		getProductInfo(pLink["href"], type, getNodeText(pLink), products)

def getProductType():
	html = '<div class="productSection">  <div class="public_h2_title">产品列表</div>            <div class="clear"></div>      <div class="public_h3_title">白细胞介素（Interleukins）及其受体</div>        <div class="product-type layui-row">    <div class="s-con cf">      <div class="clear"></div><a href="https://www.acrobiosystems.cn/letterlist-46.html" class="normal_cf" target="_blank" data-cl-id="3703562902">IL-1 alpha</a><a href="http://www.acrobiosystems.cn/letterlist-331.html" class="normal_cf" target="_blank" data-cl-id="2670047148">IL-1 beta</a><a href="http://www.acrobiosystems.cn/letterlist-332.html" class="normal_cf" target="_blank" data-cl-id="2407667388">IL-1 RAcP</a><a href="http://www.acrobiosystems.cn/letterlist-333.html" class="normal_cf" target="_blank" data-cl-id="1383162922">IL-1 RII</a><a href="http://www.acrobiosystems.cn/letterlist-753.html" class="normal_cf" target="_blank" data-cl-id="2876661084">IL-1 Rrp2</a><a href="http://www.acrobiosystems.cn/letterlist-344.html" class="normal_cf" target="_blank" data-cl-id="502425335">IL-1RL1</a><a href="http://www.acrobiosystems.cn/letterlist-49.html" class="normal_cf" target="_blank" data-cl-id="2941481159">IL-2</a><a href="http://www.acrobiosystems.cn/letterlist-345.html" class="normal_cf" target="_blank" data-cl-id="2685284625">IL-2 R alpha</a><a href="http://www.acrobiosystems.cn/letterlist-346.html" class="normal_cf" target="_blank" data-cl-id="633901029">IL-2 R beta</a><a href="http://www.acrobiosystems.cn/letterlist-347.html" class="normal_cf" target="_blank" data-cl-id="1892105140">IL-2 R gamma</a><a href="http://www.acrobiosystems.cn/letterlist-1107.html" class="normal_cf" target="_blank" data-cl-id="635499784">IL-2 R beta &amp; IL-2 R gamma</a><a href="http://www.acrobiosystems.cn/letterlist-1179.html" class="normal_cf" target="_blank" data-cl-id="794094440">IL-2 R beta &amp; IL-2 R alpha &amp; IL-2 R gamma</a><a href="https://www.acrobiosystems.cn/letterlist-1391.html" class="normal_cf" target="_blank" data-cl-id="2924703540">IL-3</a><a href="http://www.acrobiosystems.cn/letterlist-743.html" class="normal_cf" target="_blank" data-cl-id="3077279814">IL-3 R alpha</a><a href="http://www.acrobiosystems.cn/letterlist-50.html" class="normal_cf" target="_blank" data-cl-id="2907925921">IL-4</a><a href="http://www.acrobiosystems.cn/letterlist-134.html" class="normal_cf" target="_blank" data-cl-id="2534654239">IL-4 R alpha</a><a href="http://www.acrobiosystems.cn/letterlist-939.html" class="normal_cf" target="_blank" data-cl-id="2891148302">IL-5</a><a href="http://www.acrobiosystems.cn/letterlist-356.html" class="normal_cf" target="_blank" data-cl-id="1850179292">IL-5 R alpha</a><a href="http://www.acrobiosystems.cn/letterlist-51.html" class="normal_cf" target="_blank" data-cl-id="2874370683">IL-6</a><a href="http://www.acrobiosystems.cn/letterlist-122.html" class="normal_cf" target="_blank" data-cl-id="1808457149">IL-6 R alpha</a><a href="http://www.acrobiosystems.cn/letterlist-527.html" class="normal_cf" target="_blank" data-cl-id="4247757530">gp130</a><a href="http://www.acrobiosystems.cn/letterlist-52.html" class="normal_cf" target="_blank" data-cl-id="2857593064">IL-7</a><a href="http://www.acrobiosystems.cn/letterlist-357.html" class="normal_cf" target="_blank" data-cl-id="1655205410">IL-7 R alpha</a><a href="http://www.acrobiosystems.cn/letterlist-1104.html" class="normal_cf" target="_blank" data-cl-id="513608884">IL-7-P2A</a><a href="http://www.acrobiosystems.cn/letterlist-931.html" class="normal_cf" target="_blank" data-cl-id="2840815445">IL-8</a><a href="http://www.acrobiosystems.cn/letterlist-334.html" class="normal_cf" target="_blank" data-cl-id="4250737886">IL-10</a><a href="https://www.acrobiosystems.cn/letterlist-1490.html" class="normal_cf" target="_blank" data-cl-id="4275669420">IL-10 R alpha</a><a href="http://www.acrobiosystems.cn/letterlist-1173.html" class="normal_cf" target="_blank" data-cl-id="4267515505">IL-11</a><a href="http://www.acrobiosystems.cn/letterlist-1172.html" class="normal_cf" target="_blank" data-cl-id="121215407">IL-11 R alpha</a><a href="http://www.acrobiosystems.cn/letterlist-882.html" class="normal_cf" target="_blank" data-cl-id="1528751387">IL-12 R beta 1</a><a href="http://www.acrobiosystems.cn/letterlist-335.html" class="normal_cf" target="_blank" data-cl-id="2895110907">IL-12A</a><a href="http://www.acrobiosystems.cn/letterlist-336.html" class="normal_cf" target="_blank" data-cl-id="2911888526">IL-12B</a><a href="http://www.acrobiosystems.cn/letterlist-676.html" class="normal_cf" target="_blank" data-cl-id="4289168490">IL-12B &amp; IL-12A</a><a href="http://www.acrobiosystems.cn/letterlist-671.html" class="normal_cf" target="_blank" data-cl-id="4233960267">IL-13</a><a href="http://www.acrobiosystems.cn/letterlist-121.html" class="normal_cf" target="_blank" data-cl-id="3841924050">IL-13 R alpha 1</a><a href="http://www.acrobiosystems.cn/letterlist-337.html" class="normal_cf" target="_blank" data-cl-id="3825146431">IL-13 R alpha 2</a><a href="http://www.acrobiosystems.cn/letterlist-48.html" class="normal_cf" target="_blank" data-cl-id="39658685">IL-15</a><a href="http://www.acrobiosystems.cn/letterlist-927.html" class="normal_cf" target="_blank" data-cl-id="563224371">IL-15 R alpha</a><a href="http://www.acrobiosystems.cn/letterlist-338.html" class="normal_cf" target="_blank" data-cl-id="1755071458">IL-17A</a><a href="https://www.acrobiosystems.cn/letterlist-1511.html" class="normal_cf" target="_blank" data-cl-id="1738293839">IL-17B</a><a href="http://www.acrobiosystems.cn/letterlist-918.html" class="normal_cf" target="_blank" data-cl-id="1721516220">IL-17C</a><a href="http://www.acrobiosystems.cn/letterlist-340.html" class="normal_cf" target="_blank" data-cl-id="1687960982">IL-17E</a><a href="http://www.acrobiosystems.cn/letterlist-341.html" class="normal_cf" target="_blank" data-cl-id="1671183363">IL-17F</a><a href="http://www.acrobiosystems.cn/letterlist-696.html" class="normal_cf" target="_blank" data-cl-id="1536038998">IL-17A &amp; IL-17F</a><a href="http://www.acrobiosystems.cn/letterlist-339.html" class="normal_cf" target="_blank" data-cl-id="29448140">IL-17 RA</a><a href="http://www.acrobiosystems.cn/letterlist-932.html" class="normal_cf" target="_blank" data-cl-id="63003378">IL-17 RC</a><a href="http://www.acrobiosystems.cn/letterlist-1124.html" class="normal_cf" target="_blank" data-cl-id="1064018737">IL-17 RA &amp; IL-17 RC</a><a href="http://www.acrobiosystems.cn/letterlist-935.html" class="normal_cf" target="_blank" data-cl-id="4257304960">IL-17 RE</a><a href="https://www.acrobiosystems.cn/letterlist-1352.html" class="normal_cf" target="_blank" data-cl-id="4116516934">IL-18</a><a href="https://www.acrobiosystems.cn/letterlist-1362.html" class="normal_cf" target="_blank" data-cl-id="3292008034">IL-18 R beta</a><a href="http://www.acrobiosystems.cn/letterlist-342.html" class="normal_cf" target="_blank" data-cl-id="1775332323">IL-18 R1</a><a href="http://www.acrobiosystems.cn/letterlist-343.html" class="normal_cf" target="_blank" data-cl-id="2228563476">IL-18BP</a><a href="https://www.acrobiosystems.cn/letterlist-1374.html" class="normal_cf" target="_blank" data-cl-id="4149925077">IL-20</a><a href="https://www.acrobiosystems.cn/letterlist-1375.html" class="normal_cf" target="_blank" data-cl-id="1048604043">IL-20 R alpha</a><a href="http://www.acrobiosystems.cn/letterlist-692.html" class="normal_cf" target="_blank" data-cl-id="4133147458">IL-21</a><a href="http://www.acrobiosystems.cn/letterlist-349.html" class="normal_cf" target="_blank" data-cl-id="1421940860">IL-21 R</a><a href="http://www.acrobiosystems.cn/letterlist-350.html" class="normal_cf" target="_blank" data-cl-id="4116369839">IL-22</a><a href="http://www.acrobiosystems.cn/letterlist-351.html" class="normal_cf" target="_blank" data-cl-id="1425422552">IL23A &amp; IL12B</a><a href="http://www.acrobiosystems.cn/letterlist-528.html" class="normal_cf" target="_blank" data-cl-id="4176865994">IL-23R</a><a href="https://www.acrobiosystems.cn/letterlist-1389.html" class="normal_cf" target="_blank" data-cl-id="4049259363">IL-26</a><a href="https://www.acrobiosystems.cn/letterlist-1363.html" class="normal_cf" target="_blank" data-cl-id="4032481744">IL-27</a><a href="http://www.acrobiosystems.cn/letterlist-352.html" class="normal_cf" target="_blank" data-cl-id="4289388869">IL-27 Ra</a><a href="http://www.acrobiosystems.cn/letterlist-353.html" class="normal_cf" target="_blank" data-cl-id="4267368410">IL-29</a><a href="http://www.acrobiosystems.cn/letterlist-1067.html" class="normal_cf" target="_blank" data-cl-id="1918354655">IL-31</a><a href="http://www.acrobiosystems.cn/letterlist-933.html" class="normal_cf" target="_blank" data-cl-id="136144900">IL-31 RA</a><a href="http://www.acrobiosystems.cn/letterlist-354.html" class="normal_cf" target="_blank" data-cl-id="1951909893">IL-33</a><a href="http://www.acrobiosystems.cn/letterlist-1165.html" class="normal_cf" target="_blank" data-cl-id="1834466560">IL-34</a><a href="http://www.acrobiosystems.cn/letterlist-355.html" class="normal_cf" target="_blank" data-cl-id="1884799417">IL-37</a><a href="http://www.acrobiosystems.cn/letterlist-1223.html" class="normal_cf" target="_blank" data-cl-id="2035797988">IL-38</a>      <div class="clear"></div>    </div>   </div>    <div class="clear"></div>      <div class="public_h3_title">生长因子（Growth Factor）及其受体</div>  <div class="product-type layui-row">    <div class="s-con cf">      <div class="clear"></div>      <div class="cd20ProtocolTitle cd30ProtocolTitle">        <div class="h4-title">内皮生长因子（VEGFs）及其受体</div>      </div>      <div class="clear"></div>    <a href="http://www.acrobiosystems.cn/letterlist-60.html" class="normal_cf" target="_blank" data-cl-id="293220445">VEGF110</a><a href="http://www.acrobiosystems.cn/letterlist-64.html" class="normal_cf" target="_blank" data-cl-id="2004684678">VEGF120</a><a href="http://www.acrobiosystems.cn/letterlist-61.html" class="normal_cf" target="_blank" data-cl-id="2021462297">VEGF121</a><a href="http://www.acrobiosystems.cn/letterlist-65.html" class="normal_cf" target="_blank" data-cl-id="4218638838">VEGF164</a><a href="http://www.acrobiosystems.cn/letterlist-62.html" class="normal_cf" target="_blank" data-cl-id="4235416457">VEGF165</a><a href="https://www.acrobiosystems.cn/letterlist-1347.html" class="normal_cf" target="_blank" data-cl-id="2223779463">VEGF189</a><a href="http://www.acrobiosystems.cn/letterlist-515.html" class="normal_cf" target="_blank" data-cl-id="1278876838">VEGF-B</a><a href="http://www.acrobiosystems.cn/letterlist-516.html" class="normal_cf" target="_blank" data-cl-id="1295654457">VEGF-C</a><a href="http://www.acrobiosystems.cn/letterlist-517.html" class="normal_cf" target="_blank" data-cl-id="1312432076">VEGF-D</a><a href="http://www.acrobiosystems.cn/letterlist-514.html" class="normal_cf" target="_blank" data-cl-id="2940046154">VEGF R1</a><a href="http://www.acrobiosystems.cn/letterlist-127.html" class="normal_cf" target="_blank" data-cl-id="2923268535">VEGF R2</a><a href="http://www.acrobiosystems.cn/letterlist-667.html" class="normal_cf" target="_blank" data-cl-id="2906490916">VEGF R3</a>      <div class="clear"></div>      <div class="cd20ProtocolTitle cd30ProtocolTitle">        <div class="h4-title">转化生长因子（TGFs）及其受体</div>      </div>      <div class="clear"></div>    <a href="http://www.acrobiosystems.cn/letterlist-57.html" class="normal_cf" target="_blank" data-cl-id="2626812848">TGF-beta 1</a><a href="http://www.acrobiosystems.cn/letterlist-908.html" class="normal_cf" target="_blank" data-cl-id="4250467740">Latent TGF-beta 1</a><a href="http://www.acrobiosystems.cn/letterlist-859.html" class="normal_cf" target="_blank" data-cl-id="590307246">LAP (TGF-beta 1)</a><a href="https://www.acrobiosystems.cn/letterlist-1324.html" class="normal_cf" target="_blank" data-cl-id="2626816016">LRRC32 &amp; TGF-beta 1</a><a href="https://www.acrobiosystems.cn/letterlist-1240.html" class="normal_cf" target="_blank" data-cl-id="2677145705">TGF-beta 2</a><a href="https://www.acrobiosystems.cn/letterlist-1295.html" class="normal_cf" target="_blank" data-cl-id="5833301">Latent TGF-beta 2</a><a href="http://www.acrobiosystems.cn/letterlist-1135.html" class="normal_cf" target="_blank" data-cl-id="2660368086">TGF-beta 3</a><a href="http://www.acrobiosystems.cn/letterlist-486.html" class="normal_cf" target="_blank" data-cl-id="670132011">TGF-beta RII</a><a href="http://www.acrobiosystems.cn/letterlist-163.html" class="normal_cf" target="_blank" data-cl-id="2236721767">BMP-2</a><a href="http://www.acrobiosystems.cn/letterlist-550.html" class="normal_cf" target="_blank" data-cl-id="1424895605">Gremlin</a>      <div class="clear"></div>      <div class="cd20ProtocolTitle cd30ProtocolTitle">        <div class="h4-title">成纤维细胞生长因子（FGFs）及其受体</div>      </div>      <div class="clear"></div>   <a href="http://www.acrobiosystems.cn/letterlist-281.html" class="normal_cf" target="_blank" data-cl-id="2168495521">FGF acidic</a><a href="http://www.acrobiosystems.cn/letterlist-88.html" class="normal_cf" target="_blank" data-cl-id="2563395338">FGF basic</a><a href="https://www.acrobiosystems.cn/letterlist-1337.html" class="normal_cf" target="_blank" data-cl-id="2253997080">FGF-3</a><a href="https://www.acrobiosystems.cn/letterlist-1313.html" class="normal_cf" target="_blank" data-cl-id="2371440413">FGF-4</a><a href="https://www.acrobiosystems.cn/letterlist-1314.html" class="normal_cf" target="_blank" data-cl-id="2354662794">FGF-5</a><a href="https://www.acrobiosystems.cn/letterlist-1312.html" class="normal_cf" target="_blank" data-cl-id="2337885175">FGF-6</a><a href="https://www.acrobiosystems.cn/letterlist-1336.html" class="normal_cf" target="_blank" data-cl-id="2321107556">FGF-7</a><a href="https://www.acrobiosystems.cn/letterlist-1311.html" class="normal_cf" target="_blank" data-cl-id="460994356">FGF-8E</a><a href="http://www.acrobiosystems.cn/letterlist-285.html" class="normal_cf" target="_blank" data-cl-id="2153331366">FGF-9</a><a href="https://www.acrobiosystems.cn/letterlist-1339.html" class="normal_cf" target="_blank" data-cl-id="2995444490">FGF-10</a><a href="https://www.acrobiosystems.cn/letterlist-1247.html" class="normal_cf" target="_blank" data-cl-id="2878001157">FGF-19</a><a href="http://www.acrobiosystems.cn/letterlist-1190.html" class="normal_cf" target="_blank" data-cl-id="2877854062">FGF-21</a><a href="https://www.acrobiosystems.cn/letterlist-1321.html" class="normal_cf" target="_blank" data-cl-id="2844298824">FGF-23</a><a href="http://www.acrobiosystems.cn/letterlist-282.html" class="normal_cf" target="_blank" data-cl-id="2327975265">FGF R1</a><a href="https://www.acrobiosystems.cn/letterlist-1171.html" class="normal_cf" target="_blank" data-cl-id="1365185532">FGF R2 (IIIb) </a><a href="https://www.acrobiosystems.cn/letterlist-1175.html" class="normal_cf" target="_blank" data-cl-id="761044153">FGF R2 (IIIc) </a><a href="https://www.acrobiosystems.cn/letterlist-1496.html" class="normal_cf" target="_blank" data-cl-id="1787478089">FGF R3 (IIIb) </a><a href="https://www.acrobiosystems.cn/letterlist-1495.html" class="normal_cf" target="_blank" data-cl-id="1317851852">FGF R3 (IIIc) </a><a href="http://www.acrobiosystems.cn/letterlist-283.html" class="normal_cf" target="_blank" data-cl-id="2378308122">FGF R4</a>      <div class="clear"></div>      <div class="cd20ProtocolTitle cd30ProtocolTitle">        <div class="h4-title">表皮生长因子（EGFs）及其受体</div>      </div>      <div class="clear"></div> <a href="http://www.acrobiosystems.cn/letterlist-258.html" class="normal_cf" target="_blank" data-cl-id="2670866361">EGF</a><a href="http://www.acrobiosystems.cn/letterlist-1177.html" class="normal_cf" target="_blank" data-cl-id="41231847">HBEGF</a><a href="https://www.acrobiosystems.cn/letterlist-80.html" class="normal_cf" target="_blank" data-cl-id="2265018097">EGFR</a><a href="http://www.acrobiosystems.cn/letterlist-947.html" class="normal_cf" target="_blank" data-cl-id="555802650">EGFRVIII</a>      <div class="clear"></div>      <div class="cd20ProtocolTitle cd30ProtocolTitle">        <div class="h4-title">胰岛素样生长因子（IGFs）及其受体</div>      </div>      <div class="clear"></div> <a href="http://www.acrobiosystems.cn/letterlist-328.html" class="normal_cf" target="_blank" data-cl-id="3386480651">IGF-I</a><a href="http://www.acrobiosystems.cn/letterlist-93.html" class="normal_cf" target="_blank" data-cl-id="59420646">IGF-II</a><a href="http://www.acrobiosystems.cn/letterlist-329.html" class="normal_cf" target="_blank" data-cl-id="1092236889">IGF-I R</a><a href="http://www.acrobiosystems.cn/letterlist-325.html" class="normal_cf" target="_blank" data-cl-id="2199565223">IGFBP-3</a><a href="https://www.acrobiosystems.cn/letterlist-1510.html" class="normal_cf" target="_blank" data-cl-id="1116764533">IGFBP-3 R</a><a href="http://www.acrobiosystems.cn/letterlist-326.html" class="normal_cf" target="_blank" data-cl-id="2149232366">IGFBP-4</a><a href="http://www.acrobiosystems.cn/letterlist-327.html" class="normal_cf" target="_blank" data-cl-id="2132454747">IGFBP-7</a>      <div class="clear"></div>      <div class="cd20ProtocolTitle cd30ProtocolTitle">        <div class="h4-title">生长分化因子（GDFs）</div>      </div>      <div class="clear"></div>  <a href="http://www.acrobiosystems.cn/letterlist-750.html" class="normal_cf" target="_blank" data-cl-id="1975911909">GDF-15</a><a href="http://www.acrobiosystems.cn/letterlist-862.html" class="normal_cf" target="_blank" data-cl-id="2127363974">GFR alpha-like</a><a href="https://www.acrobiosystems.cn/letterlist-1244.html" class="normal_cf" target="_blank" data-cl-id="410531811">Latent GDF-2</a><a href="http://www.acrobiosystems.cn/letterlist-1150.html" class="normal_cf" target="_blank" data-cl-id="3892470783">GDF-2</a><a href="http://www.acrobiosystems.cn/letterlist-1188.html" class="normal_cf" target="_blank" data-cl-id="578308001">Latent GDF-8</a>      <div class="clear"></div>      <div class="cd20ProtocolTitle cd30ProtocolTitle" id="Chemokines">        <div class="h4-title">其他生长因子及其受体</div>      </div>      <div class="clear"></div>  <a href="http://www.acrobiosystems.cn/letterlist-748.html" class="normal_cf" target="_blank" data-cl-id="262039019">CTGF</a><a href="http://www.acrobiosystems.cn/letterlist-969.html" class="normal_cf" target="_blank" data-cl-id="1133132386">HGF</a><a href="http://www.acrobiosystems.cn/letterlist-120.html" class="normal_cf" target="_blank" data-cl-id="2717452892">HGF R</a><a href="http://www.acrobiosystems.cn/letterlist-909.html" class="normal_cf" target="_blank" data-cl-id="478684755">PDGF-BB</a><a href="http://www.acrobiosystems.cn/letterlist-940.html" class="normal_cf" target="_blank" data-cl-id="1092129310">PDGF R alpha</a><a href="http://www.acrobiosystems.cn/letterlist-425.html" class="normal_cf" target="_blank" data-cl-id="2769434244">PDGF R beta</a>      <div class="clear"></div>    </div>  </div>  <div class="clear"></div>  <div class="public_h3_title">肿瘤坏死因子（TNFs）及其受体</div>  <div class="product-type layui-row">    <div class="s-con cf">      <div class="clear"></div>    <a href="http://www.acrobiosystems.cn/letterlist-58.html" class="normal_cf" target="_blank" data-cl-id="2708780466">TNF-alpha</a><a href="http://www.acrobiosystems.cn/letterlist-494.html" class="normal_cf" target="_blank" data-cl-id="527243670">TNFR1</a><a href="http://www.acrobiosystems.cn/letterlist-495.html" class="normal_cf" target="_blank" data-cl-id="510466051">TNFR2</a><a href="http://www.acrobiosystems.cn/letterlist-687.html" class="normal_cf" target="_blank" data-cl-id="2866331016">TNFSF11</a><a href="http://www.acrobiosystems.cn/letterlist-207.html" class="normal_cf" target="_blank" data-cl-id="885405531">CD40 Ligand</a><a href="http://www.acrobiosystems.cn/letterlist-194.html" class="normal_cf" target="_blank" data-cl-id="3173217830">CD27 ligand</a><a href="http://www.acrobiosystems.cn/letterlist-424.html" class="normal_cf" target="_blank" data-cl-id="3389940619">OX40 Ligand</a><a href="http://www.acrobiosystems.cn/letterlist-736.html" class="normal_cf" target="_blank" data-cl-id="3647545666">CD30 Ligand</a><a href="http://www.acrobiosystems.cn/letterlist-137.html" class="normal_cf" target="_blank" data-cl-id="1085513274">4-1BB Ligand</a><a href="http://www.acrobiosystems.cn/letterlist-946.html" class="normal_cf" target="_blank" data-cl-id="563322595">TRAIL</a><a href="http://www.acrobiosystems.cn/letterlist-497.html" class="normal_cf" target="_blank" data-cl-id="632439552">TRAIL R1</a><a id="Chemokinesa" href="http://www.acrobiosystems.cn/letterlist-498.html" class="normal_cf" target="_blank" data-cl-id="682772409">TRAIL R2</a><a href="http://www.acrobiosystems.cn/letterlist-499.html" class="normal_cf" target="_blank" data-cl-id="716327647">TRAIL R4</a><a href="http://www.acrobiosystems.cn/letterlist-504.html" class="normal_cf" target="_blank" data-cl-id="604447469">TWEAK R</a><a href="http://www.acrobiosystems.cn/letterlist-843.html" class="normal_cf" target="_blank" data-cl-id="2232428695">APRIL</a><a href="http://www.acrobiosystems.cn/letterlist-159.html" class="normal_cf" target="_blank" data-cl-id="1245578664">BAFF</a><a href="https://www.acrobiosystems.cn/letterlist-553.html" class="normal_cf" target="_blank" data-cl-id="1864590206">BAFF R</a><a href="http://www.acrobiosystems.cn/letterlist-742.html" class="normal_cf" target="_blank" data-cl-id="1289838607">LIGHT</a><a href="http://www.acrobiosystems.cn/letterlist-92.html" class="normal_cf" target="_blank" data-cl-id="3738254429">HVEM</a><a href="http://www.acrobiosystems.cn/letterlist-304.html" class="normal_cf" target="_blank" data-cl-id="2520207195">GITR</a><a href="http://www.acrobiosystems.cn/letterlist-305.html" class="normal_cf" target="_blank" data-cl-id="4212514416">GITR Ligand</a></div>  </div>  <div class="clear"></div>  <div class="public_h3_title">趋化因子（Chemokines）及其受体</div>  <div class="product-type layui-row">    <div class="s-con cf">      <div class="clear"></div>   <a href="https://www.acrobiosystems.cn/letterlist-1441.html" class="normal_cf" target="_blank" data-cl-id="2950362084">CCL1</a><a href="https://www.acrobiosystems.cn/letterlist-1440.html" class="normal_cf" target="_blank" data-cl-id="3000694941">CCL2</a><a href="https://www.acrobiosystems.cn/letterlist-1378.html" class="normal_cf" target="_blank" data-cl-id="2983917322">CCL3</a><a href="https://www.acrobiosystems.cn/letterlist-1289.html" class="normal_cf" target="_blank" data-cl-id="2883251608">CCL5</a><a href="https://www.acrobiosystems.cn/letterlist-1470.html" class="normal_cf" target="_blank" data-cl-id="2916806846">CCL7</a><a href="https://www.acrobiosystems.cn/letterlist-1472.html" class="normal_cf" target="_blank" data-cl-id="2779631728">CCL14</a><a href="https://www.acrobiosystems.cn/letterlist-1466.html" class="normal_cf" target="_blank" data-cl-id="2829964585">CCL17</a><a href="http://www.acrobiosystems.cn/letterlist-1056.html" class="normal_cf" target="_blank" data-cl-id="2997740775">CCL19</a><a href="https://www.acrobiosystems.cn/letterlist-1390.html" class="normal_cf" target="_blank" data-cl-id="1001748567">CCL20</a><a href="https://www.acrobiosystems.cn/letterlist-1467.html" class="normal_cf" target="_blank" data-cl-id="1035303805">CCL22</a><a href="http://www.acrobiosystems.cn/letterlist-79.html" class="normal_cf" target="_blank" data-cl-id="2344475507">CX3CL1</a><a href="https://www.acrobiosystems.cn/letterlist-1520.html" class="normal_cf" target="_blank" data-cl-id="2292552671">CXCL4</a><a href="https://www.acrobiosystems.cn/letterlist-1513.html" class="normal_cf" target="_blank" data-cl-id="1838919568">CXCL10</a><a href="https://www.acrobiosystems.cn/letterlist-1307.html" class="normal_cf" target="_blank" data-cl-id="1240805168">CXCL12/SDF-1</a><a href="https://www.acrobiosystems.cn/letterlist-1471.html" class="normal_cf" target="_blank" data-cl-id="1889252425">CXCL13</a><a href="https://www.acrobiosystems.cn/letterlist-1543.html" class="normal_cf" target="_blank" data-cl-id="1939585282">CXCL16</a><a href="http://www.acrobiosystems.cn/letterlist-1128.html" class="normal_cf" target="_blank" data-cl-id="534782306">CCR5</a><a href="https://www.acrobiosystems.cn/letterlist-1182.html" class="normal_cf" target="_blank" data-cl-id="618670401">CCR8</a><a href="http://www.acrobiosystems.cn/letterlist-241.html" class="normal_cf" target="_blank" data-cl-id="482247105">CXCR4</a>      <div class="clear"></div>    </div>  </div>  <div class="clear"></div>  <div class="public_h3_title">集落刺激因子（CSFs）及其受体</div>  <div class="product-type layui-row">    <div class="s-con cf">      <div class="clear"></div><a href="http://www.acrobiosystems.cn/letterlist-681.html" class="normal_cf" target="_blank" data-cl-id="1785187337">M-CSF</a><a href="http://www.acrobiosystems.cn/letterlist-394.html" class="normal_cf" target="_blank" data-cl-id="405848475">M-CSF R</a><a href="http://www.acrobiosystems.cn/letterlist-44.html" class="normal_cf" target="_blank" data-cl-id="2012080572">GM-CSF</a><a href="http://www.acrobiosystems.cn/letterlist-880.html" class="normal_cf" target="_blank" data-cl-id="1782735662">GM-CSF R alpha</a><a href="http://www.acrobiosystems.cn/letterlist-42.html" class="normal_cf" target="_blank" data-cl-id="1477384283">G-CSF</a><a href="http://www.acrobiosystems.cn/letterlist-301.html" class="normal_cf" target="_blank" data-cl-id="349317513">G-CSF R</a>      <div class="clear"></div>    </div>  </div>  <div class="clear"></div>  <div class="public_h3_title">干扰素（IFNs）及其受体</div>  <div class="product-type layui-row">    <div class="s-con cf">      <div class="clear"></div>  <a href="http://www.acrobiosystems.cn/letterlist-1111.html" class="normal_cf" target="_blank" data-cl-id="2489338710">IFN-alpha 1</a><a href="http://www.acrobiosystems.cn/letterlist-1217.html" class="normal_cf" target="_blank" data-cl-id="2710825075">IFN-alpha 2b</a><a href="http://www.acrobiosystems.cn/letterlist-321.html" class="normal_cf" target="_blank" data-cl-id="2901514109">IFN-alpha/beta R1</a><a href="http://www.acrobiosystems.cn/letterlist-322.html" class="normal_cf" target="_blank" data-cl-id="2851181252">IFN-alpha/beta R2</a><a href="https://www.acrobiosystems.cn/letterlist-45.html" class="normal_cf" target="_blank" data-cl-id="1071202996">IFN-gamma </a><a href="http://www.acrobiosystems.cn/letterlist-323.html" class="normal_cf" target="_blank" data-cl-id="2741672385">IFN-gamma R1</a>      <div class="clear"></div>    </div>  </div>  <div class="clear"></div>  <div class="public_h3_title">补体因子</div>  <div class="product-type layui-row">    <div class="s-con cf">      <div class="clear"></div>    <a href="http://www.acrobiosystems.cn/letterlist-1632.html" class="normal_cf" target="_blank" data-cl-id="167196794">Complement C2</a><a href="https://www.acrobiosystems.cn/letterlist-1631.html" class="normal_cf" target="_blank" data-cl-id="183974413">Complement C3</a><a href="http://www.acrobiosystems.cn/letterlist-920.html" class="normal_cf" target="_blank" data-cl-id="83308699">Complement C5</a><a href="http://www.acrobiosystems.cn/letterlist-1633.html" class="normal_cf" target="_blank" data-cl-id="3408009614">Complement C5a</a><a href="http://www.acrobiosystems.cn/letterlist-237.html" class="normal_cf" target="_blank" data-cl-id="3759920692">Complement Factor D</a>      <div class="clear"></div>    </div>  </div>  <div class="clear"></div>  <div class="public_h3_title">其他细胞因子及其受体</div>  <div class="product-type layui-row">    <div class="s-con cf">      <div class="clear"></div>   <a href="http://www.acrobiosystems.cn/letterlist-40.html" class="normal_cf" target="_blank" data-cl-id="4207539828">Erythropoietin</a><a href="http://www.acrobiosystems.cn/letterlist-162.html" class="normal_cf" target="_blank" data-cl-id="239031469">Betacellulin</a><a href="http://www.acrobiosystems.cn/letterlist-56.html" class="normal_cf" target="_blank" data-cl-id="1525847215">SCF</a><a href="http://www.acrobiosystems.cn/letterlist-184.html" class="normal_cf" target="_blank" data-cl-id="3760462733">CD117</a><a href="http://www.acrobiosystems.cn/letterlist-971.html" class="normal_cf" target="_blank" data-cl-id="2603225724">MIS RII</a><a href="http://www.acrobiosystems.cn/letterlist-53.html" class="normal_cf" target="_blank" data-cl-id="1898369152">LIF</a><a href="http://www.acrobiosystems.cn/letterlist-379.html" class="normal_cf" target="_blank" data-cl-id="462915894">LIF R</a><a href="http://www.acrobiosystems.cn/letterlist-54.html" class="normal_cf" target="_blank" data-cl-id="1860703672">Oncostatin M</a><a href="http://www.acrobiosystems.cn/letterlist-39.html" class="normal_cf" target="_blank" data-cl-id="1830813652">Activin A</a><a href="https://www.acrobiosystems.cn/letterlist-139.html" class="normal_cf" target="_blank" data-cl-id="3137372522">Activin RIIA</a><a href="http://www.acrobiosystems.cn/letterlist-1066.html" class="normal_cf" target="_blank" data-cl-id="1767081790">NRG1 Beta 1</a><a href="https://www.acrobiosystems.cn/letterlist-408.html" class="normal_cf" target="_blank" data-cl-id="1240475690">NRG4</a><a href="http://www.acrobiosystems.cn/letterlist-233.html" class="normal_cf" target="_blank" data-cl-id="2978654936">CNTF R alpha</a>      <div class="clear"></div>    </div>  </div></div>'
	sope = BeautifulSoup(html, "html.parser",from_encoding="utf-8")
	types = sope.find_all("a")
	for type in types:
		getProductList(type["href"], getNodeText(type), products)

excelFileName="acrobiosystems.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("https://www.acrobiosystems.cn/P459-Human_IL-1_RAcP_%7C_IL-1_R3_Protein_His_Tag_MALS_verified.html","",'', products)
getProductType()


headers=[
	'link','type','cat','Product Name','分子别名（Synonym）','表达区间及表达系统（Source）','蛋白结构（Molecular Characterization）'
	,'内毒素（Endotoxin）','纯度（Purity）','制剂（Formulation）','重构方法（Reconstitution）','存储（Storage）','背景（Background）'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
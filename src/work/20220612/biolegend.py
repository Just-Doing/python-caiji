from enum import IntEnum
import requests
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import cfscrape
import json
import string
import re
import time
import math

import numpy as np

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

def getHtmlFromUrl(url, type="get", para={}):
	headers = {
		"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.82 Safari/537.36",
		"cookie":"_ga=GA1.2.1846208171.1605273760; href=https%3A%2F%2Fwww.sinobiological.com%2Fresearch%2Ftargeted-therapy; accessId=5aff5fb0-84db-11e8-a3b3-d368cce40a8e; _gcl_au=1.1.1660157260.1645016298; Hm_lvt_2d911fad88dfe99ff8bbf610824f3e67=1645016298; sbb=%252be43ohTbVTr09K%252bxQlr1%252bK0onQvF%252bMIXgZM%252bveGXMHU%253dXTaJopSyq01ihC4cD5jOfAa8AEgsWX792EAllASK%252bEKohz0p3JxSEJFavoBnvUmw4fhBpwMcWGZ2Qy%252bRRF2U%252bMSxcQdMfdwOcT%252bR%252bo7qyEU%252br8SBQuGE8GJWgDFeSTZ4QS0HvJFVazETAoyuKMwGHYRoD68%252f7qno5Bg%252bEH9sSXM4upMLtz%252f4IdNkjX6GD0JYHbiUh%252blGTwi25Iz3IKocTDD58DE1yYiY3DxeifN7Qz6OxtXX21lrBpnvgDu9ANN%252f7TTxWWMmOIjxVG772o%252bYGkE9AMxcU5O4cIrT9cubm6dAdgw6n%252fQRZpTVxNv2TGHdHZblPNcfu4dTWVsL3aqaag%253d%253d; _gid=GA1.2.832211649.1645016298; _ce.s=v11.rlc~1645016301520; pageViewNum=13; Hm_lpvt_2d911fad88dfe99ff8bbf610824f3e67=1645017042; Currency=RMB; LocationCode=CN"
	}

	scraper = cfscrape.create_scraper()
	html_code = scraper.get(url,headers=headers).text
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")

def requestJson(url):
	r = requests.post(url,data={"input":"atcc"}, headers={
		'Content-Type': 'application/x-www-form-urlencoded',
		'cookie':'visid_incap_2255650=4oBBaRPnQfCVoYEiTmjTq/NVAWEAAAAAQUIPAAAAAAD69PQHUoB0KplKq7/j0+gH; nlbi_2255650=CJKhHYlMm17tpKyoBzOViAAAAACDEjp3gL6bj6YL8j9XE0d/; incap_ses_893_2255650=m1tJIuDRUEp3FE/5GpNkDPRVAWEAAAAAM2KkDpvtARtZral+cMXSVw==; _gcl_au=1.1.76703404.1627477493; _gid=GA1.2.730047202.1627477493; BCSessionID=83af10b8-9488-4b7b-a3b1-3640f178dca2; categoryView=grid; _ga_S46FST9X1M=GS1.1.1627477492.1.1.1627478562.0; _ga=GA1.2.31731397.1627477493; _gat_UA-139934-1=1; _uetsid=69fc2d30efa411eb8818eb045f8760e5; _uetvid=69fc3a70efa411ebba3a23c153f6e477; .Nop.Customer=d664d529-d14a-44b1-86b3-cbf5373277b4',
		"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"
	})
	return BeautifulSoup(r.text, "html.parser",from_encoding="utf-8")


def getRenderdHtmlFromUrl(url):

	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	return BeautifulSoup(browser.page_source, "html.parser")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, pInfo):
	print(str(len(products))+":"+url)
	sope = getHtmlFromUrl(url)
	for br in sope.find_all("br"):
		br.replaceWith("\n")
	dts = sope.find_all("dt")
	for dt in dts:
		title = getNodeText(dt)
		if len(title) > 0:
			dd = dt.findNextSibling("dd")
			val = getNodeText(dd)
			pInfo[title] = val
			if title =="Gene ID":
				link = dd.find("a")
				pInfo["Gene ID link"] = link["href"]
			if title =="UniProt":
				link = dd.find("a")
				pInfo["UniProt link"] = link["href"]

	desc = sope.find("div", attrs={"class":"introDescription col-xs-12 noPaddingLeft"})
	pInfo["Description"] = getNodeText(desc).replace("Description","")
	products.append(pInfo.copy());

def getProductList(url):
	html_code = getHtmlFromUrl(url)
	proList = html_code.find_all("li", attrs={"class":"row list"})
	for tr in proList:
		pLink = tr.find("a", attrs={"itemprop":"name"})
		
		if pLink ==None:
			pInfo = products[len(products)-1].copy()
		else:
			pInfo = {"link": "https://www.biolegend.com"+pLink["href"]}
			name = getNodeText(pLink)
			clone = tr.find("div", attrs={"class":"col-xs-2 noPaddingLeft"})
			appl=tr.find("div", attrs={"class":"col-xs-1 noPadding wordWrap"})
			pInfo["appl"] = getNodeText(appl)
			pInfo["clone"] = getNodeText(clone)
			pInfo["name"] = name
		form = tr.find("form")
		price = form.find("span", attrs={"itemprop":"price"})
		specs = form.find_all("div",attrs={"class":"col-xs-3 noPadding"})
		if len(specs)>1:
			cat = getNodeText(specs[0])
			size = getNodeText(specs[1])
			pInfo["cat"] = cat
			pInfo["size"] = size
		pInfo["price"] = getNodeText(price)
		if pLink != None:
			getProductInfo("https://www.biolegend.com"+pLink["href"], pInfo)
		else:
			products.append(pInfo.copy());

products = []
getProductList('https://www.biolegend.com/en-us/search-results?PageSize=200&Category=PRIM_AB&Applications=FA')
# getProductInfo("https://www.biolegend.com/en-us/products/ultra-leaf-purified-anti-human-cd85g-ilt7-antibody-19175", {})

excelFileName="biolegend.xlsx"
wb = Workbook()
workSheet = wb.active

headers=[
	'link','name','clone','appl',
	'cat','size','price','Regulatory Status'
	,'Other Names','Isotype','Description','Reactivity'
	,'Antibody Type','Host Species','Formulation','Preparation','Concentration','Storage & Handling','Application','Recommended Usage',
	'Application Notes','Structure'
	,'Distribution','Function','Cell Type','Biology Area'
	,'Molecular Family','Gene ID','Gene ID link','UniProt','UniProt link'
]

for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
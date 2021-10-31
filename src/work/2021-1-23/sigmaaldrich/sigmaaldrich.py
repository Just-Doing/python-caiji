from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000
def urllib_download(IMAGE_URL, pName):
	try:
		from urllib.request import urlretrieve
		urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')   
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"}

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		if retryCount< 5:
			getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	productHtml = getHtmlFromUrl(url)
	if productHtml != None:
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
		name = sope.find("h1", attrs={"itemprop":"name"})
		img = sope.find("img", attrs={"itemprop":"image"})
		Assay = sope.find("h2", attrs={"itemprop":"description"})
		Synonym = sope.find("p", attrs={"class":"synonym"})
		spInfo = sope.find("div", attrs={"class":"productInfo"})
		pInfo["name"] = getNodeText(name)
		pInfo["CAS Number"] = ""
		for sp in spInfo.find_all("li"):
			spVal = getNodeText(sp)
			if spVal.find("CAS Number") > -1:
				pInfo["CAS Number"] = spVal.replace("CAS Number","")
			if spVal.find("Enzyme Commission (EC) Number") > -1:
				pInfo["Enzyme Commission (EC) Number"] = spVal
			if spVal.find("MDL number") > -1:
				pInfo["MDL number"] = spVal
			if spVal.find("NACRES") > -1:
				pInfo["NACRES"] = spVal
			if spVal.find("Molecular Weight ") > -1:
				pInfo["Molecular Weight"] = spVal
			if spVal.find("EC Number") > -1:
				pInfo["EC Number"] = spVal
			if spVal.find("PubChem Substance ID") > -1:
				pInfo["PubChem Substance ID"] = spVal
			if spVal.find("Linear Formula") > -1:
				pInfo["Linear Formula"] = spVal
		
		# if img != None:
			# if pInfo["CAS Number"] != "":
				# urllib_download('https://www.sigmaaldrich.com'+img["src"], pInfo["CAS Number"])
			# else:
				# urllib_download(img["src"], pInfo["name"])
		contentPage = sope.find("div",attrs={"id":"tab1Wrap"})
		contentTrs = contentPage.find_all("tr")
		for contentTr in contentTrs:
			tds = contentTr.find_all("td")
			if len(tds) ==2:
				title = getNodeText(tds[0]).strip()
				val = getNodeText(tds[1]).replace("\t","").replace("\n","")
				pInfo[title] = val
		productDescription = contentPage.find("div", attrs={"id":"productDescription"})
		pInfo["productDescription"] = getNodeText(productDescription)
		safetyInfoArea = sope.find("div", attrs={"id":"productDetailSafety"})
		
		safetyInfos = safetyInfoArea.find_all("div",attrs={"class":"safetyRow"})
		for safetyInfo in safetyInfos:
			title = getNodeText( safetyInfo.find("div", attrs={"class":"safetyLeft"})).strip()
			val = getNodeText(safetyInfo.find("div", attrs={"class":"safetyRight"}))
			pInfo[title] = val
		pInfo["Assay"] = getNodeText(Assay)
		pInfo["Synonym"] = getNodeText(Synonym)
		pInfo["link"] = url
		
		products.append(pInfo.copy())

def getProductList(url, pInfo, products):
	print(url)
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	prodTables = sope.find_all("table", attrs={"class":"opcTable"})
	for prodTable in prodTables:
		for protr in prodTable.find_all("tr"):
			tds = protr.find("td")
			if tds != None:
				linkInfo = tds.find("a")
				if linkInfo!=None:
					getProductInfo("https://www.sigmaaldrich.com"+linkInfo["href"], pInfo, products )

linkFile = "links.txt"
excelFileName="sigmaaldrich.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
with open(linkFile, 'r') as file_to_read:
	while True:
		lines = file_to_read.readline()
		if not lines:
			break
			pass
		typePart = lines.split("--")
		pInfo = {}
		urlInfoPart = typePart[len(typePart) - 1].split("@")
		if(len(typePart) == 1):
			pInfo["type1"] = urlInfoPart[0]
		if(len(typePart) == 2):
			pInfo["type1"] = typePart[0]
			pInfo["type2"] = urlInfoPart[0]
		if(len(typePart) == 3):
			pInfo["type1"] = typePart[0]
			pInfo["type2"] = typePart[1]
			pInfo["type3"] = urlInfoPart[0]
			
		getProductList(urlInfoPart[1].strip(),pInfo,products)
links=['https://www.sigmaaldrich.com/catalog/product/sigma/22178?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/c1184?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/c2605?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/C8546?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/C1794?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/C0615?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/E6412?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/g4423?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/p2401?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/p2611?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/p4716?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/x3876?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/49291?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/E2164?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/H2125?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/x3876?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/p4716?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/p2401?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/p3026?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/p5936?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/P5400?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/42603?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/38429?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/40452?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/L2157?lang=en&region=US',
'https://www.sigmaaldrich.com/catalog/product/sigma/93014?lang=en&region=US']
for link in links:
	getProductInfo(link, {"type1":'Enzymes for Alternative Energy Research'}, products)
headers=['link','type1','type2','type3','name','Assay','Synonym','CAS Number','Enzyme Commission (EC) Number','MDL number','NACRES','Molecular Weight','EC Number','PubChem Substance ID','Linear Formula',
	'Related Categories','Quality Level','form','InChI','InChI key','mp','density','bp','particle size','average diameter','refractive index','λmax','volume resistivity','viscosity',
	'solubility','surface area','composition','bulk density','grade, impurities','color',
	'specific activity','greener alternative product characteristics','color','greener alternative category','shipped in','storage temp.',
	'productDescription','Personal Protective Equipment','RIDADR','WGK Germany','Flash Point(F)','Flash Point(C)'
]
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	if rindex%100  == 0:
		wb.save(excelFileName)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
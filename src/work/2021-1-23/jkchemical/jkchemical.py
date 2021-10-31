from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36"}

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		if retryCount< 5:
			getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	productHtml = getRenderdHtmlFromUrl(url)
	if productHtml != None:
		pInfo["link"] = url
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
		enName = sope.find("div", attrs={"id":"title"})
		cnName = sope.find("span", attrs={"id":"ctl00_cph_Content_lb_ChineseName"})
		cas = sope.find("li", attrs={"id":"ctl00_cph_Content_li_hl_Cas"})
		cd = sope.find("li", attrs={"id":"ctl00_cph_Content_li_lt_Purity"})
		mdl = sope.find("li", attrs={"id":"ctl00_cph_Content_li_hl_Mdl"})
		fzs = sope.find("li", attrs={"id":"ctl00_cph_Content_li_lt_MF"})
		fzl = sope.find("li", attrs={"id":"ctl00_cph_Content_li_lt_MW"})
		pNumb = sope.find("li", attrs={"id":"ctl00_cph_Content_li_lt_OriginalId"})
		pInfo["cnName"] = getNodeText(cnName).replace("（订货以英文名称为准）","")
		pInfo["enName"] = getNodeText(enName).replace("（订货以英文名称为准）","").replace(pInfo["cnName"], "")
		pInfo["CAS"] = getNodeText(cas).replace("CAS：","").replace("\r","").replace("\n","").replace("  ","")
		pInfo["cd"]=getNodeText(cd)
		pInfo["mdl"]=getNodeText(mdl)
		pInfo["fzs"]=getNodeText(fzs)
		pInfo["fzl"]=getNodeText(fzl)
		pInfo["pNumb"]=getNodeText(pNumb)
		# img = sope.find("img", attrs={"id":"ctl00_cph_Content_ig_Structure"})
		# if img != None:
			# imgUrl = "https://www.jkchemical.com"+img["src"].split("?")[0]
			# if pInfo["CAS"] != "":
				# urllib_download(imgUrl, pInfo["CAS"])
			# else:
				# urllib_download(imgUrl, pInfo["name"])
		detailInfos = sope.find_all("tr")
		for detailInfo in detailInfos:
			tds = detailInfo.find_all("td")
			if len(tds)==2:
				title = getNodeText(tds[0])
				val = getNodeText(tds[1])
				pInfo[title] = val
		print(pInfo)
		products.append(pInfo.copy())

def getProductList(url, pInfo, products):
	print(url)
	productListHtml = getRenderdHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	productAreas = sope.find_all("div", attrs={"class":"PRODUCT_box"})
	for productArea in productAreas:
		if str(type(productArea))=="<class 'bs4.element.Tag'>":
			link = productArea.find("a", attrs={"class":"name"})
			if link != None:
				getProductInfo('https://www.jkchemical.com'+link["href"],pInfo, products)
	
					
def getProductPage(url, pInfo, products):
	pageUrl = url+"/1.html"
	productListHtml = getRenderdHtmlFromUrl(pageUrl)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	pageCount = sope.find("span", attrs={"id":"ctl00_ContentPlaceHolder1_lbPageCount"})
	if pageCount != None:
		for pageIndex in range(1, int(getNodeText(pageCount))+1):
			getProductList(url+"/"+str(pageIndex)+".html", pInfo, products)
	else:
		getProductList(pageUrl, pInfo, products)
			
					
def getProductType(url, typeCount, pInfo, products):
	productListHtml = getRenderdHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	typeArea = sope.find("ul", attrs={"class":"liststyle"})
	if typeArea != None:
		types = typeArea.find_all("li")
		typeCount = typeCount+1
		for type in types:
			typeLink = type.find("a")
			pInfo["type"+str(typeCount)] = getNodeText(typeLink)
			getProductType('https://www.jkchemical.com'+typeLink["href"], typeCount, pInfo, products )
	else:
		url = url.replace("/zh-cn/product-catalog/","/CH/products/search/productcategory/").replace("/1/10.html","")
		getProductPage(url, pInfo, products)
		
excelFileName="jkchemical.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
			
# getProductInfo('https://www.jkchemical.com/CH/products/A01325386.html',{},products)
getProductType('https://www.jkchemical.com/zh-CN/product-catalog/parent/1322.html',0, {}, products)
headers=['link','type1','type2','type3','type4','cnName','enName','CAS','pNumb','cd','mdl','fzs','fzl','Synonym','Melting\xa0Point',
'Symbol','Signal\xa0Word','Hazard\xa0Statements','Precautionary\xa0Statements','UN','TSCA'
]
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	if rindex%100  == 0:
		wb.save(excelFileName)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
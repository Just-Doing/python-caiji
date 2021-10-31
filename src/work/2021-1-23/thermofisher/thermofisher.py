from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36"}

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		if retryCount< 5:
			getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getTypeLeve1(url, typeCount, pInfo, products,totalLink):
	if pInfo["type"+str(typeCount)] not in totalLink:
		print(str(len(totalLink))+url)
		typeHtml = getHtmlFromUrl(url)
		if typeHtml!=None:
			typeSope = BeautifulSoup(typeHtml, "html.parser",from_encoding="utf-8")
			
			pInfo["title"] = getNodeText(typeSope.find("title"))
			description = typeSope.find("meta", attrs={"property":"og:description"})
			if description!=None:
				pInfo["description"] = description["content"]
			products.append(pInfo.copy())
			navList = typeSope.find_all("div", attrs={"class":"leftnav section"})
			if len(navList) == 0:
				getTypeLeve2(url, typeCount, pInfo, products,totalLink)
			else:
				totalLink.append(pInfo["type"+str(typeCount)])
				for nav in navList:
					typeHeader = nav.find("li", attrs={"class":"item-header nav-header"})
					pInfo["type"+str(typeCount+1)] = getNodeText(typeHeader)
					products.append(pInfo.copy())
					tempProinfo = pInfo.copy()
					typeChArea = nav.find("li", attrs={"class":"item-content"})
					if typeChArea!=None:
						typeChs = typeChArea.find_all("li")
						for typeCh in typeChs:
							typeChLink = typeCh.find("a")
							if typeChLink!=None:
								tempProinfo["type"+str(typeCount+2)] = getNodeText(typeChLink)
								link = "https://www.thermofisher.com"+typeChLink["href"]
								tempProinfo["link"] = link
								getTypeLeve2(tempProinfo["link"], typeCount+2, tempProinfo, products,totalLink)

def getTypeLeve2(url, typeCount, pInfo, products, totalLink):
	print(str(len(totalLink))+url)
	if pInfo["type"+str(typeCount)] not in totalLink:
		totalLink.append(pInfo["type"+str(typeCount)])
		typeHtml = getHtmlFromUrl(url)
		if typeHtml!=None:
			typeSope = BeautifulSoup(typeHtml, "html.parser",from_encoding="utf-8")
			
			pInfo["title"] = getNodeText(typeSope.find("title"))
			description = typeSope.find("meta", attrs={"property":"og:description"})
			if description!=None:
				pInfo["description"] = description["content"]
			products.append(pInfo.copy())
			tempProinfo = pInfo.copy()
			navList = typeSope.find_all("div", attrs={"class":"leftnav section"})
			if len(navList) > 0:
				getTypeLeve1(url, typeCount, pInfo, products,totalLink)
				
			childArea = typeSope.find("ul", attrs={"class":"item-content collapse in"})
			if childArea != None:
				chs = childArea.find_all("li")
				for typeCh in chs:
					typeChLink = typeCh.find("a")
					link = "https://www.thermofisher.com"+typeChLink["href"]
					tempProinfo["link"] = link
					tempProinfo["type"+str(typeCount+1)] = getNodeText(typeChLink)
					getTypeLeve2(tempProinfo["link"], typeCount+1, tempProinfo, products,totalLink)
	
def getType1(typeArea, products, totalLink):
	type1Link = typeArea.find("a")
	t1_pInfo={}
	t1_pInfo["type1"] = getNodeText(type1Link)
	t1_pInfo["link"] = type1Link["href"]
	if t1_pInfo["link"].find("https:") < 0:
		t1_pInfo["link"] = "https:" + t1_pInfo["link"]
	
	if t1_pInfo["type1"] not in totalLink:
		gettype1Html = getHtmlFromUrl(t1_pInfo["link"])
		type1sope = BeautifulSoup(gettype1Html, "html.parser",from_encoding="utf-8")
		totalLink.append(t1_pInfo["type1"])
		t1_pInfo["title"] = getNodeText(type1sope.find("title"))
		description = type1sope.find("meta", attrs={"property":"og:description"})
		if description!=None:
			t1_pInfo["description"] = description["content"]
		products.append(t1_pInfo.copy())
		
		typeChAreas = typeArea.find_all("li", attrs={"class":"nav-column"})
		for type2 in typeChAreas:
			type3Area = type2.find("ul", attrs={"class":"nav-items"})
			if type3Area!=None:
				type2Link = type3Area.previous_sibling
				t2_pInfo={}
				t2_pInfo["type1"] = t1_pInfo["type1"]
				t2_pInfo["type2"] = getNodeText(type2Link)
				if type2Link.name=="a":
					t2_pInfo["link"] = type2Link["href"]
					if t2_pInfo["link"].find("https:") < 0:
						t2_pInfo["link"] = "https:" + t2_pInfo["link"]
					if t2_pInfo["type2"] not in totalLink:
						totalLink.append(t2_pInfo["type2"])
						gettype2Html = getHtmlFromUrl(t2_pInfo["link"])
						if gettype2Html!=None:
							type2sope = BeautifulSoup(gettype2Html, "html.parser",from_encoding="utf-8")
							t2_pInfo["title"] = getNodeText(type2sope.find("title"))
							description = type2sope.find("meta", attrs={"property":"og:description"})
							if description!=None:
								t2_pInfo["description"] = description["content"]
				products.append(t2_pInfo.copy())
					
				type3s = type3Area.find_all("li")
				for type3 in type3s:
					type3Link = type3.find("a")
					t3_pInfo={}
					if type3Link!=None:
						t3_pInfo["type1"] = t1_pInfo["type1"]
						t3_pInfo["type2"] = t2_pInfo["type2"]
						t3_pInfo["type3"] = getNodeText(type3Link)
						t3_pInfo["link"] = type3Link["href"]
						if t3_pInfo["link"].find("https:") < 0:
							t3_pInfo["link"] = "https:" + t3_pInfo["link"]
						getTypeLeve1(t3_pInfo["link"],3, t3_pInfo, products,totalLink)

			
def getProductType(url, products, totalLink):
	productListHtml = getHtmlFromUrl(url)
	if productListHtml!=None:
		sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
		navArea = sope.find("div", attrs={"id":"meganav-extended"})
		type1 = navArea.find("li", attrs={"class":"nav-5-col"})
		type2 = navArea.find("li", attrs={"class":"nav-4-col"})
		type3 = navArea.find("li", attrs={"class":"nav-1-col"})
		type4 = navArea.find("li", attrs={"class":"nav-3-col"})
		
		getType1(type1, products,totalLink)
		# getType1(type2, products,totalLink)
		# getType1(type3, products,totalLink)
		# getType1(type4, products,totalLink)
	
		
excelFileName="thermofisher.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
totalLink = []
# getTypeLeve1("https://www.thermofisher.com/us/en/home/life-science/cell-culture.html", 3, {}, products,totalLink)
getProductType('https://www.thermofisher.com/us/en/home.html', products, totalLink)
headers=['link','type1','type2','type3','type4','type5','type6','type7','type8','title','description']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	if rindex%100  == 0:
		wb.save(excelFileName)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
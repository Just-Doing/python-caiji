from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36"}

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		if retryCount< 5:
			getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	tempPinfo = pInfo.copy()
	productHtml = getHtmlFromUrl(url)
	if productHtml != None:
		tempPinfo["link"] = url
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
		nameArea = sope.find("td", attrs={"width":"550"})
		tempPinfo["name"] = getNodeText(nameArea).replace(tempPinfo["cas"], "")
		
		specs = sope.find_all("tr")
		for spec in specs:
			tds = spec.find_all("td")
			if len(tds) == 2:
				title = getNodeText(tds[0])
				val = getNodeText(tds[1])
				tempPinfo[title] = val
		
		otSpecs = sope.find_all("font", attrs={"color":"#000"})
		for otSpec in otSpecs:
			title = getNodeText(otSpec)
			val = getNodeText(otSpec.parent.nextSibling)
			tempPinfo[title] = val
			
		haSpecs = sope.find_all("font", attrs={"color":"#CC6600"})
		for haSpec in haSpecs:
			title = getNodeText(haSpec)
			if title == "HAZARDOUS MATERIALS INFORMATION" or title == "HAZARDOUS MATERIALS INFORMATION":
				val = getNodeText(haSpec.parent.parent.parent.parent.nextSibling.nextSibling)
				tempPinfo[title] = val
		# print(tempPinfo)
		products.append(tempPinfo.copy())

def getProductList(url, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	productAreas = sope.find("span", attrs={"class":"group"})
	types = productAreas.find_all("li")
	pInfo = {}
	for type in types:
		typeLink = type.find("a")
		pInfo["type"] = getNodeText(typeLink)
		prods = type.find_all("tr")
		for prod in prods:
			prodLink = prod.find("a")
			if prodLink != None:
				pInfo["cas"] = getNodeText(prodLink)
				getProductInfo("https://aksci.com/"+prodLink["href"], pInfo, products)
	
excelFileName="aksci.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
			
# getProductInfo('https://aksci.com/item_detail.php?cat=A660',{"cas":"A660"},products)
getProductList('https://aksci.com/p-organosilicons.php', products)

headers=['link','type','name','CAS Number:','MDL Number:','MF:','MW:','Min. Purity Spec:',
'Physical Form:','Boiling Point:','Flash Point:','Density:','Refractive Index:',
'Long-Term Storage:','Note:','HAZARDOUS MATERIALS INFORMATION','Signal Word',
'Hazard Statements','Precautionary Statements']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	if rindex%100  == 0:
		wb.save(excelFileName)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
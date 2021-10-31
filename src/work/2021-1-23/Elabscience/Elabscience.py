from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36"}

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		if retryCount< 5:
			getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	pInfo["link"]=url
	productHtml = getHtmlFromUrl(url)
	if productHtml != None:
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
		pDetailArea = sope.find("div", attrs={"class":"details01"})
		name = pDetailArea.find("h1")
		cate = name.find("span")
		pInfo["cate"] = getNodeText(cate)
		pInfo["name"] = getNodeText(name).replace(pInfo["cate"], "")
		size = sope.find("ul",attrs={"class":"mid_xial"})
		pInfo["h-size"] = getNodeText(size)
		
		specInfoArea = sope.find("div", attrs={"class":"col-xs-12 col-md-8 base_info"})
		if specInfoArea != None:
			baseInfos = specInfoArea.find_all("li")
			for baseInfo in baseInfos:
				title = getNodeText(baseInfo.find("b"))
				val = getNodeText(baseInfo).replace(title, "")
				if len(title)>0:
					pInfo[title] = val
			
		
		specInfos = sope.find_all("tr")
		for specInfo in specInfos:
			tds = specInfo.find_all("td")
			if len(tds)==2:
				title = getNodeText(tds[0])
				val = getNodeText(tds[1])
				pInfo[title] = val
		
		
		background = sope.find("div", attrs={"id":"dt_tab4"})
		pInfo["background"] = getNodeText(background)
		if background!=None:
			bacArea =  background.find("table", attrs={"class":"dtl_table2"})
			if bacArea != None:
				pInfo["background"] = getNodeText(bacArea)
		
		Kitcomponents=sope.find_all("h4")
		for Kitcomponent in Kitcomponents:
			title = getNodeText(Kitcomponent)
			if title.find("Kit components & Storage")>-1:
				pInfo["Kitcomponents"] = getNodeText(Kitcomponent.nextSibling.nextSibling)
			if title.find("Background")>-1:
				pInfo["background"] = getNodeText(Kitcomponent.nextSibling.nextSibling)
			
		products.append(pInfo.copy())


def getProductList(url, pInfo, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	pListAreas = sope.find_all("div", attrs={"class":"products_box aniview det_con fix"})
	for pListArea in pListAreas:
		link = pListArea.find("a")
		getProductInfo("https://www.elabscience.com"+link["href"], pInfo, products)
	

def getProductPage(url, pInfo, products):
	productListHtml = getHtmlFromUrl(url+"1.html")
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	totalPageCountArea = sope.find("div", attrs={"class":"pro_page03"})
	totalPageCount = getNodeText(totalPageCountArea.find("i"))
	for pageIndex in range(1, int(totalPageCount)):
		pListUrl = url+str(pageIndex)+'.html'
		getProductList(pListUrl, pInfo, products)
		
excelFileName="Elabscience.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

getProductPage('https://www.elabscience.com/search-category=fcm%20antibodies&p=',{},products)
getProductPage('https://www.elabscience.com/search-category=elisa%20kits&p=',{},products)
getProductPage('https://www.elabscience.com/search-category=clia%20kits&p=',{},products)
# getProductInfo('https://www.elabscience.com/p-mouse_tff3_trefoil_factor_3_intestinal_clia_kit-16896.html',{},products)
headers=[
	'link','cate','name','h-size','Host','Applications:','Isotype:'
	,'Alternate Names','Swissprot','References','Form','Clone No.','Host','Isotype','Isotype Control','Reactivity'
	,'Application','Storage Buffer','Recommended Use','Shipping','Stability & Storage','Conjugation','Related Protocols',
	'Troubleshooting','Cat. No','Cat.No.:','Reactivity:','Species','Size:','How do you know Elabscience:'
	,'background','*Verification code:','Assay type','Format','Assay time','Detection range',
	'Sensitivity','Sample type &Sample volume','Specificity','Reproducibility','Kitcomponents',
	'Research Area'
]
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	if rindex%100  == 0:
		wb.save(excelFileName)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
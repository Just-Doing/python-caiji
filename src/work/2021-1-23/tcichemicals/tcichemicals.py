from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36"}

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		if retryCount< 5:
			getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	productHtml = getRenderdHtmlFromUrl(url)
	if productHtml != None:
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
		CAS = sope.find("span",attrs={"class":"cas productVal"})
		CATE = sope.find("span",attrs={"class":"code productVal"})
		pInfo["CAS"] = getNodeText(CAS)
		pInfo["CATE"] = getNodeText(CATE)
		
		img = sope.find("img", attrs={"class":"lazyOwl zoom-img"})
		if img != None:
			if pInfo["CAS"] != "":
				urllib_download("https://www.tcichemicals.com"+ img["src"], pInfo["CAS"])
			else:
				urllib_download(img["src"], pInfo["CATE"])
		
		detailArea = sope.find("div", attrs={"class":"product-details"})
		detailTitles = detailArea.find_all("b")
		for detail in detailTitles:
			title = getNodeText(detail)
			if title.find("Purity:")>-1:
				if str(type(detail.nextSibling.nextSibling))!="<class 'bs4.element.Comment'>":
					pInfo["Purity"] = getNodeText(detail.nextSibling.nextSibling)
		
		SynonymsTitles = detailArea.find_all("div", attrs={"class":"boldText"})
		for detail in SynonymsTitles:
			title = getNodeText(detail)
			if title.find("Synonyms:")>-1:
				if str(type(detail.nextSibling.nextSibling.nextSibling.nextSibling))!="<class 'bs4.element.Comment'>":
					pInfo["Synonyms"] = getNodeText(detail.nextSibling.nextSibling.nextSibling.nextSibling)
		
		specArea = sope.find("div", attrs={"id":"pdp-tabs"})
		specTrs = specArea.find_all("tr")
		for specTr in specTrs:
			tds = specTr.find_all("td")
			if(len(tds) == 2):
				title = getNodeText(tds[0])
				val = getNodeText(tds[1])
				pInfo[title] = val
			
		print(pInfo)
		products.append(pInfo.copy())

def getProductList(url, pInfo, products):
	productListHtml = getRenderdHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	pages = sope.find("ul", attrs={"class":"pagination"})
	if pages !=None:
		for pageIndex in range(0, len(pages.find_all("li"))-2):
			getProductPage(url+'?q=%3Arelevance&page='+str(pageIndex), pInfo, products)
	else:
		getProductPage(url, pInfo, products)
	
					
def getProductPage(url, pInfo, products):
	print(url)
	productListHtml = getRenderdHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	prodTables = sope.find_all("div", attrs={"class":"thumImg"})
	
	for prodTable in prodTables:
		linkInfo = prodTable.find("a")
		if linkInfo!=None:
			getProductInfo("https://www.tcichemicals.com"+linkInfo["href"], pInfo, products )

linkFile = "links.txt"
excelFileName="tcichemicals.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
with open(linkFile, 'r') as file_to_read:
	while True:
		lines = file_to_read.readline()
		if not lines:
			break
			pass
		typePart = lines.split("--")
		pInfo = {}
		urlInfoPart = typePart[len(typePart) - 1].split("@")
		if(len(typePart) == 1):
			pInfo["type1"] = urlInfoPart[0]
		if(len(typePart) == 2):
			pInfo["type1"] = typePart[0]
			pInfo["type2"] = urlInfoPart[0]
		if(len(typePart) == 3):
			pInfo["type1"] = typePart[0]
			pInfo["type2"] = typePart[1]
			pInfo["type3"] = urlInfoPart[0]
		if(len(typePart) == 4):
			pInfo["type1"] = typePart[0]
			pInfo["type2"] = typePart[1]
			pInfo["type3"] = typePart[2]
			pInfo["type4"] = urlInfoPart[0]
			
		getProductList(urlInfoPart[1].strip(),pInfo,products)
# getProductList('https://www.tcichemicals.com/CA/en/c/12760',{},products)
# getProductInfo('https://www.tcichemicals.com/CA/en/p/E0682',{},products)
headers=['link','type1','type2','type3','type4','CAS','CATE','Purity','Synonyms',
'Product Number','Molecular Formula / Molecular Weight','Physical State (20 deg.C)',
'Store Under Inert Gas','Condition to Avoid','CAS RN','Reaxys Registry Number',
'PubChem Substance ID','MDL Number','Appearance','Purity(with Total Nitrogen)','Melting Point',
'Specific Gravity (20/20)','Refractive Index','Pictogram','Signal Word','Hazard Statements','Precautionary Statements','HS Number'
]
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	if rindex%100  == 0:
		wb.save(excelFileName)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
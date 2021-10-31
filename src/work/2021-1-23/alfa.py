from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36"}

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		if retryCount< 5:
			getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	tempPinfo = pInfo.copy()
	productHtml = getHtmlFromUrl(url)
	if productHtml != None:
		tempPinfo["link"] = url
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
		specInfos = sope.find_all("strong")
		for specArea in specInfos:
			spacName = getNodeText(specArea)
			if spacName=="CAS号":
				tempPinfo["cas"] = getNodeText(specArea.parent.nextSibling.nextSibling)
			if spacName=="别名":
				tempPinfo["别名"] = getNodeText(specArea.parent.nextSibling.nextSibling)
			if spacName=="式量":
				tempPinfo["式量"] = getNodeText(specArea.parent.nextSibling.nextSibling)
			if spacName=="Storage & Sensitivity":
				tempPinfo["StorageSensitivity"] = getNodeText(specArea.parent.nextSibling.nextSibling)
			if spacName=="化学式":
				tempPinfo["化学式"] = getNodeText(specArea.parent.nextSibling.nextSibling)
		products.append(tempPinfo.copy())

def getProductList(url, pInfo, products):
	print(url)
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	productAreas = sope.find_all("li", attrs={"class":"list-group-item"})
	for productArea in productAreas:
		if str(type(productArea))=="<class 'bs4.element.Tag'>":
			linkArea = productArea.find("div", attrs={"class":"search-result-number"})
			if linkArea != None:
				link = linkArea.find("a")
				pInfo["cat"] = getNodeText(link)
				pNameArea = productArea.find("div", attrs={"class":"product-name search-result-name"})
				pName = getNodeText(pNameArea)
				cds = re.findall("\d+%", pName)
				if len(cds)>0:
					pInfo["cd"] = cds[0]
				else:
					pInfo["cd"]=""
				pInfo["name"] = pName.replace(pInfo["cat"],"").replace(","+pInfo["cd"],"")
				getProductInfo('https://www.alfa.com'+link["href"], pInfo, products)
	
excelFileName="alfa.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
			
# getProductInfo('https://www.alfa.com/zh-cn/catalog/H61276/',{},products)
getProductList('https://www.alfa.com/zh-cn/silanes/?page=1', {"type":"硅烷"}, products)
getProductList('https://www.alfa.com/zh-cn/silanes/?page=2', {"type":"硅烷"}, products)
getProductList('https://www.alfa.com/zh-cn/silanes/?page=3', {"type":"硅烷"}, products)
getProductList('https://www.alfa.com/zh-cn/silanes/?page=4', {"type":"硅烷"}, products)
getProductList('https://www.alfa.com/zh-cn/silanes/?page=5', {"type":"硅烷"}, products)
getProductList('https://www.alfa.com/zh-cn/silanes/?page=6', {"type":"硅烷"}, products)
getProductList('https://www.alfa.com/zh-cn/silanes/?page=7', {"type":"硅烷"}, products)
getProductList('https://www.alfa.com/zh-cn/silanes/?page=8', {"type":"硅烷"}, products)
getProductList('https://www.alfa.com/zh-cn/silanes/?page=9', {"type":"硅烷"}, products)
getProductList('https://www.alfa.com/zh-cn/silanols/', {"type":"硅醇"}, products)
getProductList('https://www.alfa.com/zh-cn/siloxanes/', {"type":"硅氧烷"}, products)

headers=['link','type','cat','name','cd','cas','别名','式量','StorageSensitivity','化学式']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	if rindex%100  == 0:
		wb.save(excelFileName)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
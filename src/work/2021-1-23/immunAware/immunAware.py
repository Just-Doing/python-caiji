from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):

	url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
	headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36"}

	request_obj=urllib.request.Request(url=url)
	response_obj=urllib.request.urlopen(request_obj)
	html_code=response_obj.read()
	return html_code

def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	pInfo["link"]=url
	productHtml = getHtmlFromUrl(url)
	if productHtml != None:
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
		name = sope.find("h1",attrs={"class":"product_title entry-title"})
		cate = sope.find("span",attrs={"class":"posted_in"})
		tag = sope.find("span",attrs={"class":"tagged_as"})
		introduction = sope.find("div",attrs={"class":"woocommerce-product-details__short-description"})
		imgArea = sope.find("figure",attrs={"class":"woocommerce-product-gallery__wrapper"})
		datasheetLink = sope.find("div",attrs={"style":"width: 100%; margin-bottom: 1em;"})
		pInfo["cate"] = getNodeText(cate).replace("Categories:","")
		pInfo["name"] = getNodeText(name)
		pInfo["introduction"] = getNodeText(introduction)
		pInfo["tag"] = getNodeText(tag)
		deception = sope.find("div", attrs={"id":"tab-description"})
		specInfos = deception.find_all("tr")
		for spec in specInfos:
			tds = spec.find_all("td")
			if len(tds)==3:
				title = getNodeText(tds[0])
				value = getNodeText(tds[2])
				if title=="Peptide" or title=="Formulations":
					value = value +"\n"+ getNodeText(spec.nextSibling.nextSibling)
				if len(title)>0:
					pInfo[title] = value
		if "Catalog#" in pInfo and len(pInfo["Catalog#"])>0:
			pInfo["pdf"]= pInfo["Catalog#"]+'.pdf'
			pInfo["img"]= pInfo["Catalog#"]+'.jpg'
		else:
			pInfo["pdf"]= pInfo["cate"]+str(len(products))+'.pdf'
			pInfo["img"]= pInfo["cate"]+str(len(products))+'.jpg'
			
		# if datasheetLink!=None:
			# urllib_download(datasheetLink.find("a")["href"], pInfo["pdf"])
		# if imgArea!=None:
			# imgLink = imgArea.find("a")
			# urllib_download(imgLink["href"], pInfo["img"])
		products.append(pInfo.copy())


def getProductList(url, pInfo, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	pListAreas = sope.find("ul", attrs={"class":"products columns-4"})
	productAreas = pListAreas.find_all("li")
	for pro in productAreas:
		if str(type(pro))=="<class 'bs4.element.Tag'>":
			link = pro.find("a")
			getProductInfo(link["href"], pInfo, products)
	

def getProductPage(url, pInfo, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	totalPageCountArea = sope.find("ul", attrs={"class":"products columns-4"})
	productAreas = totalPageCountArea.find_all("li")
	for pro in productAreas:
		link = pro.find("a")
		pInfo["type"]=getNodeText(link)
		getProductList(link["href"], pInfo, products)
		
excelFileName="Elabscience.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo('https://immunaware.com/product/h2-kb-siinfekl/',{} ,products)
# getProductInfo('https://immunaware.com/product/1001-01/',{} ,products)
getProductPage('https://immunaware.com/product-category/easymer/',{} ,products)
getProductList('https://immunaware.com/product-category/easymer/hla-b-easymer/page/2/',{"type":"HLA-B"} ,products)
getProductPage('https://immunaware.com/product-category/mhc-class-i/',{} ,products)
getProductPage('https://immunaware.com/product-category/mhc-class-ii/',{} ,products)
getProductList('https://immunaware.com/product-category/mhc-class-i/hla-a/page/2/',{"type":"HLA-A"} ,products)
getProductList('https://immunaware.com/product-category/mhc-class-i/hla-b/page/2/',{"type":"HLA-B"} ,products)

headers=[
	'link','type','name','cate','tag','introduction','Catalog#','Size','easYmer','Folding buffer',
	'Peptide','Organism','Epitope','HLA','Formulations','Buffer','Application','pdf','img'
]
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	if rindex%100  == 0:
		wb.save(excelFileName)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36"}

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		if retryCount< 5:
			getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, products):
	print(str(len(products)) + url)
	tempPinfo = {}
	productHtml = getRenderdHtmlFromUrl(url)
	if productHtml != None:
		tempPinfo["link"] = url
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
	
		nameArea = sope.find("div",attrs={"class":"inlineBlock"})
		name = nameArea.find("strong")
		tempPinfo["name"] = getNodeText(name)
		
		ispecs = sope.find_all("i")
		for ispec in ispecs:
			title = getNodeText(ispec)
			if title.find("Structure of")> -1:
				tempPinfo["Structure"] = title.replace("Structure of","")
				
				
		sspecs = sope.find_all("span", attrs={"class":"orange-star"})
		for sspec in sspecs:
			if str(type(sspec.nextSibling))=="<class 'bs4.element.NavigableString'>" and str(sspec.nextSibling).find("Storage:")>-1:
				tempPinfo["Storage"] = str(sspec.nextSibling).replace("Storage:","")
				# print(tempPinfo["Storage"])
		specs = sope.find_all("tr")
		for spec in specs:
			tds = spec.find_all("td")
			if len(tds) == 2:
				title = getNodeText(tds[0])
				val = getNodeText(tds[1])
				tempPinfo[title] = val
			if len(tds) == 4:
				title1 = getNodeText(tds[0])
				val1 = getNodeText(tds[1])
				title2 = getNodeText(tds[2])
				val2 = getNodeText(tds[3])
				tempPinfo[title1] = val1
				tempPinfo[title2] = val2
		# print(tempPinfo.copy())
		products.append(tempPinfo.copy())

def getProductList(url, products):
	print(url)
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	productAreas = sope.find_all("dl", attrs={"class":"sales-dl position-r"})
	for productArea in productAreas:
		link = productArea.find("a")
		# print(link["href"])
		getProductInfo("https://www.ambeed.com"+link["href"], products)
	
excelFileName="ambeed.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
			
# getProductInfo('https://www.ambeed.com/products/13183-70-5.html',{},products)
# getProductList('https://www.ambeed.com/organosilicon.html?pagesize=20&pageindex=1', {"type":"Siloxane Polymers"}, products)
for pageIndex in range(1, 14):
	getProductList("https://www.ambeed.com/organosilicon.html?pagesize=20&pageindex="+str(pageIndex)+"&product_list_limit=25", products)

headers=['link','name','Structure','Storage','CAS No. :','Formula :',
'Linear Structure Formula :','M.W :','Synonyms :','MDL No. :',
'Boiling Point :','InChI Key :','Pubchem ID :','Signal Word:','Precautionary Statements:',
'GHS Pictogram:']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	if rindex%100  == 0:
		wb.save(excelFileName)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
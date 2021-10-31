from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36"}

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		if retryCount< 5:
			getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	tempPinfo = pInfo.copy()
	productHtml = getHtmlFromUrl(url)
	if productHtml != None:
		tempPinfo["link"] = url
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
	
		nameArea = sope.find("div",attrs={"class":"inlineBlock"})
		specs = sope.find_all("tr")
		for spec in specs:
			tds = spec.find_all("td")
			if len(tds) == 2:
				title = getNodeText(tds[0])
				val = getNodeText(tds[1])
				tempPinfo[title] = val
		
		h2specs = sope.find_all("h2")
		for h2spec in h2specs:
			title = getNodeText(h2spec)
			val = getNodeText(h2spec.nextSibling.nextSibling)
			if title == tempPinfo["name"]+" Description":
				tempPinfo["Description"] = val
			if title == tempPinfo["name"]+" Application":
				tempPinfo["Application"] = val
			if title == tempPinfo["name"]+" Packaging":
				tempPinfo["Packaging"] = val
			if title == tempPinfo["name"]+" Storage":
				tempPinfo["Storage"] = val
		products.append(tempPinfo.copy())
		print(tempPinfo)

def getProductList(url, products):
	print(url)
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	productAreas = sope.find("div", attrs={"class":"product-large-list"})
	pros = productAreas.find_all("li", attrs={"class":"col-md-6"})
	for pro in pros:
		link = pro.find("a")
		if link!=None:
			pInfo={}
			pInfo["name"] = getNodeText(link)
			getProductInfo("https://www.cfmats.com"+link["href"], pInfo, products)
	
excelFileName="cfmats.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
			
# getProductInfo('https://www.cfmats.com/functional-silanes/3-isocyanatopropyltrimethoxysilane.html',{"name":"3-Isocyanatopropyltrimethoxysilane"},products)
# getProductList('https://www.cfmats.com/functional-silanes/?page=1', products)
for pageIndex in range(1, 11):
	getProductList("https://www.cfmats.com/functional-silanes/?page="+str(pageIndex), products)

headers=['link','name','Cas No','Formula','Color','Appearance',
'Purity','Description','Application','Packaging',
'Storage']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	if rindex%100  == 0:
		wb.save(excelFileName)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
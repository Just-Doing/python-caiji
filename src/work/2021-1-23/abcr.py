from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36"}

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		if retryCount< 5:
			getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	tempPinfo = pInfo.copy()
	productHtml = getHtmlFromUrl(url)
	if productHtml != None:
		tempPinfo["link"] = url
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
	
		specArea = sope.find("div",attrs={"class":"attributes-table-wrapper"})
		specTrs = specArea.find_all("tr")
		for specTr in specTrs:
			tds = specTr.find_all("td")
			if len(tds) == 2:
				title = getNodeText(tds[0])
				val = getNodeText(tds[1])
				tempPinfo[title] = val
		print(tempPinfo)	
		products.append(tempPinfo.copy())

def getProductList(url, pInfo, products):
	print(url)
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	productAreas = sope.find_all("div", attrs={"class":"product details product-item-details"})
	for productArea in productAreas:
		if str(type(productArea))=="<class 'bs4.element.Tag'>":
			linkArea = productArea.find("div", attrs={"class":"product-details-button-container"})
			cas = productArea.find("a", attrs={"class":"product-item-link"})
			casPart = cas["data-mage-init"].split('|')
			if len(casPart) > 1:
				pInfo["cas"] = casPart[1].replace('"}','')
			name = productArea.find("h3", attrs={"class":"secondary-product-name"})
			pInfo["name"] = getNodeText(name)
			link = linkArea.find("a")
			print(pInfo)
			getProductInfo(link["href"], pInfo, products)
	
excelFileName="abcr.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
			
# getProductInfo('https://abcr.com/de_en/ab426064',{},products)
# getProductList('https://abcr.com/de_en/products/silanes-siloxanes/siloxane-polymers?p=1&product_list_limit=25', {"type":"Siloxane Polymers"}, products)
for pageIndex in range(1, 17):
	getProductList('https://abcr.com/de_en/products/silanes-siloxanes/siloxane-polymers?p="+str(pageIndex)+"&product_list_limit=25', {"type":"Siloxane Polymers"}, products)
for pageIndex in range(1, 12):
	getProductList('https://abcr.com/de_en/products/silanes-siloxanes/inorganic-silicon-compounds?p="+str(pageIndex)+"&product_list_limit=25', {"type":"Inorganic Silicon Compounds"}, products)
for pageIndex in range(1, 124):
	getProductList('https://abcr.com/de_en/products/silanes-siloxanes/organic-silanes?p="+str(pageIndex)+"&product_list_limit=25', {"type":"Organic Silanes"}, products)

headers=['link','type','cas','name','Sum formula','Molecular weight','Density','Melting point','Boiling point','Flash point','Signal word','Hazard statements','Precautionary statements']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	if rindex%100  == 0:
		wb.save(excelFileName)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
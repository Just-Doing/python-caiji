from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36"}

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		if retryCount< 5:
			getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	pInfo["link"]=url
	productHtml = getHtmlFromUrl(url)
	if productHtml != None:
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
		name = sope.find("h1",attrs={"itemprop":"name headline"})
		pInfo["name"] = getNodeText(name)
		img = sope.find("img", attrs={"alt":"molecular structure"})
		specInfos = sope.find_all("h3", attrs={"class":"field-label inline-sibling"})
		for spec in specInfos:
			title = getNodeText(spec)
			val = getNodeText(spec.nextSibling)
			if title=="CAS #:":
				pInfo["CAS"] = val
			if title=="Linear Formula:":
				pInfo["Linear Formula"] = val
			if title=="MDL Number:":
				pInfo["MDL Number"] = val
			if title=="EC No.:":
				pInfo["EC No"] = val
		
		if img != None:
			if pInfo["CAS"] != "":
				urllib_download(img["src"], pInfo["CAS"])
			else:
				urllib_download(img["src"], pInfo["name"])
		
		infoTrs = sope.find_all("tr")
		for infoTr in infoTrs:
			infoThs = infoTr.find_all("th")
			infoTds = infoTr.find_all("td")
			if len(infoThs) == 1 and len(infoTds)==1:
				title = getNodeText(infoThs[0])
				val = getNodeText(infoTds[0])
				pInfo[title]=val
		AboutAluminum = sope.find("div", attrs={"class":"field-prod-para"})
		pInfo["AboutAluminum"] = getNodeText(AboutAluminum)
		AluminumBorohydride = sope.find("div", attrs={"class":"field-synonyms"})
		pInfo["AluminumBorohydride"] = getNodeText(AluminumBorohydride)	
		products.append(pInfo.copy())

def getProductList(url, pInfo, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	type1s = sope.find_all("div", attrs={"class":"panel panel-default greentech-panel"})
	for type1 in type1s:
		type1Txt = getNodeText(type1.find("h2", attrs={"class":"panel-title"}))
		pInfo["type1"]=type1Txt
		type2s = type1.find_all("h4")
		for type2 in type2s:
			type2Txt = getNodeText(type2)
			pInfo["type2"]=type2Txt
			if type2.nextSibling.nextSibling.name == "ul":
				productLis = type2.nextSibling.nextSibling.find_all("li")
				for product in productLis:
					if str(type(product))=="<class 'bs4.element.Tag'>":
						link = product.find("a")
						if link != None:
							getProductInfo("https://www.americanelements.com/"+link["href"], pInfo, products)	
			else:
				products1 = type2.nextSibling.nextSibling.find_all("li")
				products2 = type2.nextSibling.nextSibling.nextSibling.nextSibling.find_all("li")
				for product in products1:
					if str(type(product))=="<class 'bs4.element.Tag'>":
						link = product.find("a")
						if link != None:
							getProductInfo("https://www.americanelements.com/"+link["href"], pInfo, products)	
				for product in products2:
					if str(type(product))=="<class 'bs4.element.Tag'>":
						link = product.find("a")
						if link != None:
							getProductInfo("https://www.americanelements.com/"+link["href"], pInfo, products)	
				
		


excelFileName="americanelements.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

getProductList('https://www.americanelements.com/green-technology-alternative-energy.html#led-epitaxy',{},products)
# getProductInfo('https://www.americanelements.com/aluminum-borohydride-16962-07-5',{},products)
headers=['link','name','CAS','Linear Formula','MDL Number','EC No','Compound Formula','Molecular Weight','Appearance',
'Melting Point','Boiling Point','Density','Solubility in H2O','Exact Mass','Monoisotopic Mass','Charge','Signal Word','Hazard Statements',
'Hazard Codes','RTECS Number','Transport Information','MSDS / SDS','EC No.'
,'Beilstein/Reaxys No.','Pubchem CID','IUPAC Name','SMILES','InchI Identifier','InchI Key','AboutAluminum'
]
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	if rindex%100  == 0:
		wb.save(excelFileName)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
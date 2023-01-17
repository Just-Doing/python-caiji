from ast import Store
from asyncio.windows_events import NULL
from cProfile import label
import datetime
from enum import IntEnum
from pydoc import classname
from attr import attrs
import requests
from urllib.request import urlopen
import urllib
from selenium import webdriver
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import cfscrape
import json
import string
import re
import time
import math
import _thread
from dateutil.parser import parse
import ssl

import numpy as np

ssl._create_default_https_context = ssl._create_unverified_context
http.client._MAXHEADERS = 1000


def loads_str(data_str):
    try:
        result = json.loads(data_str)
        return result
    except Exception as e:
        error_index = re.findall(r"char (\d+)\)", str(e))
        if error_index:
            error_str = data_str[int(error_index[0])]
            data_str = data_str.replace(error_str, "<?>")


def urllib_download(IMAGE_URL, pName):
    try:
        opener = urllib.request.build_opener()
        opener.addheaders = [('User-agent', 'Mozilla/5.0'), ('cookie',
                                                             '__cf_bm=vbLxkjIVjoGJnWBrrRZP1GCsbs_KmubMaHNugmyyR3I-1659249299-0-AXJuOuT6vZtFzKG57pHX3sgKssRwKg9sXDpFaqkl5hXjPCCmxv95Lj76noiZ90Rm6c0kXSv/oW51uHGWcgUQlTE=')]
        urllib.request.install_opener(opener)
        urllib.request.urlretrieve(
            IMAGE_URL, pName.replace("/", "").replace("\\", ""))
    except:
        print('no')


def getNodeText(node):
    if (node == None):
        return ""
    else:
        return node.get_text().strip()


def getHtmlFromUrl(url, type="get", para={}):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.82 Safari/537.36",
        "cookie": ""
    }

    scraper = cfscrape.create_scraper()
    html_code = scraper.get(url, headers=headers).text
    return BeautifulSoup(html_code, "html.parser", from_encoding="utf-8")


def requestJson(url, pIndex):
    r = requests.post(url, data={"rpc": "100", "page": str(pIndex), "TaxonomicTags": "ELISA Pair Set"}, headers={
        'Content-Type': 'multipart/form-data;',
        'cookie': 'Hm_lvt_2d911fad88dfe99ff8bbf610824f3e67=1659233185; _gid=GA1.2.1110839465.1659233185; sbb=btzDOmbIJgtcJ%252fbBNFFakzEjiCmqYRsxckyc44YEsdI%253dXTaJopSyq01ihC4cD5jOfAa8AEgsWX792EAllASK%252bEKohz0p3JxSEJFavoBnvUmw4fhBpwMcWGZ2Qy%252bRRF2U%252bMSxcQdMfdwOcT%252bR%252bo7qyEU%252br8SBQuGE8GJWgDFeSTZ4gyQPbdRAfENtzDDqN8%252bGDluhZO1JALDLDvSQerjksoBXs7EVLgUw38PzM7oZm%252fQOJh74B1FrVGP39YP6NQ3ZjAvyki%252bPYATzUhPmLGBWfM6tKNdk2Xw0AJPu9svNuRInCKAY7cAiDRAl4QtVn%252fve5n66%252bTerGofcS4TcLhPpxKB%252f7UXS2Ukwrf%252bnxAim9gwczRgv6d4X4QsvzE9g25CJrg%253d%253d; _gcl_au=1.1.894241471.1659233186; cebs=1; __cf_bm=pxO5h0EnM.XSzjcys0ufAg42veHF8cDhJHzWM10WfUg-1659236014-0-AaB0w5RWqok9U/TT4voV03gLXrP9ZorKOzdpKao7/vF5cTxcc/ybitC+pwPfrFvJOJEspqyzKpl8LYk49W92HjU=; _ce.s=v~17fa36a1a61ac4f924605a0ea259c526716d971d~vpv~0~v11.rlc~1659236059303; _ga_HYV7JHQNBH=GS1.1.1659236057.2.1.1659236799.0; _ga=GA1.2.60586722.1659233185; _gat_gtag_UA_9748282_4=1; Hm_lpvt_2d911fad88dfe99ff8bbf610824f3e67=1659236800; Currency=RMB; LocationCode=CN; cebsp=7',
        "User-Agent": "PostmanRuntime/7.29.0"
    })
    print(r.text)

def getRenderdHtmlFromUrl(browser, url):
    browser.get(url)
    data4CommentsSope = BeautifulSoup(browser.page_source, "html.parser")
    data4CommentsStr = getNodeText(data4CommentsSope.find("body"))
    if data4CommentsStr.find("400 Bad Request") > -1 or data4CommentsStr.find("Internal Server Error") > -1 or data4CommentsStr.find("Internal Server Error, real status: 503") > -1:
        time.sleep(20)
        browser.get(url)
        data4CommentsSope = BeautifulSoup(browser.page_source, "html.parser")
        data4CommentsStr = getNodeText(data4CommentsSope.find("body"))
        if data4CommentsStr.find("400 Bad Request") > -1 or data4CommentsStr.find("Internal Server Error") > -1 or data4CommentsStr.find("Internal Server Error, real status: 503") > -1:
            time.sleep(20)
            browser.get(url)
            data4CommentsSope = BeautifulSoup(browser.page_source, "html.parser")
            data4CommentsStr = getNodeText(data4CommentsSope.find("body"))
    
    return json.loads(data4CommentsStr)


def writeExcel(workSheet, headers, rowIndex, info):
    cellIndex = 1
    for head in headers:
        try:
            if head in info:
                content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
                workSheet.cell(rowIndex, cellIndex).value = content.strip()
            else:
                workSheet.cell(rowIndex, cellIndex).value = ""
        except:
            print(info)
            print(rowIndex)
        cellIndex = cellIndex+1


def getWeiBoInfo(browser, products):
    weiBoArticles = []
    likedWeibo = []
    pinglun = []
    fensiListRes = []
    guanzhuListRes = []
    regReplace = re.compile(u'['u'\U0001F300-\U0001F64F' u'\U0001F680-\U0001F6FF' u'\u0000-\u26ff \U00100000-\U0010ffff \U00002700-\U000027ff]+')
    for index,p in enumerate(products):
        print(str(index)+"/"+str(len(products)))
        if index % 10 == 0:
            browser.get("https://weibo.com/p/100101B2094757D069A6FD4299")
        bozhuName = p["bozhuName"]
        data4UserInfo = getRenderdHtmlFromUrl(browser, "https://weibo.com/ajax/profile/info?uid="+p["userId"])
        #获取博主的所有文章
        for pIndex in range(1, 7):
            data = getRenderdHtmlFromUrl(browser, "https://weibo.com/ajax/statuses/mymblog?uid="+p["userId"]+"&page="+str(pIndex)+"&feature=0")
            if len(data["data"]["list"]) == 0:
                break;
            for article in data["data"]["list"]:
                userObj = article["user"]
                userName = ""
                if "screen_name" in userObj:
                    userName = userObj["screen_name"]
                else:
                    print(userObj)
                articleSope =  BeautifulSoup(article["text"], "html.parser")
                weiboContent =  re.sub(regReplace,'',getNodeText(articleSope))
                address = ""
                if "region_name" in article:
                    address = article["region_name"]
                #如果博文作者不是博主名称，则为赞过的微博
                if bozhuName == userName:
                    weiboId = str(len(weiBoArticles))
                    weiBoArticles.append({
                        "weiboId": weiboId,
                        "bozhuName": userName,
                        "gender": data4UserInfo["data"]["user"]["gender"],
                        "fensi": str(data4UserInfo["data"]["user"]["followers_count"]),
                        "guanzhu": str(data4UserInfo["data"]["user"]["friends_count"]),
                        "address": address, 
                        "weiboContent": weiboContent, 
                        "publishTime": parse(article["created_at"]).strftime("%y-%m-%d %H:%M:%S")
                    })
                    commentsUrl = "https://weibo.com/ajax/statuses/buildComments?is_reload=1&id="+str(article["id"])+"&is_show_bulletin=2&is_mix=0&count=20&type=feed&uid="+str(article["user"]["id"])
                    
                    data4Comments = getRenderdHtmlFromUrl(browser, commentsUrl )

                    for pinlun in data4Comments["data"]:
                        pinlunUserObj = pinlun["user"]
                        pinglun.append({
                            "weiboId": weiboId,
                            "pinlunRen": pinlunUserObj["screen_name"] if "screen_name" in  pinlunUserObj else "",
                            "pinlunContent": pinlun["text"],
                            "pinlunTime": parse(pinlun["created_at"]).strftime("%y-%m-%d %H:%M:%S")
                        })
                else:
                    if "title" in article:
                        data4LikedUser = getRenderdHtmlFromUrl(browser, "https://weibo.com/ajax/profile/info?uid="+str(article["user"]["id"]))
                        data4LikedUserDetail = getRenderdHtmlFromUrl(browser, "https://weibo.com/ajax/profile/detail?uid="+str(article["user"]["id"]))
                        labels = ""
                        data4LikedUserObj = data4LikedUserDetail["data"]
                        if "label_desc" in data4LikedUserObj and len(data4LikedUserObj["label_desc"]) >0:
                            for label in data4LikedUserObj["label_desc"]:
                                labels += label["name"]+"|"
                        title = article["title"]["text"]
                        verified_reason = ""
                        if "verified_reason" in data4LikedUser["data"]["user"]:
                            verified_reason = data4LikedUser["data"]["user"]["verified_reason"]
                        likedWeibo.append({
                            "bozhuName": bozhuName, 
                            "userName": userName, 
                            "gender": data4LikedUserObj["gender"],
                            "beiZhanRenUrl": "https://weibo.com/u/"+str(article["user"]["id"]), 
                            "address": address, 
                            "weiboContent": weiboContent, 
                            "publishTime": parse(article["created_at"]).strftime("%y-%m-%d %H:%M:%S"), 
                            "likedTime": title.replace("赞过的微博","").replace("她","").replace("他",""),
                            "renzheng": verified_reason,
                            "fenshi": str(data4LikedUser["data"]["user"]["followers_count"]),
                            "guanzhu": str(data4LikedUser["data"]["user"]["friends_count"]),
                            "labels": labels
                        })
            
        #获取粉丝标记
        for pIndex in range(1, 100):
            data = getRenderdHtmlFromUrl(browser, "https://weibo.com/ajax/friendships/friends?relate=fans&page="+str(pIndex)+"&uid="+p["userId"]+"&type=all&newFollowerCount=0")
            if "users" in data:
                users = data["users"]
                for user in users:
                    data4FenShiUserDetail = getRenderdHtmlFromUrl(browser, "https://weibo.com/ajax/profile/detail?uid="+str(user["id"]))
                    data4FenShiUserInfo = getRenderdHtmlFromUrl(browser, "https://weibo.com/ajax/profile/info?uid="+str(user["id"]))

                    data4FenShiUserObj = data4FenShiUserDetail["data"]
                    data4FenShiUserInfoObj = data4FenShiUserInfo["data"]
                    labels = ""
                    if "label_desc" in data4FenShiUserObj and len(data4FenShiUserObj["label_desc"]) >0:
                        for label in data4FenShiUserObj["label_desc"]:
                            labels += label["name"]+"|"
                    fensiListRes.append({
                        "bozhuName": bozhuName,
                        "fensiName": user["name"],
                        "profile_url": user["profile_url"],
                        "gender": data4FenShiUserInfoObj["user"]["gender"],
                        "fensi": str(user["followers_count"]),
                        "guanzhu": str(user["friends_count"]),
                        'labels': labels,
                        'renzheng': data4FenShiUserObj["desc_text"] if "desc_text" in data4FenShiUserObj else ""
                    })
                if len(users) < 20: break;

        #获取关注标记
        for pIndex in range(1, 100):
            data = getRenderdHtmlFromUrl(browser, "https://weibo.com/ajax/friendships/friends?page="+str(pIndex)+"&uid="+p["userId"])
            if "users" in data:
                users = data["users"]
                for user in users:
                    data4GuanZhuUserDetail = getRenderdHtmlFromUrl(browser, "https://weibo.com/ajax/profile/detail?uid="+str(user["id"]))
                    data4GuanZhuUserInfo = getRenderdHtmlFromUrl(browser, "https://weibo.com/ajax/profile/info?uid="+str(user["id"]))
                    data4GuanZhuUserObj = data4GuanZhuUserDetail["data"]
                    data4GuanZhuUserInfoObj = data4GuanZhuUserInfo["data"]
                    labels = ""
                    if "label_desc" in data4GuanZhuUserObj and len(data4GuanZhuUserObj["label_desc"]) >0:
                        for label in data4GuanZhuUserObj["label_desc"]:
                            labels += label["name"]+"|"
                    guanzhuListRes.append({
                        "bozhuName": bozhuName,
                        "guanzhuName": user["name"],
                        "profile_url": user["profile_url"],
                        "gender": data4GuanZhuUserInfoObj["user"]["gender"],
                        "fensi": str(user["followers_count"]),
                        "guanzhu": str(user["friends_count"]),
                        'labels': labels,
                        'renzheng': data4GuanZhuUserObj["desc_text"] if "desc_text" in data4GuanZhuUserObj else ""
                    })
                if len(users) < 20: break;


    return {"weiBoArticles":weiBoArticles, "likedWeibo": likedWeibo, "pinglun":pinglun, "fensiListRes":fensiListRes,"guanzhuListRes":guanzhuListRes}


def getAuthorList(url,  products):
    chrome_options = webdriver.ChromeOptions()
    # chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument("window-size=1024,768")

    chrome_options.add_argument("--no-sandbox")
    browser = webdriver.Chrome(chrome_options=chrome_options)
    browser.maximize_window()
    browser.get(url)
    # browser.add_cookie({'domain': '.sina.com.cn', 'httpOnly': False, 'name': 'ALF', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': '1695213744'})
    # browser.add_cookie({'domain': '.sina.com.cn', 'httpOnly': False, 'name': 'SUBP', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': '0033WrSXqPxfM725Ws9jqgMF55529P9D9Wh67rqn0jCF8RVZkpC54nWZ5NHD95Qceh5Xeon7SKe7Ws4Dqcj3i--Ni-iWi-2Ei--ciK.RiKLsi--4iK.Ni-8Wi--4iK.Ni-8WIrHjIgHX'})
    # browser.add_cookie({'domain': '.sina.com.cn', 'httpOnly': True, 'name': 'SCF', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': 'AjUvWuiha4N_hVf6kE9Shy6oTw93Q8gi5jAiLCjseMCxNRYNeE8g48T87JwdVmIt_TZgZX9Lxeamm-bhLVP45Mk.'})
    # browser.add_cookie({'domain': 'place.weibo.com', 'httpOnly': False, 'name': 'PHPSESSID', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': 'c1f007db93f862677d898e02dca75238'})
    browser.add_cookie({'domain': '.weibo.com', 'httpOnly': False, 'name': 'SUB', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': '_2A25OQrNYDeRhGeBO61YT-C7JyDyIHXVtOaOQrDV8PUNbmtAfLXHwkW9NSlmktIkmJfBDMuyjHrhbo18MUpRpN3T-'})
    browser.add_cookie({'domain': '.weibo.com', 'httpOnly': False, 'name': 'PC_TOKEN', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': '4815cba76b'})
    browser.add_cookie({'domain': '.weibo.com', 'httpOnly': False, 'name': 'WBPSESS', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': '-slJzTKhzDzKK5KM1fl5TcM--I2sb6AcKhCyrd1qF-vBbOlAz13CO9r8MJ_5vuhCfI5eXErjZnRoihmmEAw4d53KSz-7DVl-dkWlj0RqHN4kQzZPvuKXcNQreiROghIj'})
    browser.add_cookie({'domain': '.weibo.com', 'httpOnly': False, 'name': 'XSRF-TOKEN', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': 'pZ_wUrQ1NLgct4-XRZAlUzSe'})
    browser.add_cookie({'domain': '.weibo.com', 'httpOnly': False, 'name': 'SUBP', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': '0033WrSXqPxfM725Ws9jqgMF55529P9D9Wh67rqn0jCF8RVZkpC54nWZ5JpX5KMhUgL.Foq7ehBE1h5fe052dJLoIEQLxKMLB.2LBKzLxKqL1KnL1-qLxK.L1KMLB-2LxK.L1KMLB-83UgpD9PBt'})
    browser.add_cookie({'domain': '.weibo.com', 'httpOnly': False, 'name': 'webim_unReadCount', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': '%7B%22time%22%3A1665234784317%2C%22dm_pub_total%22%3A4%2C%22chat_group_client%22%3A0%2C%22chat_group_notice%22%3A0%2C%22allcountNum%22%3A46%2C%22msgbox%22%3A0%7D'})
    browser.add_cookie({'domain': '.weibo.com', 'httpOnly': False, 'name': '_s_tentry', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': 'login.sina.com.cn'})
    browser.add_cookie({'domain': '.weibo.com', 'httpOnly': False, 'name': 'Apache', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': '4283487584599.3394.1665311825975'})
    browser.add_cookie({'domain': '.weibo.com', 'httpOnly': False, 'name': 'UOR', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': ',,login.sina.com.cn'})
    browser.add_cookie({'domain': '.weibo.com', 'httpOnly': False, 'name': 'ALF', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': '1697117829'})
    browser.add_cookie({'domain': '.weibo.com', 'httpOnly': False, 'name': 'SSOLoginState', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': '1665232284'})
    browser.add_cookie({'domain': '.weibo.com', 'httpOnly': False, 'name': 'SCF', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': 'AjUvWuiha4N_hVf6kE9Shy6oTw93Q8gi5jAiLCjseMCxgrJKgerONTNNrRf2p99W0dLKWiH75cjTcFwYu3EzeIM.'})
    browser.add_cookie({'domain': '.weibo.com', 'httpOnly': False, 'name': 'ULV', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': '1665311825978:11:2:1:4283487584599.3394.1665311825975:1665232286884'})
    browser.add_cookie({'domain': '.weibo.com', 'httpOnly': False, 'name': 'wb_view_log_6004280530', 'path': '/', 'sameSite': 'None', 'secure': True, 'value': '1920*10801'})
    browser.add_cookie({'domain': '.weibo.com', 'httpOnly': False, 'name': 'SINAGLOBAL', 'path': '/', 'sameSite': 'None', 'secure': False, 'value': '2243265361707.869.1663339298312'})

    browser.get(url)
    time.sleep(10)
    #获取所有 人员
    sope = BeautifulSoup(browser.page_source, "html.parser")
    users = sope.find_all("div", attrs={"class":"WB_info"})
    for user in users:
        # userDetail = user.find("div", attrs={"class":"WB_info"})
        userLink = user.find("a", attrs={"class":"W_f14 W_fb S_txt1"})
        if userLink != None:
            name = getNodeText(userLink)
            if name != "金华职业技术学院":
                products.append({
                    'bozhuName': name, 'url': userLink["href"], 
                    'userId': userLink["usercard"].replace("id=","").split("&")[0]
                })
    #下拉滚动条至最底部   列表页共3次
    browser.execute_script("window.scrollTo(0,document.body.scrollHeight)")
    time.sleep(5)
    sope = BeautifulSoup(browser.page_source, "html.parser")
    users = sope.find_all("div", attrs={"class":"WB_info"})
    for user in users:
        # userDetail = user.find("div", attrs={"class":"WB_info"})
        userLink = user.find("a", attrs={"class":"W_f14 W_fb S_txt1"})
        if userLink != None:
            name = getNodeText(userLink)
            if name != "金华职业技术学院":
                products.append({
                    'bozhuName': name, 'url': userLink["href"], 
                    'userId': userLink["usercard"].replace("id=","").split("&")[0]
                })
    #下拉滚动条至最底部   列表页共3次
    browser.execute_script("window.scrollTo(0,document.body.scrollHeight)")
    time.sleep(5)
    sope = BeautifulSoup(browser.page_source, "html.parser")
    users = sope.find_all("div", attrs={"class":"WB_info"})
    for user in users:
        # userDetail = user.find("div", attrs={"class":"WB_info"})
        userLink = user.find("a", attrs={"class":"W_f14 W_fb S_txt1"})
        if userLink != None:
            name = getNodeText(userLink)
            if name != "金华职业技术学院":
                products.append({
                    'bozhuName': name, 'url': userLink["href"], 
                    'userId': userLink["usercard"].replace("id=","").split("&")[0]
                })
    return browser



excelFname = datetime.datetime.now().strftime("%y%m%d%H%M%S")+".xlsx"
products = []
browser = NULL
for pIndex in range(1, 25):
    browser=getAuthorList('https://weibo.com/p/100101B2094757D069A6FD4299?feed_filter=filter&feed_sort=filter&current_page=' +
                   str(pIndex*3-1)+'&since_id=&page='+str(pIndex), products)
weobo = getWeiBoInfo(browser, products)

#测试代码
# browser = getAuthorList('https://weibo.com/p/100101B2094757D069A6FD4299?feed_filter=filter&feed_sort=filter&current_page=0&since_id=&page=1', products)
# weobo = getWeiBoInfo(browser, [{"bozhuName":"See4amsun","userId":"7541086833"}])
#测试代码

wb = Workbook()
sheet1 = wb.create_sheet(title="微博")
sheet2 =wb.create_sheet(title="评论")
sheet3 =wb.create_sheet(title="赞过的微博")
fensiSheet =wb.create_sheet(title="粉丝")
guanzhuSheet =wb.create_sheet(title="关注")


headers1 = [
    'weiboId', 'bozhuName','gender','fensi','guanzhu','address','weiboContent','publishTime'
]
headers2 = [
    'weiboId', 'pinlunRen','pinlunContent','pinlunTime'
]
headers3 = [
    'bozhuName', 'userName','gender', 'address','weiboContent','publishTime','likedTime','beiZhanRenUrl','renzheng','fenshi','guanzhu','labels'
]

fensiHeader = ['bozhuName','fensiName','profile_url','fensi','guanzhu','gender', 'labels','renzheng']
guanzhuHeader = ['bozhuName','guanzhuName','profile_url','fensi','guanzhu','gender', 'labels','renzheng']

for index, head in enumerate(headers1):
    sheet1.cell(1, index+1).value = head.strip()
for index, p in enumerate(weobo["weiBoArticles"]):
    writeExcel(sheet1, headers1, index + 2, p)


for index, head in enumerate(headers2):
    sheet2.cell(1, index+1).value = head.strip()
for index, p in enumerate(weobo["pinglun"]):
    writeExcel(sheet2, headers2, index + 2, p)



for index, head in enumerate(headers3):
    sheet3.cell(1, index+1).value = head.strip()
for index, p in enumerate(weobo["likedWeibo"]):
    writeExcel(sheet3, headers3, index + 2, p)



for index, head in enumerate(fensiHeader):
    fensiSheet.cell(1, index+1).value = head.strip()
for index, p in enumerate(weobo["fensiListRes"]):
    writeExcel(fensiSheet, fensiHeader, index + 2, p)



for index, head in enumerate(guanzhuHeader):
    guanzhuSheet.cell(1, index+1).value = head.strip()
for index, p in enumerate(weobo["guanzhuListRes"]):
    writeExcel(guanzhuSheet, guanzhuHeader, index + 2, p)

print("flish")

wb.save(excelFname)

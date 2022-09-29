from ast import Store
from enum import IntEnum
from attr import attrs
import requests
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import cfscrape
import json
import string
import re
import time
import math
import _thread

import numpy as np

http.client._MAXHEADERS = 1000


def urllib_download(IMAGE_URL, pName):
    try:
        opener = urllib.request.build_opener()
        opener.addheaders = [('User-agent', 'Mozilla/5.0'), ('cookie',
                                                             '__cf_bm=vbLxkjIVjoGJnWBrrRZP1GCsbs_KmubMaHNugmyyR3I-1659249299-0-AXJuOuT6vZtFzKG57pHX3sgKssRwKg9sXDpFaqkl5hXjPCCmxv95Lj76noiZ90Rm6c0kXSv/oW51uHGWcgUQlTE=')]
        urllib.request.install_opener(opener)
        urllib.request.urlretrieve(
            IMAGE_URL, pName.replace("/", "").replace("\\", ""))
    except:
        print('no')


def getNodeText(node):
    if (node == None):
        return ""
    else:
        return node.get_text().strip()


def getHtmlFromUrl(url, type="get", para={}):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.82 Safari/537.36",
        "cookie": "_ga=GA1.2.1846208171.1605273760; href=https%3A%2F%2Fwww.sinobiological.com%2Fresearch%2Ftargeted-therapy; accessId=5aff5fb0-84db-11e8-a3b3-d368cce40a8e; _gcl_au=1.1.1660157260.1645016298; Hm_lvt_2d911fad88dfe99ff8bbf610824f3e67=1645016298; sbb=%252be43ohTbVTr09K%252bxQlr1%252bK0onQvF%252bMIXgZM%252bveGXMHU%253dXTaJopSyq01ihC4cD5jOfAa8AEgsWX792EAllASK%252bEKohz0p3JxSEJFavoBnvUmw4fhBpwMcWGZ2Qy%252bRRF2U%252bMSxcQdMfdwOcT%252bR%252bo7qyEU%252br8SBQuGE8GJWgDFeSTZ4QS0HvJFVazETAoyuKMwGHYRoD68%252f7qno5Bg%252bEH9sSXM4upMLtz%252f4IdNkjX6GD0JYHbiUh%252blGTwi25Iz3IKocTDD58DE1yYiY3DxeifN7Qz6OxtXX21lrBpnvgDu9ANN%252f7TTxWWMmOIjxVG772o%252bYGkE9AMxcU5O4cIrT9cubm6dAdgw6n%252fQRZpTVxNv2TGHdHZblPNcfu4dTWVsL3aqaag%253d%253d; _gid=GA1.2.832211649.1645016298; _ce.s=v11.rlc~1645016301520; pageViewNum=13; Hm_lpvt_2d911fad88dfe99ff8bbf610824f3e67=1645017042; Currency=RMB; LocationCode=CN"
    }

    scraper = cfscrape.create_scraper()
    html_code = scraper.get(url, headers=headers).text
    return BeautifulSoup(html_code, "html.parser", from_encoding="utf-8")


def requestJson(url, pIndex):
    r = requests.post(url, data={"rpc": "100", "page": str(pIndex), "TaxonomicTags": "ELISA Pair Set"}, headers={
        'Content-Type': 'multipart/form-data;',
        'cookie': 'Hm_lvt_2d911fad88dfe99ff8bbf610824f3e67=1659233185; _gid=GA1.2.1110839465.1659233185; sbb=btzDOmbIJgtcJ%252fbBNFFakzEjiCmqYRsxckyc44YEsdI%253dXTaJopSyq01ihC4cD5jOfAa8AEgsWX792EAllASK%252bEKohz0p3JxSEJFavoBnvUmw4fhBpwMcWGZ2Qy%252bRRF2U%252bMSxcQdMfdwOcT%252bR%252bo7qyEU%252br8SBQuGE8GJWgDFeSTZ4gyQPbdRAfENtzDDqN8%252bGDluhZO1JALDLDvSQerjksoBXs7EVLgUw38PzM7oZm%252fQOJh74B1FrVGP39YP6NQ3ZjAvyki%252bPYATzUhPmLGBWfM6tKNdk2Xw0AJPu9svNuRInCKAY7cAiDRAl4QtVn%252fve5n66%252bTerGofcS4TcLhPpxKB%252f7UXS2Ukwrf%252bnxAim9gwczRgv6d4X4QsvzE9g25CJrg%253d%253d; _gcl_au=1.1.894241471.1659233186; cebs=1; __cf_bm=pxO5h0EnM.XSzjcys0ufAg42veHF8cDhJHzWM10WfUg-1659236014-0-AaB0w5RWqok9U/TT4voV03gLXrP9ZorKOzdpKao7/vF5cTxcc/ybitC+pwPfrFvJOJEspqyzKpl8LYk49W92HjU=; _ce.s=v~17fa36a1a61ac4f924605a0ea259c526716d971d~vpv~0~v11.rlc~1659236059303; _ga_HYV7JHQNBH=GS1.1.1659236057.2.1.1659236799.0; _ga=GA1.2.60586722.1659233185; _gat_gtag_UA_9748282_4=1; Hm_lpvt_2d911fad88dfe99ff8bbf610824f3e67=1659236800; Currency=RMB; LocationCode=CN; cebsp=7',
        "User-Agent": "PostmanRuntime/7.29.0"
    })
    print(r.text)
    # datas = json.loads(r.text)
    # return datas
    # return BeautifulSoup(r.text, "html.parser",from_encoding="utf-8")


def getRenderdHtmlFromUrl(url):
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument("window-size=1024,768")

    chrome_options.add_argument("--no-sandbox")
    browser = webdriver.Chrome(chrome_options=chrome_options)
    browser.get(url)
    return BeautifulSoup(browser.page_source, "html.parser")


def writeExcel(workSheet, headers, rowIndex, info):
    cellIndex = 1
    for head in headers:
        try:
            if head in info:
                content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
                workSheet.cell(rowIndex, cellIndex).value = content.strip()
            else:
                workSheet.cell(rowIndex, cellIndex).value = ""
            cellIndex = cellIndex+1
        except:
            print(rowIndex)


def getProductInfo(url, products, headers):
    print(str(len(products))+":"+url)
    sope = getHtmlFromUrl(url)

    pInfo = {}
    specAreas = sope.find_all(
        "div", attrs={"class": "col-xs-12 col-sm-12 col-md-3"})
    for specArea in specAreas:
        title = getNodeText(specArea.find("h3"))
        if title == "Theoretical Analysis":
            specs = getNodeText(specArea.find("p"))
            specStrs = specs.split("\n")
            for specStr in specStrs:
                if len(specStr) > 0:
                    specStrPart = specStr.split(":")
                    pInfo[specStrPart[0]] = specStrPart[1]
    imgArea = sope.find("div", attrs={"class": "text-center product-img"})
    if imgArea != None:
        img = imgArea.find("img")
        if img != None:
            urllib_download("https://medkoo.com" +
                            img["src"], pInfo["CAS#"]+".jpg")
    technincalData = sope.find("div", attrs={"id": "technical_data"})
    biological = sope.find("div", attrs={"id": "biological"})
    solubility = sope.find("div", attrs={"id": "solubility"})
    if technincalData != None:
        tecDatas = technincalData.find_all("p")
        for tecData in tecDatas:
            title = tecData.find("strong")
            if title != None:
                titleStr = getNodeText(title)
                val = getNodeText(tecData)
                pInfo[titleStr] = val.replace(titleStr, "")
                if titleStr not in headers:
                    headers.append(titleStr)
    if biological != None:
        bioDatas = biological.find_all("tr")
        for bioData in bioDatas:
            tds = bioData.find_all("td")
            if (len(tds) == 2):
                titleStr = getNodeText(tds[0])
                val = getNodeText(tds[1])
                pInfo[titleStr] = val
                if titleStr not in headers:
                    headers.append(titleStr)
    if solubility != None:
        solTanle = solubility.find("table")
        soTrs = solTanle.find_all("tr")
        soluStr = ""
        for inx, tr in enumerate(soTrs):
            if inx == 0:
                ths = tr.find_all("th")
                for th in ths:
                    soluStr += getNodeText(th)+"|"
                soluStr += "\r\n"
            else:
                tds = tr.find_all("td")
                for td in tds:
                    soluStr += getNodeText(td)+"|"
                soluStr += "\r\n"

        pInfo["solubility"] = soluStr
        headers.append("solubility")
    ps = sope.find_all("p")
    for p in ps:
        title = getNodeText(p.find("strong"))
        if title == "Description:":
            pInfo["Description"] = getNodeText(p).replace("Description:", "")
            headers.append("Description")
    # print(pInfo["Description"])
    products.append(pInfo.copy())


def getProductList(url,  products, headers):
    sope = getHtmlFromUrl(url)
    productAres = sope.find(
        "div", attrs={"class": "form-group col-lg-12 col-sm-12 col-md-12 col-xs-12"})
    productAreas = productAres.find_all("div", attrs={"class": "row"})
    for product in productAreas:
        pLink = product.find("a")
        if pLink != None:
            getProductInfo("https://medkoo.com" +
                           pLink["href"], products, headers)


headers = [
    'link', 'Name', 'CAS#', 'Chemical Formula', 'Exact Mass', 'Molecular Weight', 'Elemental Analysis'
]
excelFname = "medkoo.xlsx"
products = []

for pageInx in range(1, 20):
    url = "https://medkoo.com/products/newest?page="+str(pageInx)
    getProductList(url, products, headers)
# getProductList("https://medkoo.com/products/newest?page=1", products, headers)
# getProductInfo("https://medkoo.com/products/13246", products, headers)
wb = Workbook()
workSheet = wb.active


for index, head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index, p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")

wb.save(excelFname)

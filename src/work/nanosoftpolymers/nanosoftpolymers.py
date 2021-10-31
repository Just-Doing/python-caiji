from urllib.request import urlopen
from selenium import webdriver
import urllib
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy

http.client._MAXHEADERS = 1000


retryCount = 0
loadCount = 0
def urllib_download(IMAGE_URL, imageName):
	global retryCount
	try:
		from urllib.request import urlretrieve
		urlretrieve(IMAGE_URL, imageName)   
	except:
		print("重试图片下载"+IMAGE_URL)
		retryCount += 1
		if(retryCount <= 5):
			urllib_download(IMAGE_URL, imageName)
		else:
			retryCount=0
			return None
		
def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

def getHtmlFromUrl(url):
	global retryCount
	try:
		html = urlopen(url).read()
		return BeautifulSoup(html, "html.parser",from_encoding="utf-8")
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getHtmlFromUrl(url)
		else:
			retryCount=0
			return None
		
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1


def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	pInfo["link"]=url
	scope = getHtmlFromUrl(url)
	if scope != None:
		descriptArea = scope.find(name="div", attrs={"class":"woocommerce-Tabs-panel woocommerce-Tabs-panel--description panel entry-content wc-tab"})
		pInfo["description"]=getNodeText(descriptArea)
		
		imageArea = scope.find(name="div", attrs={"class":"woocommerce-product-details__short-description"})
		if imageArea != None:
			urllib_download(imageArea.find("img")["src"], pInfo["name"].replace('(','').replace(')','').replace('/','').replace(':','-')+'.jpg')
		
		specArea = scope.find(name="div", attrs={"id":"tab-local_1"})
		if specArea!=None:
			specContent = getNodeText(specArea.find("p"))
			for info in specContent.split("\n"):
				infoPart = info.split(':')
				if len(infoPart) > 1:
					title = infoPart[0].strip()
					value = info.split(':')[1].strip()
					if(title == "CAS#"):
						pInfo["CAS"]=value
					if(title == "Chemical Formula"):
						pInfo["ChemicalFormula"]=value
					if(title == "Exact Mass"):
						pInfo["ExactMass"]=value
					if(title == "Molecular Weight"):
						pInfo["MolecularWeight"]=value
					if(title == "Elemental Analysis"):
						pInfo["ElementalAnalysis"]=value
					if(title == "Appearance"):
						pInfo["Appearance"]=value
					if(title == "Purity"):
						pInfo["Purity"]=value
		infoTable = scope.find(name="table", attrs={"class":"woocommerce-product-attributes shop_attributes"})
		for tr in infoTable.find_all("tr"):
			title = getNodeText(tr.find("th"))
			value = getNodeText(tr.find("td"))
			if(title == "Package size"):
				pInfo["Packagesize"]=value
		products.append(pInfo.copy())

def getProductList(url, products):
	sope = getHtmlFromUrl(url)
	proLinks = sope.find_all(name="a", attrs={"class":"woocommerce-LoopProduct-link woocommerce-loop-product__link" })
	for link in proLinks:
		prodInfo = {}
		prodInfo["name"] = getNodeText(link.find("h2"))
		getProductInfo(link["href"], prodInfo, products)


excelFileName="nanosoftpolymers.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
urls=[
	"https://www.nanosoftpolymers.com/product-category/cationic-lipids/",
	"https://www.nanosoftpolymers.com/product-category/lipids/page/1/",
	"https://www.nanosoftpolymers.com/product-category/lipids/page/2/",
	"https://www.nanosoftpolymers.com/product-category/fluorescent-dyes/page/1/",
	"https://www.nanosoftpolymers.com/product-category/fluorescent-dyes/page/2/"
]
for url in urls:
	getProductList(url, products)
	
headers=[
	'link','name','CAS','ChemicalFormula','ExactMass','MolecularWeight','ElementalAnalysis','Appearance','Purity','Packagesize',"description"
]
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
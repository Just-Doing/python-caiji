from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		if str(html_code).find("403 ERROR")>-1:
			time.sleep(360)
			return getHtmlFromUrl(url)
		else:
			return html_code
	except:
		retryCount += 1
		if retryCount < 5:
			print("retry index"+str(retryCount)+url)
			time.sleep(360)
			return getHtmlFromUrl(url)
		else:
			retryCount = 0
			return ""
def getRenderdHtmlFromUrl(url, isTry):
	global retryCount
	try:
		chrome_options = webdriver.ChromeOptions()
		chrome_options.add_argument('--headless')
		chrome_options.add_argument('--disable-gpu')
		chrome_options.add_argument("window-size=1024,768")

		chrome_options.add_argument("--no-sandbox")
		browser = webdriver.Chrome(chrome_options=chrome_options)
		browser.get(url)
		if str(browser.page_source).find("403 ERROR")>-1:
			time.sleep(360)
			return getRenderdHtmlFromUrl(url, True)
		else:
			return browser.page_source
	except:
		retryCount += 1
		if retryCount < 5:
			print("retry index"+str(retryCount)+url)
			time.sleep(360)
			return getRenderdHtmlFromUrl(url, True)
		else:
			retryCount = 0
			return ""
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, type1, type2,tag, products):
	print(str(len(products)) + url)
	productListHtml = getHtmlFromUrl(url, False)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	pInfo = {
		"link": url,
		"tag": tag,
		"type1": type1,
		"type2": type2
	}
	pName = sope.find("h1", attrs={"class": "product_title entry-title"})
	sku = sope.find("span", attrs={"class": "sku_wrapper"})
	price = sope.find("span", attrs={"class":"woocommerce-Price-amount amount"})
	IntendedUse = sope.find("div", attrs={"class":"woocommerce-product-details__short-description"})
	pInfo["product name"] = getNodeText(pName)
	pInfo["Supplier Info"] = getNodeText(sku)
	pInfo["price"] = getNodeText(price)
	pInfo["Intended Use"] = getNodeText(IntendedUse)
	
	specArea = sope.find("div", attrs={"class":"woocommerce-tabs wc-tabs-wrapper"})
	specs = specArea.find_all("strong")
	for spec in specs:
		title = getNodeText(spec)
		val = spec.nextSibling
		if title =="STORAGE AND STABILITY":
			pInfo["storge"] = spec.nextSibling.nextSibling
		pInfo[title] = val

	backs = specArea.find_all("h3")
	for back in backs:
		title = getNodeText(back)
		if back.nextSibling!=None:
			value = getNodeText(back.nextSibling.nextSibling)
			if title=="Assay Background":
				if len(value) == 0:
					pInfo["Assay Background"] = getNodeText(back.nextSibling.nextSibling.nextSibling.nextSibling)
				else:
					pInfo["Assay Background"] =value

	pdfLinkArea = sope.find("ul", attrs={"class":"products columns-2"})
	if pdfLinkArea!=None:
		pdfLinks = pdfLinkArea.find_all("a")
		DatasheetLink=""
		for pdfLink in pdfLinks:
			DatasheetLink+="https://eaglebio.com"+pdfLink["href"]+","
		pInfo["Datasheet Link"] = DatasheetLink

	products.append(pInfo.copy())
	

def getProductList(url, type1, type2,tag, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	pList = sope.find_all("a", attrs={"class":"woocommerce-LoopProduct-link woocommerce-loop-product__link"})
	for p in pList:
		getProductInfo(p["href"],type1,type2,tag,products)	

excelFileName="eaglebio.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("https://eaglebio.com/product/109-foods-igg-allergy-elisa/", "111", "222",'', products)

getProductList("https://eaglebio.com/product-category/all-products/assay-kits/allergy-assay-kits/",'Assay Kits',"Allergy",'Allergy Assay Kits', products)
getProductList("https://eaglebio.com/product-category/all-products/assay-kits/bone-metabolism-kits/",'Assay Kits',"Bone Metabolism",'Bone Metabolism Kits', products)

for page in range(0,4):
	getProductList("https://eaglebio.com/product-category/all-products/assay-kits/cancer-biomarker-kits/page/"+str(page)+"/",'Assay Kits',"Cancer Biomarker",'Cancer Biomarker Kits', products)

for page in range(0,4):
	getProductList("https://eaglebio.com/product-category/all-products/assay-kits/cardiovascular-assay-kits/page/"+str(page)+"/",'Assay Kits',"Cardiovascular & Oxidative Stress",'Cardiovascular Assay Kits', products)

for page in range(0,6):
	getProductList("https://eaglebio.com/product-category/all-products/assay-kits/endocrine-assay-kits/page/"+str(page)+"/",'Assay Kits',"Endocrinology",'Endocrine Assay Kits', products)

for page in range(0,3):
	getProductList("https://eaglebio.com/product-category/all-products/assay-kits/gastrointestinal-assays/page/"+str(page)+"/",'Assay Kits',"GastroIntestinal",'GastroIntestinal Assays', products)

for page in range(0,16):
	getProductList("https://eaglebio.com/product-category/all-products/assay-kits/immunology-assay-kits/page/"+str(page)+"/",'Assay Kits',"Immunology",'Immunology Assay Kits', products)
getProductList("https://eaglebio.com/product-category/all-products/assay-kits/nephrology-assay-kits/",'Assay Kits',"Nephrology",'Nephrology Assay Kits', products)

for page in range(0,3):
	getProductList("https://eaglebio.com/product-category/all-products/assay-kits/neurobiology-assay-kits/page/"+str(page)+"/",'Assay Kits',"Neurobiology",'Neurobiology Assay Kits', products)

for page in range(0,3):
	getProductList("https://eaglebio.com/product-category/all-products/assay-kits/steroid-assay-kits/page/"+str(page)+"/",'Assay Kits',"Steroid",'Steroid Assay Kits', products)
for page in range(0,4):
	getProductList("https://eaglebio.com/product-category/all-products/assay-kits/drug-monitoring-elisa-assays/page/"+str(page)+"/",'Assay Kits',"Therapeutic Drug Monitoring",'Drug Monitoring Assay Kits', products)
for page in range(0,8):
	getProductList("https://eaglebio.com/product-category/all-products/assay-kits/veterinarian-assay-kits/page/"+str(page)+"/",'Assay Kits',"Veterinarian",'Veterinarian Assay Kits', products)
getProductList("https://eaglebio.com/product-category/all-products/assay-kits/vitamin-assay-kits/",'Assay Kits',"Vitamin",'Vitamin Assay Kits', products)

headers=[
	'link','tag', 'type1','type2','product name','Supplier Info','price','Intended Use','Size','Sensitivity','Incubation Time','Sample Type',
	'Sample Size','Alternate Name','storge','Assay Background','Datasheet Link'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
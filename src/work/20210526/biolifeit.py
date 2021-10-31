from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):

	url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
	request_obj=urllib.request.Request(url=url)
	response_obj=urllib.request.urlopen(request_obj)
	html_code=response_obj.read()
	return html_code
	
def getRenderdHtmlFromUrl(url):

	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	return browser.page_source
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, type1, type2, products):
	print(str(len(products)) + url)
	productListHtml = getRenderdHtmlFromUrl(url, False)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	pInfo = {
		"link": url,
		"type1": type1,
		"type2": type2
	}
	
	products.append(pInfo.copy())

def getProductList(url, type1, type2, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	pList = sope.find_all("div", attrs={"class":"riga-prodotto"})
	for p in pList:
		pNameNode = p.find("span", attrs={"class":"nome-prodotto"})
		pName=getNodeText(pNameNode)
		if len(pName)>0:
			introduction = p.find("span", attrs={"class":"desc"})
			size = p.find("span", attrs={"class":"formato-new"})
			cat = p.find("span", attrs={"class":"codice"})
			Datasheet = pName+"---Data sheet.pdf"
			Materialsafetydatasheet = pName+"---Material safety data sheet.pdf"
			IFU = pName+"---IFU.pdf"
			pInfo = {
				"Product Name": pName,
				"type1": type1,
				"type2": type2,
				"introduction": getNodeText(introduction),
				"size": getNodeText(size),
				"cat": getNodeText(cat)
			}

			pdfLinks = p.find_all("a")
			for pdfLink in pdfLinks:
				if getNodeText(pdfLink) == "Data sheet":
					pInfo["Data sheet"]=Datasheet
					urllib_download("http://www.biolifeit.com/"+pdfLink["href"], Datasheet)
				if getNodeText(pdfLink) == "Material safety data sheet":
					pInfo["Material safety data sheet"]=Materialsafetydatasheet
					urllib_download("http://www.biolifeit.com/"+pdfLink["href"], Materialsafetydatasheet)
				if getNodeText(pdfLink) == "IFU":
					pInfo["IFU"]=IFU
					urllib_download("http://www.biolifeit.com/"+pdfLink["href"], IFU)

			products.append(pInfo.copy())

excelFileName="biolifeit.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("https://www.lgcstandards.com/US/en/Dexamethasone-Sodium-Phosphate/p/MM0210.00", "111", "222", products)

getProductList("http://www.biolifeit.com/prodotti.asp?id1=01&id2=40&UseLng=en",'Microbiology Culture Media',"Dehydrated Culture Media", products)
getProductList("http://www.biolifeit.com/prodotti.asp?id1=01&id2=43&UseLng=en",'Microbiology Culture Media',"Chromogenic Media", products)
getProductList("http://www.biolifeit.com/prodotti.asp?id1=01&id2=41&UseLng=en",'Microbiology Culture Media',"Raw Materials – Ingredients", products)
getProductList("http://www.biolifeit.com/prodotti.asp?id1=01&id2=42&UseLng=en",'Microbiology Culture Media',"Selective Supplements / Enrichments", products)

getProductList("http://www.biolifeit.com/prodotti.asp?id1=02&id2=54&UseLng=en",'Microbiology Ready to use Culture Media',"90 mm plates", products)
getProductList("http://www.biolifeit.com/prodotti.asp?id1=02&id2=50&UseLng=en",'Microbiology Ready to use Culture Media',"150 mm plates", products)
getProductList("http://www.biolifeit.com/prodotti.asp?id1=02&id2=49&UseLng=en",'Microbiology Ready to use Culture Media',"55 mm plates (membrane filtration)", products)
getProductList("http://www.biolifeit.com/prodotti.asp?id1=02&id2=491&UseLng=en",'Microbiology Ready to use Culture Media',"90 mm plates with two/three media", products)
getProductList("http://www.biolifeit.com/prodotti.asp?id1=02&id2=55&UseLng=en",'Microbiology Ready to use Culture Media',"Tubes", products)
getProductList("http://www.biolifeit.com/prodotti.asp?id1=02&id2=51&UseLng=en",'Microbiology Ready to use Culture Media',"Bottles", products)

getProductList("http://www.biolifeit.com/prodotti.asp?id1=03&id2=301&UseLng=en",'Colony confirmation tests',"Colony confirmation tests", products)

getProductList("http://www.biolifeit.com/prodotti.asp?id1=04&id2=401&UseLng=en",'Infectious Diseases Rapid tests',"Gastroenterology", products)
getProductList("http://www.biolifeit.com/prodotti.asp?id1=04&id2=402&UseLng=en",'Infectious Diseases Rapid tests',"Respiratory", products)
getProductList("http://www.biolifeit.com/prodotti.asp?id1=04&id2=403&UseLng=en",'Infectious Diseases Rapid tests',"Sexually Transmitted Diseases", products)
getProductList("http://www.biolifeit.com/prodotti.asp?id1=04&id2=404&UseLng=en",'Infectious Diseases Rapid tests',"Parasitology", products)
getProductList("http://www.biolifeit.com/prodotti.asp?id1=04&id2=405&UseLng=en",'Infectious Diseases Rapid tests',"Others", products)

getProductList("http://www.biolifeit.com/prodotti.asp?id1=05&id2=501&UseLng=en",'Fecal Occult Blood / Intestinal diseases',"Fecal Occult Blood / Intestinal diseases", products)

getProductList("http://www.biolifeit.com/prodotti.asp?id1=06&id2=601&UseLng=en",'Rheumatology',"Rheumatology", products)

getProductList("http://www.biolifeit.com/prodotti.asp?id1=07&id2=25&UseLng=en",'Parasitology',"Concentration System", products)
getProductList("http://www.biolifeit.com/prodotti.asp?id1=07&id2=701&UseLng=en",'Parasitology',"Rapid Tests", products)

getProductList("http://www.biolifeit.com/prodotti.asp?id1=08&id2=801&UseLng=en",'Platelet Aggregation'," Platelet Aggregation", products)


headers=[
	'Product Name','type1','type2','introduction','size','cat','Data sheet','Material safety data sheet','IFU'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36"}

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		if retryCount< 5:
			getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, products):
	print(str(len(products)) + url)
	tempPinfo = {}
	productHtml = getRenderdHtmlFromUrl(url)
	if productHtml != None:
		tempPinfo["link"] = url
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
	
		nameArea = sope.find("h3",attrs={"itemprop":"name"})
		imgArea = sope.find("img",attrs={"itemprop":"image"})
		tempPinfo["name"] = getNodeText(nameArea)
		if imgArea!=None:
			src = imgArea["src"]
			if src.find("https:") < 0:
				src = "https:"+src
			urllib_download(src, tempPinfo["name"])
			tempPinfo["img"] = tempPinfo["name"]+".jpg"
		
		specs = sope.find_all("label",attrs={"class":"pull-left"})
		for spec in specs:
			title = getNodeText(spec)
			tempPinfo[title] = getNodeText(spec.nextSibling.nextSibling)
		
		products.append(tempPinfo.copy())

def getProductList(url, products):
	print(url)
	productListHtml = getRenderdHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	pros = sope.find_all("article")
	for pro in pros:
		link = pro.find("a")
		if link!=None:
			getProductInfo("https://www.001chemical.com"+link["href"], products)

excelFileName="cfmats.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
			
# getProductInfo('https://www.001chemical.com/chem/204580-28-9', products)
# getProductList('https://www.001chemical.com/chem/organic-siliconeslist1', products)
for pageIndex in range(1, 20):
	getProductList("https://www.001chemical.com/chem/organic-siliconeslist"+str(pageIndex), products)

headers=['link','name','img','CAS Number','Catalog Number','Molecular Formula',
'Molecular Weight']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	if rindex%100  == 0:
		wb.save(excelFileName)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
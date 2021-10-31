from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):

	url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
	headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36"}

	request_obj=urllib.request.Request(url=url)
	response_obj=urllib.request.urlopen(request_obj)
	html_code=response_obj.read()
	return html_code

def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	tempPinfo = pInfo.copy()
	pInfo["link"]=url
	productHtml = getRenderdHtmlFromUrl(url)
	if productHtml != None:
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
		specs = sope.find_all("tr")
		for spec in specs:
			tds = spec.find_all("td")
			if len(tds) == 3:
				title = getNodeText(tds[0])
				val = getNodeText(tds[2])
				tempPinfo[title] = val
		products.append(tempPinfo)


def getProductList(url, pInfo, products):
	productListHtml = getRenderdHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	pListArea = sope.find("div", attrs={"id":"content"})
	pListAreas = pListArea.find_all("td", attrs={"height":"25"})
	for pro in pListAreas:
		link = pro.find("a")
		if link != None:
			getProductInfo(link["href"], pInfo, products)
	
excelFileName="Elabscience.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo('https://www.chemicalsfinder.com/moreinfos.html?id=297183',{} ,products)
# getProductList('https://www.chemicalsfinder.com/categoryinformation.html?child=3&catname=Siloxanes&&lmt=0,20&page=1',{"type":"Siloxanes"} ,products)
for pageIndex in range(1, 140):
	getProductList('https://www.chemicalsfinder.com/categoryinformation.html?child=4&catname=Silanes&&lmt=0,20&page='+str(pageIndex), {"type":"Silanes "} ,products)
for pageIndex in range(1, 42):
	getProductList('https://www.chemicalsfinder.com/categoryinformation.html?child=3&catname=Siloxanes&&lmt=0,20&page='+str(pageIndex), {"type":"Siloxanes  "} ,products)

headers=[
	'link','type','Product Name','Catalog Number','Molecular Formula','CAS No.','Chemical Supplier']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	if rindex%100  == 0:
		wb.save(excelFileName)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
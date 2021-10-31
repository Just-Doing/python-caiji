from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.103 Safari/537.36'),('Accept','text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36"}

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		if retryCount< 5:
			getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	tempPinfo = pInfo.copy()
	tempPinfo["Chemical Name:"]=""
	productHtml = getHtmlFromUrl(url)
	if productHtml != None:
		tempPinfo["link"] = url
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
	
		specs = sope.find_all("tr")
		for spec in specs:
			tds = spec.find_all("td")
			if len(tds) == 2:
				title = getNodeText(tds[0])
				val = getNodeText(tds[1])
				tempPinfo[title] = val
		
		otSpecs = sope.find_all("td", attrs={"class":"itemhead"})
		for otSpec in otSpecs:
			title = getNodeText(otSpec)
			if title =="Introduction":
				tempPinfo["Introduction"] = getNodeText(otSpec.parent.parent.parent.nextSibling.nextSibling)
			if title =="Packings":
				tempPinfo["Packings"] = getNodeText(otSpec.parent.parent.parent.nextSibling.nextSibling)
			if title =="Applications":
				tempPinfo["Applications"] = getNodeText(otSpec.parent.parent.parent.nextSibling.nextSibling.find("td", attrs={"class":"pccjustify"}))
			if title =="Chemical Structure":
				imgArea = otSpec.parent.parent.parent.find("img")
				tempPinfo["Chemical Structure"] = ""
				if imgArea != None:
					if len(tempPinfo["Chemical Name:"])>0:
						tempPinfo["Chemical Structure"] = tempPinfo["Chemical Name:"]+".jpg"
					else:
						if len(tempPinfo["CAS No.:"])>0:
							tempPinfo["Chemical Structure"] = tempPinfo["CAS No.:"]+".jpg"
					if len(tempPinfo["Chemical Structure"])>0:
						if imgArea["src"].find("..") == 0:
							src = "https://orgsi.com" + imgArea["src"].replace("..","")
							# urllib_download(src, tempPinfo["Chemical Structure"])
						else:
							src = "https://orgsi.com/" + imgArea["src"]
							# urllib_download(src, tempPinfo["Chemical Structure"])
					
		solubilityAreas = sope.find_all("td", attrs={"class":"pccleft"})
		for solubilityArea in solubilityAreas:
			val = getNodeText(solubilityArea)
			if val.find("Solubility")==0:
				tempPinfo["Solubility"] = val
		
		products.append(tempPinfo.copy())

def getProductList(url, pInfo, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	productAreas = sope.find_all("tr", attrs={"bgcolor":"#F2F2F2"})
	for productArea in productAreas:
		typeLink = productArea.find("a")
		if typeLink!=None:
			tempPinfo = pInfo.copy()
			getProductInfo("https://orgsi.com/"+typeLink["href"],tempPinfo,products)
		

def getProductType(url, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	productAreas = sope.find("td", attrs={"class":"pccjustify"})
	types = productAreas.find_all("a")
	pInfo = {}
	for typeLink in types:
		pInfo["type"] = getNodeText(typeLink)
		if typeLink["href"].find("..")<0:
			getProductList("https://orgsi.com/"+typeLink["href"], pInfo, products)
			
excelFileName="orgsi.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
			
# getProductInfo('https://orgsi.com/silanes/1100.html',{"cas":"A660"},products)
getProductType('https://orgsi.com/silanes/alfasilanes.html', products)

headers=['link','type','Chemical Structure','Introduction','Chemical Name:','CAS No.:',
'EINECS No.:','Empirical Formula:','Molecular Weight:','Boiling Point:','Flash Point:',
'Viscosity 25°C:','Color and Appearance:','Density [25°C]:',
'Refractive Index [25°C]:','Min. Purity:','Technical Grade:','Solubility','Packings']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	if rindex%100  == 0:
		wb.save(excelFileName)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
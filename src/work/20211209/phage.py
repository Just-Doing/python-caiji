from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):

	url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
	request_obj=urllib.request.Request(url=url)
	response_obj=urllib.request.urlopen(request_obj)
	html_code=response_obj.read()
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	
def getRenderdHtmlFromUrl(url):
	try:
		chrome_options = webdriver.ChromeOptions()
		chrome_options.add_argument('--headless')
		chrome_options.add_argument('--disable-gpu')
		chrome_options.add_argument("window-size=1024,768")
		chrome_options.add_argument("--no-sandbox")
		browser = webdriver.Chrome(chrome_options=chrome_options)
		browser.get(url)
		html = browser.page_source
		browser.close()
		return BeautifulSoup(html, "html.parser",from_encoding="utf-8")
	except:
		return None
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, PhageTitle, Bacterialhost):
	print(str(len(products))+url)
	sope = getHtmlFromUrl(url)
	if sope != None:
		tables = sope.find_all("table")
		pInfo = {
			"link":url,
			"Phage":PhageTitle,
			"Bacterial host":Bacterialhost,
		}
		for table in tables:
			trs = table.find_all("tr")
			haveLbl_General = table.find("span", attrs={"class":"lbl_General"})
			if len(trs) > 1 and haveLbl_General != None:
				fileds = []
				for tr in trs:
					tds = tr.find_all("td")
					lbl_General = tr.find("span", attrs={"class":"lbl_General"})
					if lbl_General != None:
						for td in tds:
							fileds.append(getNodeText(td))
					else:
						for inx,filed in enumerate(fileds):
							if len(filed) >0:
								if len(tds) >= inx+1:
									pInfo[filed] = getNodeText(tds[inx])
						fileds = []

		print(pInfo)
		products.append(pInfo.copy())


def getProductList(url):
	sope = getRenderdHtmlFromUrl(url)
	if sope!=None:
		pList = sope.find("table", attrs={"id":"tableauPhages"})
		pLinks = pList.find_all("tr")
		for pLink in pLinks:
			tds = pLink.find_all("td")
			if len(tds) == 3:
				PhageTitle = getNodeText(tds[1])
				PhageLink = tds[1].find("a")
				Bacterialhost = getNodeText(tds[2])
				if len(PhageTitle) > 0 and PhageLink != None:
					getProductInfo("https://www.phage.ulaval.ca/"+PhageLink["href"],PhageTitle, Bacterialhost)

excelFileName="phage.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo('https://www.phage.ulaval.ca/?pageDemandee=phage&noPhage=227&id=41&L=1','')
getProductList('https://www.phage.ulaval.ca/en/phages-catalog')

headers=[
	'link','Phage','Bacterial host','HER Number','Name','Morphotype','Order','Family','Genus','Species','Other designations','Characteristics','Complete genome sequence'
	,'Host strain HER','Temp.','Medium','Aerobic/Anaerobic','Agitation','Others','Isolé par','Received from'
	,'Original source'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
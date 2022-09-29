from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import requests
from requests.cookies import RequestsCookieJar

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):

	url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
	request_obj=urllib.request.Request(url=url)
	response_obj=urllib.request.urlopen(request_obj)
	html_code=response_obj.read()
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")
	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	cookies = browser.get_cookies()
	session = requests.Session()
	jar = RequestsCookieJar()
	for cookie in cookies:
		jar.set(cookie['name'], cookie['value'])
	session.cookies = jar
	resp = session.get(url)
	print(resp.content)
	return BeautifulSoup(resp.content, "html.parser",from_encoding="utf-8")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductList(url, type, products):
	sope = getRenderdHtmlFromUrl(url)
	pListArea = sope.find("table", attrs={"class":"main-product table table-condensed table-hover "})
	if pListArea!=None:
		pList = pListArea.find("tbody").find_all("tr")
		for p in pList:
			tds = p.find_all("td")
			pInfo = {
				"type": type,
				"Cas No": getNodeText(tds[1]),
				"Product": getNodeText(tds[2])
			}
			print(pInfo)
			products.append(pInfo.copy())


excelFileName="carbosynth.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

getProductList("https://www.carbosynth.com/carbosynth/website.nsf/all-products-by-sub-category!openview&restricttocategory=Impurities-iAbiraterone&cat=3&start=1&count=50&all=yes&subcat=yes", "iAbiraterone", products)
# getProductList("https://www.carbosynth.com/carbosynth/website.nsf/all-products-by-sub-category!openview&restricttocategory=Impurities-iDipyridamole&cat=3&start=1&count=50&all=yes&subcat=yes", "iDipyridamole", products)
# getProductList("https://www.carbosynth.com/carbosynth/website.nsf/all-products-by-sub-category!openview&restricttocategory=Impurities-iNilotinib&cat=3&start=1&count=50&all=yes&subcat=yes", "iNilotinib", products)
# getProductList("https://www.carbosynth.com/carbosynth/website.nsf/all-products-by-sub-category!openview&restricttocategory=Impurities-iAcarbose&cat=3&start=1&count=50&all=yes&subcat=yes", "iAcarbose", products)
# getProductList("https://www.carbosynth.com/carbosynth/website.nsf/all-products-by-sub-category!openview&restricttocategory=Impurities-iDithranol&cat=3&start=1&count=50&all=yes&subcat=yes", "iDithranol", products)
# getProductList("https://www.carbosynth.com/carbosynth/website.nsf/all-products-by-sub-category!openview&restricttocategory=Impurities-iNimesulide&cat=3&start=1&count=50&all=yes&subcat=yes", "iNimesulide", products)
# getProductList("https://www.carbosynth.com/carbosynth/website.nsf/all-products-by-sub-category!openview&restricttocategory=Impurities-iAcebutolol&cat=3&start=1&count=50&all=yes&subcat=yes", "iAcebutolol", products)
# getProductList("https://www.carbosynth.com/carbosynth/website.nsf/all-products-by-sub-category!openview&restricttocategory=Impurities-iDobutamine&cat=3&start=1&count=50&all=yes&subcat=yes", "iDobutamine", products)
# getProductList("https://www.carbosynth.com/carbosynth/website.nsf/all-products-by-sub-category!openview&restricttocategory=Impurities-iNitrendipine&cat=3&start=1&count=50&all=yes&subcat=yes", "iNitrendipine", products)
# getProductList("https://www.carbosynth.com/carbosynth/website.nsf/all-products-by-sub-category!openview&restricttocategory=Impurities-iAceclofenac&cat=3&start=1&count=50&all=yes&subcat=yes", "iAceclofenac", products)
# getProductList("https://www.carbosynth.com/carbosynth/website.nsf/all-products-by-sub-category!openview&restricttocategory=Impurities-iDocetaxel&cat=3&start=1&count=50&all=yes&subcat=yes", "iDocetaxel", products)
# getProductList("https://www.carbosynth.com/carbosynth/website.nsf/all-products-by-sub-category!openview&restricttocategory=Impurities-iNizatidine&cat=3&start=1&count=50&all=yes&subcat=yes", "iNizatidine", products)
# getProductList("https://www.carbosynth.com/carbosynth/website.nsf/all-products-by-sub-category!openview&restricttocategory=Impurities-iAcetaminophen&cat=3&start=1&count=50&all=yes&subcat=yes", "iAcetaminophen", products)
# getProductList("https://www.carbosynth.com/carbosynth/website.nsf/all-products-by-sub-category!openview&restricttocategory=Impurities-iDomperidone&cat=3&start=1&count=50&all=yes&subcat=yes", "iDomperidone", products)
# getProductList("https://www.carbosynth.com/carbosynth/website.nsf/all-products-by-sub-category!openview&restricttocategory=Impurities-iNorethindrone&cat=3&start=1&count=50&all=yes&subcat=yes", "iNorethindrone", products)
# getProductList("https://www.carbosynth.com/carbosynth/website.nsf/all-products-by-sub-category!openview&restricttocategory=Impurities-iAcetamiprid&cat=3&start=1&count=50&all=yes&subcat=yes", "iAcetamiprid", products)
# getProductList("https://www.carbosynth.com/carbosynth/website.nsf/all-products-by-sub-category!openview&restricttocategory=Impurities-iDonepezil&cat=3&start=1&count=50&all=yes&subcat=yes", "iDonepezil", products)
# getProductList("https://www.carbosynth.com/carbosynth/website.nsf/all-products-by-sub-category!openview&restricttocategory=Impurities-iNorfloxacin&cat=3&start=1&count=50&all=yes&subcat=yes", "iNorfloxacin", products)
# getProductList("https://www.carbosynth.com/carbosynth/website.nsf/all-products-by-sub-category!openview&restricttocategory=Impurities-iAciclovir&cat=3&start=1&count=50&all=yes&subcat=yes", "iAciclovir", products)
# getProductList("https://www.carbosynth.com/carbosynth/website.nsf/all-products-by-sub-category!openview&restricttocategory=Impurities-iDorzolamide&cat=3&start=1&count=50&all=yes&subcat=yes", "iDorzolamide", products)
# getProductList("https://www.carbosynth.com/carbosynth/website.nsf/all-products-by-sub-category!openview&restricttocategory=Impurities-iNorgestrel&cat=3&start=1&count=50&all=yes&subcat=yes", "iNorgestrel", products)


headers=[
	'link','type','Cas No','Product'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
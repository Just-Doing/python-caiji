from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):

	url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
	request_obj=urllib.request.Request(url=url)
	response_obj=urllib.request.urlopen(request_obj)
	html_code=response_obj.read()
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	
def getRenderdHtmlFromUrl(url):
	try:
		chrome_options = webdriver.ChromeOptions()
		chrome_options.add_argument('--headless')
		chrome_options.add_argument('--disable-gpu')
		chrome_options.add_argument("window-size=1024,768")
		chrome_options.add_argument("--no-sandbox")
		browser = webdriver.Chrome(chrome_options=chrome_options)
		browser.get(url)
		html = browser.page_source
		browser.close()
		return BeautifulSoup(html, "html.parser",from_encoding="utf-8")
	except:
		return None
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, type, products):
	sope = getRenderdHtmlFromUrl(url)
	if sope != None:
		prods = sope.find_all("div", attrs={"class":"hb-product-meta-wrapper clearfix"})
		for prod in prods:
			cas = prod.find("li", attrs = {"class":"cas-no"})
			pName = prod.find("h3", attrs = {"class":"woocommerce-loop-product__title"})
			pInfo ={
				"type":type,
				"Cas No":getNodeText(cas),
				"Product Name": getNodeText(pName)
			}
			print(str(len(products))+"==="+pInfo["Cas No"])
			products.append(pInfo.copy())

def getPage(url, type, products):
	sope = getRenderdHtmlFromUrl(url+"?product_count=96")
	if sope!=None:
		prods = sope.find_all("div", attrs={"class":"hb-product-meta-wrapper clearfix"})
		for prod in prods:
			cas = prod.find("li", attrs = {"class":"cas-no"})
			pName = prod.find("h3", attrs = {"class":"woocommerce-loop-product__title"})
			pInfo ={
				"type":type,
				"Cas No":getNodeText(cas),
				"Product Name": getNodeText(pName)
			}
			print(str(len(products)) + "==="+getNodeText(cas))
			products.append(pInfo.copy())
		pageInfo = sope.find("ul", attrs={"class":"page-numbers"})
		if pageInfo != None:
			getProductInfo(url+"page/2"+"/?product_count=96", type, products)
	

def getProductList(url, products):
	html = '''<a href="https://www.aozeal.com/product-category/phenytoin/">Phenytoin</a>
<a href="https://www.aozeal.com/product-category/phloroglucinol/">Phloroglucinol</a>
<a href="https://www.aozeal.com/product-category/pholcodine/">Pholcodine</a>
<a href="https://www.aozeal.com/product-category/phorate/">Phorate</a>
<a href="https://www.aozeal.com/product-category/phosphorylcholine/">Phosphorylcholine</a>
<a href="https://www.aozeal.com/product-category/phthalazine/">Phthalazine</a>
<a href="https://www.aozeal.com/product-category/phthalic-acid/">Phthalic Acid</a>
<a href="https://www.aozeal.com/product-category/phytomenadione/">Phytomenadione</a>
<a href="https://www.aozeal.com/product-category/picaridin/">Picaridin</a>
<a href="https://www.aozeal.com/product-category/pidotimod/">Pidotimod</a>
<a href="https://www.aozeal.com/product-category/pilocarpine/">Pilocarpine</a>
<a href="https://www.aozeal.com/product-category/pimavanserin/">Pimavanserin</a>
<a href="https://www.aozeal.com/product-category/pimecrolimus/">Pimecrolimus</a>
<a href="https://www.aozeal.com/product-category/pimobendan/">Pimobendan</a>
<a href="https://www.aozeal.com/product-category/pimozide/">Pimozide</a>
<a href="https://www.aozeal.com/product-category/pinaverium-bromide/">Pinaverium Bromide</a>
<a href="https://www.aozeal.com/product-category/pindolol/">Pindolol</a>
<a href="https://www.aozeal.com/product-category/pioglitazone/">Pioglitazone</a>
<a href="https://www.aozeal.com/product-category/pipemidic-acid/">Pipemidic Acid</a>
<a href="https://www.aozeal.com/product-category/pipenzolate/">Pipenzolate</a>
<a href="https://www.aozeal.com/product-category/piperacilin/">Piperacilin</a>
<a href="https://www.aozeal.com/product-category/piperacillin/">Piperacillin</a>
<a href="https://www.aozeal.com/product-category/piperaquine/">Piperaquine</a>
<a href="https://www.aozeal.com/product-category/piperazine/">Piperazine</a>
<a href="https://www.aozeal.com/product-category/piperidine/">Piperidine</a>
<a href="https://www.aozeal.com/product-category/pipracil/">Pipracil</a>
<a href="https://www.aozeal.com/product-category/piracetam/">Piracetam</a>
<a href="https://www.aozeal.com/product-category/pirfenidone/">Pirfenidone</a>
<a href="https://www.aozeal.com/product-category/piribedil/">Piribedil</a>
<a href="https://www.aozeal.com/product-category/piroxicam/">Piroxicam</a>
<a href="https://www.aozeal.com/product-category/pitavastatin/">Pitavastatin</a>
<a href="https://www.aozeal.com/product-category/pitofenone/">Pitofenone</a>
<a href="https://www.aozeal.com/product-category/pixantrone/">Pixantrone</a>
<a href="https://www.aozeal.com/product-category/plantainoside/">Plantainoside</a>
<a href="https://www.aozeal.com/product-category/plazomicin/">Plazomicin</a>
<a href="https://www.aozeal.com/product-category/plerixafor/">Plerixafor</a>
<a href="https://www.aozeal.com/product-category/pneumocandin/">Pneumocandin</a>
<a href="https://www.aozeal.com/product-category/pocaine/">Pocaine</a>
<a href="https://www.aozeal.com/product-category/polacrilin-potassium/">Polacrilin Potassium</a>
<a href="https://www.aozeal.com/product-category/polaprezinc/">Polaprezinc</a>
<a href="https://www.aozeal.com/product-category/policresulen/">Policresulen</a>
<a href="https://www.aozeal.com/product-category/polydatin/">Polydatin</a>
<a href="https://www.aozeal.com/product-category/polylactic-acid/">Polylactic Acid</a>
<a href="https://www.aozeal.com/product-category/pomalidomide/">Pomalidomide</a>
<a href="https://www.aozeal.com/product-category/ponatinib/">Ponatinib</a>
<a href="https://www.aozeal.com/product-category/porphobilinogen/">Porphobilinogen</a>
<a href="https://www.aozeal.com/product-category/posaconazole/">posaconazole</a>
<a href="https://www.aozeal.com/product-category/potassium/">Potassium</a>
<a href="https://www.aozeal.com/product-category/potassium-hydrogen-phthalate/">Potassium Hydrogen Phthalate</a>
<a href="https://www.aozeal.com/product-category/potassium-otiraci/">Potassium Otiraci</a>
<a href="https://www.aozeal.com/product-category/potassium-sorbate/">Potassium Sorbate</a>
<a href="https://www.aozeal.com/product-category/pralatrexate/">Pralatrexate</a>
<a href="https://www.aozeal.com/product-category/pramipexole/">Pramipexole</a>
<a href="https://www.aozeal.com/product-category/pramoxine/">Pramoxine</a>
<a href="https://www.aozeal.com/product-category/pranoprofen/">Pranoprofen</a>
<a href="https://www.aozeal.com/product-category/prasugrel/">Prasugrel</a>
<a href="https://www.aozeal.com/product-category/pravastatin/">Pravastatin</a>
<a href="https://www.aozeal.com/product-category/prazepam/">Prazepam</a>
<a href="https://www.aozeal.com/product-category/praziquantel/">Praziquantel</a>
<a href="https://www.aozeal.com/product-category/prazosin/">Prazosin</a>
<a href="https://www.aozeal.com/product-category/prednisolone/">Prednisolone</a>
<a href="https://www.aozeal.com/product-category/prednisone/">Prednisone</a>
<a href="https://www.aozeal.com/product-category/pregabalin/">Pregabalin</a>
<a href="https://www.aozeal.com/product-category/pregnenolone/">Pregnenolone</a>
<a href="https://www.aozeal.com/product-category/pregnenolone-isobutyrate/">Pregnenolone Isobutyrate</a>
<a href="https://www.aozeal.com/product-category/pretomanid/">Pretomanid</a>
<a href="https://www.aozeal.com/product-category/pridinol/">Pridinol</a>
<a href="https://www.aozeal.com/product-category/prilocaine/">Prilocaine</a>
<a href="https://www.aozeal.com/product-category/primaquine/">Primaquine</a>
<a href="https://www.aozeal.com/product-category/primidone/">Primidone</a>
<a href="https://www.aozeal.com/product-category/proadifen/">Proadifen</a>
<a href="https://www.aozeal.com/product-category/probenecid/">Probenecid</a>
<a href="https://www.aozeal.com/product-category/procaine-hydrochloride/">Procaine Hydrochloride</a>
<a href="https://www.aozeal.com/product-category/procaterol/">Procaterol</a>
<a href="https://www.aozeal.com/product-category/prochlorperazine/">Prochlorperazine</a>
<a href="https://www.aozeal.com/product-category/procyanidin/">Procyanidin</a>
<a href="https://www.aozeal.com/product-category/progesterone/">Progesterone</a>
<a href="https://www.aozeal.com/product-category/proglumetacin/">Proglumetacin</a>
<a href="https://www.aozeal.com/product-category/proguanil/">Proguanil</a>
<a href="https://www.aozeal.com/product-category/proline/">Proline</a>
<a href="https://www.aozeal.com/product-category/promazine/">Promazine</a>
<a href="https://www.aozeal.com/product-category/promegestone/">Promegestone</a>
<a href="https://www.aozeal.com/product-category/promestriene/">Promestriene</a>
<a href="https://www.aozeal.com/product-category/promethazine/">Promethazine</a>
<a href="https://www.aozeal.com/product-category/propafenone/">Propafenone</a>
<a href="https://www.aozeal.com/product-category/propamocarb/">Propamocarb</a>
<a href="https://www.aozeal.com/product-category/propanolamine/">Propanolamine</a>
<a href="https://www.aozeal.com/product-category/propargite/">Propargite</a>
<a href="https://www.aozeal.com/product-category/propiverine/">Propiverine</a>
<a href="https://www.aozeal.com/product-category/propofol/">Propofol</a>
<a href="https://www.aozeal.com/product-category/propranolol/">Propranolol</a>
<a href="https://www.aozeal.com/product-category/propyl-gallate/">Propyl Gallate</a>
<a href="https://www.aozeal.com/product-category/propylene-glycol/">Propylene Glycol</a>
<a href="https://www.aozeal.com/product-category/propylthiouracil/">Propylthiouracil</a>
<a href="https://www.aozeal.com/product-category/propyphenazone/">Propyphenazone</a>
<a href="https://www.aozeal.com/product-category/prostacyclin/">Prostacyclin</a>
<a href="https://www.aozeal.com/product-category/prostaglandin/">Prostaglandin</a>
<a href="https://www.aozeal.com/product-category/prothioconazole/">Prothioconazole</a>
<a href="https://www.aozeal.com/product-category/prothionamide/">Prothionamide</a>
<a href="https://www.aozeal.com/product-category/protriptyline/">Protriptyline</a>
<a href="https://www.aozeal.com/product-category/prucalopride/">Prucalopride</a>
<a href="https://www.aozeal.com/product-category/prucalopridesuccinate/">PrucaloprideSuccinate</a>
<a href="https://www.aozeal.com/product-category/prulifloxacin/">Prulifloxacin</a>
<a href="https://www.aozeal.com/product-category/puerarin/">Puerarin</a>
<a href="https://www.aozeal.com/product-category/pyrantel/">Pyrantel</a>
<a href="https://www.aozeal.com/product-category/pyrazinamide/">Pyrazinamide</a>
<a href="https://www.aozeal.com/product-category/pyrazine/">Pyrazine</a>
<a href="https://www.aozeal.com/product-category/pyrazinoic-acid/">Pyrazinoic Acid</a>
<a href="https://www.aozeal.com/product-category/pyridinoline/">Pyridinoline</a>
<a href="https://www.aozeal.com/product-category/pyridostigmine/">Pyridostigmine</a>
<a href="https://www.aozeal.com/product-category/pyridoxal/">Pyridoxal</a>
<a href="https://www.aozeal.com/product-category/pyridoxine/">Pyridoxine</a>
<a href="https://www.aozeal.com/product-category/pyrimethamine/">Pyrimethamine</a>
<a href="https://www.aozeal.com/product-category/pyrimido5/">Pyrimido[5</a>
<a href="https://www.aozeal.com/product-category/pyruvic-acid/">Pyruvic Acid</a>
<a href="https://www.aozeal.com/product-category/quercetin/">Quercetin</a>
<a href="https://www.aozeal.com/product-category/quetiapine/">Quetiapine</a>
<a href="https://www.aozeal.com/product-category/quinapril/">Quinapril</a>
<a href="https://www.aozeal.com/product-category/quinfamide/">Quinfamide</a>
<a href="https://www.aozeal.com/product-category/quinidine/">Quinidine</a>
<a href="https://www.aozeal.com/product-category/quinine/">Quinine</a>
<a href="https://www.aozeal.com/product-category/quinine-benzoate/">Quinine Benzoate</a>
<a href="https://www.aozeal.com/product-category/quinine-dihydrochloride/">Quinine Dihydrochloride</a>
<a href="https://www.aozeal.com/product-category/quinine-sulphate/">Quinine Sulphate</a>
<a href="https://www.aozeal.com/product-category/quinolinic-acid/">Quinolinic Acid</a>
<a href="https://www.aozeal.com/product-category/quinolone/">Quinolone</a>
<a href="https://www.aozeal.com/product-category/rabeprazole/">Rabeprazole</a>
<a href="https://www.aozeal.com/product-category/racecadotril/">Racecadotril</a>
<a href="https://www.aozeal.com/product-category/ractopamine/">Ractopamine</a>
<a href="https://www.aozeal.com/product-category/raffinose/">Raffinose</a>
<a href="https://www.aozeal.com/product-category/rafoxanide/">Rafoxanide</a>
<a href="https://www.aozeal.com/product-category/raloxifene/">Raloxifene</a>
<a href="https://www.aozeal.com/product-category/raltegravir/">Raltegravir</a>
<a href="https://www.aozeal.com/product-category/ramelteon/">Ramelteon</a>
<a href="https://www.aozeal.com/product-category/ramipril/">Ramipril</a>
<a href="https://www.aozeal.com/product-category/ramosetron/">Ramosetron</a>
<a href="https://www.aozeal.com/product-category/ranitidine/">Ranitidine</a>
<a href="https://www.aozeal.com/product-category/ranolazine/">Ranolazine</a>
<a href="https://www.aozeal.com/product-category/rapamycin/">Rapamycin</a>
<a href="https://www.aozeal.com/product-category/rasagiline/">Rasagiline</a>
<a href="https://www.aozeal.com/product-category/ravidasvir/">Ravidasvir</a>
<a href="https://www.aozeal.com/product-category/rebamipide/">Rebamipide</a>
<a href="https://www.aozeal.com/product-category/reboxetine/">Reboxetine</a>
<a href="https://www.aozeal.com/product-category/regadenoson/">Regadenoson</a>
<a href="https://www.aozeal.com/product-category/regorafenib/">Regorafenib</a>
<a href="https://www.aozeal.com/product-category/relugolix/">Relugolix</a>
<a href="https://www.aozeal.com/product-category/remdesivir/">Remdesivir</a>
<a href="https://www.aozeal.com/product-category/remifentanil/">Remifentanil</a>
<a href="https://www.aozeal.com/product-category/repaglinide/">Repaglinide</a>
<a href="https://www.aozeal.com/product-category/reserpine/">Reserpine</a>
<a href="https://www.aozeal.com/product-category/resiniferatoxin/">Resiniferatoxin</a>
<a href="https://www.aozeal.com/product-category/resveratrol/">Resveratrol</a>
<a href="https://www.aozeal.com/product-category/retinol/">Retinol</a>
<a href="https://www.aozeal.com/product-category/retinyl/">Retinyl</a>
<a href="https://www.aozeal.com/product-category/retinyl-palmitate/">Retinyl Palmitate</a>
<a href="https://www.aozeal.com/product-category/retrorsine/">Retrorsine</a>
<a href="https://www.aozeal.com/product-category/rhodamine/">Rhodamine</a>
<a href="https://www.aozeal.com/product-category/ribavarin/">Ribavarin</a>
<a href="https://www.aozeal.com/product-category/riboflavin/">Riboflavin</a>
<a href="https://www.aozeal.com/product-category/ridaforolimus/">Ridaforolimus</a>
<a href="https://www.aozeal.com/product-category/rifabutin/">Rifabutin</a>
<a href="https://www.aozeal.com/product-category/rifampicin/">Rifampicin</a>
<a href="https://www.aozeal.com/product-category/rifapentine/">Rifapentine</a>
<a href="https://www.aozeal.com/product-category/rifaximin/">Rifaximin</a>
<a href="https://www.aozeal.com/product-category/rilpivirine/">Rilpivirine</a>
<a href="https://www.aozeal.com/product-category/riluzole/">Riluzole</a>
<a href="https://www.aozeal.com/product-category/rimantadine/">Rimantadine</a>
<a href="https://www.aozeal.com/product-category/riociguat/">Riociguat</a>
<a href="https://www.aozeal.com/product-category/ripasudil/">Ripasudil</a>
<a href="https://www.aozeal.com/product-category/risedronate/">Risedronate</a>
<a href="https://www.aozeal.com/product-category/risperidone/">Risperidone</a>
<a href="https://www.aozeal.com/product-category/ritalin/">Ritalin</a>
<a href="https://www.aozeal.com/product-category/ritalinic-acid/">Ritalinic Acid</a>
<a href="https://www.aozeal.com/product-category/ritonavir/">Ritonavir</a>
<a href="https://www.aozeal.com/product-category/rivaroxaban/">Rivaroxaban</a>
<a href="https://www.aozeal.com/product-category/rivastigmine/">Rivastigmine</a>
<a href="https://www.aozeal.com/product-category/rizatriptan/">Rizatriptan</a>
<a href="https://www.aozeal.com/product-category/rocuronium/">Rocuronium</a>
<a href="https://www.aozeal.com/product-category/roflumilast/">Roflumilast</a>
<a href="https://www.aozeal.com/product-category/rolapitant/">Rolapitant</a>
<a href="https://www.aozeal.com/product-category/ropinirole/">Ropinirole</a>
<a href="https://www.aozeal.com/product-category/ropivacaine/">Ropivacaine</a>
<a href="https://www.aozeal.com/product-category/rosiglitazone/">Rosiglitazone</a>
<a href="https://www.aozeal.com/product-category/rosuvastatin/">Rosuvastatin</a>
<a href="https://www.aozeal.com/product-category/rotigotine/">Rotigotine</a>
<a href="https://www.aozeal.com/product-category/roxadustat/">Roxadustat</a>
<a href="https://www.aozeal.com/product-category/roxatidine/">Roxatidine</a>
<a href="https://www.aozeal.com/product-category/roxithromycin/">Roxithromycin</a>
<a href="https://www.aozeal.com/product-category/rucaparib/">Rucaparib</a>
<a href="https://www.aozeal.com/product-category/rufinamide/">Rufinamide</a>
<a href="https://www.aozeal.com/product-category/rupatadine/">Rupatadine</a>
<a href="https://www.aozeal.com/product-category/ruscogenin/">Ruscogenin</a>
<a href="https://www.aozeal.com/product-category/ruxolitinib/">Ruxolitinib</a>
<a href="https://www.aozeal.com/product-category/sabinene/">Sabinene</a>
<a href="https://www.aozeal.com/product-category/saccharin/">Saccharin</a>
<a href="https://www.aozeal.com/product-category/saccharopine/">Saccharopine</a>
<a href="https://www.aozeal.com/product-category/sacubitril/">Sacubitril</a>
<a href="https://www.aozeal.com/product-category/safinamide/">Safinamide</a>
<a href="https://www.aozeal.com/product-category/salacinol/">Salacinol</a>
<a href="https://www.aozeal.com/product-category/salbutamol/">Salbutamol</a>
<a href="https://www.aozeal.com/product-category/salicin/">Salicin</a>
<a href="https://www.aozeal.com/product-category/salicylic-acid/">Salicylic Acid</a>
<a href="https://www.aozeal.com/product-category/salinomycin/">Salinomycin</a>
<a href="https://www.aozeal.com/product-category/salmeterol/">Salmeterol</a>
<a href="https://www.aozeal.com/product-category/santalol/">Santalol</a>
<a href="https://www.aozeal.com/product-category/sapropterin/">Sapropterin</a>
<a href="https://www.aozeal.com/product-category/sarafloxacin/">Sarafloxacin</a>
<a href="https://www.aozeal.com/product-category/sarcosine/">Sarcosine</a>
<a href="https://www.aozeal.com/product-category/sarpogrelate/">Sarpogrelate</a>
<a href="https://www.aozeal.com/product-category/saxagliptin/">Saxagliptin</a>
<a href="https://www.aozeal.com/product-category/scabertopin/">Scabertopin</a>
<a href="https://www.aozeal.com/product-category/schaftoside/">Schaftoside</a>
<a href="https://www.aozeal.com/product-category/secnidazole/">Secnidazole</a>
<a href="https://www.aozeal.com/product-category/selamectin/">Selamectin</a>
<a href="https://www.aozeal.com/product-category/selegiline-hydrochloride/">Selegiline Hydrochloride</a>
<a href="https://www.aozeal.com/product-category/selexipag/">Selexipag</a>
<a href="https://www.aozeal.com/product-category/senecionine/">Senecionine</a>
<a href="https://www.aozeal.com/product-category/seneciphylline/">Seneciphylline</a>
<a href="https://www.aozeal.com/product-category/senecivernine/">Senecivernine</a>
<a href="https://www.aozeal.com/product-category/senkirkine/">Senkirkine</a>
<a href="https://www.aozeal.com/product-category/senkyunolide/">Senkyunolide</a>
<a href="https://www.aozeal.com/product-category/sennosides/">Sennosides</a>
<a href="https://www.aozeal.com/product-category/sepiapterin/">Sepiapterin</a>
<a href="https://www.aozeal.com/product-category/sertaconazole/">Sertaconazole</a>
<a href="https://www.aozeal.com/product-category/sertraline/">Sertraline</a>
<a href="https://www.aozeal.com/product-category/sevelamer/">Sevelamer</a>
<a href="https://www.aozeal.com/product-category/sevoflurane/">Sevoflurane</a>
<a href="https://www.aozeal.com/product-category/shikimic-acid/">Shikimic Acid</a>
<a href="https://www.aozeal.com/product-category/sibutramine/">Sibutramine</a>
<a href="https://www.aozeal.com/product-category/sildenafil/">Sildenafil</a>
<a href="https://www.aozeal.com/product-category/silodosin/">Silodosin</a>
<a href="https://www.aozeal.com/product-category/silver-sulfadiazine/">Silver Sulfadiazine</a>
<a href="https://www.aozeal.com/product-category/silybin/">Silybin</a>
<a href="https://www.aozeal.com/product-category/silymarin/">Silymarin</a>
<a href="https://www.aozeal.com/product-category/simethicone/">Simethicone</a>
<a href="https://www.aozeal.com/product-category/simvastatin/">Simvastatin</a>
<a href="https://www.aozeal.com/product-category/siponimod/">Siponimod</a>
<a href="https://www.aozeal.com/product-category/sirolimus/">Sirolimus</a>
<a href="https://www.aozeal.com/product-category/sisomicin/">Sisomicin</a>
<a href="https://www.aozeal.com/product-category/sisomicin-sulfate/">Sisomicin Sulfate</a>
<a href="https://www.aozeal.com/product-category/sitafloxacin/">Sitafloxacin</a>
<a href="https://www.aozeal.com/product-category/sitagliptin/">Sitagliptin</a>
<a href="https://www.aozeal.com/product-category/sitosterol/">Sitosterol</a>
<a href="https://www.aozeal.com/product-category/skimmianine/">skimmianine</a>
<a href="https://www.aozeal.com/product-category/sodium-aminosalicylate-dihydrate/">Sodium Aminosalicylate Dihydrate</a>
<a href="https://www.aozeal.com/product-category/sodium-benzoate/">Sodium Benzoate</a>
<a href="https://www.aozeal.com/product-category/sodium-carbonate/">Sodium Carbonate</a>
<a href="https://www.aozeal.com/product-category/sodium-chloride/">Sodium Chloride</a>
<a href="https://www.aozeal.com/product-category/sodium-cromoglicate/">Sodium Cromoglicate</a>
<a href="https://www.aozeal.com/product-category/sodium-lactate/">Sodium Lactate</a>
<a href="https://www.aozeal.com/product-category/sodium-methylparaben/">Sodium Methylparaben</a>
<a href="https://www.aozeal.com/product-category/sodium-nitroprusside/">Sodium Nitroprusside</a>
<a href="https://www.aozeal.com/product-category/sodium-picosulfate/">Sodium Picosulfate</a>
<a href="https://www.aozeal.com/product-category/sodium-propylparaben/">Sodium Propylparaben</a>
<a href="https://www.aozeal.com/product-category/sodium-salicylate/">Sodium Salicylate</a>
<a href="https://www.aozeal.com/product-category/sodium-stearyl-fumarate/">Sodium Stearyl Fumarate</a>
<a href="https://www.aozeal.com/product-category/sodium-valprote/">Sodium Valprote</a>
<a href="https://www.aozeal.com/product-category/sofosbuvir/">Sofosbuvir</a>
<a href="https://www.aozeal.com/product-category/solifenacin/">Solifenacin</a>
<a href="https://www.aozeal.com/product-category/sophocarpine/">Sophocarpine</a>
<a href="https://www.aozeal.com/product-category/sorafenib/">Sorafenib</a>
<a href="https://www.aozeal.com/product-category/sotalol/">Sotalol</a>
<a href="https://www.aozeal.com/product-category/soy-isoflavione/">Soy Isoflavione</a>
<a href="https://www.aozeal.com/product-category/sparfloxacin/">Sparfloxacin</a>
<a href="https://www.aozeal.com/product-category/spectinomycin/">Spectinomycin</a>
<a href="https://www.aozeal.com/product-category/spinosyn/">Spinosyn</a>
<a href="https://www.aozeal.com/product-category/spirodiclofen/">Spirodiclofen</a>
<a href="https://www.aozeal.com/product-category/spironolactone/">Spironolactone</a>
<a href="https://www.aozeal.com/product-category/stanozolol/">Stanozolol</a>
<a href="https://www.aozeal.com/product-category/stavudine/">Stavudine</a>
<a href="https://www.aozeal.com/product-category/stearyl-alcohol/">Stearyl Alcohol</a>
<a href="https://www.aozeal.com/product-category/stiripentol/">Stiripentol</a>
<a href="https://www.aozeal.com/product-category/streptomycin-sulfate/">Streptomycin Sulfate</a>
<a href="https://www.aozeal.com/product-category/strigol/">Strigol</a>
<a href="https://www.aozeal.com/product-category/strontium-ranelate/">Strontium Ranelate</a>
<a href="https://www.aozeal.com/product-category/succinylcholine/">Succinylcholine</a>
<a href="https://www.aozeal.com/product-category/sucralose/">Sucralose</a>
<a href="https://www.aozeal.com/product-category/sucrose/">Sucrose</a>
<a href="https://www.aozeal.com/product-category/sufentanil/">Sufentanil</a>
<a href="https://www.aozeal.com/product-category/sugammadex/">Sugammadex</a>
<a href="https://www.aozeal.com/product-category/sulbactam/">Sulbactam</a>
<a href="https://www.aozeal.com/product-category/sulbactam-sodium/">Sulbactam Sodium</a>
<a href="https://www.aozeal.com/product-category/sulconazole-nitrate/">Sulconazole Nitrate</a>
<a href="https://www.aozeal.com/product-category/sulcotrione/">Sulcotrione</a>
<a href="https://www.aozeal.com/product-category/sulfacetamide/">Sulfacetamide</a>
<a href="https://www.aozeal.com/product-category/sulfadiazine/">Sulfadiazine</a>
<a href="https://www.aozeal.com/product-category/sulfadimethoxine/">Sulfadimethoxine</a>
<a href="https://www.aozeal.com/product-category/sulfadoxine/">Sulfadoxine</a>
<a href="https://www.aozeal.com/product-category/sulfamerazine/">Sulfamerazine</a>
<a href="https://www.aozeal.com/product-category/sulfamethazine/">Sulfamethazine</a>
<a href="https://www.aozeal.com/product-category/sulfamethoxazole/">Sulfamethoxazole</a>
<a href="https://www.aozeal.com/product-category/sulfamic/">Sulfamic</a>
<a href="https://www.aozeal.com/product-category/sulfamic-acid/">Sulfamic Acid</a>
<a href="https://www.aozeal.com/product-category/sulfanilamide/">Sulfanilamide</a>
<a href="https://www.aozeal.com/product-category/sulfapyridine/">Sulfapyridine</a>
<a href="https://www.aozeal.com/product-category/sulfaquinoxaline/">Sulfaquinoxaline</a>
<a href="https://www.aozeal.com/product-category/sulfasalazine/">Sulfasalazine</a>
<a href="https://www.aozeal.com/product-category/sulfathiazole/">Sulfathiazole</a>
<a href="https://www.aozeal.com/product-category/sulfide/">Sulfide</a>
<a href="https://www.aozeal.com/product-category/sulindac/">Sulindac</a>
<a href="https://www.aozeal.com/product-category/sulphan-blue/">Sulphan Blue</a>
<a href="https://www.aozeal.com/product-category/sulpiride/">Sulpiride</a>
<a href="https://www.aozeal.com/product-category/sultamicillin/">Sultamicillin</a>
<a href="https://www.aozeal.com/product-category/sumatriptan/">Sumatriptan</a>
<a href="https://www.aozeal.com/product-category/sunitinib/">Sunitinib</a>
<a href="https://www.aozeal.com/product-category/suvorexant/">Suvorexant</a>
<a href="https://www.aozeal.com/product-category/suxamethonium/">Suxamethonium</a>
<a href="https://www.aozeal.com/product-category/swertisin/">Swertisin</a>
<a href="https://www.aozeal.com/product-category/syringetin/">Syringetin</a>
<a href="https://www.aozeal.com/product-category/syringic-acid/">Syringic Acid</a>
<a href="https://www.aozeal.com/product-category/tacrolimus/">Tacrolimus</a>
<a href="https://www.aozeal.com/product-category/tadalafil/">Tadalafil</a>
<a href="https://www.aozeal.com/product-category/tafamidis/">Tafamidis</a>
<a href="https://www.aozeal.com/product-category/tafluprost/">Tafluprost</a>
<a href="https://www.aozeal.com/product-category/tamibarotene/">Tamibarotene</a>
<a href="https://www.aozeal.com/product-category/tamoxifen-citrate/">Tamoxifen Citrate</a>
<a href="https://www.aozeal.com/product-category/tamsulosin/">Tamsulosin</a>
<a href="https://www.aozeal.com/product-category/tamsulosin-hydrochloride/">Tamsulosin Hydrochloride</a>
<a href="https://www.aozeal.com/product-category/tandospirone/">Tandospirone</a>
<a href="https://www.aozeal.com/product-category/tapentadol/">Tapentadol</a>
<a href="https://www.aozeal.com/product-category/tartrazine/">Tartrazine</a>
<a href="https://www.aozeal.com/product-category/tasimelteon/">Tasimelteon</a>
<a href="https://www.aozeal.com/product-category/tavaborole/">Tavaborole</a>
<a href="https://www.aozeal.com/product-category/tazarotene/">Tazarotene</a>
<a href="https://www.aozeal.com/product-category/tazobactam/">Tazobactam</a>
<a href="https://www.aozeal.com/product-category/tebipenem/">Tebipenem</a>
<a href="https://www.aozeal.com/product-category/tebipenem-pivoxil/">Tebipenem Pivoxil</a>
<a href="https://www.aozeal.com/product-category/tecarfarin/">Tecarfarin</a>
<a href="https://www.aozeal.com/product-category/tedizolid/">Tedizolid</a>
<a href="https://www.aozeal.com/product-category/teicoplanin/">Teicoplanin</a>
<a href="https://www.aozeal.com/product-category/telmisartan/">Telmisartan</a>
<a href="https://www.aozeal.com/product-category/tembotrione/">Tembotrione</a>
<a href="https://www.aozeal.com/product-category/temozolomide/">Temozolomide</a>
<a href="https://www.aozeal.com/product-category/temsirolimus/">Temsirolimus</a>
<a href="https://www.aozeal.com/product-category/teneligliptin/">Teneligliptin</a>
<a href="https://www.aozeal.com/product-category/tenivastatin/">Tenivastatin</a>
<a href="https://www.aozeal.com/product-category/tenofovir/">Tenofovir</a>
<a href="https://www.aozeal.com/product-category/tenofovir-alafenamide/">Tenofovir alafenamide</a>
<a href="https://www.aozeal.com/product-category/tenofovir-disoproxil/">Tenofovir disoproxil</a>
<a href="https://www.aozeal.com/product-category/tenoxicam/">Tenoxicam</a>
<a href="https://www.aozeal.com/product-category/tenuifolin/">Tenuifolin</a>
<a href="https://www.aozeal.com/product-category/tepraloxydim/">Tepraloxydim</a>
<a href="https://www.aozeal.com/product-category/terazosin/">Terazosin</a>
<a href="https://www.aozeal.com/product-category/terbinafine/">Terbinafine</a>
<a href="https://www.aozeal.com/product-category/terbutaline-sulphate/">Terbutaline Sulphate</a>
<a href="https://www.aozeal.com/product-category/terconazole/">Terconazole</a>
<a href="https://www.aozeal.com/product-category/terpineol/">Terpineol</a>
<a href="https://www.aozeal.com/product-category/terpinyl/">Terpinyl</a>
<a href="https://www.aozeal.com/product-category/testosterone/">Testosterone</a>
<a href="https://www.aozeal.com/product-category/testosterone-benzoate/">Testosterone Benzoate</a>
<a href="https://www.aozeal.com/product-category/testosterone-propionate/">Testosterone Propionate</a>
<a href="https://www.aozeal.com/product-category/tetrabenazine/">Tetrabenazine</a>
<a href="https://www.aozeal.com/product-category/tetracaine/">Tetracaine</a>
<a href="https://www.aozeal.com/product-category/tetracycline/">Tetracycline</a>
<a href="https://www.aozeal.com/product-category/tetrafluoroethylene/">Tetrafluoroethylene</a>
<a href="https://www.aozeal.com/product-category/tetrahydrouridine/">Tetrahydrouridine</a>
<a href="https://www.aozeal.com/product-category/tetryzoline/">Tetryzoline</a>
<a href="https://www.aozeal.com/product-category/tezacaftor/">Tezacaftor</a>
<a href="https://www.aozeal.com/product-category/thalidomide/">Thalidomide</a>
<a href="https://www.aozeal.com/product-category/theobromine/">Theobromine</a>
<a href="https://www.aozeal.com/product-category/thiamazole/">Thiamazole</a>
<a href="https://www.aozeal.com/product-category/thiamine/">Thiamine</a>
<a href="https://www.aozeal.com/product-category/thiamphenicol/">Thiamphenicol</a>
<a href="https://www.aozeal.com/product-category/thiencarbazone/">Thiencarbazone</a>
<a href="https://www.aozeal.com/product-category/thimerosal/">Thimerosal</a>
<a href="https://www.aozeal.com/product-category/thiocolchicoside/">Thiocolchicoside</a>
<a href="https://www.aozeal.com/product-category/thioctic-acid/">Thioctic Acid</a>
<a href="https://www.aozeal.com/product-category/thiopental/">Thiopental</a>
<a href="https://www.aozeal.com/product-category/thioridazine/">Thioridazine</a>
<a href="https://www.aozeal.com/product-category/thiotepa/">Thiotepa</a>
<a href="https://www.aozeal.com/product-category/thiothixene/">Thiothixene</a>
<a href="https://www.aozeal.com/product-category/thymol/">Thymol</a>
<a href="https://www.aozeal.com/product-category/tiagabine/">Tiagabine</a>
<a href="https://www.aozeal.com/product-category/tiamulin/">Tiamulin</a>
<a href="https://www.aozeal.com/product-category/tianeptine/">Tianeptine</a>
<a href="https://www.aozeal.com/product-category/tiapride/">Tiapride</a>
<a href="https://www.aozeal.com/product-category/tiaprofenic/">Tiaprofenic</a>
<a href="https://www.aozeal.com/product-category/tibolone/">Tibolone</a>
<a href="https://www.aozeal.com/product-category/ticagrelor/">Ticagrelor</a>
<a href="https://www.aozeal.com/product-category/ticarcillin/">Ticarcillin</a>
<a href="https://www.aozeal.com/product-category/ticlopidine/">Ticlopidine</a>
<a href="https://www.aozeal.com/product-category/tigecycline/">Tigecycline</a>
<a href="https://www.aozeal.com/product-category/tildipirosin/">Tildipirosin</a>
<a href="https://www.aozeal.com/product-category/timolol/">Timolol</a>
<a href="https://www.aozeal.com/product-category/tinidazole/">Tinidazole</a>
<a href="https://www.aozeal.com/product-category/tiopronin/">Tiopronin</a>
<a href="https://www.aozeal.com/product-category/tiotropium-bromide/">Tiotropium Bromide</a>
<a href="https://www.aozeal.com/product-category/tipiracil-hcl/">Tipiracil HCL</a>
<a href="https://www.aozeal.com/product-category/tirofiban/">Tirofiban</a>
<a href="https://www.aozeal.com/product-category/tixocortol/">Tixocortol</a>
<a href="https://www.aozeal.com/product-category/tizanidine/">Tizanidine</a>
<a href="https://www.aozeal.com/product-category/tobramycin/">Tobramycin</a>
<a href="https://www.aozeal.com/product-category/tocopherol/">Tocopherol</a>
<a href="https://www.aozeal.com/product-category/tofacitinib/">Tofacitinib</a>
<a href="https://www.aozeal.com/product-category/tofogliflozin/">Tofogliflozin</a>
<a href="https://www.aozeal.com/product-category/tolazoline/">Tolazoline</a>
<a href="https://www.aozeal.com/product-category/tolbutamide/">Tolbutamide</a>
<a href="https://www.aozeal.com/product-category/tolcapone/">Tolcapone</a>
<a href="https://www.aozeal.com/product-category/toldimphos-sodium/">Toldimphos Sodium</a>
<a href="https://www.aozeal.com/product-category/tolfenamic-acid/">Tolfenamic Acid</a>
<a href="https://www.aozeal.com/product-category/tolnaftate/">Tolnaftate</a>
<a href="https://www.aozeal.com/product-category/tolperisone/">Tolperisone</a>
<a href="https://www.aozeal.com/product-category/tolterodine/">Tolterodine</a>
<a href="https://www.aozeal.com/product-category/toltrazuril/">Toltrazuril</a>
<a href="https://www.aozeal.com/product-category/tolvaptan/">Tolvaptan</a>
<a href="https://www.aozeal.com/product-category/topiramate/">Topiramate</a>
<a href="https://www.aozeal.com/product-category/topiroxostat/">Topiroxostat</a>
<a href="https://www.aozeal.com/product-category/topotecan/">Topotecan</a>
<a href="https://www.aozeal.com/product-category/torasemide/">Torasemide</a>
<a href="https://www.aozeal.com/product-category/toremifene/">Toremifene</a>
<a href="https://www.aozeal.com/product-category/torezolid/">Torezolid</a>
<a href="https://www.aozeal.com/product-category/torsemide/">Torsemide</a>
<a href="https://www.aozeal.com/product-category/tramadol/">Tramadol</a>
<a href="https://www.aozeal.com/product-category/trametinib/">Trametinib</a>
<a href="https://www.aozeal.com/product-category/trandolapril/">Trandolapril</a>
<a href="https://www.aozeal.com/product-category/tranexamic-acid/">Tranexamic Acid</a>
<a href="https://www.aozeal.com/product-category/tranilast/">Tranilast</a>
<a href="https://www.aozeal.com/product-category/travoprost/">Travoprost</a>
<a href="https://www.aozeal.com/product-category/trazodone/">Trazodone</a>
<a href="https://www.aozeal.com/product-category/trelagliptin/">Trelagliptin</a>
<a href="https://www.aozeal.com/product-category/treprostinil/">Treprostinil</a>
<a href="https://www.aozeal.com/product-category/triadimefon/">Triadimefon</a>
<a href="https://www.aozeal.com/product-category/triamcinolone/">Triamcinolone</a>
<a href="https://www.aozeal.com/product-category/triamterene/">Triamterene</a>
<a href="https://www.aozeal.com/product-category/tribenoside/">Tribenoside</a>
<a href="https://www.aozeal.com/product-category/tribenuron/">Tribenuron</a>
<a href="https://www.aozeal.com/product-category/tributyl-acetylcitrate/">Tributyl Acetylcitrate</a>
<a href="https://www.aozeal.com/product-category/tributylamine/">Tributylamine</a>
<a href="https://www.aozeal.com/product-category/trichodesmine/">Trichodesmine</a>
<a href="https://www.aozeal.com/product-category/triclosan/">Triclosan</a>
<a href="https://www.aozeal.com/product-category/trientine/">Trientine</a>
<a href="https://www.aozeal.com/product-category/triethyl-aconitate/">Triethyl Aconitate</a>
<a href="https://www.aozeal.com/product-category/trifluoperazine-dihydrochloride/">Trifluoperazine Dihydrochloride</a>
<a href="https://www.aozeal.com/product-category/trifluorothymidine/">Trifluorothymidine</a>
<a href="https://www.aozeal.com/product-category/trifluridine/">Trifluridine</a>
<a href="https://www.aozeal.com/product-category/trihexyphenidyl/">Trihexyphenidyl</a>
<a href="https://www.aozeal.com/product-category/trilostane/">Trilostane</a>
<a href="https://www.aozeal.com/product-category/trimebutine/">Trimebutine</a>
<a href="https://www.aozeal.com/product-category/trimetazidine/">Trimetazidine</a>
<a href="https://www.aozeal.com/product-category/trimethobenzamide/">Trimethobenzamide</a>
<a href="https://www.aozeal.com/product-category/trimethoprim/">Trimethoprim</a>
<a href="https://www.aozeal.com/product-category/trimethylamine-hydrochloride/">Trimethylamine Hydrochloride</a>
<a href="https://www.aozeal.com/product-category/trimethylammonium-hydrochloride/">Trimethylammonium Hydrochloride</a>
<a href="https://www.aozeal.com/product-category/trimipramine-maleate/">Trimipramine Maleate</a>
<a href="https://www.aozeal.com/product-category/trioxysalen/">Trioxysalen</a>
<a href="https://www.aozeal.com/product-category/triphenylmethanol/">Triphenylmethanol</a>
<a href="https://www.aozeal.com/product-category/triprolidine-hydrochloride/">Triprolidine Hydrochloride</a>
<a href="https://www.aozeal.com/product-category/triptolide/">Triptolide</a>
<a href="https://www.aozeal.com/product-category/triptorelin/">Triptorelin</a>
<a href="https://www.aozeal.com/product-category/tropicamide/">Tropicamide</a>
<a href="https://www.aozeal.com/product-category/tropidine/">Tropidine</a>
<a href="https://www.aozeal.com/product-category/tropisetron/">Tropisetron</a>
<a href="https://www.aozeal.com/product-category/trospium/">Trospium</a>
<a href="https://www.aozeal.com/product-category/trospium-chloride/">Trospium Chloride</a>
<a href="https://www.aozeal.com/product-category/trovafloxacin/">Trovafloxacin</a>
<a href="https://www.aozeal.com/product-category/troxipide/">Troxipide</a>
<a href="https://www.aozeal.com/product-category/tryptophan/">Tryptophan</a>
<a href="https://www.aozeal.com/product-category/tubocurarine/">Tubocurarine</a>
<a href="https://www.aozeal.com/product-category/turmerone/">Turmerone</a>
<a href="https://www.aozeal.com/product-category/tylosin/">Tylosin</a>
<a href="https://www.aozeal.com/product-category/tyloxapol/">Tyloxapol</a>
<a href="https://www.aozeal.com/product-category/tyrosine/">Tyrosine</a>
<a href="https://www.aozeal.com/product-category/ubiquinone/">Ubiquinone</a>
<a href="https://www.aozeal.com/product-category/udenafil/">Udenafil</a>
<a href="https://www.aozeal.com/product-category/ufiprazole/">Ufiprazole</a>
<a href="https://www.aozeal.com/product-category/uracil-arabinoside/">Uracil Arabinoside</a>
<a href="https://www.aozeal.com/product-category/urapidil/">Urapidil</a>
<a href="https://www.aozeal.com/product-category/uridine/">Uridine</a>
<a href="https://www.aozeal.com/product-category/urolithin/">Urolithin</a>
<a href="https://www.aozeal.com/product-category/ursodeoxycholic-acid/">Ursodeoxycholic Acid</a>
<a href="https://www.aozeal.com/product-category/ursolic-acid/">Ursolic Acid</a>
<a href="https://www.aozeal.com/product-category/valaciclovir/">Valaciclovir</a>
<a href="https://www.aozeal.com/product-category/valacyclovir/">Valacyclovir</a>
<a href="https://www.aozeal.com/product-category/valdecoxib/">Valdecoxib</a>
<a href="https://www.aozeal.com/product-category/valerenic-acid/">Valerenic Acid</a>
<a href="https://www.aozeal.com/product-category/valganciclovir/">Valganciclovir</a>
<a href="https://www.aozeal.com/product-category/valienamine/">Valienamine</a>
<a href="https://www.aozeal.com/product-category/valproic-acid/">Valproic Acid</a>
<a href="https://www.aozeal.com/product-category/valsartan/">Valsartan</a>
<a href="https://www.aozeal.com/product-category/vancomycin/">Vancomycin</a>
<a href="https://www.aozeal.com/product-category/vanillin/">Vanillin</a>
<a href="https://www.aozeal.com/product-category/vardenafil/">Vardenafil</a>
<a href="https://www.aozeal.com/product-category/varenicline/">Varenicline</a>
<a href="https://www.aozeal.com/product-category/vasopressin/">Vasopressin</a>
<a href="https://www.aozeal.com/product-category/vecuronium/">Vecuronium</a>
<a href="https://www.aozeal.com/product-category/velpatasvir/">Velpatasvir</a>
<a href="https://www.aozeal.com/product-category/venclexta/">Venclexta</a>
<a href="https://www.aozeal.com/product-category/venlafaxine/">Venlafaxine</a>
<a href="https://www.aozeal.com/product-category/verapamil/">Verapamil</a>
<a href="https://www.aozeal.com/product-category/verbascoside/">Verbascoside</a>
<a href="https://www.aozeal.com/product-category/vernakalant/">Vernakalant</a>
<a href="https://www.aozeal.com/product-category/verrucarin/">Verrucarin</a>
<a href="https://www.aozeal.com/product-category/vigabatrin/">Vigabatrin</a>
<a href="https://www.aozeal.com/product-category/vilanterol/">Vilanterol</a>
<a href="https://www.aozeal.com/product-category/vilazodone/">Vilazodone</a>
<a href="https://www.aozeal.com/product-category/vildagliptin/">Vildagliptin</a>
<a href="https://www.aozeal.com/product-category/vincamine/">Vincamine</a>
<a href="https://www.aozeal.com/product-category/vincristine/">Vincristine</a>
<a href="https://www.aozeal.com/product-category/vindesine/">Vindesine</a>
<a href="https://www.aozeal.com/product-category/viniferin/">Viniferin</a>
<a href="https://www.aozeal.com/product-category/vinorelbine/">Vinorelbine</a>
<a href="https://www.aozeal.com/product-category/vinpocetine/">Vinpocetine</a>
<a href="https://www.aozeal.com/product-category/virginiamycin/">Virginiamycin</a>
<a href="https://www.aozeal.com/product-category/vitamin/">Vitamin</a>
<a href="https://www.aozeal.com/product-category/vitamin-b6/">Vitamin B6</a>
<a href="https://www.aozeal.com/product-category/vitexin/">Vitexin</a>
<a href="https://www.aozeal.com/product-category/vitispirane/">Vitispirane</a>
<a href="https://www.aozeal.com/product-category/voglibose/">Voglibose</a>
<a href="https://www.aozeal.com/product-category/vonoprazan/">Vonoprazan</a>
<a href="https://www.aozeal.com/product-category/voriconazole/">Voriconazole</a>
<a href="https://www.aozeal.com/product-category/vortioxetine/">Vortioxetine</a>
<a href="https://www.aozeal.com/product-category/voxilaprevir/">Voxilaprevir</a>
<a href="https://www.aozeal.com/product-category/warfarin/">Warfarin</a>
<a href="https://www.aozeal.com/product-category/withaferin/">Withaferin</a>
<a href="https://www.aozeal.com/product-category/xanomeline/">Xanomeline</a>
<a href="https://www.aozeal.com/product-category/xanthinol/">Xanthinol</a>
<a href="https://www.aozeal.com/product-category/xanthinosin/">Xanthinosin</a>
<a href="https://www.aozeal.com/product-category/xanthofulvin/">Xanthofulvin</a>
<a href="https://www.aozeal.com/product-category/xanthone/">Xanthone</a>
<a href="https://www.aozeal.com/product-category/xanthosine/">Xanthosine</a>
<a href="https://www.aozeal.com/product-category/xylazine/">Xylazine</a>
<a href="https://www.aozeal.com/product-category/xylometazoline/">Xylometazoline</a>
<a href="https://www.aozeal.com/product-category/ylangene/">Ylangene</a>
<a href="https://www.aozeal.com/product-category/yohimbine/">Yohimbine</a>
<a href="https://www.aozeal.com/product-category/yohimbine-hydrochloride/">Yohimbine Hydrochloride</a>
<a href="https://www.aozeal.com/product-category/zafirlukast/">Zafirlukast</a>
<a href="https://www.aozeal.com/product-category/zaleplon/">Zaleplon</a>
<a href="https://www.aozeal.com/product-category/zearalanone/">Zearalanone</a>
<a href="https://www.aozeal.com/product-category/zephirol/">Zephirol</a>
<a href="https://www.aozeal.com/product-category/zerumbol/">Zerumbol</a>
<a href="https://www.aozeal.com/product-category/zidovudine/">Zidovudine</a>
<a href="https://www.aozeal.com/product-category/zilpaterol-hydrochloride/">Zilpaterol Hydrochloride</a>
<a href="https://www.aozeal.com/product-category/ziprasidone/">Ziprasidone</a>
<a href="https://www.aozeal.com/product-category/zofenopril/">Zofenopril</a>
<a href="https://www.aozeal.com/product-category/zoledronic-acid/">Zoledronic Acid</a>
<a href="https://www.aozeal.com/product-category/zolmitriptan/">Zolmitriptan</a>
<a href="https://www.aozeal.com/product-category/zolpidem/">Zolpidem</a>
<a href="https://www.aozeal.com/product-category/zonisamide/">Zonisamide</a>
<a href="https://www.aozeal.com/product-category/zopiclone/">Zopiclone</a>
<a href="https://www.aozeal.com/product-category/zotepine/">Zotepine</a>
<a href="https://www.aozeal.com/product-category/zuclopenthixol-decanoate/">Zuclopenthixol Decanoate</a>'''
	sope = BeautifulSoup(html, "html.parser",from_encoding="utf-8")
	if sope!=None:
		pList = sope.find_all("a")
		for pLink in pList:
			if pLink !=None:
				getPage(pLink["href"], getNodeText(pLink), products)


excelFileName="aozeal1.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

getProductList('https://www.aozeal.com/shop-2', products)
# getPage("https://www.aozeal.com/product-category/abiraterone", "", products)
headers=[
	'link','type','Cas No','Product Name'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
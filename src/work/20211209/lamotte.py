from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):

	url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
	request_obj=urllib.request.Request(url=url)
	response_obj=urllib.request.urlopen(request_obj)
	html_code=response_obj.read()
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	
def getRenderdHtmlFromUrl(url):
	try:
		chrome_options = webdriver.ChromeOptions()
		chrome_options.add_argument('--headless')
		chrome_options.add_argument('--disable-gpu')
		chrome_options.add_argument("window-size=1024,768")
		chrome_options.add_argument("--no-sandbox")
		browser = webdriver.Chrome(chrome_options=chrome_options)
		browser.get(url)
		html = browser.page_source
		browser.close()
		return BeautifulSoup(html, "html.parser",from_encoding="utf-8")
	except:
		return None
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, type):
	sope = getRenderdHtmlFromUrl(url)
	if sope != None:
		desc = sope.find("div", attrs = {"itemprop":"description"})
		Features = sope.find("div", attrs = {"class":"product attribute description"})
		pName = sope.find("span", attrs = {"data-ui-id":"page-title-wrapper"})
		shipcode = sope.find("div", attrs = {"class":"ship-code"})
		shipweight = sope.find("div", attrs = {"class":"ship-weight"})
		pInfo ={
			"link":url,
			"type":type,
			"Product Name": getNodeText(pName),
			"description":getNodeText(desc),
			"Features": getNodeText(Features),
			"ship code": getNodeText(shipcode),
			"ship weight": getNodeText(shipweight)
		}
		print(str(len(products))+"==="+pInfo["Product Name"])
		products.append(pInfo.copy())


def getProductList(url, type):
	sope = getRenderdHtmlFromUrl(url)
	if sope!=None:
		pList = sope.find_all("li", attrs={"class":"item product product-item"})
		for p in pList:
			pLink = p.find("a")
			if pLink !=None:
				getProductInfo(pLink["href"], type)


excelFileName="lamotte.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

getProductList('https://lamotte.com/products/environmental-science-education/water-monitoring-kits/water-monitoring/?p=1','Water Monitoring')
getProductList('https://lamotte.com/products/environmental-science-education/water-monitoring-kits/water-monitoring/?p=2','Water Monitoring')
getProductList('https://lamotte.com/products/environmental-science-education/water-monitoring-kits/tablet-test-kits/','Water Monitoring Tablet Test Kits')

getProductList('https://lamotte.com/products/environmental-science-education/water-monitoring-kits/bacteria-studies/biopaddles/','Bacteria Studies')
getProductList('https://lamotte.com/products/environmental-science-education/water-monitoring-kits/bacteria-studies/coliform/','Bacteria Studies')
getProductList('https://lamotte.com/products/environmental-science-education/water-monitoring-kits/bacteria-studies/microbe-hunter-trade/','Bacteria Studies')

getProductList('https://lamotte.com/products/environmental-science-education/water-monitoring-kits/individual-test-kits/?p=1','Individual Test Kits')
getProductList('https://lamotte.com/products/environmental-science-education/water-monitoring-kits/individual-test-kits/?p=2','Individual Test Kits')
getProductList('https://lamotte.com/products/environmental-science-education/water-monitoring-kits/individual-test-kits/?p=3','Individual Test Kits')

getProductList('https://lamotte.com/products/environmental-science-education/soil-testing/soil-test-kits/','Soil Test Kits')

headers=[
	'link','type','Product Name','description','Features','ship code','ship weight'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
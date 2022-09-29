from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):

	url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
	request_obj=urllib.request.Request(url=url)
	response_obj=urllib.request.urlopen(request_obj)
	html_code=response_obj.read()
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	
def getRenderdHtmlFromUrl(url):
	try:
		chrome_options = webdriver.ChromeOptions()
		chrome_options.add_argument('--headless')
		chrome_options.add_argument('--disable-gpu')
		chrome_options.add_argument("window-size=1024,768")
		chrome_options.add_argument("--no-sandbox")
		browser = webdriver.Chrome(chrome_options=chrome_options)
		browser.get(url)
		html = browser.page_source
		browser.close()
		return BeautifulSoup(html, "html.parser",from_encoding="utf-8")
	except:
		return None
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, type, products):
	sope = getRenderdHtmlFromUrl(url)
	if sope != None:
		prods = sope.find_all("div", attrs={"class":"hb-product-meta-wrapper clearfix"})
		for prod in prods:
			cas = prod.find("li", attrs = {"class":"cas-no"})
			pName = prod.find("h3", attrs = {"class":"woocommerce-loop-product__title"})
			pInfo ={
				"type":type,
				"Cas No":getNodeText(cas),
				"Product Name": getNodeText(pName)
			}
			print(str(len(products))+"==="+pInfo["Cas No"])
			products.append(pInfo.copy())

def getPage(url, type, products):
	sope = getRenderdHtmlFromUrl(url+"?product_count=96")
	if sope!=None:
		prods = sope.find_all("div", attrs={"class":"hb-product-meta-wrapper clearfix"})
		for prod in prods:
			cas = prod.find("li", attrs = {"class":"cas-no"})
			pName = prod.find("h3", attrs = {"class":"woocommerce-loop-product__title"})
			pInfo ={
				"type":type,
				"Cas No":getNodeText(cas),
				"Product Name": getNodeText(pName)
			}
			print(str(len(products)) + "==="+getNodeText(cas))
			products.append(pInfo.copy())
		pageInfo = sope.find("ul", attrs={"class":"page-numbers"})
		if pageInfo != None:
			getProductInfo(url+"page/2"+"/?product_count=96", type, products)
	

def getProductList(url, products):
	sope = getRenderdHtmlFromUrl(url)
	if sope!=None:
		pList = sope.find_all("li", attrs={"class":"cat-item"})
		for p in pList:
			pLink = p.find("a")
			print(pLink)
			if pLink !=None:
				getPage(pLink["href"], getNodeText(pLink), products)


excelFileName="aozeal.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

getProductList('https://www.aozeal.com/shop-2', products)
# getPage("https://www.aozeal.com/product-category/abiraterone", "", products)
headers=[
	'link','type','Cas No','Product Name'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
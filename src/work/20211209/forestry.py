import requests
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):

	url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
	request_obj=urllib.request.Request(url=url)
	response_obj=urllib.request.urlopen(request_obj)
	html_code=response_obj.read()
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	
def getRenderdHtmlFromUrl(url):

	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	html = browser.page_source
	browser.close()
	return BeautifulSoup(html, "html.parser",from_encoding="utf-8")
	
def getJson(url, type="get", para={}):
	if type =="get":
		r = requests.get(url)
		datas = json.loads(r.text)
		return datas
	if type == "post":
		r = requests.post(url, data=para, headers={
			'Content-Type': 'application/x-www-form-urlencoded'
		})	
		datas = json.loads(r.text)
		return datas
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, type):
	sope = getRenderdHtmlFromUrl(url)
	
	desc = sope.find("div", attrs = {"id":"ProductDesc"})
	pName = sope.find("h1", attrs = {"class":"product-title"})
	if pName == None:
		pName = sope.find("h1", attrs = {"itemprop":"name"})
	pInfo ={
		"type":type,
		"link": url,
		"products name": getNodeText(pName),
		"descriptiom": getNodeText(desc)
	}
	print(str(len(products))+"==="+pInfo["products name"])
	products.append(pInfo.copy())


def getProductList(url, type):
	data = getJson(url)
	for p in data["result"]:
		getProductInfo(p["url"], type)
	


excelFileName="forestry.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
getProductList('https://www.forestry-suppliers.com/Klevu/resources/cloud-search/n-search/search/?ticket=klevu-15507897181989345&term=Water%20Quality%20Management&paginationStartsFrom=0&sortPrice=false&ipAddress=undefined&analyticsApiKey=klevu-15507897181989345&showOutOfStockProducts=true&klevuFetchPopularTerms=false&klevu_priceInterval=500&fetchMinMaxPrice=true&klevu_multiSelectFilters=true&noOfResults=36&klevuSort=rel&enableFilters=true&filterResults=category%3ATest%20Kits%20and%20Strips%3B%3BmarketCategory%3AWater%20Quality%20Management&visibility=search&category=KLEVU_PRODUCT&klevu_filterLimit=50&sv=2001&lsqt=WILDCARD_AND&responseType=json&resultForZero=1&klevu_loginCustomerGroup=', 'Water Test kit')
getProductList('https://www.forestry-suppliers.com/Klevu/resources/cloud-search/n-search/search/?ticket=klevu-15507897181989345&term=Water%20Quality%20Management&paginationStartsFrom=36&sortPrice=false&ipAddress=undefined&analyticsApiKey=klevu-15507897181989345&showOutOfStockProducts=true&klevuFetchPopularTerms=false&klevu_priceInterval=500&fetchMinMaxPrice=true&klevu_multiSelectFilters=true&noOfResults=36&klevuSort=rel&enableFilters=true&filterResults=category%3ATest%20Kits%20and%20Strips%3B%3BmarketCategory%3AWater%20Quality%20Management&visibility=search&category=KLEVU_PRODUCT&klevu_filterLimit=50&sv=2001&lsqt=WILDCARD_AND&responseType=json&resultForZero=1&klevu_loginCustomerGroup=','Water Test kit')
getProductList('https://www.forestry-suppliers.com/Klevu/resources/cloud-search/n-search/search/?ticket=klevu-15507897181989345&term=Water%20Quality%20Management&paginationStartsFrom=72&sortPrice=false&ipAddress=undefined&analyticsApiKey=klevu-15507897181989345&showOutOfStockProducts=true&klevuFetchPopularTerms=false&klevu_priceInterval=500&fetchMinMaxPrice=true&klevu_multiSelectFilters=true&noOfResults=36&klevuSort=rel&enableFilters=true&filterResults=category%3ATest%20Kits%20and%20Strips%3B%3BmarketCategory%3AWater%20Quality%20Management&visibility=search&category=KLEVU_PRODUCT&klevu_filterLimit=50&sv=2001&lsqt=WILDCARD_AND&responseType=json&resultForZero=1&klevu_loginCustomerGroup=','Water Test kit')
getProductList('https://www.forestry-suppliers.com/Klevu/resources/cloud-search/n-search/search/?ticket=klevu-15507897181989345&term=Soil%20Management&paginationStartsFrom=0&sortPrice=false&ipAddress=undefined&analyticsApiKey=klevu-15507897181989345&showOutOfStockProducts=true&klevuFetchPopularTerms=false&klevu_priceInterval=500&fetchMinMaxPrice=true&klevu_multiSelectFilters=true&noOfResults=36&klevuSort=rel&enableFilters=true&filterResults=category%3ATest%20Kits%20and%20Strips%3B%3BmarketCategory%3ASoil%20Management&visibility=search&category=KLEVU_PRODUCT&klevu_filterLimit=50&sv=2001&lsqt=&responseType=json&resultForZero=1&recentCategory=Test%20Kits%20and%20Strips&klevu_loginCustomerGroup=', 'Soil Test kit')

headers=[
	'link','type','products name','descriptiom'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import time

http.client._MAXHEADERS = 1000
headers=['link','type','title','Introduction','size','Product Description','Technical Library','Reference Literature','Downloads']
def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry index"+str(retryCount)+url)
		retryCount += 1
		if retryCount< 5:
			getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	time.sleep(2)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, typeInfo, products):
	global headers
	print(str(len(products)) + url)
	productListHtml = getRenderdHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	pInfo = {
		"link":url,
		"type":typeInfo,
		"size":""
	}
	skuList = sope.find_all("div", attrs={"class":"variations-form__item-sku"})
	sizeList = sope.find_all("div", attrs={"class":"variations-form__item-attr"})
	for index,sizeInfo in enumerate( sizeList):
		pInfo["size"] += getNodeText(skuList[index]) + ":" + getNodeText(sizeInfo)+","
	title = getNodeText(sope.find("h1",attrs={"class":"product_title entry-title"}))
	pInfo["title"] = title
	Introduction = getNodeText(sope.find("div",attrs={"class":"woocommerce-product-details__short-description"}))
	pInfo["Introduction"] = Introduction

	description = sope.find("div", attrs={"id":"tab-description"})
	pInfo["Product Description"] = getNodeText(description)
	
	TechnicalLibrary = sope.find("div", attrs={"class":"product-card card faq-card technical-library-card"})
	pInfo["Technical Library"] = getNodeText(TechnicalLibrary)
	
	ReferenceLiterature = sope.find("div", attrs={"id":"tab-reference_literature"})
	pInfo["Reference Literature"] = getNodeText(ReferenceLiterature)
	
	Downloads = sope.find("div", attrs={"id":"tab-downloads"})
	pInfo["Downloads"] = getNodeText(Downloads)
	


	specTitles = sope.find_all("tr")
	for spectitle in specTitles:
		ths = spectitle.find_all("th")
		tds = spectitle.find_all("td")
		if len(ths)==1 and len(tds) == 1:
			title = getNodeText(ths[0])
			val = getNodeText(tds[0])
			if title not in headers:
				headers.append(title)
			pInfo[title] = val
	products.append(pInfo.copy())

def getProductList(url, typeInfo, products):
	productListHtml = getHtmlFromUrl(url)
	data = json.loads(productListHtml)
	hits = data["hits"]["hits"]
	for hit in hits:
		link = hit["_source"]["link"]
		getProductInfo("https://promocell.com"+link, typeInfo, products)

excelFileName="promocell.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("https://promocell.com/product/free-carboxylic-acids", "111", products)
# getProductList("https://promocell.com/searchapp/search?page=1&filters=11269&sortBy=name&type=products","3D Cell Culture", products)

for page in range(1,3):
	getProductList("https://promocell.com/searchapp/search?page="+str(page)+"&filters=11269&sortBy=name&type=products","3D Cell Culture", products)
for page in range(1,3):
	getProductList("https://promocell.com/searchapp/search?page="+str(page)+"&filters=11280&sortBy=name&type=products",'Aging & Epigenetics', products)
for page in range(1,408):
	getProductList("https://promocell.com/searchapp/search?page="+str(page)+"&filters=11262&sortBy=name&type=products",'Antibodies & ELISAs', products)
for page in range(1,15):
	getProductList("https://promocell.com/searchapp/search?page="+str(page)+"&filters=10299&sortBy=name&type=products",'Apoptosis', products)
for page in range(1,29):
	getProductList("https://promocell.com/searchapp/search?page="+str(page)+"&filters=11264&sortBy=name&type=products",'Cell Metabolism & Signaling', products)
getProductList("https://promocell.com/searchapp/search?filters=11309&sortBy=name&type=products",'Cell Model Systems', products)
for page in range(1,25):
	getProductList("https://promocell.com/searchapp/search?page="+str(page)+"&filters=11273&sortBy=name&type=products",'Cell Staining & Fluorescent Labeling', products)
for page in range(1,5):
	getProductList("https://promocell.com/searchapp/search?page="+str(page)+"&filters=10619&sortBy=name&type=products",'Cell Transfection', products)
for page in range(1,11):
	getProductList("https://promocell.com/searchapp/search?page="+str(page)+"&filters=11286&sortBy=name&type=products",'Drug Discovery / ADME', products)
for page in range(1,6):
	getProductList("https://promocell.com/searchapp/search?page="+str(page)+"&filters=11292&sortBy=name&type=products",'Oxidative Stress & Cell Damage', products)
for page in range(1,64):
	getProductList("https://promocell.com/searchapp/search?page="+str(page)+"&filters=11260&sortBy=name&type=products",'Supplementary Assay Reagents', products)



for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
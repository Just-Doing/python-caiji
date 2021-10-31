from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry index"+str(retryCount)+url)
		retryCount += 1
		if retryCount< 5:
			getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, typeInfo, products):
	print(str(len(products)) + url)
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	decArea = sope.find("div", attrs={"class":"m-left"})
	pInfo = {
		"link":url,
		"type1":typeInfo["t1"],
		"type2":typeInfo["t2"]
	}
	title = getNodeText(decArea.find("h2"))
	pInfo["title"] = title
	specTitles = sope.find_all("p")
	pInfo["SAFE HANDLING PRECAUTIONS"]=""
	pInfo["PRODUCT DESCRIPTION"]=""
	pInfo["MECHANISM"]=""
	pInfo["APPLICATION RECOMMENDATION"] =""
	pInfo["WARNINGS"]=""
	pInfo["Package"]=""
	pInfo["Storage"]=""
	pInfo["Shelf life"]=""
	for spectitle in specTitles:
		title = getNodeText(spectitle)
		if title.upper() == "PRODUCT DESCRIPTION" and len(pInfo["PRODUCT DESCRIPTION"])==0:
			pInfo["PRODUCT DESCRIPTION"] = getNodeText(spectitle.nextSibling.nextSibling)
		if title.upper() == "MECHANISM" and len(pInfo["MECHANISM"])==0:
			pInfo["MECHANISM"] = getNodeText(spectitle.nextSibling.nextSibling)
		if title.upper() == "APPLICATION RECOMMENDATION" and len(pInfo["APPLICATION RECOMMENDATION"])==0:
			pInfo["APPLICATION RECOMMENDATION"] = getNodeText(spectitle.nextSibling.nextSibling)
		if title.upper() == "SAFE HANDLING PRECAUTIONS":
			if len(getNodeText(spectitle.nextSibling.nextSibling)) > 0:
				pInfo["SAFE HANDLING PRECAUTIONS"] = getNodeText(spectitle.nextSibling.nextSibling)
			if len(pInfo["SAFE HANDLING PRECAUTIONS"]) == 0:
				pInfo["SAFE HANDLING PRECAUTIONS"] = getNodeText(spectitle.nextSibling.nextSibling.nextSibling.nextSibling)
		if title.upper() == "WARNINGS":
			if len(getNodeText(spectitle.nextSibling.nextSibling))>0:
				pInfo["WARNINGS"] = getNodeText(spectitle.nextSibling.nextSibling)
			if len(pInfo["WARNINGS"]) == 0:
				pInfo["WARNINGS"] = getNodeText(spectitle.nextSibling.nextSibling.nextSibling.nextSibling)
		if title.upper().find("PACKAGE")>-1:
			pInfo["Package"] = title.replace("Package：","").replace("PACKAGE","").replace("Package","")
		if title.upper().find("PACKING")>-1 and len(pInfo["Package"])==0:
			pInfo["Package"] = title.replace("Packing specification","")
		if title.find("Package")>-1 and len(pInfo["Package"])==0:
			pInfo["Package"] = getNodeText(spectitle.nextSibling.nextSibling)
			
		if title.upper().find("STORAGE")>-1:
			pInfo["Storage"] = title.replace("Storage：","").replace("STORAGE","").replace("Storage","")
		if title.upper().find("STORAGE")>-1 and len(pInfo["Storage"]) == 0:
			pInfo["Storage"] = getNodeText(spectitle.nextSibling.nextSibling)

		if title.find("Shelf life:")>-1:
			pInfo["Shelf life"] = title.replace("Shelf life:","").replace("Shelf life","")
		if title.find("SHELFLIFE")>-1:
			pInfo["Shelf life"] = title.replace("SHELFLIFE","")
		if title.find("Shelf life")>-1 and len(pInfo["Shelf life"] ) == 0:
			pInfo["Shelf life"] = getNodeText(spectitle.nextSibling.nextSibling)
	trs = sope.find_all("tr")
	for tr in trs:
		tds = tr.find_all("td")
		if len(tds) > 1:
			title1 = getNodeText(tds[0])
			title2 = getNodeText(tds[1])
			if len(title2)==0 and len(tds)>2:
				title2= getNodeText(tds[2])
			if title1 == "Activity Temperature":
				pInfo["Activity Temperature"] = title2
			if title1 == "Optimum Temperature":
				pInfo["Optimum Temperature"] = title2
			if title1 == "Activity pH":
				pInfo["Activity pH"] = title2
			if title1 == "Optimum pH":
				pInfo["Optimum pH"] = title2
			if title1 == "Declared Activity*":
				pInfo["Declared Activity*"] = title2
			if title1 == "Physical Form":
				pInfo["Physical Form"] = title2
			if title1 == "Color**":
				pInfo["Color**"] = title2
			if title1 == "Odour":
				pInfo["Odour"] = title2
			if title2 == "Particle size (%<40 mesh)":
				pInfo["Particle size (%<40 mesh)"] = getNodeText(tds[2])
			if title2 == "Loss on drying/(%)":
				pInfo["Loss on drying/(%)"] = getNodeText(tds[2])
			if title2 == "Lead/(mg/kg)":
				pInfo["Lead/(mg/kg)"] = getNodeText(tds[2])
			if title2 == "Arsenic/(mg/kg)":
				pInfo["Arsenic/(mg/kg)"] = getNodeText(tds[2])
			if title2 == "Total viable count/(CFU/g)":
				pInfo["Total viable count/(CFU/g)"] = getNodeText(tds[2])
			if title2 == "Coliform Bacteria/(CFU/g)":
				pInfo["Coliform Bacteria/(CFU/g)"] = getNodeText(tds[2])
			if title2 == "Salmonella/(25g)":
				pInfo["Salmonella/(25g)"] = getNodeText(tds[2])
	products.append(pInfo.copy())
	# print(pInfo)
def getProductList(url, typeInfo, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	pros = sope.find_all("dd")
	for pro in pros:
		link = pro.find("a")
		getProductInfo("http://www.chinaenzymes.com"+link["href"], typeInfo, products)

excelFileName="chinaenzymes.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("http://www.chinaenzymes.com/Products/Paper_enzymes/2020/0423/286.html", {"t1":"111","t2":"22"}, products)

for pageIndex in range(1, 3):
	typeInfo = {"t1":"Food&Beverage", "t2":"Wine"}
	getProductList("http://www.chinaenzymes.com/Products/Food&Beverage/Wine/list_31_"+str(pageIndex)+".html",typeInfo, products)

for pageIndex in range(1, 6):
	typeInfo = {"t1":"Food&Beverage","t2":"Bakery"}
	getProductList("http://www.chinaenzymes.com/Products/Food&Beverage/Bakery/list_32_"+str(pageIndex)+".html",typeInfo, products)

for pageIndex in range(1, 3):
	typeInfo = {"t1":"Food&Beverage","t2":"Fruit_Juice"}
	getProductList("http://www.chinaenzymes.com/Products/Food&Beverage/Fruit_Juice/list_30_"+str(pageIndex)+".html",typeInfo, products)

for pageIndex in range(1, 7):
	typeInfo = {"t1":"Ethanol & Alcohol","t2":""}
	getProductList("http://www.chinaenzymes.com/Products/Ethanol%20&%20Alcohol/list_13_"+str(pageIndex)+".html",typeInfo, products)

for pageIndex in range(1, 4):
	typeInfo = {"t1":"Animal Feed","t2":"Complex feed enzyme"}
	getProductList("http://www.chinaenzymes.com/Products/Ethanol%20&%20Alcohol/Complex_feed_enzyme/list_33_"+str(pageIndex)+".html",typeInfo, products)

for pageIndex in range(1, 6):
	typeInfo = {"t1":"Animal Feed","t2":"Single feed enzyme"}
	getProductList("http://www.chinaenzymes.com/Products/Ethanol%20&%20Alcohol/Single_feed_enzyme/list_34_"+str(pageIndex)+".html",typeInfo, products)

for pageIndex in range(1, 4):
	typeInfo = {"t1":"Textile","t2":"Denim Finishing"}
	getProductList("http://www.chinaenzymes.com/Products/Textile/Denim_Finishing/list_35_"+str(pageIndex)+".html",typeInfo, products)

for pageIndex in range(1, 2):
	typeInfo = {"t1":"Textile","t2":"Bleach Clean up"}
	getProductList("http://www.chinaenzymes.com/Products/Textile/Bleach_Clean_up/",typeInfo, products)

for pageIndex in range(1, 2):
	typeInfo = {"t1":"Textile","t2":"Bio scouring"}
	getProductList("http://www.chinaenzymes.com/Products/Textile/Bleach_Clean_up/",typeInfo, products)

for pageIndex in range(1, 5):
	typeInfo = {"t1":"Textile","t2":"Bio polishing"}
	getProductList("http://www.chinaenzymes.com/Products/Textile/Bio_polishing/list_38_"+str(pageIndex)+".html",typeInfo, products)

for pageIndex in range(1, 3):
	typeInfo = {"t1":"Textile","t2":"Desizing"}
	getProductList("http://www.chinaenzymes.com/Products/Textile/Desizing/list_39_"+str(pageIndex)+".html",typeInfo, products)

for pageIndex in range(1, 4):
	typeInfo = {"t1":"Leather","t2":""}
	getProductList("http://www.chinaenzymes.com/Products/Leather/list_54_"+str(pageIndex)+".html",typeInfo, products)

for pageIndex in range(1, 4):
	typeInfo = {"t1":"Detergent","t2":""}
	getProductList("http://www.chinaenzymes.com/Products/Detergent/list_45_"+str(pageIndex)+".html",typeInfo, products)

for pageIndex in range(1, 8):
	typeInfo = {"t1":"Single enzyme","t2":""}
	getProductList("http://www.chinaenzymes.com/Products/Single_enzyme/list_46_"+str(pageIndex)+".html",typeInfo, products)

for pageIndex in range(1, 3):
	typeInfo = {"t1":"Brewery","t2":""}
	getProductList("http://www.chinaenzymes.com/Products/Brewery/list_52_"+str(pageIndex)+".html",typeInfo, products)

for pageIndex in range(1, 2):
	typeInfo = {"t1":"Starch","t2":""}
	getProductList("http://www.chinaenzymes.com/Products/Starch___Sugar/",typeInfo, products)


for pageIndex in range(1, 6):
	typeInfo = {"t1":"Paper enzymes","t2":""}
	getProductList("http://www.chinaenzymes.com/Products/Paper_enzymes/list_58_"+str(pageIndex)+".html",typeInfo, products)


headers=['link','type1','type2','title','PRODUCT DESCRIPTION','MECHANISM','APPLICATION RECOMMENDATION','SAFE HANDLING PRECAUTIONS','WARNINGS',
'Package','Storage','Shelf life','Activity Temperature','Optimum Temperature','Activity pH','Optimum pH','Declared Activity*','Physical Form',
'Color**','Odour','Particle size (%<40 mesh)','Loss on drying/(%)','Lead/(mg/kg)','Arsenic/(mg/kg)','Total viable count/(CFU/g)','Coliform Bacteria/(CFU/g)','Salmonella/(25g)']

for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
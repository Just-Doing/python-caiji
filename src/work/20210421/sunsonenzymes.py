from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		print(html_code)
		return html_code
	except:
		print("retry index"+str(retryCount)+url)
		retryCount += 1
		if retryCount< 5:
			getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, typeInfo, products):
	print(str(len(products)) + url)
	productListHtml = getRenderdHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	specArea = sope.find("span", attrs={"id":"ShoveWebControl_Texttestcon82"})
	
	pInfo = {
		"link":url,
		"type":typeInfo,
	}
	pInfo["Introduction"]=""
	if specArea.contents[0].name =="p":
		txt=getNodeText(specArea.contents[0])
		if len(txt)>0:
			pInfo["Introduction"]=txt
	title = getNodeText(sope.find("h2",attrs={"id":"EIMS_C_40008_4_AliaName"}))
	pInfo["title"] = title

	specTitles = sope.find_all("strong")
	for spectitle in specTitles:
		title = getNodeText(spectitle)
		if title == "Introduction" or title == "Introduction:":
			var = getNodeText(spectitle.parent).replace("Introduction:","").replace("Introduction","")
			if len(var)==0:
				var = getNodeText(spectitle.parent.nextSibling)
			pInfo["Introduction"] = var
		if title == "Benefit" or title == "Benefit:":
			var = getNodeText(spectitle.parent).replace("Benefit:","").replace("Benefit","")
			if len(var)==0:
				var = getNodeText(spectitle.parent.nextSibling)
			pInfo["Benefit"] = var

		if title == "Storage" or title == "Storage:":
			var = getNodeText(spectitle.parent).replace("Storage:","").replace("Storage","")
			if len(var)==0:
				var = getNodeText(spectitle.parent.nextSibling)
			pInfo["Storage"] = var
		if title == "Packaging" or title == "Packaging:" or title == "PACKAGE":
			var = getNodeText(spectitle.parent).replace("Packaging:","").replace("Packaging","")
			if len(var)==0:
				var = getNodeText(spectitle.parent.nextSibling)
			pInfo["Packaging"] = var
		if title == "Shelf life" or title == "Shelf life:":
			var = getNodeText(spectitle.parent).replace("Shelf life:","").replace("Shelf life","")
			if len(var)==0:
				var = getNodeText(spectitle.parent.nextSibling)
			pInfo["Shelf life"] = var
		if title == "Safety" or title == "Safety:":
			var = getNodeText(spectitle.parent).replace("Safety:","").replace("Safety","")
			if len(var)==0:
				var = getNodeText(spectitle.parent.nextSibling)
			pInfo["Safety"] = var
		if title == "PACKAGE & STORAGE" or title == "PACKAGE & STORAGE:":
			var = getNodeText(spectitle.parent).replace("PACKAGE & STORAGE:","").replace("PACKAGE & STORAGE","")
			if len(var)==0:
				var = getNodeText(spectitle.parent.nextSibling)
			pInfo["PACKAGE & STORAGE"] = var
		if title == "PACKAGE&STORAGE" or title == "PACKAGE&STORAGE:":
			var = getNodeText(spectitle.parent).replace("PACKAGE&STORAGE:","").replace("PACKAGE&STORAGE","")
			if len(var)==0:
				var = getNodeText(spectitle.parent.nextSibling)
			pInfo["PACKAGE & STORAGE"] = var
		if title == "ADVANTAGES" or title == "ADVANTAGES:":
			var = getNodeText(spectitle.parent).replace("ADVANTAGES:","").replace("ADVANTAGES","")
			if len(var)==0:
				var = getNodeText(spectitle.parent.nextSibling)
			pInfo["ADVANTAGES"] = var
		if title == "FUNCTIONS & BENEFITS" or title == "FUNCTIONS & BENEFITS:" :
			var = getNodeText(spectitle.parent).replace("FUNCTIONS & BENEFITS:","").replace("FUNCTIONS & BENEFITS","")
			if len(var)==0:
				var = getNodeText(spectitle.parent.nextSibling)
			pInfo["FUNCTIONS & BENEFITS"] = var
		if title == "FUNCTIONS&BENEFITS" or title == "FUNCTIONS&BENEFITS:":
			var = getNodeText(spectitle.parent).replace("FUNCTIONS&BENEFITS:","").replace("FUNCTIONS&BENEFITS","")
			if len(var)==0:
				var = getNodeText(spectitle.parent.nextSibling)
			pInfo["FUNCTIONS & BENEFITS"] = var
		if title == "Appearance" or title == "Appearance:":
			var = getNodeText(spectitle.parent).replace("Appearance:","").replace("Appearance","")
			if len(var)==0:
				var = getNodeText(spectitle.parent.nextSibling)
			pInfo["Appearance"] = var
		if title == "DEFINITION OF UNIT" or title == "DEFINITION OF UNIT:":
			var = getNodeText(spectitle.parent).replace("DEFINITION OF UNIT:","").replace("DEFINITION OF UNIT","")
			if len(var)==0:
				var = getNodeText(spectitle.parent.nextSibling)
			pInfo["DEFINITION OF UNIT"] = var
		if title == "Dosage:" or title == "Dosage":
			var = getNodeText(spectitle.parent).replace("Dosage:","").replace("Dosage","")
			if len(var)==0:
				var = getNodeText(spectitle.parent.nextSibling)
			if len(var) ==0 and spectitle.nextSibling!=None:
				var = getNodeText(spectitle.nextSibling.nextSibling)
			pInfo["Dosage"] = var

	pSpecInfos = sope.find_all("p")
	for pSpecinfo in pSpecInfos:
		title = getNodeText(pSpecinfo)
		if title=="Introduction:":
			pInfo["Introduction"] = getNodeText(pSpecinfo.nextSibling)
		if title=="DEFINITION OF UNIT:":
			pInfo["DEFINITION OF UNIT"] = getNodeText(pSpecinfo.nextSibling)
		if title=="FUNCTIONS & BENEFITS:":
			pInfo["FUNCTIONS & BENEFITS"] = getNodeText(pSpecinfo.nextSibling)
		if title=="ADVANTAGES:":
			pInfo["ADVANTAGES"] = getNodeText(pSpecinfo.nextSibling)

	specTitles = sope.find_all("tr")
	for spectitle in specTitles:
		tds = spectitle.find_all("td")
		if len(tds)>1:
			title = getNodeText(tds[0])
			val = getNodeText(tds[1])
			pInfo[title] = val
	products.append(pInfo.copy())
	# print(pInfo)

def getProductList(url, type1, products):
	productListHtml = getRenderdHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	plArea = sope.find("div", attrs={"id":"EIMS_C_40018_1_prolist"})
	pros = plArea.find_all("li", attrs={"class":"xn_c_products_50_proli"})
	for pro in pros:
		link = pro.find("a")
		getProductInfo("http://en.sunsonenzymes.com/"+link["href"], type1, products)

excelFileName="sunsonenzymes.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("http://en.sunsonenzymes.com/prod_view.aspx?TypeId=65&Id=197&FId=t3:65:3", "", products)

getProductList("http://en.sunsonenzymes.com/products.aspx?Id=65&TypeId=65&fid=t3:65:3","Animal health", products)
getProductList("http://en.sunsonenzymes.com/products.aspx?Id=66&TypeId=66&fid=t3:66:3","Brewing", products)
getProductList("http://en.sunsonenzymes.com/products.aspx?Id=68&TypeId=68&fid=t3:68:3","Textile", products)
getProductList("http://en.sunsonenzymes.com/products.aspx?Id=73&TypeId=73&fid=t3:73:3","Alcohol and ethanol", products)
getProductList("http://en.sunsonenzymes.com/products.aspx?Id=78&TypeId=78&fid=t3:78:3","Baking", products)
getProductList("http://en.sunsonenzymes.com/products.aspx?Id=80&TypeId=80&fid=t3:80:3","Starch&sugar", products)
getProductList("http://en.sunsonenzymes.com/products.aspx?Id=125&TypeId=125&fid=t3:125:3","Leather", products)
getProductList("http://en.sunsonenzymes.com/products.aspx?Id=84&TypeId=84&fid=t3:84:3","Meat and protein", products)

headers=['link','type','title','Introduction','Benefit','Storage','Packaging','Shelf life','Safety','PACKAGE & STORAGE','FUNCTIONS & BENEFITS','ADVANTAGES','Appearance','Color'
,'Odor','Solubility','CAS NO.IUB NO.','Temperature','pH','Liquor ratio','Dosage','Time','Inactivation','DEFINITION OF UNIT','APLu/g','CMC']

for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
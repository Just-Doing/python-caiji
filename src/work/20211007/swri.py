from enum import IntEnum
import requests
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import string
import re
import time
import math

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	
	opener = urllib.request.build_opener()
	opener.addheaders = [('User-agent', 'Mozilla/5.0')]
	urllib.request.install_opener(opener)
	urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		request_obj=urllib.request.Request(url=url,  headers={
			'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
			'cookie':'_ga=GA1.2.510794350.1632490112; _gid=GA1.2.1808662766.1632490112; __stripe_mid=1840413b-ed41-44e0-a11f-ff7c582fe2d4d861a3; __stripe_sid=26d2d2f8-4025-4ddc-a3ce-9959ee1e079c000594; PHPSESSID=ef2a208c339d0e9e15f810221e50fd85',
			"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.82 Safari/537.36"
		})
		htmlHeader = requests.head(url)
		print(htmlHeader)
		if htmlHeader.status_code ==200:
			response_obj=urllib.request.urlopen(request_obj)
			html_code=response_obj.read()
			return html_code
		else:
			return ''
	except:
		retryCount = retryCount + 1
		if retryCount < 5:
			print("retry index"+str(retryCount)+url)
			time.sleep(60)
			return getHtmlFromUrl(url)
		else:
			retryCount = 0
			return ""

def requestJson(url, page):
	r = requests.post(url, headers={
		'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
		'cookie':'_ga=GA1.2.510794350.1632490112; _gid=GA1.2.1808662766.1632490112; __stripe_mid=1840413b-ed41-44e0-a11f-ff7c582fe2d4d861a3; __stripe_sid=26d2d2f8-4025-4ddc-a3ce-9959ee1e079c000594; PHPSESSID=ef2a208c339d0e9e15f810221e50fd85',
		"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"
	},data={
		'page':page,
		'view_name':'patents',
		'view_display_id':'block_1',
		'view_dom_id':'42f98ae28b6b8d60c7be0dffcd57a7cc',
		'RecordsPerPage':'10',
		'SearchCriteria':'All Words',
		'sort':'desc',
	})
	datas = json.loads(r.text)
	return datas

def getRenderdHtmlFromUrl(url):

	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	return BeautifulSoup(browser.page_source, "html.parser",from_encoding="utf-8")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, products):
	print(str(len(products)) + url)
	sope = getHtmlFromUrl(url)
	if len(sope)>0:
		titleArea = sope.find("div", attrs={"class":"breadcrumbs"})

		pInfo={}
		products.append(pInfo.copy())

def getProductList(url,pageIndex, products):
	print(url)
	sope = requestJson(url, pageIndex)
	if len(sope)>1:
		pListAreaStr = sope[1]["data"]
		pListSope = BeautifulSoup(pListAreaStr, "html.parser",from_encoding="utf-8")
		tBody = pListSope.find("tbody")
		trs = tBody.find_all("tr")
		for tr in trs:
			pInfo={}
			tds = tr.find_all("td")
			if len(tds) == 4:
				pLink = tr.find("a")
				pInfo["Technology Name"] = getNodeText(pLink)
				pInfo["Webpage"] = "https://www.swri.org"+pLink["href"]
				pInfo["Published Time"] = getNodeText(tds[1])
				pInfo["Inventor(s)"] = getNodeText(tds[3])
				pInfo["Patent Status"] = "Patent Number："+getNodeText(tds[0])
				print(len(products))
				products.append(pInfo.copy())
excelFileName="swri.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

headers=[
	'Ref. No.','University','Technology Name','是否相关','Webpage','Published Time','Inventor(s)','Technology Categories','Keywords'
	,'Patent Status','Licensing Contact Person','Licensing Contact Email','备注'
]

# getProductList("https://www.swri.org/views/ajax", 0, products)
# getProductList("https://esschemco.com/product-category/amines/",'amines', products)

for pageIndex in range(0, 141):
	getProductList("https://www.swri.org/views/ajax", pageIndex, products)


for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
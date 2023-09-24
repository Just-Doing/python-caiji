from openpyxl import load_workbook
import xlrd

workbook = xlrd.open_workbook('tissuearray.xlsx')
sheet = workbook.sheet_by_name('tissuearray')
for i in range(sheet.nrows):
    if i>0:
        row_data = sheet.row_values(i)
        print(row_data[0])
        wb = load_workbook(row_data[0])#打开excel
        ws = wb[wb.sheetnames[0]]#读取第一个sheet
        ws.insert_rows(8) #在8行处插入新数据   插入不是替换
        ws.cell(row=8, column=1, value='Thickness (µm)')
        ws.cell(row=8, column=2, value=row_data[1])
        wb.save(row_data[0])

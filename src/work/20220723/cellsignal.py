from ast import Store
from enum import IntEnum
import requests
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import cfscrape
import json
import string
import re
import time
import math
import _thread

import numpy as np

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

def getHtmlFromUrl(url, type="get", para={}):
	headers = {
		"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.82 Safari/537.36",
		"cookie":"_ga=GA1.2.1846208171.1605273760; href=https%3A%2F%2Fwww.sinobiological.com%2Fresearch%2Ftargeted-therapy; accessId=5aff5fb0-84db-11e8-a3b3-d368cce40a8e; _gcl_au=1.1.1660157260.1645016298; Hm_lvt_2d911fad88dfe99ff8bbf610824f3e67=1645016298; sbb=%252be43ohTbVTr09K%252bxQlr1%252bK0onQvF%252bMIXgZM%252bveGXMHU%253dXTaJopSyq01ihC4cD5jOfAa8AEgsWX792EAllASK%252bEKohz0p3JxSEJFavoBnvUmw4fhBpwMcWGZ2Qy%252bRRF2U%252bMSxcQdMfdwOcT%252bR%252bo7qyEU%252br8SBQuGE8GJWgDFeSTZ4QS0HvJFVazETAoyuKMwGHYRoD68%252f7qno5Bg%252bEH9sSXM4upMLtz%252f4IdNkjX6GD0JYHbiUh%252blGTwi25Iz3IKocTDD58DE1yYiY3DxeifN7Qz6OxtXX21lrBpnvgDu9ANN%252f7TTxWWMmOIjxVG772o%252bYGkE9AMxcU5O4cIrT9cubm6dAdgw6n%252fQRZpTVxNv2TGHdHZblPNcfu4dTWVsL3aqaag%253d%253d; _gid=GA1.2.832211649.1645016298; _ce.s=v11.rlc~1645016301520; pageViewNum=13; Hm_lpvt_2d911fad88dfe99ff8bbf610824f3e67=1645017042; Currency=RMB; LocationCode=CN"
	}

	scraper = cfscrape.create_scraper()
	html_code = scraper.get(url,headers=headers).text
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")

def requestJson(url):
	r = requests.post(url,data={"input":"atcc"}, headers={
		'Content-Type': 'application/x-www-form-urlencoded',
		'cookie':'visid_incap_2255650=4oBBaRPnQfCVoYEiTmjTq/NVAWEAAAAAQUIPAAAAAAD69PQHUoB0KplKq7/j0+gH; nlbi_2255650=CJKhHYlMm17tpKyoBzOViAAAAACDEjp3gL6bj6YL8j9XE0d/; incap_ses_893_2255650=m1tJIuDRUEp3FE/5GpNkDPRVAWEAAAAAM2KkDpvtARtZral+cMXSVw==; _gcl_au=1.1.76703404.1627477493; _gid=GA1.2.730047202.1627477493; BCSessionID=83af10b8-9488-4b7b-a3b1-3640f178dca2; categoryView=grid; _ga_S46FST9X1M=GS1.1.1627477492.1.1.1627478562.0; _ga=GA1.2.31731397.1627477493; _gat_UA-139934-1=1; _uetsid=69fc2d30efa411eb8818eb045f8760e5; _uetvid=69fc3a70efa411ebba3a23c153f6e477; .Nop.Customer=d664d529-d14a-44b1-86b3-cbf5373277b4',
		"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"
	})
	return BeautifulSoup(r.text, "html.parser",from_encoding="utf-8")


def getRenderdHtmlFromUrl(url):

	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	return BeautifulSoup(browser.page_source, "html.parser")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, products):
	print(str(len(products))+":"+url)
	sope = getHtmlFromUrl(url)
	sku = sope.find("tr", attrs={"class":"selected pointer"})
	appArea = sope.find("ol", attrs={"id":"applications-list"})
	appStr = ""
	if appArea != None:
		apps = appArea.find_all("li")
		for app in apps:
			appStr += getNodeText(app) + ","
	
	pInfo = {
		"link":url,
		"Product Name": getNodeText(sope.find("h1", attrs={"itemprop":"name"})),
		"Filter": appStr
	}
	if sku!=None:
		skuTds = sku.find_all("td")
		pInfo["cat"]=getNodeText(skuTds[0])
		pInfo["size"]=getNodeText(skuTds[1])

	specData = sope.find("div", attrs={"id":"validation-data"})
	if specData != None:
		specs = specData.find_all("tr")
		for spec in specs:
			tds = spec.find_all("td")
			if len(tds) == 2:
				title = getNodeText(tds[0])
				value = getNodeText(tds[1])
				pInfo[title] = value
	storage = sope.find("div", attrs={"id":"storage"})
	pInfo["storage"]=getNodeText(storage).replace("Storage","")

	
	specificity = sope.find("div", attrs={"id":"specificity"})
	pInfo["specificity"]=getNodeText(specificity).replace("Specificity / Sensitivity","")


	source = sope.find("div", attrs={"id":"source"})
	pInfo["source"]=getNodeText(source).replace("Source / Purification","")
	print(pInfo)
	products.append(pInfo.copy());

def getProductList(url, products):
	html_code = getRenderdHtmlFromUrl(url)
	proListArea = html_code.find("table", attrs={"id":"product-list"}).find("tbody")
	proList = proListArea.find_all("tr")
	for tr in proList:
		pLink = tr.find("a")
		getProductInfo("https://www.cellsignal.com"+pLink["href"], products)
		




def theardFun(startPage, endPage,excelFname):
	products = []
	# getProductList('https://www.cellsignal.com/browse/primary-antibodies/neuroscience/monoclonal-antibody?N=102236+782416468+4294967218+4294956287&No=0&Nrpp=30')
	for pIndex in range(startPage, endPage):
		getProductList('https://www.cellsignal.com/browse/primary-antibodies/neuroscience/monoclonal-antibody?N=102236+782416468+4294967218+4294956287&No='+str(pIndex*200)+'&Nrpp=200', products)


	wb = Workbook()
	workSheet = wb.active

	headers=[
		'link','Product Name','cat','size','Filter','REACTIVITY','SENSITIVITY','MW (kDa)','Source/Isotype','storage','specificity'
		,'source'
	]

	for index,head in enumerate(headers):
		workSheet.cell(1, index+1).value = head.strip()
	for index,p in enumerate(products):
		writeExcel(workSheet, headers, index + 2, p)
	print("flish")	

	wb.save(excelFname)
	
try:
	# _thread.start_new_thread( theardFun, (0, 3, "product1" ) )
	# _thread.start_new_thread( theardFun, (3, 6, "product2" ) )
	# _thread.start_new_thread( theardFun, (6, 9, "product3" ) )
	# _thread.start_new_thread( theardFun, (9, 12, "product4" ) )
	# _thread.start_new_thread( theardFun, (12, 15, "product5" ) )
	_thread.start_new_thread( theardFun, (15, 19, "product6" ) )
except:
	print ("Error: 无法启动线程")
while 1:
   pass
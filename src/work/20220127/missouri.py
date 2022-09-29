from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import cfscrape
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')
	

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36"}

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		if retryCount < 5:
			getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url, isScreenShotName=""):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")
	chrome_options.add_argument('user-agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.130 Safari/537.36"')
	chrome_options.add_argument("cookie=__hstc=240517788.35783e1d438e8f99e34188727b050107.1648374652176.1648374652176.1648374652176.1; hubspotutk=35783e1d438e8f99e34188727b050107; __hssrc=1; _gid=GA1.2.122553597.1648374652; _gcl_au=1.1.15847883.1648374654; _fbp=fb.1.1648374655847.2011294960; __hssc=240517788.6.1648374652176; _ga=GA1.1.457442853.1648374652; _ga_SBEXK5LM3N=GS1.1.1648374653.1.1.1648376932.0")
	chrome_options.add_argument("--no-sandbox")
	
	# chrome_options.add_argument("--proxy-server=http://127.0.0.1:7890")

	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	time.sleep(10)
	if len(isScreenShotName) > 0:
		imgEle = browser.find_element_by_xpath('//body/img[1]')
		if imgEle !=None:
			imgEle.screenshot(isScreenShotName)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(products):
	for pInfo in products:
		productListHtml = getRenderdHtmlFromUrl(pInfo["link"])
		sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
		tds = sope.find_all("td")
		for td in tds:
			title = getNodeText(td.find("u"))
			title2 = getNodeText(td.find("b"))
			if title.find("Manager") > -1:
				pInfo["Contact email"] = getNodeText(td.find("a"))
				pInfo["Contact person"] = td.find("b").nextSibling.nextSibling.nextSibling
			
			if title2.find("Title") > -1:
				pInfo["Title"] = td.find("b").nextSibling.nextSibling
				print(pInfo["Title"])
		spans = sope.find_all("span")
		for span in spans:
			title = getNodeText(span).replace("\n", "").replace("\r", "")
			if title.find("Patent Status") > -1:
				pInfo["Patent Status"] = span.nextSibling
			if title.find("Inventors") > -1:
				pInfo["Inventors"] = span.nextSibling

def getProductLink(url, type,  products):
	print(str(len(products)) + url)
	pInfo={
		"link": url,
		"type":type
	}
	products.append(pInfo.copy())
def getProductList(url, type, products):
	print(url)
	productListHtml = getRenderdHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	proArea = sope.find("table", attrs={"style":"width:100%;margin-top:0px;margin-left:0px;margin-right:0px;padding-top:0px;"})
	pros = proArea.find_all("tr")
	for pro in pros:
		link = pro.find("a")
		if link!=None:
			getProductLink("https://tech.missouri.edu"+link["href"], type, products)

excelFileName="missouri-Other Software.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("https://tech.missouri.edu/UREF045FF4C1CB69F944AD9642AFA4568BE2/TF_Technology_Input/TSAA6290080312340349/A11137930331",'', products)

# getProductList("https://tech.missouri.edu/UREFB47C3C9A0C633B498FE46EF36314AE76/TF_Home/V14EXEC/B_Nothing-B_SetCVar_form~searchtype_Taxonomy-B_SetCVar_Form~Taxonomy~Item_TSAA6716131839419764-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.09033141667473088",'Copyright -> Other Copyright', products)
# getProductList("https://tech.missouri.edu/UREF6D3CF18C69C9A342AD64B66F6D53E400/TF_Home/V14EXEC/B_Nothing-B_SetCVar_form~searchtype_Taxonomy-B_SetCVar_Form~Taxonomy~Item_TSAA6716131828408114-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.3421182607830018",'Copyright -> Training Materials', products)

# getProductList("https://tech.missouri.edu/UREF383FEA27EF433C4BB4333CE40EAE9F27/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_3_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.577617520082129",'Engineering -> Chemical', products)
# getProductList("https://tech.missouri.edu/UREF6215EE0403F8A5409F114DBA6A00A200/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_2_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.7258459607476451",'Engineering -> Chemical', products)
# getProductList("https://tech.missouri.edu/UREF2365F56C4592DC4BB4C461C7DFD4D9E6/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_3_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.44387231878402567",'Engineering -> Chemical', products)

# getProductList("https://tech.missouri.edu/UREFED571DF711328A4F896DA89FFF1E98AF/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_1_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.885377228236935",'Engineering -> Materials Science', products)
# getProductList("https://tech.missouri.edu/UREFED571DF711328A4F896DA89FFF1E98AF/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_2_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.885377228236935",'Engineering -> Materials Science', products)
# getProductList("https://tech.missouri.edu/UREFED571DF711328A4F896DA89FFF1E98AF/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_3_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.885377228236935",'Engineering -> Materials Science', products)
# getProductList("https://tech.missouri.edu/UREFED571DF711328A4F896DA89FFF1E98AF/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_4_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.885377228236935",'Engineering -> Materials Science', products)
# getProductList("https://tech.missouri.edu/UREFED571DF711328A4F896DA89FFF1E98AF/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_5_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.885377228236935",'Engineering -> Materials Science', products)
# getProductList("https://tech.missouri.edu/UREFED571DF711328A4F896DA89FFF1E98AF/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_6_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.885377228236935",'Engineering -> Materials Science', products)
# getProductList("https://tech.missouri.edu/UREFED571DF711328A4F896DA89FFF1E98AF/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_7_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.885377228236935",'Engineering -> Materials Science', products)


# getProductList("https://tech.missouri.edu/UREFB7F90AE169C02547B5A4E4CA47C30F65/TF_Home/V14EXEC/B_Nothing-B_SetCVar_form~searchtype_Taxonomy-B_SetCVar_Form~Taxonomy~Item_TSAA6716132003504057-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.07636543525838468",'Engineering -> Mechanical', products)

# getProductList("https://tech.missouri.edu/UREFB7F90AE169C02547B5A4E4CA47C30F65/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_1_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.518468075891573",'Engineering -> Other Engineering', products)
# getProductList("https://tech.missouri.edu/UREFB7F90AE169C02547B5A4E4CA47C30F65/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_2_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.518468075891573",'Engineering -> Other Engineering', products)
# getProductList("https://tech.missouri.edu/UREFB7F90AE169C02547B5A4E4CA47C30F65/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_3_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.518468075891573",'Engineering -> Other Engineering', products)


# getProductList("https://tech.missouri.edu/UREF6225F6E9F791C54DBFD56724DF88AABD/TF_Home/V14EXEC/B_Nothing-B_SetCVar_form~searchtype_Taxonomy-B_SetCVar_Form~Taxonomy~Item_TSAA6716132034534549-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.6464323559465619",'Life Sciences -> Animal Agriculture', products)

# getProductList("https://tech.missouri.edu/UREF6225F6E9F791C54DBFD56724DF88AABD/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_1_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.29751949514200837",'Life Sciences -> Other Life Sciences', products)
# getProductList("https://tech.missouri.edu/UREF6225F6E9F791C54DBFD56724DF88AABD/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_2_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.29751949514200837",'Life Sciences -> Other Life Sciences', products)

# getProductList("https://tech.missouri.edu/UREF6225F6E9F791C54DBFD56724DF88AABD/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_1_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.5403285400683817",'Life Sciences -> Plant Biotechnology', products)
# getProductList("https://tech.missouri.edu/UREF6225F6E9F791C54DBFD56724DF88AABD/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_2_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.5403285400683817",'Life Sciences -> Plant Biotechnology', products)


# getProductList("https://tech.missouri.edu/UREF6225F6E9F791C54DBFD56724DF88AABD/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_1_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.5723123471349976",'Life Sciences -> Research Tools', products)
# getProductList("https://tech.missouri.edu/UREF6225F6E9F791C54DBFD56724DF88AABD/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_2_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.5723123471349976",'Life Sciences -> Research Tools', products)
# getProductList("https://tech.missouri.edu/UREF6225F6E9F791C54DBFD56724DF88AABD/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_3_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.5723123471349976",'Life Sciences -> Research Tools', products)
# getProductList("https://tech.missouri.edu/UREF6225F6E9F791C54DBFD56724DF88AABD/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_4_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.5723123471349976",'Life Sciences -> Research Tools', products)

# getProductList("https://tech.missouri.edu/UREF6225F6E9F791C54DBFD56724DF88AABD/TF_Home/V14EXEC/B_Nothing-B_SetCVar_form~searchtype_Taxonomy-B_SetCVar_Form~Taxonomy~Item_TSAA6716132141602129-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.2339664628273599",'Life Sciences -> Veterinary Medicine', products)

# getProductList("https://tech.missouri.edu/UREFC8BC848563476A439FAC7825274EC908/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_1_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.8576052889580295",'Medical -> Diagnostics', products)
# getProductList("https://tech.missouri.edu/UREFC8BC848563476A439FAC7825274EC908/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_2_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.8576052889580295",'Medical -> Diagnostics', products)
# getProductList("https://tech.missouri.edu/UREFC8BC848563476A439FAC7825274EC908/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_3_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.8576052889580295",'Medical -> Diagnostics', products)
# getProductList("https://tech.missouri.edu/UREFC8BC848563476A439FAC7825274EC908/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_4_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.8576052889580295",'Medical -> Diagnostics', products)
# getProductList("https://tech.missouri.edu/UREFC8BC848563476A439FAC7825274EC908/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_5_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.8576052889580295",'Medical -> Diagnostics', products)



# getProductList("https://tech.missouri.edu/UREF3F6CA2C030477C46A79DC10981B047B5/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_1_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.22703734748214854",'Medical -> Drug Delivery', products)
# getProductList("https://tech.missouri.edu/UREF3F6CA2C030477C46A79DC10981B047B5/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_2_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.22703734748214854",'Medical -> Drug Delivery', products)

# getProductList("https://tech.missouri.edu/UREF3F6CA2C030477C46A79DC10981B047B5/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_1_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.6198739699962372",'Medical -> Medical Devices', products)
# getProductList("https://tech.missouri.edu/UREF3F6CA2C030477C46A79DC10981B047B5/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_2_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.6198739699962372",'Medical -> Medical Devices', products)
# getProductList("https://tech.missouri.edu/UREF3F6CA2C030477C46A79DC10981B047B5/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_3_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.6198739699962372",'Medical -> Medical Devices', products)
# getProductList("https://tech.missouri.edu/UREF3F6CA2C030477C46A79DC10981B047B5/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_4_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.6198739699962372",'Medical -> Medical Devices', products)
# getProductList("https://tech.missouri.edu/UREF3F6CA2C030477C46A79DC10981B047B5/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_5_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.6198739699962372",'Medical -> Medical Devices', products)

# getProductList("https://tech.missouri.edu/UREF3F6CA2C030477C46A79DC10981B047B5/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_1_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.03600912779575505",'Medical -> Other Medical', products)
# getProductList("https://tech.missouri.edu/UREF3F6CA2C030477C46A79DC10981B047B5/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_2_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.03600912779575505",'Medical -> Other Medical', products)
# getProductList("https://tech.missouri.edu/UREF3F6CA2C030477C46A79DC10981B047B5/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_3_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.03600912779575505",'Medical -> Other Medical', products)
# getProductList("https://tech.missouri.edu/UREF3F6CA2C030477C46A79DC10981B047B5/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_4_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.03600912779575505",'Medical -> Other Medical', products)

getProductList("https://tech.missouri.edu/UREF11E2532CBB70714DAEFF9597B0B76E17/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_1_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.8112260315147568",'Medical -> Therapeutics', products)
getProductList("https://tech.missouri.edu/UREF11E2532CBB70714DAEFF9597B0B76E17/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_2_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.8112260315147568",'Medical -> Therapeutics', products)
getProductList("https://tech.missouri.edu/UREF11E2532CBB70714DAEFF9597B0B76E17/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_3_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.8112260315147568",'Medical -> Therapeutics', products)
getProductList("https://tech.missouri.edu/UREF11E2532CBB70714DAEFF9597B0B76E17/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_4_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.8112260315147568",'Medical -> Therapeutics', products)
getProductList("https://tech.missouri.edu/UREF11E2532CBB70714DAEFF9597B0B76E17/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_5_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.8112260315147568",'Medical -> Therapeutics', products)
getProductList("https://tech.missouri.edu/UREF11E2532CBB70714DAEFF9597B0B76E17/TF_Home/V14EXEC/B_Nothing-B_SetCVar_BackButton_True-B_SetProp_Output%3AData%3APage_6_K2AB6683113640718159-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.8112260315147568",'Medical -> Therapeutics', products)

getProductList("https://tech.missouri.edu/UREF11E2532CBB70714DAEFF9597B0B76E17/TF_Home/V14EXEC/B_Nothing-B_SetCVar_form~searchtype_Taxonomy-B_SetCVar_Form~Taxonomy~Item_TSAA6716150635897430-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.226567397343721",'Software -> Engineering Software', products)

getProductList("https://tech.missouri.edu/UREF11E2532CBB70714DAEFF9597B0B76E17/TF_Home/V14EXEC/B_Nothing-B_SetCVar_form~searchtype_Taxonomy-B_SetCVar_Form~Taxonomy~Item_TSAA6716151652515591-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.7361749796184973",'Software -> Life Sciences Software', products)

getProductList("https://tech.missouri.edu/UREF11E2532CBB70714DAEFF9597B0B76E17/TF_Home/V14EXEC/B_Nothing-B_SetCVar_form~searchtype_Taxonomy-B_SetCVar_Form~Taxonomy~Item_TSAA6716151701524271-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.2160941807629062",'Software -> Medical Software', products)

getProductList("https://tech.missouri.edu/UREF11E2532CBB70714DAEFF9597B0B76E17/TF_Home/V14EXEC/B_Nothing-B_SetCVar_form~searchtype_Taxonomy-B_SetCVar_Form~Taxonomy~Item_TSAA6716151709532330-B_SetProp_Visibility_Visible_K2AB6715124953811689_%3A%3A%3A%3A/DIVTAG/K2AB6715124953811689/NoLoadResultDiv/0.07656843622269127",'Software -> Other Software', products)



# urllib_download("https://www.severnbiotech.com/images/DMF 2.5L.JPG", "test.jpg")

getProductInfo(products)

headers=['link', 'Title', 'type', 'Contact person','Contact email','Patent Status','Inventors']

for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
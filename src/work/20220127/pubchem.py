from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36"}

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		if retryCount < 5:
			getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', str(info[head]))
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, cid, cas, products):
	print(str(len(products))+ cas + url)
	
	productListHtml = getHtmlFromUrl(url)
	tempPinfo = {
		"cas":cas,
		"ComputedPropertiesLink": "https://pubchem.ncbi.nlm.nih.gov/compound/"+str(cid)+"#section=Computed-Properties&fullscreen=true",
	}
	data = json.loads(productListHtml)
	
	Section = data["Record"]["Section"]
	try:
		TOCHeading1 = next(filter(lambda i: i["TOCHeading"]=="Safety and Hazards", Section))
		TOCHeading1Section = TOCHeading1["Section"]
		TOCHeading2 = next(filter(lambda i : i["TOCHeading"]=="Hazards Identification", TOCHeading1Section))
		TOCHeading2Section = TOCHeading2["Section"]
		
		GHSClassification = next(filter(lambda i: i["TOCHeading"]=="GHS Classification", TOCHeading2Section))
		GHSClassificationSection = GHSClassification["Information"]
		Pictogram = next(filter(lambda i: i["Name"]=="Pictogram(s)", GHSClassificationSection))
		Signal = next(filter(lambda i: i["Name"]=="Signal", GHSClassificationSection))
		GHSHazardStatements = next(filter(lambda i: i["Name"]=="GHS Hazard Statements", GHSClassificationSection))
		PrecautionaryStatementCodes = next(filter(lambda i: i["Name"]=="Precautionary Statement Codes", GHSClassificationSection))

		tempPinfo["Pictogram"] = Pictogram["Value"]["StringWithMarkup"][0]["Markup"][0]["URL"]
		tempPinfo["Signal"] = Signal["Value"]["StringWithMarkup"][0]["String"]
		tempPinfo["GHSHazardStatements"] = GHSHazardStatements["Value"]["StringWithMarkup"][0]["String"]
		tempPinfo["PrecautionaryStatementCodes"] = PrecautionaryStatementCodes["Value"]["StringWithMarkup"][0]["String"]
	except:
		tempPinfo["Pictogram"] = ''
		tempPinfo["Signal"] = ''
		tempPinfo["GHSHazardStatements"] = ''
		tempPinfo["PrecautionaryStatementCodes"] = ''
	try:
		pumbUrl1 = "https://pubchem.ncbi.nlm.nih.gov/sdq/sdqagent.cgi?infmt=json&outfmt=json&query={%22select%22:%22*%22,%22collection%22:%22pubmed%22,%22where%22:{%22ands%22:[{%22cid%22:%22"+str(cid)+"%22},{%22pmidsrcs%22:%22xref%22}]},%22order%22:[%22articlepubdate,desc%22],%22start%22:1,%22limit%22:5,%22width%22:1000000,%22listids%22:0}"	
		dataStr1 = getHtmlFromUrl(pumbUrl1)
		
		
		pubData1 = json.loads(dataStr1)
		rows1 = pubData1["SDQOutputSet"][0]["rows"]
		dataRows1 = list(map(lambda x: {"pmid": x["pmid"], "Publication Date":x["articlepubdate"], "Title":x["articletitle"], "Journal": x["articlejourname"]}, rows1))
		tempPinfo["Literature"]=str(dataRows1)
	except:
		tempPinfo["Literature"]=''
	try:
		pumbUrl = "https://pubchem.ncbi.nlm.nih.gov/sdq/sdqagent.cgi?infmt=json&outfmt=json&query={%22select%22:%22*%22,%22collection%22:%22patent%22,%22where%22:{%22ands%22:[{%22cid%22:%22"+str(cid)+"%22}]},%22order%22:[%22prioritydate,desc%22],%22start%22:1,%22limit%22:5,%22width%22:1000000,%22listids%22:0}"	
		dataStr = getHtmlFromUrl(pumbUrl)
		pubData = json.loads(dataStr)
		rows = pubData["SDQOutputSet"][0]["rows"]
		dataRows = list(map(lambda x: {"Publication Number":x["publicationnumber"], "Title":x["title"], "Priority Date":x["prioritydate"]}, rows))
		tempPinfo["Patents"]=str(dataRows)
	except:
		tempPinfo["Patents"]=''
	products.append(tempPinfo.copy())

def getProductList(cas, products):
	productListHtml = getHtmlFromUrl("https://pubchem.ncbi.nlm.nih.gov/rest/pug_view/data/compound/"+cas+"/JSON/?heading=Computed+Properties")
	if productListHtml!=None:
		data = json.loads(productListHtml)
		try:
			pInfo = {"cas":cas}
			Section = data["Record"]["Section"]
			TOCHeading1 = next(filter(lambda i: i["TOCHeading"]=="Chemical and Physical Properties", Section))
			TOCHeading1Section = TOCHeading1["Section"]
			TOCHeading2 = next(filter(lambda i : i["TOCHeading"]=="Computed Properties", TOCHeading1Section))
			TOCHeading2Section = TOCHeading2["Section"]
			for sec in TOCHeading2Section:
				value = sec["Information"][0]["Value"]
				if "StringWithMarkup" in value:
					pInfo[sec["TOCHeading"]] = value["StringWithMarkup"][0]["String"]
				if "Number" in value:
					pInfo[sec["TOCHeading"]] = value["Number"][0]
				
			products.append(pInfo.copy())
			print(str(len(products)))
		except:
			products.append({"cas":cas})
	else:
		products.append({"cas":cas})

products = []

excelFileName="pubchem.xlsx"
wb = Workbook()
workSheet = wb.active
fileName="cid.txt"
with open(fileName,'r') as file_to_read:
	index = 1
	type=1
	while True:
		lines = file_to_read.readline()
		if not lines:
				break
		getProductList(lines.replace("\n",""), products)

headers=['cas',
 'Molecular Weight', 
  'XLogP3',
 'Hydrogen Bond Donor Count', 
 'Hydrogen Bond Acceptor Count', 
 'Rotatable Bond Count', 
 'Exact Mass', 
 'Monoisotopic Mass',
 'Topological Polar Surface Area', 
 'Heavy Atom Count', 
 'Formal Charge', 
 'Complexity', 
 'Isotope Atom Count', 
 'Defined Atom Stereocenter Count', 
 'Undefined Atom Stereocenter Count',
 'Defined Bond Stereocenter Count', 
 'Undefined Bond Stereocenter Count', 
'Covalently-Bonded Unit Count', 
'Compound Is Canonicalized'
]

for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
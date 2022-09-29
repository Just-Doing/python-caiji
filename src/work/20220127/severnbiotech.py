from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')
	

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry index"+str(retryCount)+url)
		retryCount += 1
		if retryCount< 5:
			getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url, isScreenShotName=""):
	print(url)
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	if len(isScreenShotName) > 0:
		imgEle = browser.find_element_by_xpath('//body/img[1]')
		if imgEle !=None:
			imgEle.screenshot(isScreenShotName)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, products):
	print(str(len(products)) + url)
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	pInfo={
		"link": url,
		"nav": getNodeText(sope.find("div", attrs={"id":"navBreadCrumb"})).replace("\xa0","").replace("\n","")
	}
	imageArea = sope.find("div", attrs={"id":"productMainImage"})
	pName = sope.find("h1", attrs={"id":"productName"})
	image = imageArea.find("img")
	pInfo["img"] = image["src"].replace("images/", "").replace("/","").replace("\\","").replace("*","").replace("%","").replace("$","").replace("#","").replace("(","").replace(")","")
	if pInfo["img"].find("no_picture.gif") < 0:
		getRenderdHtmlFromUrl("http://www.severnbiotech.com/"+image["src"], pInfo["img"])
	pInfo["pName"] = getNodeText(pName)
	specs = sope.find_all("h4", attrs={"class":"optionName back"})
	for spec in specs:
		title = getNodeText(spec)
		if title=="Pack Size":
			pInfo["Pack Size"] = getNodeText(spec.nextSibling.nextSibling)
	desc = sope.find("div", attrs={"id":"productDescription"})
	decChild = desc.find("p", attrs={"class":"MsoNormal"})
	if decChild!=None:
		desc = decChild
	# pInfo["Description"] = str(desc.prettify()).replace("<br/>","").replace("<b>","").replace("</b>","").replace("</div>","").replace('<div class="productGeneral biggerText" id="productDescription">',"")
	# pInfo["Description"] = pInfo["Description"].replace("<p>","").replace("</p>","").replace("<strong>","").replace("</strong>","")
	# pInfo["Description"] = pInfo["Description"].replace('<p class="MsoNormal" style="margin: 0cm 0cm 0.0001pt; font-family: Calibri, sans-serif;">',"")
	# pInfo["Description"] = pInfo["Description"].replace('<span style="font-size: small;">',"").replace('<font size="2">',"").replace('</font>',"").replace('<o:p>',"").replace('</o:p>',"")
	# pInfo["Description"] = pInfo["Description"].replace('<font face="Swiss721BT-Roman" size="1">',"")
	# pInfo["Description"] = pInfo["Description"].replace('<p class="MsoNormal">',"")
	pInfo["Description"] = getNodeText(desc)
	pdfs = sope.find_all("a")
	for pdf in pdfs:
		img = pdf.find("img")
		if img!= None:
			imgSrc = img["src"]
			if imgSrc =="includes/templates/severn/images/ms1.gif":
				src = pdf["href"]
				if src.find(".pdf")>-1 or src.find(".PDF")>-1:
					pInfo["pdf"]=pInfo["pName"].replace("/","").replace("\\","").replace("*","").replace("%","").replace("$","").replace("#","").replace(".","").replace("(","").replace(")","").replace("{","").replace("}","")+".pdf"
					urllib_download(src, pInfo["pdf"])

	products.append(pInfo.copy())
def getProductList(url, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	pros = sope.find_all("div", attrs={"class":"innerColumn"})
	for pro in pros:
		link = pro.find("a")
		getProductInfo(link["href"], products)

excelFileName="severnbiotech.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("http://www.severnbiotech.com/index.php?main_page=product_info&cPath=59&products_id=237", products)
# getProductList("http://www.severnbiotech.com/index.php?main_page=products_all&disp_order=1&page=14", products)
for pageIndex in range(1, 15):
	getProductList("http://www.severnbiotech.com/index.php?main_page=products_all&disp_order=1&page="+str(pageIndex)+"&zenid=dl4n3q2ofqum2m06s027h48hd7", products)

# urllib_download("https://www.severnbiotech.com/images/DMF 2.5L.JPG", "test.jpg")



headers=['link','nav','pName','Pack Size','Description','img','pdf']

for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
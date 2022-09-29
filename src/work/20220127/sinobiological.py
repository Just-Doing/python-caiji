from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import time
import cfscrape


http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	headers = {
		"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.82 Safari/537.36",
		"cookie":"_ga=GA1.2.1846208171.1605273760; href=https%3A%2F%2Fwww.sinobiological.com%2Fresearch%2Ftargeted-therapy; accessId=5aff5fb0-84db-11e8-a3b3-d368cce40a8e; _gcl_au=1.1.1660157260.1645016298; Hm_lvt_2d911fad88dfe99ff8bbf610824f3e67=1645016298; sbb=%252be43ohTbVTr09K%252bxQlr1%252bK0onQvF%252bMIXgZM%252bveGXMHU%253dXTaJopSyq01ihC4cD5jOfAa8AEgsWX792EAllASK%252bEKohz0p3JxSEJFavoBnvUmw4fhBpwMcWGZ2Qy%252bRRF2U%252bMSxcQdMfdwOcT%252bR%252bo7qyEU%252br8SBQuGE8GJWgDFeSTZ4QS0HvJFVazETAoyuKMwGHYRoD68%252f7qno5Bg%252bEH9sSXM4upMLtz%252f4IdNkjX6GD0JYHbiUh%252blGTwi25Iz3IKocTDD58DE1yYiY3DxeifN7Qz6OxtXX21lrBpnvgDu9ANN%252f7TTxWWMmOIjxVG772o%252bYGkE9AMxcU5O4cIrT9cubm6dAdgw6n%252fQRZpTVxNv2TGHdHZblPNcfu4dTWVsL3aqaag%253d%253d; _gid=GA1.2.832211649.1645016298; _ce.s=v11.rlc~1645016301520; pageViewNum=13; Hm_lpvt_2d911fad88dfe99ff8bbf610824f3e67=1645017042; Currency=RMB; LocationCode=CN"
	}

	
	scraper = cfscrape.create_scraper()
	html_code = scraper.get(url,headers=headers).text
	return html_code
	
	
def getRenderdHtmlFromUrl(url):

	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("cookie=_ga=GA1.2.1846208171.1605273760; __cf_bm=izqCRiCxXxdO2k6bC5_oXBf5ua2By93wN57WYZ02Ix4-1645016302-0-AZD2hIGrFZttPsJO99Tyxx/ERKy8EoGfASkvywGfTqzGX/N0aGX5QblxV7G711Hh4UNUvVWKdXZhdjwzPSp3gCo=; qimo_seosource_5aff5fb0-84db-11e8-a3b3-d368cce40a8e=%E7%AB%99%E5%86%85; qimo_seokeywords_5aff5fb0-84db-11e8-a3b3-d368cce40a8e=; href=https%3A%2F%2Fwww.sinobiological.com%2Fresearch%2Ftargeted-therapy; qimo_xstKeywords_5aff5fb0-84db-11e8-a3b3-d368cce40a8e=; accessId=5aff5fb0-84db-11e8-a3b3-d368cce40a8e; _gcl_au=1.1.1660157260.1645016298; Hm_lvt_2d911fad88dfe99ff8bbf610824f3e67=1645016298; sbb=%252be43ohTbVTr09K%252bxQlr1%252bK0onQvF%252bMIXgZM%252bveGXMHU%253dXTaJopSyq01ihC4cD5jOfAa8AEgsWX792EAllASK%252bEKohz0p3JxSEJFavoBnvUmw4fhBpwMcWGZ2Qy%252bRRF2U%252bMSxcQdMfdwOcT%252bR%252bo7qyEU%252br8SBQuGE8GJWgDFeSTZ4QS0HvJFVazETAoyuKMwGHYRoD68%252f7qno5Bg%252bEH9sSXM4upMLtz%252f4IdNkjX6GD0JYHbiUh%252blGTwi25Iz3IKocTDD58DE1yYiY3DxeifN7Qz6OxtXX21lrBpnvgDu9ANN%252f7TTxWWMmOIjxVG772o%252bYGkE9AMxcU5O4cIrT9cubm6dAdgw6n%252fQRZpTVxNv2TGHdHZblPNcfu4dTWVsL3aqaag%253d%253d; _gid=GA1.2.832211649.1645016298; _ce.s=v11.rlc~1645016301520; pageViewNum=12; Hm_lpvt_2d911fad88dfe99ff8bbf610824f3e67=1645016885; Currency=RMB; LocationCode=CN")
	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	return BeautifulSoup(browser.page_source, "html.parser",from_encoding="utf-8")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, products):
	print(str(len(products)) + url)
	htmlStr = getHtmlFromUrl(url)
	sope = BeautifulSoup(htmlStr, "html.parser",from_encoding="utf-8")
	scripts = sope.find_all("script")
	for script in scripts:
		scriptStr = getNodeText(script)
		if(scriptStr.find("var proteinData =")> -1):
			scriptStr = scriptStr.replace("var proteinData =", "").replace(";","").replace("'","\"")
			scriptStr = scriptStr.replace("Species","\"Species\"").replace("ProductUrl","\"ProductUrl\"").replace("ProductName","\"ProductName\"")
			scriptStr = scriptStr.replace("Catalog","\"Catalog\"").replace("ExpressionHost","\"ExpressionHost\"").replace("Image:","\"Image\":")
			scriptStr = scriptStr.replace("Burkitt\"s","Burkitt's").replace("\r\n","").replace("},            ]","}            ]")
			# print(scriptStr)
			datas = json.loads(scriptStr)
			for data in datas:
				products.append({
					"Catalog": data["Catalog"],
					"ProductName": data["ProductName"]
				}.copy())

def getProductList(url, products):
	htmlStr = getHtmlFromUrl(url)
	sope = BeautifulSoup(htmlStr, "html.parser",from_encoding="utf-8")
	col_fc = sope.find("div", attrs={"class":"col_fc"})
	lis = col_fc.find_all("li")
	for li in lis:
		link = li.find("a")
		if link !=None:
			getProductInfo("https://www.sinobiological.com"+link["href"], products)


excelFileName="conjuprobe.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("https://www.sinobiological.com/category/cd40-ligand", products)

getProductList("https://www.sinobiological.com/research/targeted-therapy", products)

headers=[
	'Catalog','ProductName'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
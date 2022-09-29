from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import cfscrape

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')
	

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	headers = {
		"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.82 Safari/537.36",
		"cookie":"_ga=GA1.2.1846208171.1605273760; href=https%3A%2F%2Fwww.sinobiological.com%2Fresearch%2Ftargeted-therapy; accessId=5aff5fb0-84db-11e8-a3b3-d368cce40a8e; _gcl_au=1.1.1660157260.1645016298; Hm_lvt_2d911fad88dfe99ff8bbf610824f3e67=1645016298; sbb=%252be43ohTbVTr09K%252bxQlr1%252bK0onQvF%252bMIXgZM%252bveGXMHU%253dXTaJopSyq01ihC4cD5jOfAa8AEgsWX792EAllASK%252bEKohz0p3JxSEJFavoBnvUmw4fhBpwMcWGZ2Qy%252bRRF2U%252bMSxcQdMfdwOcT%252bR%252bo7qyEU%252br8SBQuGE8GJWgDFeSTZ4QS0HvJFVazETAoyuKMwGHYRoD68%252f7qno5Bg%252bEH9sSXM4upMLtz%252f4IdNkjX6GD0JYHbiUh%252blGTwi25Iz3IKocTDD58DE1yYiY3DxeifN7Qz6OxtXX21lrBpnvgDu9ANN%252f7TTxWWMmOIjxVG772o%252bYGkE9AMxcU5O4cIrT9cubm6dAdgw6n%252fQRZpTVxNv2TGHdHZblPNcfu4dTWVsL3aqaag%253d%253d; _gid=GA1.2.832211649.1645016298; _ce.s=v11.rlc~1645016301520; pageViewNum=13; Hm_lpvt_2d911fad88dfe99ff8bbf610824f3e67=1645017042; Currency=RMB; LocationCode=CN"
	}

	
	scraper = cfscrape.create_scraper()
	html_code = scraper.get(url,headers=headers).text
	return html_code
def getRenderdHtmlFromUrl(url, isScreenShotName=""):
	print(url)
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	if len(isScreenShotName) > 0:
		imgEle = browser.find_element_by_xpath('//body/img[1]')
		if imgEle !=None:
			imgEle.screenshot(isScreenShotName)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, cat, products):
	print(str(len(products)) + url)
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	keyWord = sope.find("ul", attrs={"class":"keywords"})
	researchers = sope.find("div", attrs={"class":"researchers-title"})
	contactName = sope.find("div", attrs={"class":"contact-name"})
	contactEmail = sope.find("div", attrs={"class":"contact-email"})
	title = sope.find("div", attrs={"class":"grid-x ip-title"})
	keyWordStr = ""
	if keyWord != None:
		lis = keyWord.find_all("li")
		for li in lis:
			keyWordStr+=getNodeText(li)+"\n"
	pInfo={
		"link": url,
		"cat": cat,
		"keyWord": keyWordStr,
		"title": getNodeText(title.find("h2")),
		"research": getNodeText(researchers.find("div", attrs={"class":"r-name-0"})),
		"Contact person": getNodeText(contactName),
		"Contact email": getNodeText(contactEmail)
	}

	products.append(pInfo.copy())
def getProductList(url, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	pros = sope.find_all("div", attrs={"class":"cell card-wrap"})
	for pro in pros:
		link = pro.find("a")
		cat = pro.find("p", attrs={"class":"industries"})
		getProductInfo(link["href"], getNodeText(cat), products)

excelFileName="invent.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("https://invent.psu.edu/ip_item/a-method-to-enhance-skin-immune-response-against-infection/",'', products)
# getProductList("https://invent.psu.edu/tools/intellectual-property-navigator/?fwp_industry_type=engineering-materials%2Cfood-agriculture%2Cit-edtech%2Cmedical-sciences&fwp_paged=1", products)
for pageIndex in range(1, 18):
	getProductList("https://invent.psu.edu/tools/intellectual-property-navigator/?fwp_industry_type=engineering-materials%2Cfood-agriculture%2Cit-edtech%2Cmedical-sciences&fwp_paged="+str(pageIndex), products)

# urllib_download("https://www.severnbiotech.com/images/DMF 2.5L.JPG", "test.jpg")



headers=['link','title','cat','keyWord','research','Contact person','Contact email']

for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
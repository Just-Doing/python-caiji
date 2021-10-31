from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy
import string
import ssl
ssl._create_default_https_context = ssl._create_unverified_context

http.client._MAXHEADERS = 1000


def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"}

		request_obj=urllib.request.Request(url=url,headers=headers)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getHtmlFromUrl(url)
		else:
			retryCount=0
			return None
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1


def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	productHtml = getRenderdHtmlFromUrl(url)
	if productHtml != None:
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
		
		pInfo["link"] = url
		cat = sope.find(name="span", attrs={"class": "catalog"})
		price = sope.find(name="div", attrs={"class": "price price_now"})
		size = sope.find(name="div", attrs={"class": "unit active"})
		decArea = sope.find(name="div", attrs={"class": "panel-body-inner pd-0 clearfix"})
		pInfo["cat"] = getNodeText(cat)
		pInfo["price"] = getNodeText(price)
		pInfo["size"] = getNodeText(size)
		pInfo["RelatedPathways"] = ""
		if decArea !=None:
			attrAreas = decArea.find_all(name="div", attrs={"class":"col-md-12 product_details"})
			for attrArea in attrAreas:
				attrLabelArea = attrArea.find(name="div", attrs={"class":"col-md-3"})
				attrValArea = attrArea.find(name="div", attrs={"class":"col-md-9"})
				attrLabel = getNodeText( attrLabelArea)
				attrVal = getNodeText( attrValArea)
				pInfo[attrLabel] = attrVal
			backInfoArea = sope.find_all(name="div", attrs={"class":"product_details_wrap"})
			for info in backInfoArea:
				title = getNodeText(info.find(name="div", attrs={"class":"title"}))
				if title =="Full Name":
					pInfo["fullname"] = getNodeText(info.find(name="div", attrs={"class":"cnt"}))
			relateArea = sope.find(name="div", attrs={"id":"relatedPathways"})
			if relateArea != None:
				relatePro = relateArea.find_all("li")
				for li in relatePro:
					pInfo["RelatedPathways"] = pInfo["RelatedPathways"]  + getNodeText(li)+","
		
		products.append(pInfo.copy())
				

def getProductList(url, type1 ,type2 ,type3 , products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	productListArea = sope.find("div", attrs={"class":"tab-content"})
	if productListArea != None:
		links = productListArea.find_all("a")
		for link in links:
			pInfo={
				"name":getNodeText(link),
				"nav":type1+" => "+type2+" => "+type3
			}
			getProductInfo("https://www.sinobiological.com"+link["href"], pInfo, products)



def getProductType2(url, type, products):
	type2ListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(type2ListHtml, "html.parser",from_encoding="utf-8")
	childType = sope.find(name="div", attrs={"id":"popup1" })
	type2list = childType.previous_sibling.previous_sibling.previous_sibling.previous_sibling
	typeImages = type2list.find_all("img");
	imgIndex = 1
	provType = ''
	for img in typeImages:
		currentType = re.findall(r'[^\\/:*?"<>|\r\n]+$',img["src"])[0].replace('.png', '')
		if currentType != provType:
			provType = currentType
			typeArea = sope.find(name="div",attrs={"id":"popup"+str(imgIndex)})
			if typeArea !=None:
				type3s = typeArea.find_all("li")
				for type3 in type3s:
					link = type3.find("a")
					type3str = getNodeText(link)
					if link!= None and len(type3str) > 0:
						getProductList("https://www.sinobiological.com"+link['href'], type, currentType, type3str, products)
			imgIndex = imgIndex+1
			
excelFileName="sinobiological7.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
# getProductInfo('https://www.sinobiological.com/cdna-clone/mouse-14-3-3-beta-mg50861-g', {}, products)

getProductType2("https://www.sinobiological.com/pathways/cytokine-receptor-g-chain-pathway", "Common Cytokine Receptor Gamma-Chain Signaling Pathways", products)

headers=['link','nav','name','cat','price','size','Species','NCBI Ref Seq','RefSeq ORF Size','Sequence Description','Description','Promoter','Vector',
'Restriction Sites','Tag Sequence','Sequencing Primers',
'Quality Control','Antibiotic in E.coli','Antibiotic in Mammalian cell','Application','Shipping','Storage']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
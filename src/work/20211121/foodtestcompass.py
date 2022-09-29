from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):

	url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
	request_obj=urllib.request.Request(url=url)
	response_obj=urllib.request.urlopen(request_obj)
	html_code=response_obj.read()
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	
def getRenderdHtmlFromUrl(url):

	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	return BeautifulSoup(browser.page_source, "html.parser",from_encoding="utf-8")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, type1, type2, products):
	print(str(len(products)) + url)
	sope = getHtmlFromUrl(url)
	produceName = sope.find("h2", attrs={"class":"product_title entry-title"})
	
	pInfo = {
		"link": url,
		"type1": type1,
		"type2": type2,
		"Product Name": getNodeText(produceName)
	}
	specArea = sope.find("div", attrs={"class":"tabbable"})
	specs = specArea.find_all("div", attrs={"class":"row"})
	for spec in specs:
		title = getNodeText(spec.find("div", attrs={"class":"col-sm-4"}))
		value = getNodeText(spec.find("div", attrs={"class":"col-sm-8"}))
		pInfo[title] = value
	products.append(pInfo.copy())

def getProductList(url, type1, type2, products):
	sope = getHtmlFromUrl(url)
	pList = sope.find_all("div", attrs={"class":"product-list-item prod-layout-classic"})
	for p in pList:
		pLink = p.find("a")
		getProductInfo(pLink["href"], type1, type2, products)


excelFileName="foodtestcompass.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("https://foodtestcompass.com/alertox-sticks-egg",'','', products)

getProductList("https://foodtestcompass.com/product/egg",'Food allergens','Egg', products)
getProductList("https://foodtestcompass.com/product/soy",'Food allergens','Soy', products)
getProductList("https://foodtestcompass.com/product/beta_lactoglobulin",'Food allergens','Beta-lactoglobulin', products)
getProductList("https://foodtestcompass.com/product/casein",'Food allergens','Casein', products)
getProductList("https://foodtestcompass.com/product/milk",'Food allergens','Milk', products)
getProductList("https://foodtestcompass.com/product/peanut",'Food allergens','Peanut', products)
getProductList("https://foodtestcompass.com/product/fish",'Food allergens','Fish', products)
getProductList("https://foodtestcompass.com/product/crustacea",'Food allergens','Crustacea', products)
getProductList("https://foodtestcompass.com/product/shellfish",'Food allergens','Shellfish', products)
getProductList("https://foodtestcompass.com/product/mollusc",'Food allergens','Mollusc', products)
getProductList("https://foodtestcompass.com/product/hazelnut",'Food allergens','Hazelnut', products)
getProductList("https://foodtestcompass.com/product/almond",'Food allergens','Almond', products)
getProductList("https://foodtestcompass.com/product/mustard",'Food allergens','Mustard', products)
getProductList("https://foodtestcompass.com/product/lysozyme",'Food allergens','Lysozyme', products)
getProductList("https://foodtestcompass.com/product/cashew",'Food allergens','Cashew', products)
getProductList("https://foodtestcompass.com/product/pistachio",'Food allergens','Pistachio', products)
getProductList("https://foodtestcompass.com/product/walnut",'Food allergens','Walnut', products)
getProductList("https://foodtestcompass.com/product/sesame",'Food allergens','Sesame', products)
getProductList("https://foodtestcompass.com/product/lupin",'Food allergens','Lupin', products)
getProductList("https://foodtestcompass.com/product/coconut",'Food allergens','Coconut', products)
getProductList("https://foodtestcompass.com/product/brazil_nut",'Food allergens','Brazil nut', products)
getProductList("https://foodtestcompass.com/product/tree_nuts",'Food allergens','Tree nuts', products)

getProductList("https://foodtestcompass.com/product/gluten",'Gluten','Gluten', products)
getProductList("https://foodtestcompass.com/product/total_aflatoxins",'Mycotoxins','Total aflatoxins', products)
getProductList("https://foodtestcompass.com/product/aflatoxin_B1",'Mycotoxins','Aflatoxin B1', products)
getProductList("https://foodtestcompass.com/product/Deoxynivalenol_(DON)",'Mycotoxins','Deoxynivalenol (DON)', products)

headers=[
	'link','type1','type2','Product Name','Applicable to liquids','Applicable to solid fluid matr','Applicable to surface',
	'Assay time','ASSAY TYPE','CERT','Disposable supplied','Dosing range','Limit of detection lod a','MANUFACT',
	'Number of determinations','Number of standards rm for cal','Principle of the assay','Product code','Product line',
	'Sample preparation solid','Sample prep liquid','Sample prep surface','Sample size',
	'Shelf life','Storage','Target','Tech','Test format',
	'Validation Report Available'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
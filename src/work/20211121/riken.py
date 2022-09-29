from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):

	url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
	request_obj=urllib.request.Request(url=url)
	response_obj=urllib.request.urlopen(request_obj)
	html_code=response_obj.read()
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	
def getRenderdHtmlFromUrl(url):

	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	return BeautifulSoup(browser.page_source, "html.parser",from_encoding="utf-8")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, type, products):
	print(str(len(products)) + url)
	sope = getRenderdHtmlFromUrl(url)
	infoTab = sope.find("table", attrs={"border":"1"})
	cellNo = infoTab.find("td", attrs={"class":"tdcolor-head"})
	cellNoPart = getNodeText(cellNo).split("\n")[1].split(":")
	pInfo = {
		"link": url,
		"type": type,
		"Cell No":cellNoPart[0],
		"Cell Name":cellNoPart[1],
	}
	trs = infoTab.find_all("tr")
	for tr in trs[1:]:
		tds = tr.find_all("td")
		if len(tds)>0:
			value = getNodeText(tds[-1]).replace("\n","").replace("\t","").replace("\r","").replace("\xa0"," ")
			title = getNodeText(tr).replace("\n","").replace("\t","").replace("\r","").replace("\xa0"," ").replace(value, "")
			pInfo[title] = value
			if len(tds) == 1:
				if "Order Form" in pInfo:
					pInfo["Order Form"] += value
	products.append(pInfo.copy())

def getProductList(url, type1, products):
	sope = getHtmlFromUrl(url)
	pList = sope.find_all("td", attrs={"rowspan":"4"})
	for p in pList:
		pLink = p.find("a")
		getProductInfo("https://cellbank.brc.riken.jp/"+pLink["href"], type1, products)


excelFileName="riken.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("https://cellbank.brc.riken.jp/cell_bank/CellInfo/?cellNo=RCB0009&lang=En",'', products)

# getProductList("https://cellbank.brc.riken.jp/cell_bank/CellList/?lang=En&cellCategory=RCB",'General cell lines (RCB)', products)
# getProductList("https://cellbank.brc.riken.jp/cell_bank/CellList/?lang=En&cellCategory=HPS",'Human iPS cells (HPS)', products)
# getProductList("https://cellbank.brc.riken.jp/cell_bank/CellList/?lang=En&cellCategory=HES",'Human ES cells (HES)', products)
getProductList("https://cellbank.brc.riken.jp/cell_bank/CellList/?lang=En&cellCategory=APS",'Animal iPS cells (APS)', products)
# getProductList("https://cellbank.brc.riken.jp/cell_bank/CellList/?lang=En&cellCategory=AES",'Animal ES cell lines and germline stem cell lines (AES)', products)
# getProductList("https://mus.brc.riken.jp/en/wp-content/mouse_es/index.html",'The Gene-trap Mouse ES cell clones (AES1001-AES2770)', products)
# getProductList("https://cell.brc.riken.jp/en/hcb",'Cord blood stem cells (HCB, CBF, CFD, C34)', products)
# getProductList("https://cellbank.brc.riken.jp/cell_bank/CellList/?lang=En&cellCategory=HEV",'EBV transformed B cell lines derived from Japanese (HEV)', products)
# getProductList("https://cellbank.brc.riken.jp/cell_bank/CellList/?lang=En&cellCategory=HSC",'Sonoda-Tajima Collection (HSC)', products)
# getProductList("https://cellbank.brc.riken.jp/cell_bank/CellList/?lang=En&cellCategory=GMC",'Goto Collection : Werner syndrome patients (GMC)', products)


headers=[
	'link','type','Cell No','Cell Name','Comment','Comment from the depositor','Terms and conditions','Remarks',
	'Order Form','Basic informationDepositor','Originator','Year of deposit','Animal  ','Genus','Species','Tissue',
	'Classification','History','Lifespan','Morphology','Cellosaurus(Expasy)','deposit info',
	'Medium','Culture type','Medium and additives','Antibiotics','Passage method',
	'Culture informationPassage ratio','SC frequency','Temperature','CO2 concentration','Freeze medium','Freezing method',
	'Mycoplasma','Animal PCR','Imagesdeposit info',
	'Reference informationReference',"User's Publication"
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
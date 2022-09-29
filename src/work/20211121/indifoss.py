from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):

	url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
	request_obj=urllib.request.Request(url=url)
	response_obj=urllib.request.urlopen(request_obj)
	html_code=response_obj.read()
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	
def getRenderdHtmlFromUrl(url):

	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	return BeautifulSoup(browser.page_source, "html.parser",from_encoding="utf-8")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, type, products):
	print(str(len(products)) + url)
	sope = getHtmlFromUrl(url)
	dec = sope.find("div", attrs={"class":"row mt-3 responsive-mt product_description d-block"})
	Benefits = sope.find("div", attrs={"class":"row mt-3 responsive-mt product_description"})
	
	pInfo = {
		"link": url,
		"type": type,
		"Description": getNodeText(dec),
		"Benefits": getNodeText(Benefits)
	}
	tableArea = sope.find("div", attrs={"class":"table-responsive"})
	if tableArea!= None:
		specTables = tableArea.find_all("table")
		if len(specTables) > 0:
			specStr = ""
			specTrs = specTables[0].find_all("tr")
			for tr in specTrs:
				tds = tr.find_all("td")
				if len(tds) >= 2:
					specStr += getNodeText(tds[-2]) + ":" + getNodeText(tds[-1])+";"
		if len(specTables) > 1:
			orderTrs = specTables[1].find_all("tr")
			KitComponents = ""
			for orderTr in orderTrs:
				orderTds = orderTr.find_all("td")
				if len(orderTds) == 3:
					pInfo["Specification"] = getNodeText(orderTds[1])
				if len(orderTds)>0:
					KitComponents += getNodeText(orderTds[-1])+";"
			pInfo["Kit Components"] = KitComponents
	products.append(pInfo.copy())

def getProductList(url, type1, products):
	sope = getHtmlFromUrl(url)
	pList = sope.find_all("div", attrs={"class":"col-lg-4 col-md-4 col-sm-12 col-6 col-xs-12 product-card mb-4"})
	for p in pList:
		pLink = p.find("a")
		getProductInfo(pLink["href"], type1, products)


excelFileName="indifoss.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("https://www.indifoss.com/chemical-safety/milk-antibiotic-rapid-test/4in1-melamine-beta-lactams-cephalexin-tetracyclines",'', products)

getProductList("https://www.indifoss.com/chemical-safety/milk-antibiotic-rapid-test",'Milk Antibiotic Rapid Test', products)
getProductList("https://www.indifoss.com/chemical-safety/milk-antibiotic-rapid-test/9",'Milk Antibiotic Rapid Test', products)
getProductList("https://www.indifoss.com/chemical-safety/milk-antibiotic-rapid-test/18",'Milk Antibiotic Rapid Test', products)

getProductList("https://www.indifoss.com/chemical-safety/milk-pesticides-screening",'Milk Pesticides Screening', products)

getProductList("https://www.indifoss.com/chemical-safety/pesticides-rapid-test",'Pesticides Rapid Test', products)
getProductList("https://www.indifoss.com/chemical-safety/pesticides-rapid-test/9",'Pesticides Rapid Test', products)

getProductList("https://www.indifoss.com/chemical-safety/cereals-feed-rapid-test",'Cereals & Feed Rapid Test', products)

getProductList("https://www.indifoss.com/chemical-safety/honey-antibiotic-rapid-test",'Honey Antibiotic Rapid Test', products)

getProductList("https://www.indifoss.com/chemical-safety/meatseafood-rapid-test",'Meat & Seafood Rapid Test', products)
getProductList("https://www.indifoss.com/chemical-safety/meatseafood-rapid-test/9",'Meat & Seafood Rapid Test', products)

getProductList("https://www.indifoss.com/chemical-safety/elisa-test-kits-for-seafood",'ELISA Test Kits for Seafood', products)


headers=[
	'link','type','Description','Benefits','Specification','Kit Components'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
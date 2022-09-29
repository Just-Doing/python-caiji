from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):

	url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
	request_obj=urllib.request.Request(url=url)
	response_obj=urllib.request.urlopen(request_obj)
	html_code=response_obj.read()
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	
def getRenderdHtmlFromUrl(url):

	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	return BeautifulSoup(browser.page_source, "html.parser",from_encoding="utf-8")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, type1,type2, products):
	print(str(len(products)) + url)
	sope = getRenderdHtmlFromUrl(url)
	
	pInfo = {
		"link": url,
		"type1": type1,
		"type2": type2,
		"Product Name": getNodeText(sope.find("h1", attrs={"itemprop":"name"})),
		"Package Size": getNodeText(sope.find("td", attrs={"data-label":"Package Size"})),
		"description": getNodeText(sope.find("div", attrs={"id":"tab-description"})),
		"CATEGORIES": getNodeText(sope.find("span", attrs={"class":"posted_in"})),
		"TAGS": getNodeText(sope.find("span", attrs={"class":"tagged_as"}))
	}
	specInfoStr = getNodeText(sope.find("div", attrs={"class":"wpb_text_column wpb_content_element"}))
	specInfos = specInfoStr.split("\n")
	for specInfo in specInfos:
		titleAndval = specInfo.split(":")
		if len(titleAndval) == 1:
			titleAndval = specInfo.split("\xa0– ")
		if len(titleAndval) == 2:
			title = titleAndval[0]
			val = titleAndval[1]
			pInfo[title] = val
	products.append(pInfo.copy())

def getProductList(url, type1, type2, products):
	print(url)
	sope = getRenderdHtmlFromUrl(url)
	pList = sope.find("table", attrs={"class":"wcpt-table wcpt-table-10009"})
	if pList == None:
		pList = sope.find("table", attrs={"class":"wcpt-table wcpt-table-10004"})
	if pList == None:
		pList = sope.find("table", attrs={"class":"wcpt-table wcpt-table-9226"})
	if pList == None:
		pList = sope.find("table", attrs={"class":"wcpt-table wcpt-table-9231"})
	if pList !=None:
		for tr in pList.find("tbody").find_all("tr"):
			pLink = tr.find("a")
			getProductInfo(pLink["href"], type1, type2, products)


excelFileName="conjuprobe.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("https://conju-probe.com/product/amino-bis-peg3-bcn-exo/",'','', products)

getProductList("https://conju-probe.com/product-category-click-chemistry-reagents-bcn-reagents/",'Copper-Free Click Chemistry','BCN linkers', products)
getProductList("https://conju-probe.com/product-category-click-chemistry-reagents-dbco-reagents/",'Copper-Free Click Chemistry','DBCO linkers', products)
getProductList("https://conju-probe.com/product-category-copper-free-click-chemistry/product-category-click-chemistry-reagents-peg-azide/",'Copper-Free Click Chemistry','PEG-azide', products)
getProductList("https://conju-probe.com/product-category-click-chemistry-reagents-peg-alkyne/",'Copper-Free Click Chemistry','PEG-alkyne', products)


headers=[
	'link','type1','type2','Product Name','Package Size','description','CATEGORIES','TAGS','Chemical Formula',
	'Molecular Weight','CAS','Purity','Physical Form','Solubility',
	'Storage at','Excellent biocompatibility','High chemoselectivity','Hydrophilic PEG3 Spacer',
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
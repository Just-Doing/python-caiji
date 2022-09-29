from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):

	url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
	request_obj=urllib.request.Request(url=url)
	response_obj=urllib.request.urlopen(request_obj)
	html_code=response_obj.read()
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	
def getRenderdHtmlFromUrl(url):

	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	return BeautifulSoup(browser.page_source, "html.parser",from_encoding="utf-8")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, products):
	print(str(len(products)) + url)
	sope = getHtmlFromUrl(url)
	
	pInfo = {
		"link": url
	}
	trs = sope.find_all("tr")
	for tr in trs:
		ths = tr.find_all("th")
		tds = tr.find_all("td")
		if len(ths) >0 and len(tds) > 0:
			for inx, th in enumerate(ths):
				title = getNodeText(th)
				value = getNodeText(tds[inx])
				pInfo[title] = value
	products.append(pInfo.copy())

def getProductList(url, products):
	sope = getHtmlFromUrl(url)
	table = sope.find("table")
	trs = table.find_all("tr")
	for tr in trs:
		pLinks = tr.find_all("a")
		for pLink in pLinks:
			if pLink != None and pLink.has_key("href"):
				href = pLink["href"]
				if href.find("search_res_det.cgi?ID=") > -1:
					getProductInfo("https://cellbank.nibiohn.go.jp/~cellbank/en/"+pLink["href"],  products)

excelFileName="nibiohn.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("https://cellbank.nibiohn.go.jp/~cellbank/en/search_res_det.cgi?ID=1990", products)

getProductList("https://cellbank.nibiohn.go.jp/~cellbank/en/search_res_list_combi.cgi", products)


headers=[
	'link','JCRB No.','Cell Name','Profile','Other Name','Animal'
	,'Strain','Genus','Species','Sex','Identity','Tissue for Primary Cancer'
	,'Case history','Metastasis','Tissue Metastasized','Genetics','Life Span','Crisis PDL','Morphology'
	,'Character','Classify','Established by','Registered by','Regulation for Distribution','Comment'
	,'Year','Medium','Methods for Passages','Cell Number on Passage','Race'
	,'CO2 Conc.','Tissue Sampling','Tissue Type','Detected DNA Virus','Detected RNA Virus','Pubmed id:3469020'
	,'Pubmed id:6572557','Pubmed id:6572557','Cell No.','LOT No.'
	,'Lot Specification','Temperature','Cell Density at Seeding','Doubling Time'
	,'Cell Number in Vial (cells/1ml)','Viability at cell freezing (%)','Antibiotics Used','Passage Number','PDL'
	,'Sterility: MYCOPLASMA','Sterility: BACTERIA','Sterility: FUNGI','Isozyme Analysis','Chromosome Mode'
	,'Chromosome Information','Surface Antigen','DNA Profile (STR)','Adhesion','Exoteric Gene'
	,'Medium for Freezing','RFLP','Viability immediately after thawing (%)','Additional information'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):

	url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
	request_obj=urllib.request.Request(url=url)
	response_obj=urllib.request.urlopen(request_obj)
	html_code=response_obj.read()
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	
def getRenderdHtmlFromUrl(url):

	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	return BeautifulSoup(browser.page_source, "html.parser",from_encoding="utf-8")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, type, products):
	print(str(len(products)) + url)
	sope = getHtmlFromUrl(url)
	pName = ""
	spec=""
	infoTb = sope.find("table")
	if infoTb != None:
		infoTr = infoTb.find_all("tr")[1]
		infoTds = infoTr.find_all("td")
		infoThs = infoTr.find_all("th")
		if len(infoTds) > 1:
			pName = getNodeText(infoTds[1])
			spec = getNodeText(infoTds[2])
		else :
			if len(infoThs) > 1:
				pName = getNodeText(infoThs[1])
				spec = getNodeText(infoThs[2])
	pInfo = {
		"link": url,
		"type": type,
		"Product Name": pName,
		"specification": spec,
	}
	h3s = sope.find_all("div", attrs={"style":"margin-bottom: 15px; font-size: 20px; color: #ac3d12"})
	for h3 in h3s:
		title = getNodeText(h3)
		specTable = h3.find_next_sibling("table")
	
		KitInformationStr = ""
		CrossStr = ""
		if title == "Product Introduction":
			pInfo["Advantages"] = getNodeText(h3.find_next_sibling("div"))
		if title == "Kit Information" and specTable!=None:
			trs = specTable.find_all("tr")
			for tr in trs:
				tds = tr.find_all("td")
				if len(tds) == 2:
					KitInformationStr = getNodeText(tds[0]) + ":" + getNodeText(tds[1])
			pInfo["Kit Information"] = KitInformationStr
			
		if title == "Specificity" and specTable!=None:
			trs = specTable.find_all("tr")
			for tr in trs:
				tds = tr.find_all("td")
				if len(tds) == 2:
					CrossStr += getNodeText(tds[0]) + ":" + getNodeText(tds[1]) + ";"
			pInfo["Cross-Reactivity(%)"] = CrossStr

		if title == "Contents":
			pInfo["Contents"] = getNodeText(h3.find_next_sibling("div"))

	products.append(pInfo.copy())

def getProductList(url, type1, products):
	sope = getHtmlFromUrl(url)
	pList = sope.find_all("li", attrs={"class":"py-1"})
	for p in pList:
		pLink = p.find("a")
		getProductInfo("https://www.abc-residue.com/"+pLink["href"], type1, products)


excelFileName="residue.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("https://www.abc-residue.com/product_detail.php?cate_index=9&item=156&%20set_lang=en",'', products)

getProductList("https://www.abc-residue.com/product.php?cate_index=9&set_lang=en",'Food Safety Test Kit', products)
getProductList("https://www.abc-residue.com/product.php?cate_index=15&set_lang=en",'Animal Related Test Kit', products)
getProductList("https://www.abc-residue.com/product.php?cate_index=10&set_lang=en",'Pesticide Test kit', products)
getProductList("https://www.abc-residue.com/product.php?cate_index=12&set_lang=en",'Mycotoxin Test Kit', products)


headers=[
	'link','type','Product Name','specification','Advantages','Kit Information','Cross-Reactivity(%)','Contents'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
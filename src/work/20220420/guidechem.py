from time import time
from urllib.request import urlopen
import urllib
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import requests
from requests.cookies import RequestsCookieJar
import cfscrape
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	print('download:'+IMAGE_URL)
	
	opener = urllib.request.build_opener()
	opener.addheaders = [('User-agent', 'Mozilla/5.0')]
	urllib.request.install_opener(opener)
	urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	headers = {
		"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.82 Safari/537.36",
		"cookie":"_ga=GA1.2.1846208171.1605273760; href=https%3A%2F%2Fwww.sinobiological.com%2Fresearch%2Ftargeted-therapy; accessId=5aff5fb0-84db-11e8-a3b3-d368cce40a8e; _gcl_au=1.1.1660157260.1645016298; Hm_lvt_2d911fad88dfe99ff8bbf610824f3e67=1645016298; sbb=%252be43ohTbVTr09K%252bxQlr1%252bK0onQvF%252bMIXgZM%252bveGXMHU%253dXTaJopSyq01ihC4cD5jOfAa8AEgsWX792EAllASK%252bEKohz0p3JxSEJFavoBnvUmw4fhBpwMcWGZ2Qy%252bRRF2U%252bMSxcQdMfdwOcT%252bR%252bo7qyEU%252br8SBQuGE8GJWgDFeSTZ4QS0HvJFVazETAoyuKMwGHYRoD68%252f7qno5Bg%252bEH9sSXM4upMLtz%252f4IdNkjX6GD0JYHbiUh%252blGTwi25Iz3IKocTDD58DE1yYiY3DxeifN7Qz6OxtXX21lrBpnvgDu9ANN%252f7TTxWWMmOIjxVG772o%252bYGkE9AMxcU5O4cIrT9cubm6dAdgw6n%252fQRZpTVxNv2TGHdHZblPNcfu4dTWVsL3aqaag%253d%253d; _gid=GA1.2.832211649.1645016298; _ce.s=v11.rlc~1645016301520; pageViewNum=13; Hm_lpvt_2d911fad88dfe99ff8bbf610824f3e67=1645017042; Currency=RMB; LocationCode=CN"
	}

	scraper = cfscrape.create_scraper()
	html_code = scraper.get(url,headers=headers).text
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	
def getRenderdHtmlFromUrl(url, isScreenShotName=""):
	chrome_options = Options()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")
	chrome_options.add_argument('user-agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.130 Safari/537.36"')
	chrome_options.add_argument("cookie=__hstc=240517788.35783e1d438e8f99e34188727b050107.1648374652176.1648374652176.1648374652176.1; hubspotutk=35783e1d438e8f99e34188727b050107; __hssrc=1; _gid=GA1.2.122553597.1648374652; _gcl_au=1.1.15847883.1648374654; _fbp=fb.1.1648374655847.2011294960; __hssc=240517788.6.1648374652176; _ga=GA1.1.457442853.1648374652; _ga_SBEXK5LM3N=GS1.1.1648374653.1.1.1648376932.0")
	chrome_options.add_argument("--no-sandbox")

	browser = webdriver.Chrome(options=chrome_options)
	browser.get(url)
	if len(isScreenShotName) > 0:
		imgEle = browser.find_element_by_xpath('//body/img[1]')
		if imgEle !=None:
			imgEle.screenshot(isScreenShotName)
	return browser.page_source
	

def getRenderdHtmlFromUrl1(url):

	chrome_options = Options()
	# chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")
	chrome_options.add_argument('User-Agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.127 Safari/537.36')
	chrome_options.add_argument('Cookie=__utmz=196486330.1639217227.1.1.utmcsr=(direct)|utmccn=(direct)|utmcmd=(none); clientkey=1639217244433_2131; _clientkey_=1650553025945; JSESSIONID=aaa9w1a1Iu4eJ60N43qby; search_key=; __utma=196486330.305675233.1639217227.1650558235.1650634379.6; __utmc=196486330; visittimes=51; __utmt=1; __utmb=196486330.9.10.1650634379; __atuvc=21%7C16; __atuvs=6262ae8bff220521007; view=770; code=0')
	# chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(options=chrome_options)
	browser.get(url)
	time.sleep(2)
	browser.execute_script("window.scrollTo(0,200)")
	time.sleep(4)
	browser.execute_script("window.scrollTo(0,500)")
	time.sleep(2)
	browser.execute_script("window.scrollTo(0,1000)")
	time.sleep(4)
	browser.execute_script("window.scrollTo(0,1500)")
	time.sleep(2)
	browser.execute_script("window.scrollTo(0,document.body.scrollHeight)")
	return BeautifulSoup(browser.page_source, "html.parser",from_encoding="utf-8")
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, type, products):
	print(str(len(products)) + url)
	sope = getHtmlFromUrl(url)
	name = getNodeText(sope.find("div", attrs={"class":"eng_he"}))
	pInfo={
		"link": url,
		"type":type
		
	}
	dds = sope.find_all("dd")
	for dd in dds:
		span = dd.find("span")
		if span ==None:
			title = getNodeText(dd.find("em"))
			if title == "":
				title = getNodeText(dd.find("strong"))
			val = getNodeText(dd).replace(title, "")
			pInfo[title] = val
			if len(title)>0 and title not in headers:
				headers.append(title)
	
	trs = sope.find_all("tr")
	for tr in trs:
		tds = tr.find_all("td")
		if len(tds) == 2:
			title = getNodeText(tds[0])
			val = getNodeText(tds[1])
			pInfo[title] = val
			if len(title)>0 and title not in headers:
				headers.append(title)
		# if len(tds) ==1:
		# 	img = tds[0].find("img")
		# 	if img != None:
		# 		if "src" in img.attrs and img["src"].find("https:")>-1:
		# 			getRenderdHtmlFromUrl(img["src"], name+".jpg")
		# 			pInfo["image"] =  name+".jpg"
	desc = sope.find("td", attrs={"class":"TDAmarginR10"})
	pInfo["Description"] = getNodeText(desc)

	products.append(pInfo.copy())

def getProductList(url,type, products):
	sope = getHtmlFromUrl(url)
	
	pListAreas = sope.find_all("a", attrs={"class":"utr_lk_table"})
	for pLink in pListAreas:
		if pLink["href"]!="javascript:void(0);":
			getProductInfo("https://www.guidechem.com"+pLink["href"],type, products)


headers=[
	'link','type','Product Name','Description','image'
]
excelFileName="guidechem.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("https://www.guidechem.com/trade/diethylhexyl-butamido-triazone-id4748212.html",'',products )
# getProductList("https://www.guidechem.com/product/listc_catid-7312-p1.html",'Cosmetic Raw Materials', products)
# 101  41后没有详情
for pIndex in range(1,42):
	getProductList("https://www.guidechem.com/product/listc_catid-7312-p"+str(pIndex)+".html",'Cosmetic Raw Materials', products)
	#101 71后没有详情
for pIndex in range(1,71):
	getProductList("https://www.guidechem.com/product/listc_catid-7313-p"+str(pIndex)+".html",'Others', products)
#68   31后无详情
for pIndex in range(1,31):
	getProductList("https://www.guidechem.com/product/listc_catid-7315-p"+str(pIndex)+".html",'Daily Chemicals', products)

#20   11后无详情
for pIndex in range(1,11):
	getProductList("https://www.guidechem.com/product/listc_catid-7316-p"+str(pIndex)+".html",'Detergent Raw Materials', products)
#24  11后无详情
for pIndex in range(1,11):
	getProductList("https://www.guidechem.com/product/listc_catid-7317-p"+str(pIndex)+".html",'Hair Care Chemicals', products)
#20  16后无详情
for pIndex in range(1,16):
	getProductList("https://www.guidechem.com/product/listc_catid-7311-p"+str(pIndex)+".html",'Detergent', products)

#9 4后无详情
for pIndex in range(1,4):
	getProductList("https://www.guidechem.com/product/listc_catid-7314-p"+str(pIndex)+".html",'Air Fresheners', products)

#7 4后无详情
for pIndex in range(1,4):
	getProductList("https://www.guidechem.com/product/listc_catid-7318-p"+str(pIndex)+".html",'Oral Care Chemicals', products)

for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
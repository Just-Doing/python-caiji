from time import time
from urllib.request import urlopen
import urllib
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import requests
from requests.cookies import RequestsCookieJar
import cfscrape
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	print('download:'+IMAGE_URL)
	
	opener = urllib.request.build_opener()
	opener.addheaders = [('User-agent', 'Mozilla/5.0')]
	urllib.request.install_opener(opener)
	urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	headers = {
		"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.82 Safari/537.36",
		"cookie":"_ga=GA1.2.1846208171.1605273760; href=https%3A%2F%2Fwww.sinobiological.com%2Fresearch%2Ftargeted-therapy; accessId=5aff5fb0-84db-11e8-a3b3-d368cce40a8e; _gcl_au=1.1.1660157260.1645016298; Hm_lvt_2d911fad88dfe99ff8bbf610824f3e67=1645016298; sbb=%252be43ohTbVTr09K%252bxQlr1%252bK0onQvF%252bMIXgZM%252bveGXMHU%253dXTaJopSyq01ihC4cD5jOfAa8AEgsWX792EAllASK%252bEKohz0p3JxSEJFavoBnvUmw4fhBpwMcWGZ2Qy%252bRRF2U%252bMSxcQdMfdwOcT%252bR%252bo7qyEU%252br8SBQuGE8GJWgDFeSTZ4QS0HvJFVazETAoyuKMwGHYRoD68%252f7qno5Bg%252bEH9sSXM4upMLtz%252f4IdNkjX6GD0JYHbiUh%252blGTwi25Iz3IKocTDD58DE1yYiY3DxeifN7Qz6OxtXX21lrBpnvgDu9ANN%252f7TTxWWMmOIjxVG772o%252bYGkE9AMxcU5O4cIrT9cubm6dAdgw6n%252fQRZpTVxNv2TGHdHZblPNcfu4dTWVsL3aqaag%253d%253d; _gid=GA1.2.832211649.1645016298; _ce.s=v11.rlc~1645016301520; pageViewNum=13; Hm_lpvt_2d911fad88dfe99ff8bbf610824f3e67=1645017042; Currency=RMB; LocationCode=CN"
	}

	scraper = cfscrape.create_scraper()
	html_code = scraper.get(url,headers=headers).text
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	
def getRenderdHtmlFromUrl(url, isScreenShotName=""):
	chrome_options = Options()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")
	chrome_options.add_argument('user-agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.130 Safari/537.36"')
	chrome_options.add_argument("cookie=__hstc=240517788.35783e1d438e8f99e34188727b050107.1648374652176.1648374652176.1648374652176.1; hubspotutk=35783e1d438e8f99e34188727b050107; __hssrc=1; _gid=GA1.2.122553597.1648374652; _gcl_au=1.1.15847883.1648374654; _fbp=fb.1.1648374655847.2011294960; __hssc=240517788.6.1648374652176; _ga=GA1.1.457442853.1648374652; _ga_SBEXK5LM3N=GS1.1.1648374653.1.1.1648376932.0")
	chrome_options.add_argument("--no-sandbox")

	browser = webdriver.Chrome(options=chrome_options)
	browser.get(url)
	if len(isScreenShotName) > 0:
		imgEle = browser.find_element_by_xpath('//body/img[1]')
		if imgEle !=None:
			imgEle.screenshot(isScreenShotName)
	return browser.page_source
	

def getRenderdHtmlFromUrl1(url):

	chrome_options = Options()
	# chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")
	chrome_options.add_argument('User-Agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.127 Safari/537.36')
	chrome_options.add_argument('Cookie=__utmz=196486330.1639217227.1.1.utmcsr=(direct)|utmccn=(direct)|utmcmd=(none); clientkey=1639217244433_2131; _clientkey_=1650553025945; JSESSIONID=aaa9w1a1Iu4eJ60N43qby; search_key=; __utma=196486330.305675233.1639217227.1650558235.1650634379.6; __utmc=196486330; visittimes=51; __utmt=1; __utmb=196486330.9.10.1650634379; __atuvc=21%7C16; __atuvs=6262ae8bff220521007; view=770; code=0')
	# chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(options=chrome_options)
	browser.get(url)
	time.sleep(2)
	browser.execute_script("window.scrollTo(0,200)")
	time.sleep(4)
	browser.execute_script("window.scrollTo(0,500)")
	time.sleep(2)
	browser.execute_script("window.scrollTo(0,1000)")
	time.sleep(4)
	browser.execute_script("window.scrollTo(0,1500)")
	time.sleep(2)
	browser.execute_script("window.scrollTo(0,document.body.scrollHeight)")
	return BeautifulSoup(browser.page_source, "html.parser",from_encoding="utf-8")
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, type,name,proten,unit,price, products):
	print(str(len(products)) + url)
	sope = getHtmlFromUrl(url)
	pInfo={
		"link": url,
		"type":type,
		"Product Name": name,
		"Protein Family":proten,
		"unit": unit,
		"price":price
	}
	desc = sope.find("div", attrs={"property":"content:encoded"})
	pInfo["Description"] = getNodeText(desc)
	specs = sope.find_all("div", attrs={"class":"field"})
	for spec in specs:
		title = getNodeText(spec.find("div", attrs={"class":"field-label"}))
		val = getNodeText(spec.find("div", attrs={"class":"field-items"}))
		pInfo[title] = val
	print(pInfo)
	products.append(pInfo.copy())

def getProductList(url,type, products):
	sope = getHtmlFromUrl(url)
	
	pListAreas = sope.find("table", attrs={"class":"views-table cols-5"})
	if pListAreas!=None:
		tbody = pListAreas.find("tbody")
		if tbody!=None:
			trs = tbody.find_all("tr")
			for p in trs:
				pLink = p.find("a")
				tds = p.find_all("td")
				getProductInfo("https://www.abbiotec.com"+pLink["href"],
					type,
					getNodeText(tds[1]), 
					getNodeText(tds[2]), 
					getNodeText(tds[3]), 
					getNodeText(tds[4]), 
					products
				)


headers=[
	'link','type','Product Name','Protein Family','unit','price','Description','Composition','Alternate Names'
	,'Purity','Solubility','MW','Storage','Sequence'
]
excelFileName="abbiotec.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("https://www.abbiotec.com/peptides/rgdv-peptide",'','','','','',products )

getProductList("https://www.abbiotec.com/peptides?populate=acth&field_protein_family_tid=All",'ACTH & Analogs', products)
getProductList("https://www.abbiotec.com/peptides?populate=Adrenomedullin&field_protein_family_tid=All",'Adrenomedullins', products)
getProductList("https://www.abbiotec.com/peptides?populate=amylin&field_protein_family_tid=All",'Amylin Peptides', products)
getProductList("https://www.abbiotec.com/peptides?populate=amyloid&field_protein_family_tid=All",'Amyloid Peptides', products)
getProductList("https://www.abbiotec.com/peptides?populate=angiotensins&field_protein_family_tid=All",'Angiotensins', products)
getProductList("https://www.abbiotec.com/peptides?populate=Antimicrobial&field_protein_family_tid=All",'Antimicrobial Peptides', products)
getProductList("https://www.abbiotec.com/peptides?populate=Bombesins&field_protein_family_tid=All",'Bombesins', products)
getProductList("https://www.abbiotec.com/peptides?populate=Bradykinins&field_protein_family_tid=All",'Bradykinins', products)
getProductList("https://www.abbiotec.com/peptides?populate=Calcitonin&field_protein_family_tid=All",'Calcitonin & Analogs', products)
getProductList("https://www.abbiotec.com/peptides?populate=adhesion&field_protein_family_tid=All",'CAM & Ligands', products)
getProductList("https://www.abbiotec.com/peptides?populate=Corticotropin&field_protein_family_tid=All",'Corticotropin & Analogs', products)
getProductList("https://www.abbiotec.com/peptides?populate=Endorphin&field_protein_family_tid=All",'Endorphins', products)
getProductList("https://www.abbiotec.com/peptides?populate=Endothelin&field_protein_family_tid=All",'Endothelins', products)
getProductList("https://www.abbiotec.com/peptides?populate=Exendin&field_protein_family_tid=All",'Exendins', products)
getProductList("https://www.abbiotec.com/peptides?populate=gastrin&field_protein_family_tid=All",'Gastrin & Analogs', products)
getProductList("https://www.abbiotec.com/peptides?populate=Gastrointestinal&field_protein_family_tid=All",'Gastrointestinal Peptides', products)
getProductList("https://www.abbiotec.com/peptides?populate=GHRF&field_protein_family_tid=All",'GHRP & Analogs', products)
getProductList("https://www.abbiotec.com/peptides?populate=Glucagon&field_protein_family_tid=All",'Glucagon & Analogs', products)
getProductList("https://www.abbiotec.com/peptides?populate=lh-rh&field_protein_family_tid=All",'LH-RH & Analogs', products)
getProductList("https://www.abbiotec.com/peptides?populate=msh&field_protein_family_tid=All",'MSH and Analogs', products)
getProductList("https://www.abbiotec.com/peptides?populate=Natriuretic&field_protein_family_tid=All",'Natriuretic Peptides', products)
getProductList("https://www.abbiotec.com/peptides?populate=Neurokinin&field_protein_family_tid=All",'Neurokinins & Related', products)
getProductList("https://www.abbiotec.com/peptides?populate=Neuropeptide&field_protein_family_tid=All",'Neuropeptides', products)
getProductList("https://www.abbiotec.com/peptides?populate=Neurotensin&field_protein_family_tid=All",'Neurotensins & Analogs', products)
getProductList("https://www.abbiotec.com/peptides?populate=Opioid&field_protein_family_tid=All",'Opioid Related Peptides', products)
getProductList("https://www.abbiotec.com/peptides?populate=Gastrointestinal&field_protein_family_tid=All",'Gastrointestinal Peptides', products)
getProductList("https://www.abbiotec.com/peptides?populate=PACAP&field_protein_family_tid=All",'PACAPs', products)
getProductList("https://www.abbiotec.com/peptides?populate=Pancreatic&field_protein_family_tid=All",'Pancreatic Polypeptides', products)
getProductList("https://www.abbiotec.com/peptides?populate=protease&field_protein_family_tid=16",'Peptidase Substrates', products)
getProductList("https://www.abbiotec.com/peptides?populate=&field_protein_family_tid=18",'Protein Kinases & Related', products)
getProductList("https://www.abbiotec.com/peptides?populate=Somatostatin&field_protein_family_tid=All",'Somatostatins', products)
getProductList("https://www.abbiotec.com/peptides?populate=Neuropeptide&field_protein_family_tid=All",'Neuropeptides', products)
getProductList("https://www.abbiotec.com/peptides?populate=tag&field_protein_family_tid=All",'Tag Peptides', products)
getProductList("https://www.abbiotec.com/peptides?populate=Thrombin&field_protein_family_tid=All",'Thrombins & Related', products)
getProductList("https://www.abbiotec.com/peptides?populate=Toxin&field_protein_family_tid=All",'Toxins', products)
getProductList("https://www.abbiotec.com/peptides?populate=Urocortin&field_protein_family_tid=All",'Urocortins', products)
getProductList("https://www.abbiotec.com/peptides?populate=Vasopressin&field_protein_family_tid=All",'Vasopressins & Related', products)
getProductList("https://www.abbiotec.com/peptides?populate=Viral&field_protein_family_tid=All",'Viral Peptides', products)


for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
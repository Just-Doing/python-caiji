from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import requests
from requests.cookies import RequestsCookieJar
import cfscrape

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	print('download:'+IMAGE_URL)
	
	opener = urllib.request.build_opener()
	opener.addheaders = [('User-agent', 'Mozilla/5.0')]
	urllib.request.install_opener(opener)
	urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	headers = {
		"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.82 Safari/537.36",
		"cookie":"_ga=GA1.2.1846208171.1605273760; href=https%3A%2F%2Fwww.sinobiological.com%2Fresearch%2Ftargeted-therapy; accessId=5aff5fb0-84db-11e8-a3b3-d368cce40a8e; _gcl_au=1.1.1660157260.1645016298; Hm_lvt_2d911fad88dfe99ff8bbf610824f3e67=1645016298; sbb=%252be43ohTbVTr09K%252bxQlr1%252bK0onQvF%252bMIXgZM%252bveGXMHU%253dXTaJopSyq01ihC4cD5jOfAa8AEgsWX792EAllASK%252bEKohz0p3JxSEJFavoBnvUmw4fhBpwMcWGZ2Qy%252bRRF2U%252bMSxcQdMfdwOcT%252bR%252bo7qyEU%252br8SBQuGE8GJWgDFeSTZ4QS0HvJFVazETAoyuKMwGHYRoD68%252f7qno5Bg%252bEH9sSXM4upMLtz%252f4IdNkjX6GD0JYHbiUh%252blGTwi25Iz3IKocTDD58DE1yYiY3DxeifN7Qz6OxtXX21lrBpnvgDu9ANN%252f7TTxWWMmOIjxVG772o%252bYGkE9AMxcU5O4cIrT9cubm6dAdgw6n%252fQRZpTVxNv2TGHdHZblPNcfu4dTWVsL3aqaag%253d%253d; _gid=GA1.2.832211649.1645016298; _ce.s=v11.rlc~1645016301520; pageViewNum=13; Hm_lpvt_2d911fad88dfe99ff8bbf610824f3e67=1645017042; Currency=RMB; LocationCode=CN"
	}

	scraper = cfscrape.create_scraper()
	html_code = scraper.get(url,headers=headers).text
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	
def getRenderdHtmlFromUrl(url, isScreenShotName=""):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")
	chrome_options.add_argument('user-agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.130 Safari/537.36"')
	chrome_options.add_argument("cookie=__hstc=240517788.35783e1d438e8f99e34188727b050107.1648374652176.1648374652176.1648374652176.1; hubspotutk=35783e1d438e8f99e34188727b050107; __hssrc=1; _gid=GA1.2.122553597.1648374652; _gcl_au=1.1.15847883.1648374654; _fbp=fb.1.1648374655847.2011294960; __hssc=240517788.6.1648374652176; _ga=GA1.1.457442853.1648374652; _ga_SBEXK5LM3N=GS1.1.1648374653.1.1.1648376932.0")
	chrome_options.add_argument("--no-sandbox")

	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	if len(isScreenShotName) > 0:
		try:
			imgEle = browser.find_element_by_xpath('//body/img[1]')
			if imgEle !=None:
				imgEle.screenshot(isScreenShotName)
		except:
				print("download error")
	return browser.page_source
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, products):
	print(str(len(products)) + url)
	sope = getHtmlFromUrl(url)
	descArea = sope.find("div", attrs={"id":"Description"})
	imgAreas = sope.find_all("img", attrs={"width":"400"})
	imgSrc = ""
	for imgArea in imgAreas:
		if imgArea["src"].find("https://") > -1:
			imgSrc = imgArea["src"]

	if imgSrc=="":
		imgArea = sope.find("img", attrs={"class":"swiper-slide-image lazyloaded"})
		if imgArea != None:
			imgSrc = imgArea["src"]
	if imgSrc=="":
		imgArea = sope.find("img", attrs={"class":"attachment-full size-full lazyloaded"})
		if imgArea != None:
			imgSrc = imgArea["src"]
	pInfo={
		"link": url,
		"Name": getNodeText(sope.find("h1", attrs={"class":"item name fn"})).replace("\xa0","").replace("\n","")
	}
	if descArea != None:
		pInfo["Description"] = getNodeText(descArea.next_sibling)
	trs = sope.find_all("tr")
	for tr in trs:
		tds = tr.find_all("td")
		if len(tds) == 2:
			title = getNodeText(tds[0])
			val = getNodeText(tds[1])
			pInfo[title] = val
	if imgSrc!="":
		imgName = ""
		if "CAS No." in pInfo:
			imgName=pInfo["CAS No."]
		
		if "CAS No.:" in pInfo:
			imgName=pInfo["CAS No.:"]
		# getRenderdHtmlFromUrl(imgSrc, imgName+".jpg")
		pInfo["image"] = imgName+".jpg"
	print(pInfo)
	products.append(pInfo.copy())

def getProductList(url, products):
	sope = getHtmlFromUrl(url)
	pListAreas = sope.find_all("article")
	for pArea in pListAreas:
		pLink = pArea.find("a")
		getProductInfo(pLink["href"], products)


excelFileName="finechem5.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("https://www.fine-chem.cn/products/bromoethane-cas-74-96-4",products )
# getProductList("https://www.fine-chem.cn/chemicals/daily-chemicals/"+str(1), products)
for pIndex in range(1,255):
	getProductList("https://www.fine-chem.cn/chemicals/daily-chemicals/"+str(pIndex), products)


headers=[
	'link','image','Description','Chemical Name','Synonyms','CAS No.','CAS No.:','Molecular Formula','Molecular Weight','PSA'
	,'LogP','EINECS number','MDL number','Appearance & Physical State','Density','Boiling Point','Melting Point','Flash Point'
	,'Refractive Index','Water Solubility','Stability','Storage Condition','Vapor Density'
	,'Vapor Pressure','RTECS','Hazard Class','Safety Statements','HS Code'
	,'Packing Group','WGK Germany','RIDADR','Risk Statements','Hazard Codes'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
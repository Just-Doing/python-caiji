from enum import IntEnum
import requests
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import cfscrape
import json
import string
import re
import time
import math

import numpy as np

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

def getHtmlFromUrl(url, type="get", para={}):
	headers = {
		"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.82 Safari/537.36",
		"cookie":"_ga=GA1.2.1846208171.1605273760; href=https%3A%2F%2Fwww.sinobiological.com%2Fresearch%2Ftargeted-therapy; accessId=5aff5fb0-84db-11e8-a3b3-d368cce40a8e; _gcl_au=1.1.1660157260.1645016298; Hm_lvt_2d911fad88dfe99ff8bbf610824f3e67=1645016298; sbb=%252be43ohTbVTr09K%252bxQlr1%252bK0onQvF%252bMIXgZM%252bveGXMHU%253dXTaJopSyq01ihC4cD5jOfAa8AEgsWX792EAllASK%252bEKohz0p3JxSEJFavoBnvUmw4fhBpwMcWGZ2Qy%252bRRF2U%252bMSxcQdMfdwOcT%252bR%252bo7qyEU%252br8SBQuGE8GJWgDFeSTZ4QS0HvJFVazETAoyuKMwGHYRoD68%252f7qno5Bg%252bEH9sSXM4upMLtz%252f4IdNkjX6GD0JYHbiUh%252blGTwi25Iz3IKocTDD58DE1yYiY3DxeifN7Qz6OxtXX21lrBpnvgDu9ANN%252f7TTxWWMmOIjxVG772o%252bYGkE9AMxcU5O4cIrT9cubm6dAdgw6n%252fQRZpTVxNv2TGHdHZblPNcfu4dTWVsL3aqaag%253d%253d; _gid=GA1.2.832211649.1645016298; _ce.s=v11.rlc~1645016301520; pageViewNum=13; Hm_lpvt_2d911fad88dfe99ff8bbf610824f3e67=1645017042; Currency=RMB; LocationCode=CN"
	}

	scraper = cfscrape.create_scraper()
	html_code = scraper.get(url,headers=headers).text
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")

def requestJson(url):
	r = requests.post(url,data={"input":"atcc"}, headers={
		'Content-Type': 'application/x-www-form-urlencoded',
		'cookie':'visid_incap_2255650=4oBBaRPnQfCVoYEiTmjTq/NVAWEAAAAAQUIPAAAAAAD69PQHUoB0KplKq7/j0+gH; nlbi_2255650=CJKhHYlMm17tpKyoBzOViAAAAACDEjp3gL6bj6YL8j9XE0d/; incap_ses_893_2255650=m1tJIuDRUEp3FE/5GpNkDPRVAWEAAAAAM2KkDpvtARtZral+cMXSVw==; _gcl_au=1.1.76703404.1627477493; _gid=GA1.2.730047202.1627477493; BCSessionID=83af10b8-9488-4b7b-a3b1-3640f178dca2; categoryView=grid; _ga_S46FST9X1M=GS1.1.1627477492.1.1.1627478562.0; _ga=GA1.2.31731397.1627477493; _gat_UA-139934-1=1; _uetsid=69fc2d30efa411eb8818eb045f8760e5; _uetvid=69fc3a70efa411ebba3a23c153f6e477; .Nop.Customer=d664d529-d14a-44b1-86b3-cbf5373277b4',
		"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"
	})
	return BeautifulSoup(r.text, "html.parser",from_encoding="utf-8")


def getRenderdHtmlFromUrl(url):

	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	return BeautifulSoup(browser.page_source, "html.parser")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, type):
	print(str(len(products))+":"+url)
	sope = getHtmlFromUrl(url)
	for br in sope.find_all("br"):
		br.replaceWith("\n")
	pInfo = {"link":url,"Category": type}
	pName = sope.find("h1", attrs={"class":"j-prod-info_title"})
	pInfo["Product Name"] = getNodeText(pName)
	dts = sope.find_all("dt")
	for dt in dts:
		title = getNodeText(dt)
		value = getNodeText(dt.findNextSibling("dd"))
		pInfo[title] = value
	descs = sope.find("div", attrs={"class":"j-prod-info_desc"}).children
	for desc in descs:
		if desc.name != "br":
			descValues = desc.split(':')
			if len(descValues) == 2:
				pInfo[descValues[0].replace('\n','').strip()] = descValues[1]
	h2s = sope.find_all("h2")
	ps = sope.find_all("p")
	h2s.extend(ps)
	for strong in h2s:
		title =  getNodeText(strong.find("strong"))
		value = getNodeText(strong)

		if title.find("Introduction") == 0:
			if title == value:
				pInfo["Introduction"] = getNodeText(strong.findNextSibling("p"))
			else:
				pInfo["Introduction"] = getNodeText(strong)
			pInfo["Introduction"] = pInfo["Introduction"].replace(title, "")
		if title.find("Function") == 0:
			if title == value:
				pInfo["Function"] = getNodeText(strong.findNextSibling("p"))
			else:
				pInfo["Function"] = getNodeText(strong)
			pInfo["Function"] = pInfo["Function"].replace(title, "")
		if title.find("Application") == 0:
			if title == value:
				pInfo["Application"] = getNodeText(strong.findNextSibling("p"))
			else:
				pInfo["Application"] = getNodeText(strong)
			pInfo["Application"] = pInfo["Application"].replace(title, "")
	products.append(pInfo.copy());
	print(pInfo)

def getProductList(url, type):
	html_code = getHtmlFromUrl(url)
	proList = html_code.find_all("div", attrs={"class":"j-prod-list-itemdivs"})
	for tr in proList:
		pLink = tr.find("a")
		getProductInfo(pLink["href"], type)

products = []
# getProductList('https://www.herealth.com/page1/organic-extract.html','Organic Product')
# getProductInfo("https://www.herealth.com/products/organic-reishi-mushroom-extract.html", "ttt")
for pIndex in range(1, 3):
	getProductList('https://www.herealth.com/page'+str(pIndex)+'/organic-extract.html','Organic Product')
for pIndex in range(1, 9):
	getProductList('https://www.herealth.com/plant-extract.html','Plant Extract')
for pIndex in range(1, 2):
	getProductList('https://www.herealth.com/ginseng-product.html','Ginseng Family')
for pIndex in range(1, 2):
	getProductList('https://www.herealth.com/mushroom-product.html','Mushroom Family')
for pIndex in range(1, 3):
	getProductList('https://www.herealth.com/fruit-powder.html','Fruit Powder')
for pIndex in range(1, 2):
	getProductList('https://www.herealth.com/protein-family.html','Protein Family')
for pIndex in range(1, 2):
	getProductList('https://www.herealth.com/spices-powder.html','Spices Powder')

excelFileName="herealth.xlsx"
wb = Workbook()
workSheet = wb.active

headers=[
	'link','Category','Product Name','Introduction',
	'Function','Application','Model NO.','Delivery:','Minimum order quantity:','Supply Ability:','Country of Origin:'
	,'Stock Time:','Product Code','Specification'
	,'Assay Method','Botanical Source','Botanical Part Used','Character','Certification','Process Flow'
]

for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
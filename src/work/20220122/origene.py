from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import requests
from requests.cookies import RequestsCookieJar

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):

	url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
	request_obj=urllib.request.Request(url=url)
	response_obj=urllib.request.urlopen(request_obj)
	html_code=response_obj.read()
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")
	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	cookies = browser.get_cookies()
	session = requests.Session()
	jar = RequestsCookieJar()
	for cookie in cookies:
		jar.set(cookie['name'], cookie['value'])
	session.cookies = jar
	resp = session.get(url)
	print(resp.content)
	return BeautifulSoup(resp.content, "html.parser",from_encoding="utf-8")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, type, type2):
	print(str(len(products))+ "---" +url)
	sope = getHtmlFromUrl(url)
	pInfo = {
		"link": url,
		"categry": type,
		"class": type2,
		"product name": getNodeText(sope.find("h1", attrs={"class":"name"})),
		"description": getNodeText(sope.find("p", attrs={"class":"long-description mt-3"}))
	}
	trs = sope.find_all("tr", attrs={"class":"attribute"})
	for tr in trs:
		tds = tr.find_all("td")
		if len(tds) == 2:
			title = getNodeText(tds[0])
			value = getNodeText(tds[1])
			pInfo[title] = value
			if title == "Sequence Data":
				pInfo["ORF Nucleotide Sequence"] = getNodeText(tds[1].find("div", attrs={"id":"sequence"}))
				pInfo["Protein Sequence"] = getNodeText(tds[1].find("div", attrs={"id":"protein-sequence"}))
			if title == "RefSeq":
				pInfo["RefSeq"] = ""
				refLinks = tds[1].find_all("a")
				for refLink in refLinks:
					pInfo["RefSeq"] += refLink["href"]+";"
			if title == "UniProt ID":
				pInfo["UniProt ID"] = ""
				refLinks = tds[1].find_all("a")
				for refLink in refLinks:
					pInfo["UniProt ID"] += refLink["href"]+";"
	products.append(pInfo.copy())

def getProductList(url, type, type2):
	sope = getHtmlFromUrl(url)
	pList = sope.find_all("article", attrs={"class":"container"})
	for p in pList:
		pLink = p.find("a")
		getProductInfo("https://www.origene.com"+pLink["href"], type, type2)

def getPage(url, type,type2, maxpage):
	for pIndex in range(1, maxpage):
		getProductList(url + "&page="+str(pIndex), type, type2)

excelFileName="origene.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
# getProductList("https://www.origene.com/search?q=mhc&sub_category=Expression+Plasmids", '')
# getProductInfo("https://www.origene.com/catalog/cdna-clones/expression-plasmids/rg205216/hlag-hla-g-nm_002127-human-tagged-orf-clone", '' )

getPage("https://www.origene.com/search?q=mhc&sub_category=Primary+Antibodies", "Antibodies",'Primary Antibodies', 24)
# getPage("https://www.origene.com/search?q=mhc&sub_category=Expression+Plasmids", "Expression Plasmids", 28)
# getPage("https://www.origene.com/search?q=mhc&sub_category=Lentiviral+Particles", "Lentiviral Particles", 10)
# getPage("https://www.origene.com/search?q=mhc&sub_category=3%27+UTR+Clones", "3' UTR Clones", 3)

# getPage("https://www.origene.com/search?q=mhc&sub_category=Primary+Antibodies", "Primary Antibodies", 25)

# getPage("https://www.origene.com/search?q=mhc&sub_category=Recombinant+Proteins", "Recombinant Proteins", 3)
# getPage("https://www.origene.com/search?q=mhc&sub_category=Over-expression+Lysates", "Over-expression Lysates", 5)
# getPage("https://www.origene.com/search?q=mhc&sub_category=Cytosections", "Cytosections", 3)
# getPage("https://www.origene.com/search?q=mhc&sub_category=Mass+Spec+Standards", "Mass Spec Standards", 2)

# getPage("https://www.origene.com/search?q=mhc&sub_category=shRNA+Plasmids", "shRNA Plasmids", 10)
# getPage("https://www.origene.com/search?q=mhc&sub_category=shRNA+Lentiviral+Particles", "shRNA Lentiviral Particles", 6)
# getPage("https://www.origene.com/search?q=mhc&sub_category=siRNA+Oligo+Duplexes", "siRNA Oligo Duplexes", 7)

# getPage("https://www.origene.com/search?q=mhc&sub_category=Knockout+Kits+%28CRISPR%29", "Knockout Kits (CRISPR)", 3)
# getPage("https://www.origene.com/search?q=mhc&sub_category=CRISPRa+Kits", "CRISPRa Kits", 3)
# getPage("https://www.origene.com/search?q=mhc&sub_category=qPCR+Primer+Pairs", "qPCR Primer Pairs", 4)
# getPage("https://www.origene.com/search?q=mhc&sub_category=qPCR+Template+Standards", "qPCR Template Standards", 3)


headers=[
	'link','categry','class','product name','description',
	'Clone Name',
	'Applications',
	'Recommended Dilution',
	'Reactivity',
	'Host',
	'Isotype',
	'Clonality',
	'Immunogen',
	'Specificity',
	'Formulation',
	'Concentration',
	'Conjugation',
	'Storage',
	'Stability',
	'Background',
	'Synonyms'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
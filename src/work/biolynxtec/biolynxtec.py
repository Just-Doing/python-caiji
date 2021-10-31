from urllib.request import urlopen
from selenium import webdriver
import urllib
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy

http.client._MAXHEADERS = 1000


def urllib_download(IMAGE_URL, imageName):
	try:
		from urllib.request import urlretrieve
		urlretrieve(IMAGE_URL, imageName)   
	except:
		print("重试图片下载"+IMAGE_URL)
		urllib_download(IMAGE_URL, imageName)
		
def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url):
	global retryCount
	try:
		html = urlopen(url).read()
		return BeautifulSoup(html, "html.parser",from_encoding="utf-8")
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getHtmlFromUrl(url)
		else:
			retryCount=0
			return None
		
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1


def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	pInfo["link"]=url
	scope = getHtmlFromUrl(url)
	if scope != None:
		pInfo["Description"]=getNodeText(scope.find(name="div", attrs={"class":"summary"}))
		specArea = scope.find(name="div", attrs={"class":"spec"})
		imgArea = scope.find(name="div", attrs={"class":"gallery"})
		images = imgArea.find_all("img")
		imgIndex = 1
		for img in images:
			urllib_download("http://www.biolynxtec.com"+img["src"], pInfo["name"]+str(imgIndex)+'.jpg')
			
			pInfo["imgtitle"+str(imgIndex)] = img["alt"]
			pInfo["imgname"+str(imgIndex)] = pInfo["name"]+str(imgIndex)
			imgIndex = imgIndex + 1
			
		sizeDls = specArea.find_all("dl")
		for dl in sizeDls:
			dt = dl.find("dt")
			dd = dl.find("dd")
			if(getNodeText(dt) == "Catalog No."):
				pInfo["CatalogNo"]=getNodeText(dd)
			if(getNodeText(dt) == "Clone No."):
				pInfo["CloneNo"]=getNodeText(dd)
			if(getNodeText(dt) == "Application"):
				pInfo["Application"]=getNodeText(dd)
			if(getNodeText(dt) == "Subcellular location"):
				pInfo["Subcellularlocation"]=getNodeText(dd)
			if(getNodeText(dt) == "Control"):
				pInfo["Control"]=getNodeText(dd)
			if(getNodeText(dt) == "Recommended method"):
				pInfo["Recommendedmethod"]=getNodeText(dd)
			if(getNodeText(dt) == "Volume"):
				pInfo["Volume"]=getNodeText(dd)
			if(getNodeText(dt) == "Dilution"):
				pInfo["Dilution"]=getNodeText(dd)
			if(getNodeText(dt) == "Immunogen"):
				pInfo["Immunogen"]=getNodeText(dd)
		products.append(pInfo.copy())
	

def getProductList(url, products):
	sope = getHtmlFromUrl(url)
	tableArea = sope.find(name="table", attrs={"class":"table table-striped table-hover" })
	for tr in tableArea.find_all("tr"):
		link = tr.find(name="a", attrs={"class":"btn"})
		if link != None:
			prodInfo = {}
			prodInfo["name"] = getNodeText(tr.find_all("td")[1])
			getProductInfo("http://www.biolynxtec.com"+link["href"], prodInfo, products)




excelFileName="mouser.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
url = "http://www.biolynxtec.com/products/antibody/?layout=list&page="
pageIndex = 1
while pageIndex < 11:
	getProductList(url+str(pageIndex), products)
	
	pageIndex = pageIndex + 1

	
headers=[
	'link','imgtitle1','imgname1','name','Description','CatalogNo','CloneNo','Application',
	'Subcellularlocation','Control',
	'Recommendedmethod',
	'Volume','Dilution','Immunogen','imgtitle2','imgname2'
]
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
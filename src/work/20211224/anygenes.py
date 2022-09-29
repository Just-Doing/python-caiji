import requests
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def requestJson(url, para):
	r = requests.post(url, headers={
		'accept': '*/*',
		'cookie':'PHPSESSID=de050780bac240e034b428a371485127; _ga=GA1.2.1746286200.1640352434; _gid=GA1.2.1216257325.1640352434',
		"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"
	},data=para)
	return r.text

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):

	url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
	request_obj=urllib.request.Request(url=url)
	response_obj=urllib.request.urlopen(request_obj)
	html_code=response_obj.read()
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	
def getRenderdHtmlFromUrl(url):
	try:
		chrome_options = webdriver.ChromeOptions()
		chrome_options.add_argument('--headless')
		chrome_options.add_argument('--disable-gpu')
		chrome_options.add_argument("window-size=1024,768")
		chrome_options.add_argument("--no-sandbox")
		browser = webdriver.Chrome(chrome_options=chrome_options)
		browser.get(url)
		html = browser.page_source
		browser.close()
		return BeautifulSoup(html, "html.parser",from_encoding="utf-8")
	except:
		return None
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(typepath, idlien, type, nomProduit):
	sope = requestJson("https://www.anygenes.com/catalogue_traitementPDO.php",{
		"idProduit": idlien,
		"nomProduit": nomProduit,
		"typePath": typepath,
		"descProduit":""
	})
	if sope != None:
		sope = BeautifulSoup(sope, "html.parser",from_encoding="utf-8")
		tableArea = sope.find_all("table", attrs={"class":"table-bordered table text-center table-striped"})
		
			
		pInfo ={
			"type":type,
			"human_"+"species": "Homo sapiens (human)",
			"mouse_"+"species": "Mus musculus (mouse)",
			"rat_"+"species": "Rattus norvegicus (rat)",
			"human_"+type: "",
			"mouse_"+type: "",
			"rat_"+type: "",
			"human_"+"Product ref":"",
			"human_"+"Informations":"",
			"mouse_"+"Product ref":"",
			"mouse_"+"Informations":"",
			"rat_"+"Product ref":"",
			"rat_"+"Informations":""
		}
		trs1 = tableArea[0].find_all("tr")
		for index,tr in enumerate(trs1):
			tds = tr.find_all("td")
			if len(tds) > 2:
				sheetFile = tds[2].find("a")
				pdfName = "human_"+ nomProduit + "_" + getNodeText(tds[2])+ str(index)+".pdf"
				pInfo["human_"+type] += getNodeText(tds[0]) + ";"
				pInfo["human_"+"Product ref"] = pInfo["human_"+"Product ref"]+ getNodeText(tds[1]) + ";"
				pInfo["human_"+"Informations"] = pInfo["human_"+"Informations"] + pdfName+ ";"
				urllib_download(sheetFile["href"], pdfName)

		trs2 = tableArea[1].find_all("tr")
		for index,tr in enumerate(trs2):
			tds = tr.find_all("td")
			if len(tds) > 2:
				sheetFile = tds[2].find("a")
				pdfName = "mouse_"+ nomProduit + "_" + getNodeText(tds[2])+ str(index)+".pdf"
				pInfo["mouse_"+type] += getNodeText(tds[0]) + ";"
				pInfo["mouse_"+"Product ref"] = pInfo["mouse_"+"Product ref"]+ getNodeText(tds[1]) + ";"
				pInfo["mouse_"+"Informations"] = pInfo["mouse_"+"Informations"] + pdfName+ ";"
				urllib_download(sheetFile["href"], pdfName)
		
		trs3 = tableArea[2].find_all("tr")
		for index,tr in enumerate(trs3):
			tds = tr.find_all("td")
			if len(tds) > 2:
				sheetFile = tds[2].find("a")
				pdfName = "rat_"+ nomProduit + "_" + getNodeText(tds[2])+ str(index)+".pdf"
				pInfo["rat_"+type] += getNodeText(tds[0]) + ";"
				pInfo["rat_"+"Product ref"] = pInfo["rat_"+"Product ref"]+ getNodeText(tds[1]) + ";"
				pInfo["rat_"+"Informations"] = pInfo["rat_"+"Informations"] + pdfName+ ";"
				urllib_download(sheetFile["href"], pdfName)
		
		print(str(len(products)))
		products.append(pInfo.copy())


def getProductList(url):
	sope = getRenderdHtmlFromUrl(url)
	if sope!=None:
		type1Area = sope.find("div", attrs={"id":"1_Ul"})
		type2Area = sope.find("div", attrs={"id":"2_Ul"})
		type3Area = sope.find("div", attrs={"id":"3_Ul"})
		type4Area = sope.find("div", attrs={"id":"4_Ul"})
		type5Area = sope.find("div", attrs={"id":"5_Ul"})
		type6Area = sope.find("div", attrs={"id":"7_Ul"})
		for pLink in type1Area.find_all("a"):
			getProductInfo(pLink["typepath"], pLink["idlien"], 'Signaling Pathways', getNodeText(pLink))
		for pLink in type2Area.find_all("a"):
			getProductInfo(pLink["typepath"], pLink["idlien"], 'Neurodegenerative Diseases Pathways', getNodeText(pLink))
		for pLink in type3Area.find_all("a"):
			getProductInfo(pLink["typepath"], pLink["idlien"], 'Autoimmune Diseases Pathways', getNodeText(pLink))
		for pLink in type4Area.find_all("a"):
			getProductInfo(pLink["typepath"], pLink["idlien"], 'Mental Disorders Pathways', getNodeText(pLink))
		for pLink in type5Area.find_all("a"):
			getProductInfo(pLink["typepath"], pLink["idlien"], 'Cancer Diseases', getNodeText(pLink))
		for pLink in type6Area.find_all("a"):
			getProductInfo(pLink["typepath"], pLink["idlien"], 'Pharmacological Pathways', getNodeText(pLink))


excelFileName="anygenes.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

getProductList('https://www.anygenes.com/home/products/signaling-pathways')
# getProductInfo('Signaling Pathways','2', 'Signaling Pathways', 'Adipogenesis')
headers=[
	'type',
	'human_species',
	'mouse_species',
	'rat_species',
	'human_Signaling Pathways',
	'mouse_Signaling Pathways',
	'rat_Signaling Pathways',
	'human_Neurodegenerative Diseases Pathways',
	'mouse_Neurodegenerative Diseases Pathways',
	'rat_Neurodegenerative Diseases Pathways',
	'human_Autoimmune Diseases Pathways',
	'mouse_Autoimmune Diseases Pathways',
	'rat_Autoimmune Diseases Pathways',
	'human_Mental Disorders Pathways',
	'mouse_Mental Disorders Pathways',
	'rat_Mental Disorders Pathways',
	'human_Cancer Diseases',
	'mouse_Cancer Diseases',
	'rat_Cancer Diseases',
	'human_Pharmacological Pathways',
	'mouse_Pharmacological Pathways',
	'rat_Pharmacological Pathways',
	'human_Product ref',
	'mouse_Product ref',
	'rat_Product ref',
	'human_Informations',
	'mouse_Informations',
	'rat_Informations'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import cfscrape
import re
import copy
import string
import time
import _thread

http.client._MAXHEADERS = 1000


def urllib_download(IMAGE_URL, pName):
    try:
        opener = urllib.request.build_opener()
        opener.addheaders = [('User-agent', 'Mozilla/5.0')]
        urllib.request.install_opener(opener)
        urllib.request.urlretrieve(IMAGE_URL, pName.replace(
            "/", "").replace("\\", "")+'.jpg')
    except:
        print('no')


def getNodeText(node):
    if (node == None):
        return ""
    else:
        return node.get_text().strip()


retryCount = 0
loadCount = 0


def getHtmlFromUrl(url, type="get", para={}):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.82 Safari/537.36",
        "cookie": "_ga=GA1.2.1846208171.1605273760; href=https%3A%2F%2Fwww.sinobiological.com%2Fresearch%2Ftargeted-therapy; accessId=5aff5fb0-84db-11e8-a3b3-d368cce40a8e; _gcl_au=1.1.1660157260.1645016298; Hm_lvt_2d911fad88dfe99ff8bbf610824f3e67=1645016298; sbb=%252be43ohTbVTr09K%252bxQlr1%252bK0onQvF%252bMIXgZM%252bveGXMHU%253dXTaJopSyq01ihC4cD5jOfAa8AEgsWX792EAllASK%252bEKohz0p3JxSEJFavoBnvUmw4fhBpwMcWGZ2Qy%252bRRF2U%252bMSxcQdMfdwOcT%252bR%252bo7qyEU%252br8SBQuGE8GJWgDFeSTZ4QS0HvJFVazETAoyuKMwGHYRoD68%252f7qno5Bg%252bEH9sSXM4upMLtz%252f4IdNkjX6GD0JYHbiUh%252blGTwi25Iz3IKocTDD58DE1yYiY3DxeifN7Qz6OxtXX21lrBpnvgDu9ANN%252f7TTxWWMmOIjxVG772o%252bYGkE9AMxcU5O4cIrT9cubm6dAdgw6n%252fQRZpTVxNv2TGHdHZblPNcfu4dTWVsL3aqaag%253d%253d; _gid=GA1.2.832211649.1645016298; _ce.s=v11.rlc~1645016301520; pageViewNum=13; Hm_lpvt_2d911fad88dfe99ff8bbf610824f3e67=1645017042; Currency=RMB; LocationCode=CN"
    }

    scraper = cfscrape.create_scraper()
    html_code = scraper.get(url, headers=headers).text
    return html_code


def getRenderdHtmlFromUrl(browser, url):
    browser.get(url)
    time.sleep(4)
    return browser.page_source


def writeExcel(workSheet, headers, rowIndex, info):
    cellIndex = 1
    for head in headers:
        try:
            if head in info:
                content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
                workSheet.cell(rowIndex, cellIndex).value = content.strip()
            else:
                workSheet.cell(rowIndex, cellIndex).value = ""
            cellIndex = cellIndex+1
        except:
            print(rowIndex)


def getProductInfo(url, cas, products, browser):
    print(str(len(products)) + cas + url)

    productListHtml = getRenderdHtmlFromUrl(browser, url)
    tempPinfo = {
        "cas": cas
    }
    sope = BeautifulSoup(productListHtml, "html.parser", from_encoding="utf-8")
    tempPinfo["Product Name"] = getNodeText(
        sope.find("h1", attrs={"class": "m-zero p-zero"}))
    ths = sope.find_all("th")
    for th in ths:
        title = getNodeText(th)
        if title == "Molecular Formula":
            tempPinfo["Molecular Formula"] = getNodeText(
                th.findNextSibling("td"))
        if title == "Synonyms":
            tempPinfo["Synonyms"] = getNodeText(th.findNextSibling("td"))
        if title == "Molecular Weight":
            tempPinfo["Molecular Weight"] = getNodeText(
                th.findNextSibling("td"))

    IUPACName = sope.find("section", attrs={"id": "IUPAC-Name"})
    InChI = sope.find("section", attrs={"id": "InChI"})
    InChIKey = sope.find("section", attrs={"id": "InChIKey"})
    CanonicalSMILES = sope.find("section", attrs={"id": "Canonical-SMILES"})
    CAS = sope.find("section", attrs={"id": "CAS"})
    MolecularFormula = sope.find("section", attrs={"id": "Molecular-Formula"})
    if IUPACName != None:
        tempPinfo["IUPACName"] = getNodeText(
            IUPACName.find("div", attrs={"class": "section-content"}))
    if InChI != None:
        tempPinfo["InChI"] = getNodeText(InChI.find(
            "div", attrs={"class": "section-content"}))
    if InChIKey != None:
        tempPinfo["InChIKey"] = getNodeText(InChIKey.find(
            "div", attrs={"class": "section-content"}))
    if CanonicalSMILES != None:
        tempPinfo["CanonicalSMILES"] = getNodeText(
            CanonicalSMILES.find("div", attrs={"class": "section-content"}))
    if CAS != None:
        tempPinfo["CAS1"] = getNodeText(
            CAS.find("div", attrs={"class": "section-content"}))
    if MolecularFormula != None:
        tempPinfo["MolecularFormula"] = getNodeText(
            MolecularFormula.find("div", attrs={"class": "section-content"}))

    products.append(tempPinfo.copy())


def getProductList(cas, products, browser):
    productListHtml = getHtmlFromUrl(
        "https://www.ncbi.nlm.nih.gov/pccompound/?term="+cas)
    if productListHtml != None:
        sope = BeautifulSoup(
            productListHtml, "html.parser", from_encoding="utf-8")
        pros = sope.find_all("div", attrs={"class": "rprt"})
        if len(pros) > 0:
            pro = pros[0]
            title = pro.find("p", attrs={"class": "title"})
            if getNodeText(title).find(cas) > -1:
                pLink = title.find("a")
                getProductInfo(pLink["href"], cas, products,browser)
            else:
                products.append({"cas": cas})
        else:
            products.append({"cas": cas})
    else:
        products.append({"cas": cas})


chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--headless')
chrome_options.add_argument('--disable-gpu')
chrome_options.add_argument("window-size=1024,768")
chrome_options.add_argument("--no-sandbox")
browser = webdriver.Chrome(chrome_options=chrome_options)
products = []
excelFileName = "ncbi1.xlsx"
wb = Workbook()
workSheet = wb.active
fileName = "cat1.txt"
with open(fileName, 'r') as file_to_read:
    index = 1
    type = 1
    while True:
        lines = file_to_read.readline()
        if not lines:
            break
        getProductList(lines.replace("\n", ""), products, browser)

headers = ['cas', 'Product Name', 'Molecular Formula', 'Synonyms', 'Molecular Weight',
            'IUPACName', 'InChI', 'InChIKey', 'CanonicalSMILES', 'CAS1', 'MolecularFormula']
rindex = 1
for p in products:
    writeExcel(workSheet, headers, rindex, p)
    rindex = rindex + 1
print("flish")

wb.save(excelFileName)
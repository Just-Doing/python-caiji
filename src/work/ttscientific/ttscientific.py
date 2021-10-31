from urllib.request import urlopen
from selenium import webdriver
import urllib
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy

http.client._MAXHEADERS = 1000


retryCount = 0
loadCount = 0
def urllib_download(IMAGE_URL, imageName):
	global retryCount
	try:
		from urllib.request import urlretrieve
		urlretrieve(IMAGE_URL, imageName)   
	except:
		print("重试图片下载"+IMAGE_URL)
		retryCount += 1
		if(retryCount <= 5):
			urllib_download(IMAGE_URL, imageName)
		else:
			retryCount=0
			return None
		
def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

def getHtmlFromUrl(url):
	global retryCount
	try:
		html = urlopen(url).read()
		return BeautifulSoup(html, "html.parser",from_encoding="utf-8")
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getHtmlFromUrl(url)
		else:
			retryCount=0
			return None
		
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1


def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	pInfo["link"]=url
	scope = getHtmlFromUrl(url)
	if scope != None:
		descriptArea = scope.find(name="div", attrs={"id":"tabs-1"})
		image = scope.find(name="img", attrs={"id":"FeaturedImage-product-template"})
		if image != None:
			urllib_download("https:"+image["src"], pInfo["name"].replace('(','').replace(')','').replace(':','-')+'.jpg')
		sizeTrs = scope.find_all("tr")
		for tr in sizeTrs:
			tds=tr.find_all("td")
			title = getNodeText(tds[0])
			value = getNodeText(tds[1])
			if(title == "CAS"):
				pInfo["CAS"]=value
			if(title == "Formula"):
				pInfo["Formula"]=value
			if(title == "Molecular Weight"):
				pInfo["MolecularWeight"]=value
			if(title == "Purity"):
				pInfo["Purity"]=value
			if(title == "Storage"):
				pInfo["Storage"]=value
			if(title == "Solubility"):
				pInfo["Solubility"]=value
		if descriptArea != None:
			infos = descriptArea.find_all("p")
			if(len(infos) > 0):
				pInfo["description"] = getNodeText(infos[0])
			if(len(infos) > 1):
				pInfo["description"] = pInfo["description"] + "\r\n" + getNodeText(infos[1])
		products.append(pInfo.copy())
	

def getProductList(url, products):
	sope = getHtmlFromUrl(url)
	proLinks = sope.find_all(name="a", attrs={"class":"full-width-link" })
	for link in proLinks:
		print(link["href"])
		prodInfo = {}
		prodInfo["name"] = getNodeText(link)
		getProductInfo("https://ttscientific.com"+link["href"], prodInfo, products)




excelFileName="ttscientific.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
urls=[
	"https://ttscientific.com/collections/phosphatidic-acid",
	"https://ttscientific.com/collections/phosphatidylcholine-pc?page=1",
	"https://ttscientific.com/collections/phosphatidylcholine-pc?page=2",
	"https://ttscientific.com/collections/phosphatidylethanolamine-pe",
	"https://ttscientific.com/collections/phosphatidylglycerol-pg",
	"https://ttscientific.com/collections/phosphatidylserine-ps"
]
for url in urls:
	getProductList(url, products)

	
headers=[
	'link','name','CAS','Formula','MolecularWeight','Purity','Storage','Solubility',"description"
]
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
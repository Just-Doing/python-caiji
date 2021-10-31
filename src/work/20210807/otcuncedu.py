from enum import IntEnum
import requests
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
from functools import reduce
import string
import re
import time
import math

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0

def requestJson(url):
	r = requests.get(url, headers={
		'Content-Type': 'application/json; charset=utf-8',
		"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.159 Safari/537.36"
	})
	datas = json.loads(r.text)
	return datas

	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductList(url, products):
	print(url)
	data = requestJson(url)
	for pInfoData in data["data"]:
		attrInfo = pInfoData["attributes"]
		uId = attrInfo["uuid"]
		pInfoDetail = requestJson("https://unc.flintbox.com/api/v1/technologies/"+uId+"?organizationId=123&organizationAccessKey=710fd77b-1f7c-41f3-a76e-c5fa43853596")
		includes = pInfoDetail["included"]
		member = reduce(lambda m, cur: cur if cur["type"] == "member" else m, includes, None)
		connector = reduce(lambda m, cur: cur if cur["type"] == "contact" else m, includes, None)
		pInfo = {
			"Title": attrInfo["name"],
			"Published": attrInfo["publishedOn"],
			"Webpage": "https://unc.flintbox.com/technologies/"+uId,
			"Inventor(s)": member["attributes"]["fullName"] if member != None else '',
			"Licensing Contact Person": connector["attributes"]["fullName"] if connector != None else '',
			"Licensing Contact Email": connector["attributes"]["email"] if connector != None else '',

		}
		products.append(pInfo.copy())
			

excelFileName="otcuncedu.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

headers=[
	'Title','Published','Webpage','Inventor(s)','Licensing Contact Person','Licensing Contact Email','备注'
]

# getProductList("https://unc.flintbox.com/api/v1/technologies?organizationId=123&organizationAccessKey=710fd77b-1f7c-41f3-a76e-c5fa43853596&page=1&query=",  products)
for pageIndex in range(1, 14):
	getProductList("https://unc.flintbox.com/api/v1/technologies?organizationId=123&organizationAccessKey=710fd77b-1f7c-41f3-a76e-c5fa43853596&page="+str(pageIndex)+"&query=",  products)


for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
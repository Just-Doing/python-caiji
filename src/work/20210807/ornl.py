from enum import IntEnum
import requests
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import string
import re
import time
import math

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		request_obj=urllib.request.Request(url=url,  headers={
			'Content-Type': 'text/html; charset=utf-8',
			'cookie':'visid_incap_2255650=4oBBaRPnQfCVoYEiTmjTq/NVAWEAAAAAQUIPAAAAAAD69PQHUoB0KplKq7/j0+gH; _gcl_au=1.1.76703404.1627477493; BCSessionID=83af10b8-9488-4b7b-a3b1-3640f178dca2; categoryView=grid; _gid=GA1.2.1947002654.1627710541; .Nop.RecentlyViewedProducts=28110%2C14707%2C4708; .Nop.Customer=d664d529-d14a-44b1-86b3-cbf5373277b4; nlbi_2255650=1ADEGU1fM0jQGgIoBzOViAAAAADCU/81s7IlQ0yjFyM1Kmh8; incap_ses_572_2255650=7JvyE/p4ICEV3ldVVSfwByoNBmEAAAAAt7nF+IFxHG7u7Ms9cJcM6w==; _ga_S46FST9X1M=GS1.1.1627786536.11.0.1627786536.0; _ga=GA1.2.31731397.1627477493; _gat_UA-139934-1=1; ___utmvc=gnWVgF+w8L0A1xlSx94kN24GWOENyTQ2saNvP9dnGk63mOUgy3UnUJbhAZomvGvn7MO6ShqGKaIjDl35hqkW+3+Y6MDWqzsdlE89b6oVw2s82GKerPDWYVLCLpZ57sD7absNA6sZWx8uPUhj8H1KHsdpFPbjNg9DWQkCjJlR9SLqmMW9YqjAXhbM3KbvVgwbRbjj04RhPpEaUhbJ45G9oNxnmhLtz35ESPW8GnNrea2qj4DnxheBGkALQHuiLuiLElYVyuG2cCQ7O7QvYj9QQy4Zr1/r7YY8lKWuEzcNCQZnw8foV9CvuliXsfx2DJoXrCg7pYtXxhZtHQSL9MhGw/u0wYzNjs4igDvUxYWXjGcJQqN+oFCLs2m86yqu1TB6NMT9qKSZ+qkYMvLUFF1OT/90EgjYZvOHyz8SEuv6xpFqCsIXMSLCdrZsMHvzM+D5DRbrSa5g38MooeMSGnrQQtMOxZEwrB32Q9BaAuRMn5MbxBrDfFdq2cmTYHOiBTHCKtV6Bdin37eiJQDpb6fuIOWayGQj46EmFvSWY+5ZaOVyFKuTVIN1LtthjKK71J/h9ToDrPBlYxoZrsuQqq14/FxFhQnv+xKfTmzbeM7zTZ40gMnf30hDEv9P8TW0q+U605+eJ7quCK5GB68UaHtrBRo6gSdRtz3l9ATNZCPKwG1npKtH8SREBp/OOypg8yHEyDdSsckb4bPchFn1GCAUV8sdc0Af+RQlvEsMYISt4NAbVL78zSPEPofbapLU8QvTx6bEuu/V/FR81YNYFL5Mx1ykQ6aaLxM1Essvn3p1gcXGkFAYzEvqk5P/K+SCtb08eD1jvJ8oyil7pzQPuEGxWo3mjqY5YB5oZqhHdr4DJKYEwZjqFw0hSEe2tFBcsmrW3XyHB1KykQ31Qrvd+utEBVPsZ9EfUQLYrejV/n5PakZ7A5j75fx0jIQKnyjy2oFnNM/smJrV46PQpWmsGUjxABsYz1PacqhvYfmCwJhjEmN9IfXU7FWJZJpae96y1pyirOFv1cKYEx6RouZJA3kxAXBeEKFSmxcF7QYNE9wYn8v01Y03bemLev6Z/9Qp+EwjSul00AUro9E4p4OjqA5RSTIn8OPW9GTHgBqo+SO8PMkU9zPC62NVH27Vsb7DwAR9tnFewztXXnAQWHwYz7ySCoVxL45t06AZO6JdztfGULkOL55mbQ+AnFrxqsb1ivUddyhDLq/5qmjHbte6Vy5nDtYieHo7gGh2/RwPl6R8Ku9oef2T1pposWume8VvaNg6DxRrIkVbMV4mYqnNAZeuLyWySK4OmC5Ml3iManTz9zo4EocOGCqUEkaCWdvHnST4EdxlhyX0zxqCQizKHn+RNqbVMArqAu4XRuirkRR2AnvfKrpTo55ToDL6l+BRQzPtLj0FjhSacPMkWhNcBS8TgjspG6/SXvYJr9cBaZHeNxW7ykS1nKR7beRgoB10tzUugTw0D7yRHCRojli93mAzpe5F3mBH35evphkGAUHPf5ybdJHqtP7vQQRT8U8qz8IzBHW2JigejVZWPPf84klb+IUJrOiSU0kEQaMDIqjlHAA70eiC5DAO4C7O/x8G2/jOG4K5wosDvN4ng1jMCIj0EQiDE+EJ2G5l+FXaQ5lzFKLZ+Zhmm0ajcpE+jEwUKJBpBnxSy0iQv8jnKK9VbfXvnnz+NX3YjWC7Yy+CPZPYKe1zEgx1oKwIAL7w960hiZPhe9wfB2qNsYqAQnEfaWjKkTbPwbLN3rKSqz/Gp4AhonM6rb1vXA5Js4RI1/KMakokr1n66ubGyMJ2U9TavZ+fZTQxaZ3YnFT8GbQBQqT2pmu/KipWNeIOJI5o/GWJov5uXhLP7E9fLmbp65uAaEKnn+a1jYnU+lmgu3l5yshl3HFclDRSeCTkJV7xM4+mtGIFqsXgsYfJE6M/w5/lzzRV2NPtPohkZrL7aVeG2wXQCuMDfNtEBaO/kX3w1eGvrSaU9U0RC5gzg56ms+6NgQtm3eQxQfZIa7FXwiMENR9cSQk9uXraP/tL8cpQ4CxkaWdlc3Q9MTM5NzkxLHM9NjI4NGIwNmY4Yjc0YTE3NzlhYjA5MDdkODA2YjhkOWZhZGFiYTI2NzZkOGE4NDllOWRhNWEzNjk4OTlkYTA4ZDZkODQ5ZDdmODE5YTc2NzQ=; _uetsid=0f3330d0f1c311ebaf055f374dd02305; _uetvid=69fc3a70efa411ebba3a23c153f6e477',
			"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"
		})
		htmlHeader = requests.head(url)
		if htmlHeader.status_code ==200:
			response_obj=urllib.request.urlopen(request_obj)
			html_code=response_obj.read()
			return html_code
		else:
			return ''
	except:
		retryCount = retryCount + 1
		if retryCount < 5:
			print("retry index"+str(retryCount)+url)
			time.sleep(60)
			return getHtmlFromUrl(url)
		else:
			retryCount = 0
			return ""

def requestJson(url):
	r = requests.post(url, headers={
		'Content-Type': 'application/x-www-form-urlencoded',
		'cookie':'visid_incap_2255650=4oBBaRPnQfCVoYEiTmjTq/NVAWEAAAAAQUIPAAAAAAD69PQHUoB0KplKq7/j0+gH; nlbi_2255650=CJKhHYlMm17tpKyoBzOViAAAAACDEjp3gL6bj6YL8j9XE0d/; incap_ses_893_2255650=m1tJIuDRUEp3FE/5GpNkDPRVAWEAAAAAM2KkDpvtARtZral+cMXSVw==; _gcl_au=1.1.76703404.1627477493; _gid=GA1.2.730047202.1627477493; BCSessionID=83af10b8-9488-4b7b-a3b1-3640f178dca2; categoryView=grid; _ga_S46FST9X1M=GS1.1.1627477492.1.1.1627478562.0; _ga=GA1.2.31731397.1627477493; _gat_UA-139934-1=1; _uetsid=69fc2d30efa411eb8818eb045f8760e5; _uetvid=69fc3a70efa411ebba3a23c153f6e477; .Nop.Customer=d664d529-d14a-44b1-86b3-cbf5373277b4',
		"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"
	})
	datas = json.loads(r.text)
	return datas

def getRenderdHtmlFromUrl(url):

	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	return BeautifulSoup(browser.page_source, "html.parser",from_encoding="utf-8")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, products):
	print(str(len(products)) + url)
	html_code = getHtmlFromUrl(url)
	if len(html_code)>0:
		sope= BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
		pName = sope.find("h1", attrs={"class":"page-title"})
		CategoryArea = sope.find("div", attrs={"class":"pure-u-1 pure-u-sm-1 pure-u-md-1-3"})
		Categorys = CategoryArea.find_all("div", attrs={"class":"pure-u-1 field-element"})
		inventor = sope.find("div", attrs={"class":"technology-inventor"})
		Category = ""
		if len(Categorys)>0:
			for CategorysDiv in Categorys:
				Category += getNodeText(CategorysDiv)+";"
		pInfo = {
			"link": url,
			"product name": getNodeText(pName),
			"Market Category":Category,
			"inventor": getNodeText(inventor)
		}
		contanctTitles = sope.find_all("h3", attrs={"class":"field-label pure-u-1"})
		for contanctTitle in contanctTitles:
			title = getNodeText(contanctTitle)
			if title == "Licensing Contact":
				connectorArea = contanctTitle.next_sibling.next_sibling
				cName = getNodeText(connectorArea.find("div", attrs={"class":"pure-u-1-1 field-element"}))
				cMail = getNodeText(connectorArea.find("div", attrs={"class":"field field--name-field-tech-licensing-email field--type-email field--label-hidden field__item"}))
				pInfo["Contact Name"] = cName
				pInfo["Contact Email"] = cMail
		print(pInfo)
		products.append(pInfo.copy())

def getProductList(url, products):
	print(url)
	html_code = getHtmlFromUrl(url)
	if len(html_code)>0:
		sope= BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
		pListTr = sope.find_all("div", attrs={"class":"views-row pure-u-1"})
		for p in pListTr:
			pLink = p.find("a")
			getProductInfo("https://www.ornl.gov"+pLink["href"], products)
		
	

excelFileName="ornl.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

headers=[
	'link','product name','Market Category','inventor','Contact Name','Contact Email'
]

# getProductInfo("https://www.ornl.gov/technology/202104846", products)
# getProductList("https://www.ornl.gov/technology-transfer/technologies?search_api_fulltext=&page=0", products)
for pageIndex in range(0, 135):
	getProductList("https://www.ornl.gov/technology-transfer/technologies?search_api_fulltext=&page="+str(pageIndex),products)


for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
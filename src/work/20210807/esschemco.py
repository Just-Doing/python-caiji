from enum import IntEnum
import requests
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import string
import re
import time
import math

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	
	opener = urllib.request.build_opener()
	opener.addheaders = [('User-agent', 'Mozilla/5.0')]
	urllib.request.install_opener(opener)
	urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		request_obj=urllib.request.Request(url=url,  headers={
			'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
			'cookie':'_ga=GA1.2.510794350.1632490112; _gid=GA1.2.1808662766.1632490112; __stripe_mid=1840413b-ed41-44e0-a11f-ff7c582fe2d4d861a3; __stripe_sid=26d2d2f8-4025-4ddc-a3ce-9959ee1e079c000594; PHPSESSID=ef2a208c339d0e9e15f810221e50fd85',
			"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.82 Safari/537.36"
		})
		htmlHeader = requests.head(url)
		print(htmlHeader)
		if htmlHeader.status_code ==200:
			response_obj=urllib.request.urlopen(request_obj)
			html_code=response_obj.read()
			return html_code
		else:
			return ''
	except:
		retryCount = retryCount + 1
		if retryCount < 5:
			print("retry index"+str(retryCount)+url)
			time.sleep(60)
			return getHtmlFromUrl(url)
		else:
			retryCount = 0
			return ""

def requestJson(url):
	r = requests.post(url, headers={
		'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
		'cookie':'_ga=GA1.2.510794350.1632490112; _gid=GA1.2.1808662766.1632490112; __stripe_mid=1840413b-ed41-44e0-a11f-ff7c582fe2d4d861a3; __stripe_sid=26d2d2f8-4025-4ddc-a3ce-9959ee1e079c000594; PHPSESSID=ef2a208c339d0e9e15f810221e50fd85',
		"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"
	})
	datas = json.loads(r.text)
	return datas

def getRenderdHtmlFromUrl(url):

	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	return BeautifulSoup(browser.page_source, "html.parser",from_encoding="utf-8")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url,type1, products):
	print(str(len(products)) + url)
	sope = getRenderdHtmlFromUrl(url)
	if len(sope)>0:
		titleArea = sope.find("div", attrs={"class":"breadcrumbs"})
		desArea = sope.find("div", attrs={"itemprop":"description"})
		pInfo = {
			"type":type1,
			"url": url,
			"Product Name": getNodeText(titleArea.find("h1")),
			"Nav": getNodeText(titleArea.find("ul"))
		}
		descs = desArea.find_all("p")
		for desc in descs:
			titles = desc.find_all("strong")
			val = getNodeText(desc).replace("\n", "").replace('\xa0', '')
			if len(titles)>1:
				title1 = getNodeText(titles[0])
				title2 = getNodeText(titles[1])
				pInfo[title1] = val[0: val.find(title2)].replace(title1,"")
				pInfo[title2] = val[val.find(title2):].replace(title2,"")

			if len(titles) == 1:
				title = getNodeText(titles[0])
				if len(title) > 0:
					if title.find("Inventory status: ")> -1:
						pInfo["Inventory status"] = title.replace("Inventory status: ", "")
					else:
						pInfo[title] = val.replace(title, "")
		relatePro = sope.find("div", attrs={"class":"related products"})
		if relatePro!=None:
			rePros = relatePro.find_all("a")
			reproStr = ""
			for repro in rePros:
				reproStr += getNodeText(repro)+"|||"
			pInfo["Related Products"] = reproStr

		packSizeArea = sope.find("div", attrs={"id":"variations"})
		pkTrs = packSizeArea.find_all("tr")
		pkSize = ""
		price = ""
		for pkTr in pkTrs:
			tds = pkTr.find_all("td")
			if len(tds) == 4:
				pkSize  += getNodeText(tds[0]) + "|||"
				price  += getNodeText(tds[1]) + "|||"
		pInfo["Pack Size"] = pkSize
		pInfo["Price"] = price

		CertificateofAnalysis = sope.find("a", attrs={"class":"maxbutton-1 maxbutton maxbutton-certificate-of-analysis"})
		SDS = sope.find("a", attrs={"class":"maxbutton-2 maxbutton maxbutton-sds"})
		if CertificateofAnalysis != None:
			href = CertificateofAnalysis["href"]
			pInfo["CertificateofAnalysis"] = href[href.rindex('/'):]
			urllib_download(href, pInfo["CertificateofAnalysis"])
		if SDS != None:
			href = SDS["href"]
			pInfo["SDS"] = href[href.rindex('/'):]
			urllib_download(href, pInfo["SDS"])
		img = sope.find("img", attrs={"class":"attachment-shop_single size-shop_single wp-post-image"})
		if img!=None:
			src = img["src"]
			imgName = src[src.rindex('/'):]
			pInfo["img"] = imgName
			urllib_download("https:"+img["src"], imgName)
		moreInfoArea = sope.find("div", attrs={"id":"tab-more-info"})
		if moreInfoArea!=None:
			moreDescs = moreInfoArea.find_all("p")
			for moreDesc in moreDescs:
				val = getNodeText(moreDesc).replace("\n", "").replace('\xa0', '')
				title = getNodeText(moreDesc.find("strong"))
				if len(title) > 0:
					pInfo[title] = val.replace(title, "")
		products.append(pInfo.copy())

def getProductList(url,type1, products):
	print(url)
	sope = getRenderdHtmlFromUrl(url)
	if len(sope)>0:
		pListTr = sope.find_all("a", attrs={"class":"productimage"})
		for p in pListTr:
			getProductInfo("https:"+p["href"],type1, products)
		
	

excelFileName="esschemco.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

headers=[
	'type','url','Product Name','Nav','CAT#:','CAS#:','Purity:','MF:','MW:','Inventory status','Related Products','Pack Size','Price',
	'CertificateofAnalysis','SDS','img','Synonym:','Lot:','Appearance:'
]

getProductInfo("https://esschemco.com/product/steroids/10beta-hydroxy-norethindrone/",'ttt', products)
# getProductList("https://esschemco.com/product-category/amines/",'amines', products)
# getProductList("https://esschemco.com/product-category/glucuronides/",'Glucuronides', products)
# getProductList("https://esschemco.com/product-category/heterocyclic-compounds/",'Heterocyclic Compounds', products)
# getProductList("https://esschemco.com/product-category/isothiocyanates/",'Isothiocyanates', products)

# for pageIndex in range(1, 18):
# 	getProductList("https://esschemco.com/product-category/labeled-compounds/page/"+str(pageIndex)+"/",'Labeled Compounds', products)

# for pageIndex in range(1, 5):
# 	getProductList("https://esschemco.com/product-category/metabolites-impurities/page/"+str(pageIndex)+"/",'Metabolites & Impurities', products)

# for pageIndex in range(1, 10):
# 	getProductList("https://esschemco.com/product-category/steroids/page/"+str(pageIndex)+"/",'Steroids', products)

# for pageIndex in range(1, 14):
# 	getProductList("https://esschemco.com/product-category/epichem-reference-standards/page/"+str(pageIndex)+"/",'Epichem Reference Standards', products)



for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
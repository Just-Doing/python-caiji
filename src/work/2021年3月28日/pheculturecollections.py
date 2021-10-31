import requests
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36"}

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		if retryCount < 5:
			getHtmlFromUrl(url)

def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(letter, type, url, products):
	print(str(len(products)) + url)
	productHtml = getHtmlFromUrl(url)
	if productHtml != None:
		tempPinfo={
			"letter":letter,
			"type":type, 
			"url":url
		}
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
		
		specs = sope.find_all("tr", attrs={"class":"detail"})
		for spec in specs:
			title = getNodeText(spec.find("th"))
			val = spec.find("td")
			
			tempPinfo[title] = getNodeText(val)
		availableFormats = sope.find("div", attrs={"id":"availableFormats"})
		if availableFormats!=None:
			tempPinfo["availableFormats"] = getNodeText(availableFormats.find("p"))
		products.append(tempPinfo)

def getProductList(letter, type, url, pageIndex, products):
	productHtml = getHtmlFromUrl(url)
	if productHtml!=None:
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
		pageLink = sope.find("span", attrs={"class":"pagelinks"})
		if pageLink!=None and pageIndex == 1:
			linkCount = len(pageLink.find_all("a"))
			for pageIndex in range(2, linkCount-2):
				getProductList(letter, type, url+"&d-49653-p="+str(pageIndex), pageIndex, products)
		
		trs = sope.find_all("tr", attrs={"class":"odd"})
		trs1 = sope.find_all("tr", attrs={"class":"even"})
		for tr in trs+trs1:
			link = tr.find("a")
			linkUrl = link["href"].split("?")
			getProductInfo(letter, type, "https://www.phe-culturecollections.org.uk/products/bacteria/detail.jsp?"+linkUrl[1], products)
		
def getProductType(q, products):
	url = "https://www.phe-culturecollections.org.uk/products/bacteria/browse.jsp"
	d = {
		'a2z': q
	}

	r = requests.post(url, data=d, headers={ 'Content-Type':'application/x-www-form-urlencoded' })
	sope = BeautifulSoup(r.text, "html.parser",from_encoding="utf-8")
	tabarea = sope.find("table", attrs={"id":"resultItem"})
	for tr in tabarea.find_all("tr"):
		link = tr.find("a")
		if link!=None:
			getProductList(q, getNodeText(link), "https://www.phe-culturecollections.org.uk/products/bacteria/browse.jsp"+link["href"], 1, products)
		
excelFileName="pheculturecollections.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
# getProductList("https://www.phe-culturecollections.org.uk/products/bacteria/browse.jsp?level=1&val1=Corynebacterium+diphtheriae&mode=&mode=", 1 ,products)
for letter in ["A","B","C","D","E","F","G","H","I","J","K","L"
,"M","N","O","P","Q","R","S","T","U","V"
,"W","X","Y","Z"]:
	getProductType(letter ,products)


headers=['url','letter',"type","NCTC Number:","Current Name:","Other Collection No:","Previous Catalogue Name:","Type Strain:",
"Family:","Hazard Group (ACDP):","Conditions for growth on solid media:","Conditions for growth on liquid media:",
"16S rRNA Gene Sequence:","23S rRNA Gene Sequence:","Taxonomy:","availableFormats","Toxin Status:","Antigenic Properties:",
"Miscellaneous Sequence Data:","Other Names:","Variety:","Also Known As:","Isolated From:"
	]
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	if rindex%100  == 0:
		wb.save(excelFileName)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
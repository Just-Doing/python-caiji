import requests
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36"}

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		if retryCount < 5:
			getHtmlFromUrl(url)

def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(id, products):
	print(str(len(products)) +"=="+ str(id))
	if len(str(id))>0:
		url = "https://webservices.bio-aware.com/ncimbnew/api/Details/GetTemplateByIdAndRecordDetails?p_TemplateId=7&p_RecordId="+str(id)+"&p_DesignMode=1"
		r = requests.get(url,headers={ 
			'Content-Type':'application/json; charset=utf-8',
			"Accept":"application/json, text/plain, */*",
			"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.114 Safari/537.36"		,
			"WebsiteId":"14"
		})
		datas = json.loads(r.text)
		proFied = datas["Data"]
		valueObj = proFied["RecordDetails"]
		if  proFied!= None and valueObj!=None:
			tempPinfo={}
			specs = proFied["Template"]["Fields"][0]["childs"]
			for spec in specs:
				if spec["FieldType"]>-107 and spec["FieldType"]<-99:
					tempPinfo[spec["title"].strip()] = valueObj[str(spec["FieldType"])]["Value"]
				else:
					if "Value" in valueObj[str(spec["FieldKey"])]:
						if spec["title"].strip() == "Media":
							if len(valueObj[str(spec["FieldKey"])]["Value"])>0:
								tempPinfo[spec["title"].strip()] = valueObj[str(spec["FieldKey"])]["Value"][0]["Name"]["Value"]
						else:
							tempPinfo[spec["title"].strip()] = valueObj[str(spec["FieldKey"])]["Value"]
			products.append(tempPinfo)

def getProductType(startIndex, products):
	url = "https://webservices.bio-aware.com/ncimbnew/api/Search/SearchForSummaryGrid"
	d = {
		"TableKey":"2891383000000113",
		"Fields":["name","e517","e3","t2","t6","d1"],
		"iDisplayLength":100,
		"iDisplayStart":startIndex,
		"ComplexQuery":None,
		"SortColumn":None,
		"SortDirection":None,
		"LoadOwnerRecord":False
	}

	r = requests.post(url, data=json.dumps(d), headers={ 
		'Content-Type':'application/json',"WebsiteId":"14"
	})
	datas = json.loads(r.text)
	if "Data" in datas:
		for prd in datas["Data"]["RecordEntityList"]:
			getProductInfo( prd["Id"], products)
		
		
excelFileName="ncimb.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
# getProductList("https://www.phe-culturecollections.org.uk/products/bacteria/browse.jsp?level=1&val1=Corynebacterium+diphtheriae&mode=&mode=", 1 ,products)
for pageIndex in range(0, 84):
	getProductType( pageIndex*100 ,products)
# getProductInfo("25154",products)


headers=['NCIMB number','Deposit type','Type strain','GMO','Taxon name','Preservation method','Preserved as'
,'Media','Gas regime','Growth factors (and/or information)','ACDP category','Colony Edge','Colony Surface','Colony Shape'
,'Colony Elevation','Colony Colour','Colony Opacity','Cellular Shape','Cellular Size','Media','Cellular Arrangement'
,'Gram stain','Source','Other collection IDs','Genotype information','Yeast?', 'Hazard information','Additional hazard information','Cellular Motility','Straight Or Curved Axis',
'Parallel Or Curved Sides','Rounded Ends','Other names','Incubation period','Infrasubspecfic names'
	]
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	if rindex%100  == 0:
		wb.save(excelFileName)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
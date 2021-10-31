import requests
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36"}

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		if retryCount < 5:
			getHtmlFromUrl(url)

def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, products):
	print(str(len(products)) + url)
	productHtml = getHtmlFromUrl(url)
	if productHtml != None:
		tempPinfo={ 
			"url":url
		}
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
		
		specArea = sope.find("div", attrs={"class":"product-detail"})
		name = specArea.find("h1")
		tempPinfo["name"] = getNodeText(name)
		specs = specArea.find_all("div", attrs={"class":"field"})
		for spec in specs:
			title = getNodeText(spec.find("div",attrs={"class":"label"}))
			val = getNodeText(spec.find("div",attrs={"class":"value"}))
			
			tempPinfo[title] = val
		products.append(tempPinfo)

def getProductType(group, data, startIndex, products):
	url = "https://www.dsmz.de/collection/catalogue/microorganisms/catalogue?type=2601231117"
	d = {
		'controller': 'user_dsmz_catalogue_products',
		'data[filter][product_group]': group,
		'data[filter][search_term]':'',
		'data[offset]':startIndex,
		'data[jsLib]':data,
		'action':'getResults'
	}

	r = requests.post(url, data=d, headers={ 'Content-Type':'application/x-www-form-urlencoded' })
	datas = json.loads(r.text)
	sope = BeautifulSoup(datas["data"]["results"], "html.parser",from_encoding="utf-8")
	prds = sope.find_all("div", attrs={"class":"product"})
	for prd in prds:
		name = prd.find("div", attrs={"class":"name"})
		link = name.find("a")
		getProductInfo( "https://www.dsmz.de"+link["href"], products)
		
		
excelFileName="dsmz-archaea.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
# getProductList("https://www.phe-culturecollections.org.uk/products/bacteria/browse.jsp?level=1&val1=Corynebacterium+diphtheriae&mode=&mode=", 1 ,products)
for pageIndex in range(0, 62):
	getProductType('19', 'jsLib_1333', pageIndex ,products)
# getProductInfo("https://www.dsmz.de/collection/catalogue/details/culture/DSM-1" ,products)


headers=['url','name',"DSM No.:","Other collection no. or WDCM no.:","Isolated from:","Cultivation conditions:",
"Summary and additional information:","Risk group:",
"Supplied as:","Genbank accession numbers:","Synonym(s):"
	]
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	if rindex%100  == 0:
		wb.save(excelFileName)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
import requests
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.193 Safari/537.36"}

		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		if retryCount < 5:
			getHtmlFromUrl(url)

def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, products):
	print(str(len(products)) + url)
	if url.find(".ashx")<0:
		tempPinfo = {}
		tempPinfo["link"]=url
		tempPinfo["nub"]=url.replace("https://www.atcc.org/products/all/","").replace(".aspx","")
		productHtml = getHtmlFromUrl(url)
		if productHtml != None:
			sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
			titleAres = sope.find("h1", attrs={"class":"product_name"})
			if titleAres!=None:
				tempPinfo["title"]=getNodeText(titleAres.find("i"))
				specs = sope.find_all("tr")
				for spec in specs:
					tds = spec.find_all("td")
					ths = spec.find_all("th")
					if len(tds) == 1 and len(ths ) == 1:
						title = getNodeText(ths[0])
						val = getNodeText(tds[0])
						tempPinfo[title] = val
				products.append(tempPinfo)


def getProductList(aq,q,startIndex, products):
	url = "https://www.atcc.org/coveo/rest/v2/?sitecoreItemUri=sitecore%3A%2F%2Fliveweb%2F%7BF6F5050D-7B79-4599-B4A5-91361DAD7759%7D%3Flang%3Den%26ver%3D1&siteName=ATCC"

	d = {
		'aq': aq,
		'retrieveFirstSentences': 'true',
		'timezone':'Asia/Shanghai',
		'disableQuerySyntax':'false',
		'enableDuplicateFiltering':'false',
		'enableCollaborativeRating':'false',
		'debug':'false',
		'context':'{}',
		'cq':'((@fz95xlanguage14674=="en" @fz95xlatestversion14674=="1") OR @syssource==("ATCC-PROD"))',
		'searchHub':'search',
		'language':'en',
		'pipeline':'ATCC',
		'firstResult':startIndex,
		'numberOfResults':'100',
		'excerptLength':'200',
		'enableDidYouMean':'true',
		'sortCriteria':'relevancy',
		'queryFunctions':'[]',
		'rankingFunctions':'[]',
		'q':q
	}
	
	r = requests.post(url, data=d, headers={ 'Content-Type':'application/x-www-form-urlencoded' })
	datas = json.loads(r.text)
	for data in datas["results"]:
		getProductInfo(data["uri"], products)
excelFileName="atcc1.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
# getProductInfo("https://www.atcc.org/products/all/3629.aspx" ,products)
# for pageIndex in range(0, 51):
	# aq='(@productlineclassification=="Bacteria & Phages") (@syssource==("Coveo_liveweb_index - production","ATCC-PROD") ((@syssource=="ATCC-PROD")) NOT @ftemplateid14674==("adb6ca4f-03ef-4f47-b9ac-9ce2ba53ff97","fe5dd826-48c6-436d-b87a-7c4210c7413b"))'
	# q=''
	# getProductList(aq, q, pageIndex * 100 ,products)

for pageIndex in range(0, 12):
	aq='(@productlineclassification=="Bacteria & Phages") (@taxonomylistofapplications=="Antibiotic Resistance") (@syssource==("Coveo_liveweb_index - production","ATCC-PROD") ((@syssource=="ATCC-PROD")) NOT @ftemplateid14674==("adb6ca4f-03ef-4f47-b9ac-9ce2ba53ff97","fe5dd826-48c6-436d-b87a-7c4210c7413b"))'
	q=''
	getProductList(aq, q, pageIndex * 100 ,products)

for pageIndex in range(0, 10):
	aq='(@contenttype==Products) (@productlineclassification==("Bacteria & Phages","Nucleic Acids - DNA/ RNA")) (@syssource==("Coveo_liveweb_index - production","ATCC-PROD") ((@syssource=="ATCC-PROD")) NOT @ftemplateid14674==("adb6ca4f-03ef-4f47-b9ac-9ce2ba53ff97","fe5dd826-48c6-436d-b87a-7c4210c7413b"))'
	q='@taxonomylistofapplications="Sequencing"'
	getProductList(aq, q, pageIndex * 100 ,products)
	
for pageIndex in range(0, 9):
	aq='(@contenttype==Products) (@syssource==("Coveo_liveweb_index - production","ATCC-PROD") ((@syssource=="ATCC-PROD")) NOT @ftemplateid14674==("adb6ca4f-03ef-4f47-b9ac-9ce2ba53ff97","fe5dd826-48c6-436d-b87a-7c4210c7413b"))'
	q='@productline="B251"'
	getProductList(aq, q, pageIndex * 100 ,products)
	
for pageIndex in range(0, 11):
	aq='(@contenttype==Products) (@syssource==("Coveo_liveweb_index - production","ATCC-PROD") ((@syssource=="ATCC-PROD")) NOT @ftemplateid14674==("adb6ca4f-03ef-4f47-b9ac-9ce2ba53ff97","fe5dd826-48c6-436d-b87a-7c4210c7413b"))'
	q='@productline="B451"'
	getProductList(aq, q, pageIndex * 100 ,products)
	
for pageIndex in range(0, 1):
	aq='(@contenttype==Products) (@syssource==("Coveo_liveweb_index - production","ATCC-PROD") ((@syssource=="ATCC-PROD")) NOT @ftemplateid14674==("adb6ca4f-03ef-4f47-b9ac-9ce2ba53ff97","fe5dd826-48c6-436d-b87a-7c4210c7413b"))'
	q='@productline="A051"'
	getProductList(aq, q, pageIndex * 100 ,products)
		
for pageIndex in range(0, 2):
	aq='(@contenttype==Products) (@syssource==("Coveo_liveweb_index - production","ATCC-PROD") ((@syssource=="ATCC-PROD")) NOT @ftemplateid14674==("adb6ca4f-03ef-4f47-b9ac-9ce2ba53ff97","fe5dd826-48c6-436d-b87a-7c4210c7413b"))'
	q='@taxonomylistofapplications="Media_testing"'
	getProductList(aq, q, pageIndex * 100 ,products)
	
	
for pageIndex in range(0, 1):
	aq='(@contenttype==Products) (@syssource==("Coveo_liveweb_index - production","ATCC-PROD") ((@syssource=="ATCC-PROD")) NOT @ftemplateid14674==("adb6ca4f-03ef-4f47-b9ac-9ce2ba53ff97","fe5dd826-48c6-436d-b87a-7c4210c7413b"))'
	q='@taxonomylistofapplications="Water_testing"'
	getProductList(aq, q, pageIndex * 100 ,products)
	
headers=['nub','title',
		'Deposited As','Strain Designations','Application','Biosafety Level',
		'Product Format','Storage Conditions','Isolation','Type Strain','Medium',
		'Growth Conditions','Cross References',
		'Antigenic Properties','Disclosure','U.S. Patent Number','Genome Sequenced Strain',
		'Genotype','Source','Intended Use','Classification','Antibiotic Resistance',
		'Recommended Host','Effect on Host','Cryopreservation',
		'Toxins','Morphology','Sequenced Data'
	]
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	if rindex%100  == 0:
		wb.save(excelFileName)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
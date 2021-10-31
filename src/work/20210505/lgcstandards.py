from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		if str(html_code).find("403 ERROR")>-1:
			time.sleep(360)
			return getHtmlFromUrl(url)
		else:
			return html_code
	except:
		retryCount += 1
		if retryCount < 5:
			print("retry index"+str(retryCount)+url)
			return getHtmlFromUrl(url)
		else:
			retryCount = 0
			return ""
def getRenderdHtmlFromUrl(url, isTry):
	global retryCount
	try:
		chrome_options = webdriver.ChromeOptions()
		chrome_options.add_argument('--headless')
		chrome_options.add_argument('--disable-gpu')
		chrome_options.add_argument("window-size=1024,768")

		chrome_options.add_argument("--no-sandbox")
		browser = webdriver.Chrome(chrome_options=chrome_options)
		browser.get(url)
		if str(browser.page_source).find("403 ERROR")>-1:
			time.sleep(360)
			return getRenderdHtmlFromUrl(url, True)
		else:
			return browser.page_source
	except:
		retryCount += 1
		if retryCount < 5:
			print("retry index"+str(retryCount)+url)
			return getRenderdHtmlFromUrl(url, True)
		else:
			retryCount = 0
			return ""
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, type1, type2, products):
	print(str(len(products)) + url)
	productListHtml = getRenderdHtmlFromUrl(url, False)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	pInfo = {
		"link": url,
		"type1": type1,
		"type2": type2
	}
	titleArea = sope.find("div", attrs={"class": "product__title-wrapper"})
	if titleArea != None:
		pInfo["Product Name"] = getNodeText(titleArea.find("h1"))
		pInfo["Synonyms"] = getNodeText(titleArea.find("p"))

		specAreas = sope.find_all("div",attrs={"class":"product__item"})
		for specArea in specAreas:
			title = getNodeText(specArea.find("h2"))
			val = getNodeText(specArea.find("p"))
			if title == "API Family":
				val = getNodeText(specArea.parent)
			if len(title) > 0:
				pInfo[title]=val
		
		img = sope.find("img", attrs={"itemprop":"image"})
		if img != None:
			if "CAS Number" in pInfo and len(pInfo["CAS Number"])>0:
				urllib_download(img["src"], pInfo["CAS Number"])
			else:
				urllib_download(img["src"], pInfo["Product Name"])
		products.append(pInfo.copy())

def getProductList(url, type1, type2, products):
	productListHtml = getHtmlFromUrl(url)
	data = json.loads(productListHtml)
	hits = data["products"]
	for hit in hits:
		link = hit["url"]
		getProductInfo("https://www.lgcstandards.com/US/en"+link, type1, type2, products)

excelFileName="lgcstandards.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("https://www.lgcstandards.com/US/en/Dexamethasone-Sodium-Phosphate/p/MM0210.00", "111", "222", products)
# getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/322963/products?currentPage=1&q=&sort=relevance-code&pageSize=10&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical',"COVID-19 research and reference materials", products)

for page in range(0,13):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/322963/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical',"COVID-19 research and reference materials", products)
for page in range(0,16):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/323809/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','API standards', products)
for page in range(0,72):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/323810/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit",'Pharmaceutical','Impurity standards', products)
for page in range(0,42):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/324176/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Pharmacopoeial standards', products)
for page in range(0,31):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279516/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','European Pharmacopoeia (Ph. Eur.)', products)
for page in range(0,19):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279513/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','British Pharmacopoeia', products)
for page in range(0,4):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279517/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Reagents according to pharmacopoeias', products)
for page in range(0,2):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279526/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=20&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','International reference standards for antibiotics (WHO)', products)
for page in range(0,17):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/323817/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Building blocks', products)
for page in range(0,30):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/324179/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Enzyme activators, inhibitors & substrates', products)
for page in range(0,5):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/323822/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Neurochemicals', products)
for page in range(0,11):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/323824/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Carbohydrates', products)
for page in range(0,2):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/323827/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Cross-linkers', products)
for page in range(0,28):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/323816/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Chiral molecules', products)
for page in range(0,81):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/323830/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Pharmaceutical toxicology', products)
for page in range(0,7):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279509/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Elemental impurities', products)
for page in range(0,2):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279527/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Physical properties', products)
for page in range(0,2):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279507/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=20&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Pharmaceutical proficiency testing', products)

for page in range(0,6):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279543/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Food & Beverage',"Cannabis-related compounds", products)
for page in range(0,4):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279550/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Food & Beverage',"Dyes & metabolites", products)
for page in range(0,4):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279551/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=20&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Food & Beverage',"Allergens", products)
for page in range(0,4):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279552/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Food & Beverage',"Nutritional composition compounds", products)
for page in range(0,6):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279553/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Food & Beverage',"Food additives, flavours & adulterants", products)
for page in range(0,3):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279554/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Food & Beverage',"Mycotoxins", products)
for page in range(0,11):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279557/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Food & Beverage',"Environmental food contaminants", products)
for page in range(0,49):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279562/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Food & Beverage',"Pesticides & metabolites", products)
for page in range(0,20):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279568/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Food & Beverage',"Pharma & vet compounds & metabolites", products)
for page in range(0,5):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279569/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Food & Beverage',"Phytochemicals", products)
for page in range(0,8):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279599/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Food & Beverage',"Food contact materials", products)
for page in range(0,2):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279604/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=20&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Food & Beverage',"Beverage reference materials", products)
for page in range(0,3):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/280775/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Food & Beverage',"Food & beverage proficiency testing", products)
for page in range(0,4):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279622/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Food & Beverage',"Standards for food regulatory methods", products)
for page in range(0,8):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279627/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Food & Beverage',"Stable isotope labelled compounds", products)
for page in range(0,4):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279643/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=20&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Food & Beverage',"Microbiology", products)

for page in range(0,7):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279645/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Environmental',"Stable isotope labelled compounds", products)
for page in range(0,6):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279646/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Environmental',"PCBs & related compounds", products)
for page in range(0,3):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279647/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Environmental',"Flame retardants", products)
for page in range(0,4):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279649/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Environmental',"Dyes & metabolites", products)
for page in range(0,2):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279650/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=20&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Environmental',"Dioxins & furans", products)
for page in range(0,20):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279651/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Environmental',"Pharma & vet compounds & metabolites", products)
for page in range(0,4):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279653/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Environmental',"Mycotoxins", products)
for page in range(0,2):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279644/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Environmental',"Additional organic reference materials", products)
for page in range(0,7):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279654/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Environmental',"Polycyclic aromatic hydrocarbons (PAHs)", products)
for page in range(0,5):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279655/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Environmental',"Volatile organic compounds (VOCs)", products)
for page in range(0,8):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279656/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Environmental',"Hydrocarbons & petrochemicals", products)
for page in range(0,10):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279657/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Environmental',"Standards for environmental regulatory methods", products)
for page in range(0,14):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279663/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Environmental',"Aqueous inorganic", products)
for page in range(0,2):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/280826/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Environmental',"Environmental proficiency testing", products)
for page in range(0,3):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279674/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Environmental',"Physical properties", products)
for page in range(0,47):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279696/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Environmental',"Pesticides & metabolites", products)

for page in range(0,40):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279798/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Forensic & Toxicology',"New psychoactive substances (NPS)", products)
for page in range(0,7):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279811/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Forensic & Toxicology',"Sports drugs & steroids", products)
for page in range(0,4):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279812/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Forensic & Toxicology',"Ethanol & related materials", products)
getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/324096/products?pageSize=20&fields=FULL&sort=code-asc&currentPage=0&q=&country=US&lang=en&defaultB2BUnit=",'Forensic & Toxicology',"Cannabinoids", products)
for page in range(0,4):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279770/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Forensic & Toxicology',"Smoking-related substances", products)
for page in range(0,10):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279768/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Forensic & Toxicology',"Opiates & opioids", products)
for page in range(0,10):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279762/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Forensic & Toxicology',"Amphetamines", products)
for page in range(0,5):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279764/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Forensic & Toxicology',"Benzodiazepines", products)
for page in range(0,2):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279765/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Forensic & Toxicology',"Cocaine & related materials", products)
for page in range(0,3):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279767/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Forensic & Toxicology',"Hallucinogens", products)
getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279766/products?pageSize=20&fields=FULL&sort=code-asc&currentPage=0&q=&country=US&lang=en&defaultB2BUnit=",'Forensic & Toxicology',"GHB & related materials", products)
for page in range(0,2):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279763/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Forensic & Toxicology',"Barbiturates", products)
for page in range(0,84):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/324180/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Forensic & Toxicology',"Pharmaceutical toxicology", products)
for page in range(0,2):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279771/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=20&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Forensic & Toxicology',"Z-drugs", products)

getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/324099/products?pageSize=20&fields=FULL&sort=code-asc&currentPage=0&q=&country=US&lang=en&defaultB2BUnit=",'Forensic & Toxicology',"Drug metabolism", products)
for page in range(0,3):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/324100/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Forensic & Toxicology',"Enviromental toxicology", products)
for page in range(0,4):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/324102/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=20&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Forensic & Toxicology',"Food mutagens", products)
for page in range(0,2):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/324101/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Forensic & Toxicology',"Mycotoxins", products)

getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/280823/products?pageSize=20&fields=FULL&sort=code-asc&currentPage=0&q=&country=US&lang=en&defaultB2BUnit=",'Industrial',"Proficiency testing", products)
for page in range(0,13):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279705/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Industrial',"Petroleum reference materials", products)
for page in range(0,4):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279708/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Industrial',"Metallo-organic standards", products)
for page in range(0,4):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279709/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Industrial',"Sulfur, Nitrogen & Chlorine", products)
for page in range(0,2):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279711/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=20&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Industrial',"Petroleum physical test standards", products)
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279713/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=20&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Industrial',"Matrix oils & solvents", products)
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279714/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=20&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Industrial',"Biodiesel standards", products)
for page in range(0,13):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279715/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Industrial',"Aqueous inorganic", products)
for page in range(0,5):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279726/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Industrial',"Laboratory consumables", products)
for page in range(0,88):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279733/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Industrial',"Metal alloys", products)
for page in range(0,20):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279736/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Industrial',"Aluminium base", products)
for page in range(0,37):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279744/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Industrial',"Process materials, geological, cement & soils", products)
for page in range(0,3):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279750/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Industrial',"XRF monitor glasses", products)
for page in range(0,4):
	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279757/products?currentPage="+str(page)+"&q=&sort=relevance&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Industrial',"Equipment for sample preparation", products)


headers=[
	'link', 'type1','type2','Product Name','Synonyms','Product Code','CAS Number','Product Format','Matrix','Molecular Formula',
	'Molecular Weight','API Family','Product Categories','Product Type','Accurate Mass','Smiles','InChI','IUPAC','Storage Temperature',
	'Shipping Temperature','Country of Origin'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
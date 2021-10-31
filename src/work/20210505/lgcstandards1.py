from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		request_obj=urllib.request.Request(url=url)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		if str(html_code).find("403 ERROR")>-1:
			time.sleep(360)
			return getHtmlFromUrl(url)
		else:
			return html_code
	except:
		retryCount += 1
		if retryCount < 5:
			print("retry index"+str(retryCount)+url)
			time.sleep(360)
			return getHtmlFromUrl(url)
		else:
			retryCount = 0
			return ""
def getRenderdHtmlFromUrl(url, isTry):
	global retryCount
	try:
		chrome_options = webdriver.ChromeOptions()
		chrome_options.add_argument('--headless')
		chrome_options.add_argument('--disable-gpu')
		chrome_options.add_argument("window-size=1024,768")

		chrome_options.add_argument("--no-sandbox")
		browser = webdriver.Chrome(chrome_options=chrome_options)
		browser.get(url)
		if str(browser.page_source).find("403 ERROR")>-1:
			time.sleep(360)
			return getRenderdHtmlFromUrl(url, True)
		else:
			return browser.page_source
	except:
		retryCount += 1
		if retryCount < 5:
			print("retry index"+str(retryCount)+url)
			time.sleep(360)
			return getRenderdHtmlFromUrl(url, True)
		else:
			retryCount = 0
			return ""
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, type1, type2, products):
	print(str(len(products)) + url)
	productListHtml = getRenderdHtmlFromUrl(url, False)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	pInfo = {
		"link": url,
		"type1": type1,
		"type2": type2
	}
	try:
		syncDom=sope.find_all("p", attrs={"class":"product__synonyms show-more-content"})
		print(syncDom)
		titleArea = sope.find("div", attrs={"class": "product__title-wrapper"})
		navArea = sope.find("ul", attrs={"class": "breadcrumb outline"})
		if titleArea != None:
			pInfo["Product Name"] = getNodeText(titleArea.find("h1"))
			pInfo["Synonyms"] = getNodeText(titleArea.find("p"))
			pInfo["nav"] = getNodeText(navArea)

			specAreas = sope.find_all("div",attrs={"class":"product__item"})
			for specArea in specAreas:
				title = getNodeText(specArea.find("h2"))
				val = getNodeText(specArea.find("p"))
				if title == "API Family":
					val = getNodeText(specArea.parent)
				if len(title) > 0:
					pInfo[title]=val
			
			img = sope.find("img", attrs={"itemprop":"image"})
			if img != None:
				if "CAS Number" in pInfo and len(pInfo["CAS Number"])>0:
					urllib_download(img["src"], pInfo["CAS Number"])
				else:
					urllib_download(img["src"], pInfo["Product Name"])
			products.append(pInfo.copy())
	except:
		pInfo={
			"nav":"出错了"
		}
		products.append(pInfo.copy())

def getProductList(url, type1, type2, products):
	productListHtml = getHtmlFromUrl(url)
	try:
		data = json.loads(productListHtml)
		hits = data["products"]
		for hit in hits:
			link = hit["url"]
			getProductInfo("https://www.lgcstandards.com/US/en"+link, type1, type2, products)
	except:
		pInfo = {
			"link":url,
			"type1":"读取出错"
		}
		products.append(pInfo)

excelFileName="lgcstandards.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

getProductInfo("https://www.lgcstandards.com/US/en/Dexamethasone-Sodium-Phosphate/p/MM0210.00", "111", "222", products)
# getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/322963/products?currentPage=1&q=&sort=relevance-code&pageSize=10&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical',"COVID-19 research and reference materials", products)

# for page in range(0,13):
# 	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/322963/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical',"COVID-19 research and reference materials", products)
# for page in range(0,16):
# 	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/323809/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','API standards', products)
# for page in range(0,72):
# 	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/323810/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit",'Pharmaceutical','Impurity standards', products)
# for page in range(0,42):
# 	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/324176/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Pharmacopoeial standards', products)
# for page in range(0,31):
# 	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279516/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','European Pharmacopoeia (Ph. Eur.)', products)
# for page in range(0,19):
# 	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279513/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','British Pharmacopoeia', products)
# for page in range(0,4):
# 	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279517/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Reagents according to pharmacopoeias', products)
# for page in range(0,2):
# 	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279526/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=20&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','International reference standards for antibiotics (WHO)', products)
# for page in range(0,17):
# 	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/323817/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Building blocks', products)
# for page in range(0,30):
# 	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/324179/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Enzyme activators, inhibitors & substrates', products)
# for page in range(0,5):
# 	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/323822/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Neurochemicals', products)
# for page in range(0,11):
# 	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/323824/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Carbohydrates', products)
# for page in range(0,2):
# 	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/323827/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Cross-linkers', products)
# for page in range(0,28):
# 	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/323816/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Chiral molecules', products)
# for page in range(0,81):
# 	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/323830/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Pharmaceutical toxicology', products)
# for page in range(0,7):
# 	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279509/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Elemental impurities', products)
# for page in range(0,2):
# 	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279527/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=100&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Physical properties', products)
# for page in range(0,2):
# 	getProductList("https://www.lgcstandards.com/US/en/lgcwebservices/lgcstandards/categories/279507/products?currentPage="+str(page)+"&q=&sort=relevance-code&pageSize=20&country=US&lang=en&fields=FULL&defaultB2BUnit=",'Pharmaceutical','Pharmaceutical proficiency testing', products)


headers=[
	'link','nav', 'type1','type2','Product Name','Synonyms','Product Code','CAS Number','Product Format','Matrix','Molecular Formula',
	'Molecular Weight','API Family','Product Categories','Product Type','Accurate Mass','Smiles','InChI','IUPAC','Storage Temperature',
	'Shipping Temperature','Country of Origin'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
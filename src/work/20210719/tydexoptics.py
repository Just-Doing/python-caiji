import requests
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import string
import math
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		request_obj=urllib.request.Request(url=url,  headers={
			'Content-Type': 'text/html; charset=utf-8',
			"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"
		})
		htmlHeader = requests.head(url)
		if htmlHeader.status_code ==200:
			response_obj=urllib.request.urlopen(request_obj)
			html_code=response_obj.read()
			return html_code
		else:
			return ''
	except:
		retryCount = retryCount + 1
		if retryCount < 5:
			print("retry index"+str(retryCount)+url)
			time.sleep(60)
			return getHtmlFromUrl(url)
		else:
			retryCount = 0
			return ""

def requestJson(url):
	r = requests.post(url, headers={
		'Content-Type': 'application/x-www-form-urlencoded',
		'cookie':'visid_incap_2255650=4oBBaRPnQfCVoYEiTmjTq/NVAWEAAAAAQUIPAAAAAAD69PQHUoB0KplKq7/j0+gH; nlbi_2255650=CJKhHYlMm17tpKyoBzOViAAAAACDEjp3gL6bj6YL8j9XE0d/; incap_ses_893_2255650=m1tJIuDRUEp3FE/5GpNkDPRVAWEAAAAAM2KkDpvtARtZral+cMXSVw==; _gcl_au=1.1.76703404.1627477493; _gid=GA1.2.730047202.1627477493; BCSessionID=83af10b8-9488-4b7b-a3b1-3640f178dca2; categoryView=grid; _ga_S46FST9X1M=GS1.1.1627477492.1.1.1627478562.0; _ga=GA1.2.31731397.1627477493; _gat_UA-139934-1=1; _uetsid=69fc2d30efa411eb8818eb045f8760e5; _uetvid=69fc3a70efa411ebba3a23c153f6e477; .Nop.Customer=d664d529-d14a-44b1-86b3-cbf5373277b4',
		"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"
	})
	datas = json.loads(r.text)
	return datas

def getRenderdHtmlFromUrl(url):

	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	return BeautifulSoup(browser.page_source, "html.parser",from_encoding="utf-8")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, typeStr, products):
	print(str(len(products)) + url)
	html_code = getHtmlFromUrl(url)
	if len(html_code)>0:
		sope= BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
		pName = sope.find("h1", attrs={"itemprop":"name"})
		specInfos = sope.find_all("h5")
		Description = sope.find("div", attrs={"class":"full-description"})
		pInfo = {
			"link": url,
			"Product Category1": 'Fiber Optic',
			"Product Category2": typeStr,
			"Product Name": getNodeText(pName),
			"Description": getNodeText(Description)
		}
		
		for specInfo in specInfos:
			title = getNodeText(specInfo)	
			if title == "Features":
				pInfo["Features"] = getNodeText(specInfo.next_sibling.next_sibling)	
			if title == "Application":
				pInfo["Application"] = getNodeText(specInfo.next_sibling.next_sibling)
		products.append(pInfo.copy())
	

def getProductList(url, typestr, products):
	html_code = getHtmlFromUrl(url)
	if len(html_code)>0:
		sope= BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
		pLinkArea = sope.find("div", attrs={"class":"page-inner clearfix"})
		pLinks = pLinkArea.find_all("a")
		for Plink in pLinks:
			print(Plink)

def getProducType(url, products):
	html_code = getHtmlFromUrl(url)
	if len(html_code)>0:
		sope= BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
		typeArea = sope.find("li", attrs={"class":"active dropdown"})
		types = typeArea.find_all("li", attrs={"class":"dropdown-submenu"})
		for type in types:
			lLink = type.find("a")
			getProductList(lLink["href"], getNodeText(lLink), products)

excelFileName="lcom.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("http://www.tydexoptics.com/products/optics_for_detectors_and_sensors/", '', products)
getProductList('http://www.tydexoptics.com/products/spectroscopy/','', products)
# getProductInfo("http://www.tydexoptics.com/products/optics_for_detectors_and_sensors/", '', products)
# getProductInfo("http://www.tydexoptics.com/products/optics_for_meteorology_and_climatology/" '', products)
# getProductInfo("http://www.tydexoptics.com/products/libs/", '', products)
# getProductInfo("http://www.tydexoptics.com/products/atypical_components/", '', products)

headers=[
	'link','Product Category1','Product Category2','Product Name','Features','Application','Description'
]

for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
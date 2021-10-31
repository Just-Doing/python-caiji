from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
    
	url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
	request_obj=urllib.request.Request(url=url)
	response_obj=urllib.request.urlopen(request_obj)
	html_code=response_obj.read()
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	
def getRenderdHtmlFromUrl(url):

	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	return BeautifulSoup(browser.page_source, "html.parser",from_encoding="utf-8")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, typeStr, size, products):
	print(str(len(products)) + url)
	sope = getHtmlFromUrl(url, False)
	nav = sope.find("div", attrs={"class": "breadcrumbs"})
	pInfo = {
		"link": url,
		"type": typeStr,
		"Sizes Available":size,
		"nav": getNodeText(nav).replace("\r","").replace("\n","")
	}
	
	specInfos = sope.find_all("tr")
	for specInfo in specInfos:
		ths = specInfo.find_all("th")
		tds = specInfo.find_all("td")
		if len(ths)==1 and len(tds)==1:
			title = getNodeText(ths[0])
			val = getNodeText(tds[0])
			valLink = tds[0].find("a")
			if valLink!=None:
				val=val+":"+valLink["href"]
			pInfo[title] = val
	description=sope.find("p", attrs={"itemprop":"description"})
	pInfo["Product Description"] = getNodeText(description)
	products.append(pInfo.copy())
	

def getProductList(url,type, products):
	sope = getHtmlFromUrl(url)
	pList = sope.find_all("li", attrs={"class":"item product product-item"})
	for p in pList:
		pLink = p.find("a")
		size = p.find("div", attrs={"class":"sizes"})
		getProductInfo(pLink["href"], type, getNodeText(size), products)


excelFileName="biorbyt.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("https://www.biorbyt.com/maltose-microplate-assay-kit-orb759231.html", "111", "222", products)

# getProductList("https://www.biorbyt.com/kits/assay-kits.html?categoryproduct=529&p=2&product_list_limit=50", 'ELISA Kits', products)

getProductList("https://www.biorbyt.com/kits/assay-kits.html?categoryproduct=526&product_list_limit=50",'ELISA Kits', products)

for page in range(1,145):
	getProductList("https://www.biorbyt.com/kits/assay-kits.html?categoryproduct=529&p="+str(page)+"&product_list_limit=50",'ELISA Kits', products)

for page in range(1,45):
	getProductList("https://www.biorbyt.com/kits/assay-kits.html?categoryproduct=531&p="+str(page)+"&product_list_limit=50",'Assay Kits', products)

getProductList("https://www.biorbyt.com/kits/assay-kits.html?categoryproduct=533&product_list_limit=50",'Small molecules', products)

headers=[
	'link','type','nav','Product Name','Sensitivity','Alternative Names',
	'Uniprot ID','Entrez','Catalog Number','Reactivity','Range','Tested applications','Target','Storage','Note','Detection Method','Product Description','Application Notes', 'Sizes Available'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
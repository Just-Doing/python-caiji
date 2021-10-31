from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
    
	url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
	request_obj=urllib.request.Request(url=url)
	response_obj=urllib.request.urlopen(request_obj)
	html_code=response_obj.read()
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")
	
def getRenderdHtmlFromUrl(url):

	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	return BeautifulSoup(browser.page_source, "html.parser",from_encoding="utf-8")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, type1, type2, products):
	print(str(len(products)) + url)
	sope = getHtmlFromUrl(url, False)
	pName = sope.find("h1", attrs={"class":"productView-title"})
	sku = sope.find("dd", attrs={"itemprop":"sku"})
	weight = sope.find( lambda tag:tag.has_attr('data-product-weight'))
	Description = sope.find("div", attrs={"class":"Description_Description"})
	pInfo = {
		"link": url,
		"Product Category1": type1,
		"Product Category2": type2,
		"Product Name": getNodeText(pName),
		"sku": getNodeText(sku),
		"weight": getNodeText(weight),
		"Description": getNodeText(Description)
	}
	
	products.append(pInfo.copy())
	

def getProductList(url, type1, type2, products):
	sope = getHtmlFromUrl(url)
	pList = sope.find_all("li", attrs={"class":"product"})
	for p in pList:
		pLink = p.find("a")
		getProductInfo(pLink["href"], type1, type2, products)

def getTypeList(url, products):
	sope = getHtmlFromUrl(url)

excelFileName="cablesforless.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductList("https://www.cablesforless.com/fiber-optic/multimode-fiber-optic-cables/",'Fiber Optic','Fiber Optic DIY', products)
# getProductInfo("https://www.cablesforless.com/1-meter-fiber-multimode-62-5-125-duplex-st-st/", "111", "222", products)

getProductList("https://www.ipgphotonics.com/en/products/lasers",'lasers', products)
getProductList("https://www.ipgphotonics.com/en/products/beam-delivery",'beam delivery', products)
getProductList("https://www.ipgphotonics.com/en/applications/medical",'medical', products)
getProductList("https://www.ipgphotonics.com/en/products/telecom-equipment",'telecom equipment', products)
getProductList("https://www.ipgphotonics.com/en/products/components",'components', products)

headers=[
	'link','Product Category1','Product Category2','Product Name','sku','weight','Description'
]
for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
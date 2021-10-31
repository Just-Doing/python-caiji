from urllib.request import urlopen
from selenium import webdriver
import urllib
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy

http.client._MAXHEADERS = 1000


retryCount = 0
loadCount = 0
def urllib_download(IMAGE_URL, imageName):
	global retryCount
	try:
		from urllib.request import urlretrieve
		urlretrieve(IMAGE_URL, imageName)   
	except:
		print("重试图片下载"+IMAGE_URL)
		retryCount += 1
		if(retryCount <= 5):
			urllib_download(IMAGE_URL, imageName)
		else:
			retryCount=0
			return None
		
def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

def getHtmlFromUrl(url):
	global retryCount
	try:
		html = urlopen(url).read()
		return BeautifulSoup(html, "html.parser",from_encoding="utf-8")
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getHtmlFromUrl(url)
		else:
			retryCount=0
			return None
		
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1


def getProductInfo(url, products):
	print(str(len(products)) + url)
	pInfo={"link": url}
	scope = getHtmlFromUrl(url)
	if scope != None:
		descriptArea = scope.find(name="div", attrs={"class":"descriptionContent"})
		for desInfo in descriptArea.find_all("h4"):
			title = getNodeText(desInfo)
			value = getNodeText(desInfo.next_sibling.next_sibling)
			if title=="General description":
				pInfo["Generaldescription"]=value
			if title=="Application":
				pInfo["Application"]=value
			if title=="Biochem/physiol Actions":
				pInfo["BiochemphysiolActions"]=value
		
		prodInfoArea = scope.find(name="div", attrs={"class":"productInfo"})
		infos = prodInfoArea.find_all("li")
		for info in infos:
			value = getNodeText(info)
			if value.find("CAS Number") > -1:
				pInfo["CAS"] = value.replace("CAS Number ","")
			if value.find("Empirical Formula (Hill Notation)") > -1:
				pInfo["EmpiricalFormula"] = value.replace("Empirical Formula (Hill Notation) ","")
			if value.find("Molecular Weight") > -1:
				pInfo["MolecularWeight"] = value.replace("Molecular Weight ","")
			if value.find("NACRES") > -1:
				pInfo["NACRES"] = value.replace("NACRES ","")
		# pInfo["name"]= getNodeText(scope.find(name="p", attrs={"class":"product-name"}))
		# urllib_download("https://www.sigmaaldrich.com"+scope.find(name="a", attrs={"class":"thumbnail"})["href"], pInfo["name"].replace('(','').replace(')','').replace('/','').replace(':','-')+'.jpg')
		
		specArea = scope.find(name="div", attrs={"id":"productDetailProperties"})
		trSpecifications = specArea.find_all("tr")
		for info in trSpecifications:
			title = getNodeText(info.find("td"))
			value = getNodeText(info.find_all("td")[1])
			if(title == "Related Categories"):
				pInfo["RelatedCategories"]=value
			if(title == "packaging"):
				pInfo["packaging"]=value
			if(title == "mfr. no."):
				pInfo["mfrno"]=value
			if(title == "concentration"):
				pInfo["concentration"]=value
			if(title == "shipped in"):
				pInfo["shippedin"]=value
			if(title == "storage temp."):
				pInfo["storagetemp"]=value
			if(title == "SMILES string"):
				pInfo["SMILESstring"]=value
		
		products.append(pInfo.copy())

def getProductList(url, products):
	sope = getHtmlFromUrl(url)
	proLinks = sope.find_all(name="a", attrs={"class":"woocommerce-LoopProduct-link woocommerce-loop-product__link" })
	for link in proLinks:
		prodInfo = {}
		prodInfo["name"] = getNodeText(link.find("h2"))
		getProductInfo(link["href"], prodInfo, products)


excelFileName="nanosoftpolymers.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
urls=[
	"https://www.sigmaaldrich.com/catalog/product/avanti/610000?lang=zh&region=CN",
	"https://www.sigmaaldrich.com/catalog/product/avanti/890895c?lang=zh&region=CN",
	"https://www.sigmaaldrich.com/catalog/product/avanti/890895p?lang=zh&region=CN",
	"https://www.sigmaaldrich.com/catalog/product/avanti/890890c?lang=zh&region=CN",
	"https://www.sigmaaldrich.com/catalog/product/avanti/890890p?lang=zh&region=CN",
	'https://www.sigmaaldrich.com/catalog/product/avanti/890880c?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890880p?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890870c?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890870p?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890860c?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890860p?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890000p?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890893p?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890898c?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890898p?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/850310c?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/850310p?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890850c?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890850o?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890820c?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890820p?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890830c?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890830p?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890840c?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890840c?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890840p?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890704p?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890705c?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890705p?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890704c?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890703c?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890703p?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890703p?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890703p?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890717o?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890701c?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890701p?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890700p?lang=zh&region=CN',
	'https://www.sigmaaldrich.com/catalog/product/avanti/890700c?lang=zh&region=CN'
]

for url in urls:
	getProductInfo(url, products)
	
headers=[
	'link','CAS','EmpiricalFormula','MolecularWeight','NACRES','RelatedCategories','packaging','mfrno','concentration','shippedin','storagetemp',
	'SMILESstring',"Generaldescription",'Application','BiochemphysiolActions'
]
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
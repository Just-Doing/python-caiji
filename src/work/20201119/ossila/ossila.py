from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		from urllib.request import urlretrieve
		urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')   
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"}

		request_obj=urllib.request.Request(url=url,headers=headers)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1


def getProductInfo(url, products):
	print(str(len(products)) + url)
	productHtml = getRenderdHtmlFromUrl(url)
	if productHtml != None:
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
		
		name = sope.find(name="h1", attrs={"id": "product-title-label"})
		imgArea = sope.find(name="div", attrs={"class": "img-responsive text-center"})
		if imgArea != None:
			img = imgArea.find("img")
			urllib_download("https:"+img["src"], getNodeText(name))
			
		pInfo ={"link":url} 
		
		pInfo["name"] = getNodeText(name)
		
		size = sope.find(name="select", attrs={"class": "single-option-selector"})
		pInfo["size"] = getNodeText(size)
		h2Titles = sope.find_all("h2")
		for h2title in h2Titles:
			if getNodeText(h2title) == "General Information":
				infoTable = h2title.findNextSibling("table")
				trs = infoTable.find_all("tr")
				for tr in trs:
					tds = tr.find_all("td")
					if len(tds) > 1:
						pInfo[getNodeText(tds[0])] = getNodeText(tds[1])
			if getNodeText(h2title) == "Batch information" or getNodeText(h2title) == "Batch Details" or getNodeText(h2title) == "Batch details" or getNodeText(h2title) == "Batch Information":
				infoTable = h2title.findNextSibling("table")
				index_rr = 0
				index_mw = 0
				index_mn = 0
				rr = "-"
				mw = "-"
				mn = "-"
				titleTr = infoTable.find("tr")
				titleTds = titleTr.find_all("td")
				for index,infoTd in enumerate(titleTds):
					if getNodeText(infoTd) == "RR":
						index_rr = index
					if getNodeText(infoTd) == "Mw" or getNodeText(infoTd) == "MW":
						index_mw = index
					if getNodeText(infoTd) == "Mn" or getNodeText(infoTd) == "MN":
						index_mn = index
				infoTrs = infoTable.find_all("tr")
				del infoTrs[0]
				for index,infoTr in enumerate(infoTrs):
					tds = infoTr.find_all("td")
					if index_rr > 0:
						rr = getNodeText(tds[index_rr])
					if index_mw > 0:
						mw = getNodeText(tds[index_mw])
					if index_mn > 0:
						mn = getNodeText(tds[index_mn])
					pInfo["batch"+str(index+1)] = getNodeText(tds[0]) +"/"+ rr+"/"+mw+"/"+mn
		print(pInfo)
		products.append(pInfo.copy())
		
def getProductList(url, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	prods = sope.find_all("div", attrs={"class":"col-xs-12 col-sm-4 col-md-3"})
	
	for pro in prods:
		getProductInfo("https://www.ossila.com/" + pro.find("a")["href"], products )
		

excelFileName="riyngroup.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
links = [
	'https://www.ossila.com/collections/ofet-and-oled-polymers',
	'https://www.ossila.com/collections/polymer-donors',
	'https://www.ossila.com/collections/polymer-acceptors',
	'https://www.ossila.com/collections/interface-polymers',
	'https://www.ossila.com/collections/semiconducting-polymers',
	'https://www.ossila.com/collections/perovskite-inks',
	'https://www.ossila.com/collections/perovskite-precursor-materials',
	'https://www.ossila.com/collections/perovskite-interface-materials',
	'https://www.ossila.com/collections/dssc-materials',
	'https://www.ossila.com/collections/transport-layer-materials',
	'https://www.ossila.com/collections/dopant-materials',
	'https://www.ossila.com/collections/host-materials',
	'https://www.ossila.com/collections/host-materials',
	'https://www.ossila.com/collections/semiconducting-molecules',
	'https://www.ossila.com/collections/thiophene-tt-bdt-bdd-fl-cz-monomers',
	'https://www.ossila.com/collections/btd-bta-dpp-qx-tpd-monomers',
	'https://www.ossila.com/collections/naphthalenediimide-monomers',
	'https://www.ossila.com/collections/nfa-monomers',
	'https://www.ossila.com/collections/monomers'
]
# getProductInfo('https://www.ossila.com/products/f8bt?variant=20155499675744', products)
for link in links:
	getProductList(link, products)
headers=['link','name','size','Full name','Synonyms','CAS number','Chemical formula','Molecular weight','HOMO / LUMO','Solubility','Classification / Family',
'batch1','batch2','batch3','batch4','batch5','batch6','batch7','batch8','batch9','batch10'
]
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
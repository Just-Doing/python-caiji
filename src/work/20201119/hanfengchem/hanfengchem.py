from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000


def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

def urllib_download(IMAGE_URL, pName):
    from urllib.request import urlretrieve
    urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')   
	
retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"}

		request_obj=urllib.request.Request(url=url,headers=headers)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getHtmlFromUrl(url)
		else:
			retryCount=0
			return None
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1


def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	productHtml = getHtmlFromUrl(url)
	if productHtml != None:
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
		
		pInfo["link"] = url
		attrArea = sope.find(name="div", attrs={"class": "about_r_con"})
		img = attrArea.find("img")
		# urllib_download("http://www.hanfengchem.com/"+img["src"], pInfo["name"])
		
		attrInfos = attrArea.find_all("span")
		for attrName in attrInfos:
			title = getNodeText(attrName)
			if title == "Description":
				pInfo["description"] = attrName.nextSibling.nextSibling
			if title.find("Formula") > -1:
				pInfo["Formula "] = title
			if title.find("MW (g/mole)") > -1:
				pInfo["MW (g/mole)"] = title
			if title.find("TG (℃)") > -1:
				pInfo["TG (℃)"] = title
			if title.find("TGA (℃, 0.5% weight loss)") > -1:
				pInfo["TGA (℃, 0.5% weight loss)"] = title
			if title.find("Absorption (nm, THF)") > -1:
				pInfo["Absorption (nm, THF)"] = title
			if title.find("Photoluminescene (nm, THF)") > -1:
				pInfo["Photoluminescene (nm, THF)"] = title
			if title.find("EL device") > -1:
				pInfo["EL device"] = title
		specList = attrArea.find_all("li")
		for spec in specList:
			specValue = getNodeText(spec)
			specPart = specValue.split(":")
			if len(specPart) == 2:
				title = specPart[0]
				specVal = specPart[1]
				pInfo[title] = specVal
		print(pInfo)
		products.append(pInfo.copy())
				

def getProductList(url, pInfo, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	productList = sope.find_all("div", attrs={"class":"pro_pic"})
	for pro in productList:
		link = pro.find("a")
		pInfo["name"] = getNodeText(link)
		getProductInfo("http://www.hanfengchem.com/en/"+link["href"], pInfo, products)


excelFileName="hanfengchem.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
links = [
	'http://www.hanfengchem.com/en/product-201308271625435468.html',
	'http://www.hanfengchem.com/en/product-201308271625542187.html',
	'http://www.hanfengchem.com/en/product-201308271626424375.html',
	'http://www.hanfengchem.com/en/product-201308271626552031.html',
	'http://www.hanfengchem.com/en/product-201308271627073125.html',
	'http://www.hanfengchem.com/en/product-201308271627177812.html',
	'http://www.hanfengchem.com/en/product-201308271627554843.html',
	'http://www.hanfengchem.com/en/product-201309121157087656.html',
	'http://www.hanfengchem.com/en/product-201309121157190625.html',
	'http://www.hanfengchem.com/en/product-201309121157301093.html',
	'http://www.hanfengchem.com/en/product-201206221148571636.html',
	'http://www.hanfengchem.com/en/product-201206221149116431.html'
]
index = 0
for link in links:
	pInfo={}
	if index == 0:
		pInfo["type"] = "OLED materials"
		pInfo["type2"] = "Hole Blocking Layer"
	if index == 1:
		pInfo["type"] = "OLED materials"
		pInfo["type2"] = "Hole Injection Layer"
	if index == 2:
		pInfo["type"] = "OLED materials"
		pInfo["type2"] = "Hole Transport Layer"
	if index == 3:
		pInfo["type"] = "OLED materials"
		pInfo["type2"] = "Electron Injection Layer"
	if index == 4:
		pInfo["type"] = "OLED materials"
		pInfo["type2"] = "Electron Transport Layer"
	if index == 5:
		pInfo["type"] = "OLED materials"
		pInfo["type2"] = "Fluorescent Host"
	if index == 6:
		pInfo["type"] = "OLED materials"
		pInfo["type2"] = "Phosphorescent Host"
	if index == 7:
		pInfo["type"] = "OLED materials"
		pInfo["type2"] = "Green Dopant"
	if index == 8:
		pInfo["type"] = "OLED materials"
		pInfo["type2"] = "Blue Dopant"
	if index == 9:
		pInfo["type"] = "OLED materials"
		pInfo["type2"] = "Red Dopant"
	if index == 10:
		pInfo["type"] = "OTFT / OFET / Materials"
	if index == 11:
		pInfo["type"] = "OPV Cell Materials"
	
	getProductList(link, pInfo, products)
	index = index + 1

# getProductInfo('http://www.hanfengchem.com/en/product_show_558.html',{"name":"test"}, products)
headers=['link', 'type','type2', 'name', 'description','Formula ','Weight ','CAS No. ','Thermal Gravimetric Analysis ','Absorption ','Photoluminescence ','Grade ',
'MW (g/mole)', 'TG (℃)','TGA (℃, 0.5% weight loss)','Absorption (nm, THF)','Photoluminescene (nm, THF)','EL device']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
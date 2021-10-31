from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy
import string
from selenium.webdriver.support.select import Select
import time

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		from urllib.request import urlretrieve
		urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')   
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"}

		request_obj=urllib.request.Request(url=url,headers=headers)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url,isAll):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	chrome_options.add_argument("--proxy-server=http://127.0.0.1:7890")
	
	browser.get(url)
	time.sleep(5)
	if isAll == 1:
		selectControl = browser.find_element_by_name("perPage")
		Select(selectControl).select_by_value('all')
	
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1


def getProductInfo(url, products):
	print(str(len(products)) + url)
	productHtml = getHtmlFromUrl(url)
	if productHtml != None:
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
		pInfo = {"link": url}
		headerArea = sope.find(name="div", attrs={"class": "product-post__header__info grid"})
		name = headerArea.find("div", attrs={"class":"grid-1of1"})
		pInfo["name"] = getNodeText(name)
		infos = headerArea.find_all("td")
		pInfo["Clone"] = getNodeText(infos[0])
		pInfo["cat"] = getNodeText(infos[1])
		pInfo["Category"] = getNodeText(infos[2])
		detailArea = sope.find("section", attrs={"class":"product-post__details"})
		infoTrs = detailArea.find_all("tr")
		for infoTr in infoTrs:
			tds = infoTr.find_all("td")
			if len(tds) == 2:
				pInfo[getNodeText(tds[0])] = getNodeText(tds[1])
		references = sope.find("section", attrs={"id":"references"})
		if references!=None:
			refDetails = references.find_all("p")
			refDetailCount =0
			for refDetail in refDetails:
				if refDetailCount < 6:
					titleInfo = refDetail.find("strong")
					if titleInfo!=None:
						refDetailCount = refDetailCount + 1
						pInfo["referencesTitle"+str(refDetailCount)] = getNodeText(titleInfo)
						pInfo["referencesInfo"+str(refDetailCount)] = getNodeText(refDetail.nextSibling.nextSibling)
		print(pInfo)
		products.append(pInfo.copy())
		
def getProductList(url, products, isAll):
	productListHtml = getRenderdHtmlFromUrl(url, isAll)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	productArea = sope.find("section", attrs={"class":"products__list grid"})
	productList = productArea.find_all("div",attrs={"class":"product grid-1of2 grid-1of1--notebook-small"})
	for linkArea in productList:
		link = linkArea.find("a")
		print(link)
		getProductInfo("https://bxcell.com"+link["href"], products)

excelFileName="bxcell.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
# getProductInfo("https://bxcell.com/product/m-cd3e/", products)
getProductList('https://bxcell.com/shop-products/brand/invivoplus-antibodies/', products, 1)
getProductList('https://bxcell.com/shop-products/brand/invivomab-antibodies/', products, 1)
getProductList('https://bxcell.com/shop-products/brand/readytag-antibodies/', products, 0)
headers=['link','name','Clone','cat','Category','Isotype','Immunogen','Reported Applications','Formulation','Endotoxin','Purity','Purification',
'referencesTitle1','referencesInfo1','referencesTitle2','referencesInfo2','referencesTitle3','referencesInfo3','referencesTitle4','referencesInfo4','referencesTitle5','referencesInfo5'
]
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
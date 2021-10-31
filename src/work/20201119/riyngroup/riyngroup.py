from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		from urllib import request
		from request import urlretrieve
		opener = request.build_opener()
		opener.addheaders = ([("User-Agent","Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36")])
		request.install_opener(opener)
		urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')   
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"}

		request_obj=urllib.request.Request(url=url,headers=headers)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1


def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	productHtml = getHtmlFromUrl(url)
	if productHtml != None:
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
		
		pInfo["link"] = url
		name = sope.find(name="div", attrs={"itemprop": "name"})
		attrArea = sope.find(name="div", attrs={"id": "tagContent"})
		attrTr = attrArea.find_all("tr")
		for tr in attrTr:
			# img = tr.find("img")
			# if img!=None:
				# urllib_download(img["src"], getNodeText(name))
			tds = tr.find_all("td")
			if len(tds) == 2:
				title = getNodeText(tds[0]).replace(":","")
				value = getNodeText(tds[1])
				pInfo[title] = value
			if len(tds) == 3:
				title = getNodeText(tds[0])
				value = getNodeText(tds[2])
				pInfo[title] = value
				
		products.append(pInfo.copy())
def getProductList(url, pInfo, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	productArea = sope.find("div", attrs={"itemprop":"description"})
	prods = productArea.find_all("div", attrs={"class":"box"})
	for pro in prods:
		getProductInfo("http://www.riyngroup.com" + pro.find("a")["href"], pInfo, products )
		
def getProductListPage(url,pInfo, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	pageArea = sope.find("div", attrs={"class":"pageNum"})
	if pageArea!=None:
		linkAreas = pageArea.find_all("a")
		pageSize = len(linkAreas)
		if pageSize ==0:
			pageSize = 1
		pageIndex = 1
		while pageIndex <= pageSize:
			getProductList(url+"/p/"+str(pageIndex), pInfo, products)
			pageIndex = pageIndex+1

def getProductType(url, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	productTypeArea = sope.find("li", attrs={"class":"box active"})
	if productTypeArea != None:
		linkAreas = productTypeArea.find_all("li")
		for linkArea in linkAreas:
			link = linkArea.find("a")
			pInfo={
				"type":getNodeText(link)
			}
			getProductListPage("http://www.riyngroup.com/"+link["href"], pInfo, products)

excelFileName="riyngroup.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
# getProductListPage('http://www.riyngroup.com//boric-acid.html/p/1',{}, products)
# getProductInfo('http://www.riyngroup.com/1-bromopyrene-1714-29-0-c16h9br-15396534502369548.html',{}, products)
getProductType("http://www.riyngroup.com/carbazole.html", products)
headers=['link','Name','Aliase','Molecular Structure','Molecular Formula','Molecular Weight','CAS No.','Synonyms','Formula','Exact Mass','PSA','Purity','Properties','Boiling Poin','Density','Flash Point',
	'CAS Registry Number','Melting point','Hazard Symbols','Risk Codes', 'Safety Description','MSDS'
]
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
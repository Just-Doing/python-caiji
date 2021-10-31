from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')   
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"}

		request_obj=urllib.request.Request(url=url,headers=headers)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)


def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	productHtml = getHtmlFromUrl(url)
	if productHtml != None:
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
		
		pInfo["link"] = url
		attrArea = sope.find(name="dl", attrs={"id": "content_specifications"})
		atrs = attrArea.find_all("dt")
		for atr in atrs:
			atrTitle = getNodeText(atr)
			if atrTitle == "Product name":
				pInfo["name"]=getNodeText(atr.nextSibling.nextSibling)
			if atrTitle == "Catalog number":
				pInfo["Catalog"]=getNodeText(atr.nextSibling.nextSibling)
			if atrTitle == "Description":
				pInfo["Description"]=getNodeText(atr.nextSibling.nextSibling)
			if atrTitle == "Source":
				pInfo["Source"]=getNodeText(atr.nextSibling.nextSibling)
			if atrTitle == "Product category":
				pInfo["Productcategory"]=getNodeText(atr.nextSibling.nextSibling)
			if atrTitle == "Product sub category":
				pInfo["Productsubcategory"]=getNodeText(atr.nextSibling)
			if atrTitle == "Shipment info":
				pInfo["Shipmentinfo"]=getNodeText(atr.nextSibling.nextSibling)
			if atrTitle == "Nacres Codification":
				pInfo["NacresCodification"]=getNodeText(atr.nextSibling.nextSibling)
		print(pInfo)
		products.append(pInfo.copy())
		
def getProductList(url, pInfo, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	prods = sope.find_all("td", attrs={"nowrap":"nowrap"})
	for pro in prods:
		getProductInfo("https://www.tebu-bio.com"+pro.find("a")["href"], pInfo, products )


excelFileName="tebu-bio.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

# getProductInfo("https://www.tebu-bio.com/Product/089AS-61651/26Rfa_Hypothalamic_Peptide.html", {}, products)
for pageIndex in range(1,73):
	getProductList('https://www.tebu-bio.com/search/Proteins_;_Peptides/Peptides?module=search2&search_page='+str(pageIndex)+'&searchp=Array&searchp_ser=a%3A2%3A%7Bs%3A3%3A%22cat%22%3Bs%3A19%3A%22Proteins+%26+Peptides%22%3Bs%3A6%3A%22subcat%22%3Bs%3A8%3A%22Peptides%22%3B%7D&id_q=',{}, products)

headers=['link','name','Catalog','Description','Source','Productcategory','Productsubcategory','Shipmentinfo','NacresCodification']
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	if rindex%100  == 0:
		wb.save(excelFileName)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
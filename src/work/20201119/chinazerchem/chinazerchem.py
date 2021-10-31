from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import re
import copy
import string
import ssl
ssl._create_default_https_context = ssl._create_unverified_context

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	opener=urllib.request.build_opener()
	opener.addheaders=[("User-Agent","Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36")]
	urllib.request.install_opener(opener)
	urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')
	
def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"}

		request_obj=urllib.request.Request(url=url,headers=headers)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		if(retryCount <= 5):
			getHtmlFromUrl(url)
		else:
			retryCount=0
			return None
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head in info:
			workSheet.cell(rowIndex, cellIndex).value = info[head].strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1


def getProductInfo(url,type, products):
	print(str(len(products)) + url)
	productHtml = getHtmlFromUrl(url)
	if productHtml != None:
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
		
		pInfo = {"link": url,"type":type} 
		detailArea = sope.find(name="div", attrs={"class": "pro-detail"})
		name = detailArea.find("h1")
		pInfo["name"] = getNodeText(name)
		img = detailArea.find("img")
		detailStr = getNodeText(detailArea.find(name="div", attrs={"class":"fr"}))
		infos = detailStr.split("\r\n")
		casNo = ""
		for info in infos:
			if info.find("CAS No.")>-1:
				casNo = info.replace("CAS No.","")
			infoPart = info.split(":")
			if len(infoPart) == 2:
				pInfo[infoPart[0]] = infoPart[1]
		
		if img != None:
			urllib_download("https://www.chinazerchem.com"+img["src"], casNo)
		pDetailArea = sope.find("div", attrs={"class":"tagContent selectTag"})
		details = pDetailArea.find_all("p")
		for detail in details:
			detailStr = getNodeText(detail)
			if detailStr.find("Density")>-1:
				pInfo["Density"] = detailStr
			if detailStr.find("Flash point")>-1:
				pInfo["Flash point"] = detailStr
			detailPart = detailStr.split(":")
			if len(detailPart) == 2:
				pInfo[detailPart[0]] = detailPart[1]
		print(pInfo)
		products.append(pInfo.copy())
def getProductList(url,type, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	productArea = sope.find("ul", attrs={"class":"pro_lb0"})
	prods = productArea.find_all("li")
	for pro in prods:
		getProductInfo(pro.find("a")["href"],type, products )


excelFileName="chinazerchem.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
# getProductInfo('https://www.chinazerchem.com/material/liquid-crystal-intermediate/alpha-carboline-cas-244-76-8.html', products)
# getProductInfo('http://www.riyngroup.com/1-bromopyrene-1714-29-0-c16h9br-15396534502369548.html',{}, products)

links=[
	'https://www.chinazerchem.com/material/liquid-crystal-intermediate',
	'https://www.chinazerchem.com/material/liquid-crystal-intermediate/page-2/',
	'https://www.chinazerchem.com/material/liquid-crystal-intermediate/page-3/',
	'https://www.chinazerchem.com/material/liquid-crystal-intermediate/page-4/',
	'https://www.chinazerchem.com/material/liquid-crystal-intermediate/page-5/',
	'https://www.chinazerchem.com/material/liquid-crystal-intermediate/page-6/',
	'https://www.chinazerchem.com/material/liquid-crystal-intermediate/page-7/',
	'https://www.chinazerchem.com/material/liquid-crystal-intermediate/page-8/',
	'https://www.chinazerchem.com/material/liquid-crystal-intermediate/page-9/',
	'https://www.chinazerchem.com/material/liquid-crystal-intermediate/page-10/',
	'https://www.chinazerchem.com/material/liquid-crystal-intermediate/page-11/',
	'https://www.chinazerchem.com/material/liquid-crystal-intermediate/page-12/',
	'https://www.chinazerchem.com/material/liquid-crystal-intermediate/page-13/',
	'https://www.chinazerchem.com/material/oled-intermediate/',
	'https://www.chinazerchem.com/material/oled-intermediate/page-2/',
	'https://www.chinazerchem.com/material/oled-intermediate/page-3/',
	'https://www.chinazerchem.com/material/oled-intermediate/page-4/',
	'https://www.chinazerchem.com/material/oled-intermediate/page-5/',
	'https://www.chinazerchem.com/material/oled-intermediate/page-6/',
	'https://www.chinazerchem.com/material/oled-intermediate/page-7/',
	'https://www.chinazerchem.com/material/oled-intermediate/page-8/',
	'https://www.chinazerchem.com/material/oled-intermediate/page-9/',
	'https://www.chinazerchem.com/material/oled-intermediate/page-10/',
	'https://www.chinazerchem.com/material/oled-intermediate/page-11/',
	'https://www.chinazerchem.com/material/oled-intermediate/page-12/',
	'https://www.chinazerchem.com/material/oled-intermediate/page-13/',
	'https://www.chinazerchem.com/material/oled-intermediate/page-14/',
	'https://www.chinazerchem.com/material/oled-intermediate/page-15/',
	'https://www.chinazerchem.com/material/oled-intermediate/page-16/',
	'https://www.chinazerchem.com/material/oled-intermediate/page-17/',
	'https://www.chinazerchem.com/material/oled-intermediate/page-18/',
	'https://www.chinazerchem.com/material/oled-intermediate/page-19/',
	'https://www.chinazerchem.com/material/oled-intermediate/page-20/',
	'https://www.chinazerchem.com/material/oled-intermediate/page-21/',
	'https://www.chinazerchem.com/material/oled-intermediate/page-22/'
]
index = 0
for link in links:
	type = "Liquid Crystal Intermediate" if index<13 else "OLED Intermediate"
	getProductList(link, type, products)
	index = index+1
headers=['link','type','name','Molecular formula','Appearance','Purity','Product Name','Cas No','Molecular Formula','Boiling point',
	'Density','About The Introduction Of\xa0Product','●Uses'
]
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
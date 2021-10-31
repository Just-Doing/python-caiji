from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import re
import copy
import string

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\","")+'.jpg')   
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
loadCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		headers = {"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"}

		request_obj=urllib.request.Request(url=url,headers=headers)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		print("retry"+url)
		retryCount += 1
		print(retryCount)
		getHtmlFromUrl(url)
def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	
	browser.get(url)
	return browser.page_source
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, pInfo, products):
	print(str(len(products)) + url)
	productHtml = getHtmlFromUrl(url)
	if productHtml != None:
		sope = BeautifulSoup(productHtml, "html.parser",from_encoding="utf-8")
		
		pInfo["link"] = url
		name = sope.find(name="div", attrs={"class": "product-name"})
		shortdescription = sope.find(name="div", attrs={"class": "short-description"})
		cat = sope.find(name="div", attrs={"class": "sku"})
		pInfo["name"] = getNodeText(name)
		pInfo["shortdescription"] = getNodeText(shortdescription)
		pInfo["cat"] = getNodeText(cat)
		
		description = sope.find(name="div", attrs={"id": "extra_tabs_description_contents"})
		pInfo["description"] = getNodeText(description)
		
		# img = sope.find("img", attrs={"id":"image-main"})
		# if img != None:
			# if pInfo["cat"] != "":
				# urllib_download(img["src"], pInfo["cat"].replace("Catalog #: ",""))
			# else:
				# urllib_download(img["src"], pInfo["name"])
				
		attrArea = sope.find(name="div", attrs={"id": "extra_tabs_additional_contents"})
		attrTr = attrArea.find_all("tr")
		for tr in attrTr:
			title = getNodeText(tr.find("th"))
			value = getNodeText(tr.find("td"))
			pInfo[title] = value
				
		products.append(pInfo.copy())
def getProductList(url, pInfo, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	prods = sope.find_all("h2", attrs={"class":"product-name"})
	for pro in prods:
		getProductInfo(pro.find("a")["href"], pInfo, products )


def getProductType(url, products):
	productListHtml = getHtmlFromUrl(url)
	sope = BeautifulSoup(productListHtml, "html.parser",from_encoding="utf-8")
	productTypes = sope.find_all("li", attrs={"class":"apptrian-subcategories-category-wrapper"})
	for linkArea in productTypes:
		link = linkArea.find("a")
		pInfo={
			"type":getNodeText(link)
		}
		getProductList(link["href"]+"?limit=all", pInfo, products)
		

excelFileName="biovision.xlsx"
wb = Workbook()
workSheet = wb.active
products = []
# getProductListPage('http://www.riyngroup.com//boric-acid.html/p/1',{}, products)
# getProductInfo('https://www.biovision.com/blebbistatin-8433.html',{}, products)
getProductType("https://www.biovision.com/products/enzyme-inhibitors.html", products)
headers=['link','name','shortdescription','cat','description','Alternate Name','Appearance','CAS #','Molecular Formula','Molecular Weight','Purity','Solubility','SMILES',
	'InChi','InChi Key','PubChem CID','Handling','Storage Conditions','Shipping Conditions','USAGE'
]
rindex = 1
for p in products:
	writeExcel(workSheet, headers, rindex, p)
	if rindex%100  == 0:
		wb.save(excelFileName)
	rindex = rindex+1
print("flish")	

wb.save(excelFileName)
from enum import IntEnum
import requests
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import string
import re
import time
import math

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	
	opener = urllib.request.build_opener()
	opener.addheaders = [('User-agent', 'Mozilla/5.0')]
	urllib.request.install_opener(opener)
	urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		request_obj=urllib.request.Request(url=url,  headers={
			'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
			'cookie':'mage-banners-cache-storage=%7B%7D; mediav=%7B%22eid%22%3A%2274861%22%2C%22ep%22%3A%22%22%2C%22vid%22%3A%22%3ANLiZ36)Kt%3AZ%2FW%23(!KDa%22%2C%22ctn%22%3A%22%22%2C%22vvid%22%3A%22%3ANLiZ36)Kt%3AZ%2FW%23(!KDa%22%2C%22_mvnf%22%3A1%2C%22_mvctn%22%3A0%2C%22_mvck%22%3A0%2C%22_refnf%22%3A0%7D; PHPSESSID=vjadri4bmd3lj0iaaia24e70cj; Qs_lvt_31925=1636261449; Hm_lvt_8b1cb2df0051d3b34c13154f8e73ac6d=1636261450; form_key=3yhMq3L6k4v3lpRk; mage-messages=; form_key=3yhMq3L6k4v3lpRk; acw_tc=b68c82a216362673784797806e177ea045e831319fb9be90e1430963b4; mage-cache-storage=%7B%7D; mage-cache-storage-section-invalidation=%7B%7D; mage-cache-sessid=true; recently_viewed_product=%7B%7D; recently_viewed_product_previous=%7B%7D; recently_compared_product=%7B%7D; recently_compared_product_previous=%7B%7D; product_data_storage=%7B%7D; Hm_lpvt_8b1cb2df0051d3b34c13154f8e73ac6d=1636267455; Qs_pv_31925=4592131511226659300%2C1343047776903697000%2C2521513081411179000%2C2790434845695460400%2C2337970868885108700; private_content_version=9ece668a97bdefbfc2f31d1ff3539a72; section_data_ids=%7B%7D',
			"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.82 Safari/537.36"
		})
		htmlHeader = requests.head(url,allow_redirects = True)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		retryCount = retryCount + 1
		if retryCount < 5:
			print("retry index"+str(retryCount)+url)
			time.sleep(20)
			return getHtmlFromUrl(url)
		else:
			retryCount = 0
			return ""

def requestJson(url, page):
	r = requests.post(url, headers={
		'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
		'cookie':'mage-banners-cache-storage=%7B%7D; mediav=%7B%22eid%22%3A%2274861%22%2C%22ep%22%3A%22%22%2C%22vid%22%3A%22%3ANLiZ36)Kt%3AZ%2FW%23(!KDa%22%2C%22ctn%22%3A%22%22%2C%22vvid%22%3A%22%3ANLiZ36)Kt%3AZ%2FW%23(!KDa%22%2C%22_mvnf%22%3A1%2C%22_mvctn%22%3A0%2C%22_mvck%22%3A0%2C%22_refnf%22%3A0%7D; PHPSESSID=vjadri4bmd3lj0iaaia24e70cj; Qs_lvt_31925=1636261449; Hm_lvt_8b1cb2df0051d3b34c13154f8e73ac6d=1636261450; form_key=3yhMq3L6k4v3lpRk; mage-messages=; form_key=3yhMq3L6k4v3lpRk; acw_tc=b68c82a216362673784797806e177ea045e831319fb9be90e1430963b4; mage-cache-storage=%7B%7D; mage-cache-storage-section-invalidation=%7B%7D; mage-cache-sessid=true; recently_viewed_product=%7B%7D; recently_viewed_product_previous=%7B%7D; recently_compared_product=%7B%7D; recently_compared_product_previous=%7B%7D; product_data_storage=%7B%7D; Hm_lpvt_8b1cb2df0051d3b34c13154f8e73ac6d=1636267455; Qs_pv_31925=4592131511226659300%2C1343047776903697000%2C2521513081411179000%2C2790434845695460400%2C2337970868885108700; private_content_version=9ece668a97bdefbfc2f31d1ff3539a72; section_data_ids=%7B%7D',
		"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"
	},data={
		'page':page,
		'rpc':20,
		'TaxonomicTags':'Biomarker Proteins'
	})
	datas = json.loads(r.text)
	return datas

def getRenderdHtmlFromUrl(url):

	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	return BeautifulSoup(browser.page_source, "html.parser",from_encoding="utf-8")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProdInfo(sope, pInfo):
	dts = sope.find_all("dt")
	for dt in dts:
		title = getNodeText(dt)
		if title.find("Flash Point") > -1 and "Flash Point" not in pInfo:
			pInfo["Flash Point"] = getNodeText(dt.findNextSibling("dd"))
		if title.find("pKa") > -1 and "pKa" not in pInfo:
			pInfo["pKa"] = getNodeText(dt.findNextSibling("dd"))
		if title.find("Solubility") > -1 and "Solubility" not in pInfo:
			pInfo["Solubility"] = getNodeText(dt.findNextSibling("dd"))
		if title.find("Water Solubility") > -1 and "Water Solubility" not in pInfo:
			pInfo["Water Solubility"] = getNodeText(dt.findNextSibling("dd"))
		if title.find("Purification Methods") > -1 and "Purification Methods" not in pInfo:
			pInfo["Purification Methods"] = getNodeText(dt.findNextSibling("dd"))
		if title.find("Usage") > -1 and "Usage" not in pInfo:
			pInfo["Usage"] = getNodeText(dt.findNextSibling("dd"))
	

def getProdIndoFromUrl(url, keywork, pInfo):
	pInfo = {"keywork": keywork}
	print(str(len(products)) + "===" + url)
	htmlStr = getHtmlFromUrl(url)
	sope = BeautifulSoup(htmlStr, "html.parser",from_encoding="utf-8")
	link = sope.find("a", attrs={"class":"cas_lit_table"})
	if link != None:
		print("https://www.guidechem.com"+link["href"])
		htmlStr1 = getHtmlFromUrl("https://www.guidechem.com"+link["href"])
		sope1 = BeautifulSoup(htmlStr1, "html.parser",from_encoding="utf-8")
		getProdInfo(sope1, pInfo)
	else:
		getProdInfo(sope, pInfo)
	print(pInfo)
	products.append(pInfo.copy())


excelFileName="guidechem.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

headers=[
	'keywork','Flash Point','pKa','Solubility','Water Solubility','Purification Methods','Usage'
]

fileName="cat.txt"
with open(fileName,'r') as file_to_read:
	index = 1
	type=1
	while True:
		lines = file_to_read.readline()
		if not lines:
				break
		print(lines)
		getProdIndoFromUrl("https://www.guidechem.com/dictionary_keys_"+lines.replace("\r", "").replace("\n", "").strip()+".html", lines, products)
# getProductList("https://m.chemicalbook.com/Search_EN.aspx?keyword=25574-11-2&start=0", '', products)
# getProdIndoFromUrl("https://www.guidechem.com/dictionary_keys_335-67-1.html", {})

for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
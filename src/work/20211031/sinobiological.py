from enum import IntEnum
import requests
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import json
import string
import re
import time
import math

http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	
	opener = urllib.request.build_opener()
	opener.addheaders = [('User-agent', 'Mozilla/5.0')]
	urllib.request.install_opener(opener)
	urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

retryCount = 0
def getHtmlFromUrl(url, type="get", para={}):
	global retryCount
	try:
		url = urllib.parse.quote(url, safe=string.printable).replace(' ','%20')
		request_obj=urllib.request.Request(url=url,  headers={
			'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
			'cookie':'aksci=99+50+1+1+DESC; hblid=VuvO1MqlIMmhkQgy7G5Jx0HAWB6bQfao; olfsk=olfsk7317773861110564; _ga=GA1.2.345298335.1615472840; _gcl_au=1.1.1113486186.1635649360; __utma=122193563.345298335.1615472840.1617026891.1635649360.4; __utmc=122193563; __utmz=122193563.1635649360.4.1.utmcsr=(direct)|utmccn=(direct)|utmcmd=(none); PHPSESSID=a3aj2395dm47ffphvopoap20l7; aksci.com=so525hnj6bk4cr1dehhf17qog2; wcsid=e6h7QB0Uw3X0fJ8F7G5Jx0H6okj1jb4a; _gid=GA1.2.1573783117.1635649406; _okdetect=%7B%22token%22%3A%2216356495135710%22%2C%22proto%22%3A%22about%3A%22%2C%22host%22%3A%22%22%7D; _ok=3507-598-10-5787; _okbk=cd4%3Dtrue%2Cvi5%3D0%2Cvi4%3D1635649519502%2Cvi3%3Dactive%2Cvi2%3Dfalse%2Cvi1%3Dfalse%2Ccd8%3Dchat%2Ccd6%3D0%2Ccd5%3Daway%2Ccd3%3Dfalse%2Ccd2%3D0%2Ccd1%3D0%2C; _oklv=1635683673045%2Ce6h7QB0Uw3X0fJ8F7G5Jx0H6okj1jb4a; _gat=1; aksci_b=%3BJ20439+X8450+4733DP+0018AA+0043AA',
			"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/93.0.4577.82 Safari/537.36"
		})
		htmlHeader = requests.head(url,allow_redirects = True)
		response_obj=urllib.request.urlopen(request_obj)
		html_code=response_obj.read()
		return html_code
	except:
		retryCount = retryCount + 1
		if retryCount < 5:
			print("retry index"+str(retryCount)+url)
			time.sleep(60)
			return getHtmlFromUrl(url)
		else:
			retryCount = 0
			return ""

def requestJson(url, page):
	r = requests.post(url, headers={
		'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
		'cookie':'_ga=GA1.2.1846208171.1605273760; __cf_bm=iRBScpkSFqF0OWIT0GG_Go1VDezQFrob_oCXQnLabC4-1635162409-0-AZJP/3K51HNpvDA0kZl14TvvC+TFv0DObjPZhmBajMd0KNwph0OCL1nUo/jVBqBe//r3KuJpbhsZdMmLRfHLP0c=; qimo_seosource_5aff5fb0-84db-11e8-a3b3-d368cce40a8e=%E7%AB%99%E5%86%85; href=https%3A%2F%2Fwww.sinobiological.com%2Fcategory%2Fbiomarker-proteins-list; qimo_seokeywords_5aff5fb0-84db-11e8-a3b3-d368cce40a8e=; qimo_xstKeywords_5aff5fb0-84db-11e8-a3b3-d368cce40a8e=; accessId=5aff5fb0-84db-11e8-a3b3-d368cce40a8e; _gcl_au=1.1.17108153.1635162415; Hm_lvt_2d911fad88dfe99ff8bbf610824f3e67=1635162415; sbb=KGeNsFgzzNpBzobR1k4d7sprXS%252bHdCace6zs5xYWB0A%253dXTaJopSyq01ihC4cD5jOfAa8AEgsWX792EAllASK%252bEKohz0p3JxSEJFavoBnvUmw4fhBpwMcWGZ2Qy%252bRRF2U%252bMSxcQdMfdwOcT%252bR%252bo7qyEU%252br8SBQuGE8GJWgDFeSTZ44MLANMh6MuJoPjusJD60J%252bVWYvFJQEE8ibTkV5yFp%252bBO2hY5uV8gjB%252fyF0CawzwY2UiOPeoc2uumuiV8WJvsyHlZVxBVVkH8ueqd70Z0hdMOrcjEx2ONFOURwNrW4yd%252bGW8EnpM%252f1lcenHTI4w88sbFHK2fgMgUyajXvsUeKrszN9Y3wNRhnw7Qk5eU6hwPBq3SZ4AJIMb4kxxOrud9NJg%253d%253d; _gid=GA1.2.1763186893.1635162415; _fbp=fb.1.1635162417040.99186667; Hm_lpvt_2d911fad88dfe99ff8bbf610824f3e67=1635162574; Currency=RMB; LocationCode=CN; pageViewNum=3',
		"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"
	},data={
		'page':page,
		'rpc':20,
		'TaxonomicTags':'Biomarker Proteins'
	})
	datas = json.loads(r.text)
	return datas

def getRenderdHtmlFromUrl(url):

	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	return BeautifulSoup(browser.page_source, "html.parser",from_encoding="utf-8")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProdInfo(sope, pInfo):
	rightTds = sope.find_all("td", attrs={"align":"right"})
	for rTd in rightTds:
		title = getNodeText(rTd)
		if title == "Physical Form:":
			pInfo['Physical Form'] = getNodeText(rTd.nextSibling)
		if title == "Solubility:":
			pInfo['Solubility'] = getNodeText(rTd.nextSibling)

def getProdIndoFromUrl(url, pInfo):
	print(url)
	htmlStr = getHtmlFromUrl(url)
	sope = BeautifulSoup(htmlStr, "html.parser",from_encoding="utf-8")
	getProdInfo(sope, pInfo)

def getProductList(url, keywork, products):
	pInfo = {}
	htmlStr = getHtmlFromUrl(url)
	sope = BeautifulSoup(htmlStr, "html.parser")
	resTr = sope.find_all("tr", attrs={"class":"item-tr"})
	if len(resTr) == 0:
		getProdInfo(sope, pInfo)
	else:
		for tr in resTr:
			tds = tr.find_all("td")
			if len(tds) > 3:
				cas = getNodeText(tds[3])
				mfc = cas.replace(keywork, "")
				if mfc.find("MFC") ==0 or len(mfc) == 0:
					link = tds[2].find("a")
					getProdIndoFromUrl("https://aksci.com/"+link["href"], pInfo)
					break;
	products.append(pInfo.copy())

excelFileName="swri.xlsx"
wb = Workbook()
workSheet = wb.active
products = []

headers=[
	'Physical Form','Solubility'
]

fileName="cat.txt"
with open(fileName,'r') as file_to_read:
	index = 1
	type=1
	while True:
		lines = file_to_read.readline()
		if not lines:
    			break
		print(lines)
		getProductList("https://aksci.com/item_list.php?search="+lines, lines, products)


for index,head in enumerate(headers):
    workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
    writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFileName)
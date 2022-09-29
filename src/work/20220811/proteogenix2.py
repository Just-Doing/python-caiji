from ast import And, Store
from enum import IntEnum
from attr import attrs
import requests
from urllib.request import urlopen
import urllib
from selenium import webdriver
from bs4 import BeautifulSoup
import http.client
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import cfscrape
import json
import string
import re
import time
import math
import _thread
import os
import ssl

import numpy as np

ssl._create_default_https_context = ssl._create_unverified_context
http.client._MAXHEADERS = 1000

def urllib_download(IMAGE_URL, pName):
	try:
		opener = urllib.request.build_opener()
		opener.addheaders = [('User-agent', 'Mozilla/5.0'),('cookie','__cf_bm=vbLxkjIVjoGJnWBrrRZP1GCsbs_KmubMaHNugmyyR3I-1659249299-0-AXJuOuT6vZtFzKG57pHX3sgKssRwKg9sXDpFaqkl5hXjPCCmxv95Lj76noiZ90Rm6c0kXSv/oW51uHGWcgUQlTE=')]
		urllib.request.install_opener(opener)
		urllib.request.urlretrieve(IMAGE_URL, pName.replace("/","").replace("\\",""))
	except:
		print('no')

def getNodeText(node):
	if(node == None):
		return ""
	else:
		return node.get_text().strip()

def getHtmlFromUrl(url, type="get", para={}):
	headers = {
		"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.82 Safari/537.36",
		"cookie":"_ga=GA1.2.1846208171.1605273760; href=https%3A%2F%2Fwww.sinobiological.com%2Fresearch%2Ftargeted-therapy; accessId=5aff5fb0-84db-11e8-a3b3-d368cce40a8e; _gcl_au=1.1.1660157260.1645016298; Hm_lvt_2d911fad88dfe99ff8bbf610824f3e67=1645016298; sbb=%252be43ohTbVTr09K%252bxQlr1%252bK0onQvF%252bMIXgZM%252bveGXMHU%253dXTaJopSyq01ihC4cD5jOfAa8AEgsWX792EAllASK%252bEKohz0p3JxSEJFavoBnvUmw4fhBpwMcWGZ2Qy%252bRRF2U%252bMSxcQdMfdwOcT%252bR%252bo7qyEU%252br8SBQuGE8GJWgDFeSTZ4QS0HvJFVazETAoyuKMwGHYRoD68%252f7qno5Bg%252bEH9sSXM4upMLtz%252f4IdNkjX6GD0JYHbiUh%252blGTwi25Iz3IKocTDD58DE1yYiY3DxeifN7Qz6OxtXX21lrBpnvgDu9ANN%252f7TTxWWMmOIjxVG772o%252bYGkE9AMxcU5O4cIrT9cubm6dAdgw6n%252fQRZpTVxNv2TGHdHZblPNcfu4dTWVsL3aqaag%253d%253d; _gid=GA1.2.832211649.1645016298; _ce.s=v11.rlc~1645016301520; pageViewNum=13; Hm_lpvt_2d911fad88dfe99ff8bbf610824f3e67=1645017042; Currency=RMB; LocationCode=CN"
	}

	scraper = cfscrape.create_scraper()
	html_code = scraper.get(url,headers=headers).text
	return BeautifulSoup(html_code, "html.parser",from_encoding="utf-8")


def postHtmlFromUrl(url, type="get", para={}):
	r = requests.post(url, headers={
		'Content-Type': 'application/x-www-form-urlencoded',
		'cookie':'visid_incap_2255650=4oBBaRPnQfCVoYEiTmjTq/NVAWEAAAAAQUIPAAAAAAD69PQHUoB0KplKq7/j0+gH; nlbi_2255650=CJKhHYlMm17tpKyoBzOViAAAAACDEjp3gL6bj6YL8j9XE0d/; incap_ses_893_2255650=m1tJIuDRUEp3FE/5GpNkDPRVAWEAAAAAM2KkDpvtARtZral+cMXSVw==; _gcl_au=1.1.76703404.1627477493; _gid=GA1.2.730047202.1627477493; BCSessionID=83af10b8-9488-4b7b-a3b1-3640f178dca2; categoryView=grid; _ga_S46FST9X1M=GS1.1.1627477492.1.1.1627478562.0; _ga=GA1.2.31731397.1627477493; _gat_UA-139934-1=1; _uetsid=69fc2d30efa411eb8818eb045f8760e5; _uetvid=69fc3a70efa411ebba3a23c153f6e477; .Nop.Customer=d664d529-d14a-44b1-86b3-cbf5373277b4',
		"User-Agent": "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.86 Safari/537.36"
	})
	print(r.text)
	return BeautifulSoup(r.text, "html.parser",from_encoding="utf-8")

def requestJson(url):
	print(url)
	payload={'wpgb': '{"is_main_query":true,"main_query":{"product_cat":"proteins","error":"","m":"","p":0,"post_parent":"","subpost":"","subpost_id":"","attachment":"","attachment_id":0,"name":"","pagename":"","page_id":0,"second":"","minute":"","hour":"","day":0,"monthnum":0,"year":0,"w":0,"category_name":"","tag":"","cat":"","tag_id":"","author":"","author_name":"","feed":"","tb":"","paged":0,"meta_key":"","meta_value":"","preview":"","s":"","sentence":"","title":"","fields":"","menu_order":"","embed":"","category__in":[],"category__not_in":[],"category__and":[],"post__in":[],"post__not_in":[],"post_name__in":[],"tag__in":[],"tag__not_in":[],"tag__and":[],"tag_slug__in":[],"tag_slug__and":[],"post_parent__in":[],"post_parent__not_in":[],"author__in":[],"author__not_in":[],"orderby":"menu_order title","order":"ASC","meta_query":[],"tax_query":{"0":{"taxonomy":"product_visibility","field":"term_taxonomy_id","terms":[7],"operator":"NOT IN"},"relation":"AND"},"wc_query":"product_query","posts_per_page":16},"permalink":"https://www.proteogenix.science/product-category/proteins/","facets":[31,32,34,35,36,39,40,41,42,43,44,45],"lang":"en","id":3}'}
	files=[ ]
	headers = {
		'Cookie': '_gcl_au=1.1.656931413.1660218437; _gid=GA1.2.1932146283.1660218438; wp-wpml_current_language=en; _hjFirstSeen=1; _hjSession_1888702=eyJpZCI6IjliMmFiYmVlLTYxM2MtNGIyNC05NzlkLWFkZGM4NDE4M2MwNCIsImNyZWF0ZWQiOjE2NjAyMTg0NjQ1NTIsImluU2FtcGxlIjpmYWxzZX0=; _hjAbsoluteSessionInProgress=0; BTWSEC=8692f4b908b336f304a2f4fc1c97655a; __stripe_mid=54f7b82c-2f2e-482b-9a85-dd12034fe4f1a5cdf8; cluid=348AFCBE-8D5F-BB11-99FF-45BF36777868; _hjSessionUser_1888702=eyJpZCI6ImNiZTkyNGVlLTM5NzUtNWJjNi1hOGNlLWQyMmQzZjAwNTg1OCIsImNyZWF0ZWQiOjE2NjAyMTg0NDM5OTcsImV4aXN0aW5nIjp0cnVlfQ==; _hjIncludedInPageviewSample=1; _hjIncludedInSessionSample=0; lepopup-onload-get-your-free-hanbook=ilovefamily; tarteaucitron=!gtag=true; nitroCachedPage=0; _ga=GA1.2.850015843.1660218437; _ga_PGVXKWTM9X=GS1.1.1660218437.1.1.1660222371.0'
	}
	response = requests.request("POST", url, headers=headers, data=payload, files=files)
	datas = json.loads(response.text)
	return datas
	# return BeautifulSoup(r.text, "html.parser",from_encoding="utf-8")


def getRenderdHtmlFromUrl(url):
	chrome_options = webdriver.ChromeOptions()
	chrome_options.add_argument('--headless')
	chrome_options.add_argument('--disable-gpu')
	chrome_options.add_argument("window-size=1024,768")

	chrome_options.add_argument("--no-sandbox")
	browser = webdriver.Chrome(chrome_options=chrome_options)
	browser.get(url)
	return BeautifulSoup(browser.page_source, "html.parser")
	
	
	
def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		try:
			if head in info:
				content = ILLEGAL_CHARACTERS_RE.sub(r'', info[head])
				workSheet.cell(rowIndex, cellIndex).value = content.strip()
			else:
				workSheet.cell(rowIndex, cellIndex).value = ""
			cellIndex=cellIndex+1
		except:
			print(rowIndex)

def getProductInfo(url, products):
	print(str(len(products))+":"+url)
	sope= getHtmlFromUrl(url)
	nav = sope.find("div", attrs={"class":"row breadcrumbs-wrapper"})
	sizeDataEl = sope.find("form", attrs={"class":"variations_form cart"})
	productNameArea = sope.find("h1", attrs={"class":"product_title entry-title"})
	if productNameArea!=None:
		pInfo = {
			"link":url,
			"productName": getNodeText(productNameArea),
			"nav": getNodeText(nav)
		}
		if sizeDataEl!= None:
			sizeData = sizeDataEl["data-product_variations"]
			sizes = json.loads(sizeData)
			for index,size in enumerate(sizes):
				pInfo["size"+str(index)] = size["attributes"]["attribute_size"]
				pInfo["price"+str(index)] = str(size["display_price"])
				pInfo["Reference"+str(index)] = str(size["sku"])
		else:
			pInfo["size0"] = getNodeText(sope.find("div", attrs={"class":"vct-variable-container"}))
			pInfo["price0"] = getNodeText(sope.find("span", attrs={"class":"woocommerce-Price-amount amount"}))
			sku = sope.find("span", attrs={"class":"sku_wrapper"})
			pInfo["Reference0"] = getNodeText(sku).replace("Reference:","").replace("\r\n","").strip()
		specs = sope.find("table", attrs={"class":"woocommerce-product-attributes shop_attributes"})
		trs = sope.find_all("tr")
		for tr in trs:
			ths = tr.find_all("th")
			tds = tr.find_all("td")
			if len(ths) ==1 and len(tds) ==1:
				title = getNodeText(ths[0])
				value = getNodeText(tds[0]).replace("\n","\r\n")
				pInfo[title] = value
				if title =="Publications":
					links = tds[0].find_all("a")
					linkStr = ""
					for link in links:
						linkStr += link["href"]+";"
					pInfo["PublicationsLink"] = linkStr
		desc = sope.find("div", attrs={"class":"long-description-content-wrapper"})
		pInfo["Description"] = getNodeText(desc)
		datasheet = sope.find("div", attrs={"class":"data-sheet"})
		if datasheet!= None:
			datasheetLink = datasheet.find("a")
			fileName = os.path.basename(datasheetLink["href"])
			pInfo["datasheet"] = fileName
			urllib_download(datasheetLink["href"], fileName)
		eddatasheet = sope.find("div", attrs={"class":"product-extra-document"})
		if eddatasheet!= None:
			eddatasheetLink = eddatasheet.find("a")
			fileName = os.path.basename(eddatasheetLink["href"])
			pInfo["extradocument"] = fileName
			urllib_download(eddatasheetLink["href"], fileName)
		# imgs = sope.find_all("div", attrs={"class":"woocommerce-product-gallery__image"})
		# for inx,imgArea in enumerate(imgs):
		# 	img = imgArea.find("img")
		# 	imgLink = img["data-src"]
		# 	fileName = os.path.basename(imgLink)
		# 	pInfo["img"+str(inx)] = fileName
		# 	urllib_download(imgLink, fileName)

		products.append(pInfo.copy());

def getProductList(url, products):
	data = requestJson(url)
	if "posts" in data:
		sope = BeautifulSoup(data["posts"], "html.parser")
		pros = sope.find_all("article", attrs={"class":"wpgb-card-3"})
		print("product count ----- "+str(len(pros)))
		for tr in pros:
			proLink = tr.find("a", attrs={"class":"wpgb-card-layer-link"})
			if proLink != None:
				getProductInfo(proLink["href"], products)
		
headers=[
	'link','productName','nav','Size','size0','price0','Reference0','size1','price1','Reference1','size2','price2','Reference2','size3','price3','Reference3'
	,'Description','Brand','Product type','Host Species','Product name','Expression system','Molecular weight'
	,'Purity estimated','Buffer','Form','Delivery condition','Storage condition','Fragment Type','Aliases /Synonyms'
	,'Reference','Related Products','Publications','PublicationsLink','Note','datasheet','extradocument'
]

excelFname="proteogenix2.xlsx"
products = []
# getProductInfo('https://www.proteogenix.science/product/5-amp-activated-protein-kinase-catalytic-subunit-alpha-2prkaa2/', products)

for pIndex in range(1, 17):
	getProductList("https://www.proteogenix.science/?wpgb-ajax=refresh&_per_page=100&_page="+str(pIndex), products)

wb = Workbook()
workSheet = wb.active


for index,head in enumerate(headers):
	workSheet.cell(1, index+1).value = head.strip()
for index,p in enumerate(products):
	writeExcel(workSheet, headers, index + 2, p)
print("flish")	

wb.save(excelFname)
	
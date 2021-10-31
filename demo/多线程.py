from urllib.request import urlopen
from bs4 import BeautifulSoup
import http.client
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter
import json
import _thread
from urllib import parse

http.client._MAXHEADERS = 1000

def getHtmlFromUrl(url):
	try:
		html = urlopen(url).read()
		return html
	except Exception:
		print(Exception)
		print("重试"+url)
		getHtmlFromUrl(url)

def writeExcel(workSheet, headers, rowIndex, info):
	cellIndex=1
	for head in headers:
		if head.strip() in info:
			workSheet.cell(rowIndex, cellIndex).value = str(info[head.strip()]).strip()
		else:
			workSheet.cell(rowIndex, cellIndex).value = ""
		cellIndex=cellIndex+1



def theardFun(startPage, endPage,excelFname):
	excelFileName="c:\\"+excelFname+".xlsx"
	wb = Workbook()
	workSheet = wb.active
	wb.save(excelFileName)
	
try:
	_thread.start_new_thread( theardFun, (1, 45, "product_novu1" ) )
	_thread.start_new_thread( theardFun, (45, 97, "product_novu2" ) )
except:
	print ("Error: 无法启动线程")
while 1:
   pass